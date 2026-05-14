import json
import os
from pathlib import Path
import base64
import re
import zipfile

import altair as alt
import duckdb
import pandas as pd
import math
import requests
import streamlit as st
import boto3

# LOCKED STARTUP v4 - cloud startup uses speed tables only; no index shard download

from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# App environment setup
# Recommended: set APP_ENV in Streamlit Secrets for each app:
# DEV app:  APP_ENV = "DEV"
# LIVE app: APP_ENV = "LIVE"
# If no secret/environment variable is set, this file defaults to DEV for safety.
APP_ENV = os.environ.get("APP_ENV", "DEV").strip().upper()
try:
    APP_ENV = str(st.secrets.get("APP_ENV", APP_ENV)).strip().upper()
except Exception:
    pass
if APP_ENV not in {"DEV", "LIVE"}:
    APP_ENV = "DEV"

st.set_page_config(
    page_title="Candidate Connect DEV" if APP_ENV == "DEV" else "Candidate Connect",
    layout="wide"
)


st.markdown('\n<style id="cc-final-hard-fixes">\n/* Hard fixes for real logos, KPI icons, and no \'Ready\' overlay */\n.cc-donut::after,\n.cc-donut:after {\n  content: none !important;\n  display: none !important;\n}\n.logo-tss,\nimg[src*="TSS_Logo_Transparent"] {\n  max-width: 330px !important;\n  max-height: 112px !important;\n  width: auto !important;\n  height: auto !important;\n  object-fit: contain !important;\n  object-position: center center !important;\n  display: block !important;\n  filter: brightness(1.15) saturate(1.18) contrast(1.08) drop-shadow(0 10px 18px rgba(0,0,0,.58)) !important;\n}\n.brand-right {\n  overflow: visible !important;\n}\n.cc-open-metric .icon img {\n  width: 38px !important;\n  height: 38px !important;\n  object-fit: contain !important;\n  display: block !important;\n}\n</style>\n', unsafe_allow_html=True)


st.markdown('\n<style id="cc-no-unknown-and-real-tss">\n.cc-open-metrics {\n  grid-template-columns: repeat(4, minmax(0, 1fr)) !important;\n}\n.logo-tss,\nimg[src*="TSS_Logo_Transparent"] {\n  max-width: 330px !important;\n  max-height: 120px !important;\n  width: auto !important;\n  height: auto !important;\n  object-fit: contain !important;\n  display: block !important;\n  filter: drop-shadow(0 10px 18px rgba(0,0,0,.55)) !important;\n}\n</style>\n', unsafe_allow_html=True)


st.markdown('\n<style id="cc-true-black-background-fix">\n/* True black foundation for consistent display across monitors */\n:root {\n  --cc-bg: #000000 !important;\n  --cc-bg-2: #000000 !important;\n  --cc-sidebar: #000000 !important;\n  --cc-panel: #05080d !important;\n  --cc-panel-2: #07101a !important;\n  --cc-card: #07101a !important;\n  --cc-card-soft: #0a111c !important;\n}\n\nhtml,\nbody,\n[data-testid="stAppViewContainer"],\n[data-testid="stApp"],\n.stApp {\n  background: #000000 !important;\n}\n\n[data-testid="stSidebar"],\n[data-testid="stSidebarContent"] {\n  background: #000000 !important;\n}\n\n.main,\n.block-container,\nsection.main,\n[data-testid="stMain"],\n[data-testid="stMainBlockContainer"] {\n  background: #000000 !important;\n}\n\n/* Remove gray monitor-dependent page glow while keeping subtle red header depth */\n[data-testid="stAppViewContainer"] > .main {\n  background: #000000 !important;\n}\n\n/* Keep cards dark, not gray */\n.top-shell,\n.section-card,\n.chart-card,\n.table-card,\n.metric-card,\n.cc-opening-shell,\n.cc-open-card,\n.cc-open-metric,\n.cc-open-table,\n.cc-filter-shell,\n.stTabs [data-baseweb="tab-panel"] {\n  background: linear-gradient(180deg, #07101a 0%, #03070c 100%) !important;\n}\n\n/* Header stays dramatic but starts from black */\n.top-shell,\n.cc-opening-header,\n.brand-shell {\n  background:\n    radial-gradient(circle at 72% 8%, rgba(185,28,28,.22), transparent 34%),\n    linear-gradient(90deg, #000000 0%, #05070c 43%, #180405 100%) !important;\n}\n\n/* Sidebar controls on black */\n[data-testid="stSidebar"] .stButton > button,\n[data-testid="stSidebar"] [data-baseweb="select"] > div,\n[data-testid="stSidebar"] input,\n[data-testid="stSidebar"] textarea {\n  background-color: #07101a !important;\n}\n\n/* Make Streamlit white gaps less likely */\ndiv[data-testid="stVerticalBlock"],\ndiv[data-testid="stHorizontalBlock"] {\n  background: transparent !important;\n}\n</style>\n', unsafe_allow_html=True)


st.markdown('\n<style id="cc-single-panel-nav-css">\n[data-testid="stSidebar"] .cc-nav-menu-title {\n  font-size: 11px;\n  font-weight: 900;\n  letter-spacing: .08em;\n  color: #ff3b3b;\n  text-transform: uppercase;\n  margin: 16px 0 8px 2px;\n}\n[data-testid="stSidebar"] .cc-active-section-title {\n  font-size: 15px;\n  font-weight: 900;\n  color: #ffffff;\n  margin: 16px 0 10px 0;\n}\n[data-testid="stSidebar"] .stButton > button {\n  min-height: 38px !important;\n}\n[data-testid="stSidebar"] .stButton > button[kind="secondary"] {\n  justify-content: flex-start !important;\n  text-align: left !important;\n  border: 1px solid rgba(148,163,184,.20) !important;\n  background: #05080d !important;\n  color: #e5e7eb !important;\n}\n[data-testid="stSidebar"] .stButton > button[kind="secondary"]:hover {\n  border-color: rgba(248,113,113,.85) !important;\n  background: linear-gradient(90deg, rgba(185,28,28,.32), rgba(7,16,26,.95)) !important;\n  color: #ffffff !important;\n}\n</style>\n', unsafe_allow_html=True)

# R2 public-read setup
R2_BASE_BY_ENV = {
    "DEV": "https://pub-376c4497d59b4a7988a8af29700531e0.r2.dev",
    "LIVE": "https://pub-a9e33b718082407cbd85e7b86b0fcb5c.r2.dev",
}
R2_BUCKET_BY_ENV = {
    "DEV": "candidate-connect-data-dev",
    "LIVE": "candidate-connect-data",
}

R2_BASE = R2_BASE_BY_ENV[APP_ENV]
R2_BUCKET = R2_BUCKET_BY_ENV[APP_ENV]

# Optional overrides, useful if a public R2 URL changes later.
try:
    R2_BASE = str(st.secrets.get("R2_BASE", R2_BASE)).strip() or R2_BASE
    R2_BUCKET = str(st.secrets.get("R2_BUCKET", R2_BUCKET)).strip() or R2_BUCKET
except Exception:
    R2_BASE = os.environ.get("R2_BASE", R2_BASE)
    R2_BUCKET = os.environ.get("R2_BUCKET", R2_BUCKET)

LOCAL_ROOT = Path("/tmp/candidate_connect_r2")
LOCAL_MANIFEST = LOCAL_ROOT / "dataset_manifest.json"

# DEV data source setup
# DEV first tries local Athenix/Candidate Connect files, then falls back to R2.
def _truthy(value) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}

USE_LOCAL_DATA = APP_ENV == "DEV"
try:
    USE_LOCAL_DATA = _truthy(st.secrets.get("USE_LOCAL_DATA", USE_LOCAL_DATA))
except Exception:
    USE_LOCAL_DATA = _truthy(os.environ.get("USE_LOCAL_DATA", USE_LOCAL_DATA))

# Streamlit Cloud must never try to use local DEV shard folders.
# It should start from the small R2 speed tables and only touch index/detail shards
# after a user action that truly needs them.
IS_STREAMLIT_CLOUD = Path("/mount/src").exists() or bool(os.environ.get("STREAMLIT_SHARING") or os.environ.get("STREAMLIT_SERVER_PORT"))
if IS_STREAMLIT_CLOUD:
    USE_LOCAL_DATA = False

ATHENIX_OUTPUT_PATH = Path.home() / "Desktop" / "Athenix_Data_Pipeline" / "07_outputs"
LOCAL_DEV_SHARDS_DIR = Path("data/shards")
LOCAL_DEV_INDEX_SHARDS_DIR = LOCAL_DEV_SHARDS_DIR / "index"
LOCAL_DEV_DETAIL_SHARDS_DIR = LOCAL_DEV_SHARDS_DIR / "detail"
LOCAL_DEV_SPEED_DIR = LOCAL_DEV_SHARDS_DIR / "speed"
LOCAL_DEV_VOTERS_CSV = Path("data/local_data/voters.csv")
LOCAL_CANDIDATE_CONNECT_CSV = Path("data/raw/voter_file/candidate_connect_data.csv")
LOCAL_ATHENIX_FEATURED_CSV = ATHENIX_OUTPUT_PATH / "master" / "PA_VOTER_MASTER_FEATURED.csv"

CC_LOGO = Path("candidate_connect_logo.png")
TSS_LOGO = Path("TSS_Logo_Transparent.png")
TSS_REPORTING_LOGO = Path("TSS_Logo_Transparent_REPORTING.png")
SAVED_UNIVERSES_PATH = Path("saved_universes.json")
SAVED_UNIVERSES_R2_KEY = "app_state/saved_universes.json"
# Voter corrections are app-state/data, not code. Store them with the
# Athenix pipeline outputs so they can be re-applied after every SURE/CURRENT/contact refresh.
VOTER_CORRECTIONS_PATH = ATHENIX_OUTPUT_PATH / "app_state" / "voter_record_corrections.json"
VOTER_CORRECTIONS_LOCAL_FALLBACK_PATH = Path("voter_record_corrections.json")



# -----------------------------------------------------------------------------
# Candidate Connect centralized dark-theme engine
# Keep all major UI/chart colors here so future sections inherit one system.
# -----------------------------------------------------------------------------
CC_THEME = {
    "bg_primary": "#070A10",
    "bg_secondary": "#0B111A",
    "panel": "#0E1621",
    "panel_alt": "#131C28",
    "panel_soft": "#182331",
    "border": "#3A2630",
    "border_soft": "#2A3340",
    "text_primary": "#F8FAFC",
    "text_secondary": "#D7DEE8",
    "text_muted": "#98A3B3",
    "brand_red": "#D71920",
    "brand_red_dark": "#8F1118",
    "brand_blue": "#1D4ED8",
    "brand_blue_hover": "#2563EB",
    "brand_gold": "#F2B84B",
    "rep_red": "#C91F27",
    "dem_blue": "#1D4ED8",
    "other_green": "#4C9A2A",
}

# Required global political colors. Do not use lighter ad-hoc variants elsewhere.
PARTY_COLOR_MAP = {"R": CC_THEME["rep_red"], "D": CC_THEME["dem_blue"], "O": CC_THEME["other_green"]}
PARTY_NAME_COLOR_MAP = {"Republican": CC_THEME["rep_red"], "Democrat": CC_THEME["dem_blue"], "Democratic": CC_THEME["dem_blue"], "Other": CC_THEME["other_green"]}

# Muted dark-dashboard chart palettes for non-party charts.
AGE_COLOR_RANGE = ["#6E0F14", "#8F1118", "#A9161D", "#C91F27", "#E03A3E", "#B87333", "#F2B84B", "#8F1118", "#C91F27"]
GENDER_COLOR_RANGE = [CC_THEME["rep_red"], CC_THEME["dem_blue"], CC_THEME["other_green"], CC_THEME["brand_gold"], "#64748B"]
CHART_NEUTRAL_COLOR_RANGE = [CC_THEME["rep_red"], CC_THEME["dem_blue"], CC_THEME["other_green"], CC_THEME["brand_gold"], "#A78BFA", "#22D3EE", "#FB7185"]



# Global Altair dark theme so charts do not render on white panels.
def _cc_altair_dark_theme():
    return {
        "config": {
            "background": "transparent",
            "view": {"stroke": "transparent"},
            "axis": {
                "labelColor": "#E5E7EB",
                "titleColor": "#F8FAFC",
                "gridColor": "rgba(148, 163, 184, 0.14)",
                "domainColor": "rgba(148, 163, 184, 0.38)",
                "tickColor": "rgba(148, 163, 184, 0.30)",
                "labelFontSize": 12,
                "titleFontSize": 12,
            },
            "legend": {
                "labelColor": "#E5E7EB",
                "titleColor": "#F8FAFC",
                "labelFontSize": 12,
                "titleFontSize": 12,
                "symbolStrokeColor": "rgba(255,255,255,.35)",
            },
            "title": {"color": "#F8FAFC"},
            "range": {
                "category": [CC_THEME["rep_red"], CC_THEME["dem_blue"], CC_THEME["other_green"], CC_THEME["brand_gold"], "#64748B"]
            },
        }
    }

try:
    alt.themes.register("candidate_connect_dark", _cc_altair_dark_theme)
    alt.themes.enable("candidate_connect_dark")
except Exception:
    pass

GEO_DISPLAY_LABELS = {
    "USC": "Congressional",
    "STS": "State Senate",
    "STH": "State House",
}


def geo_label(col: str) -> str:
    return GEO_DISPLAY_LABELS.get(col, col)

st.markdown("""
<style>
/* --------------------------------------------------------------------------
   Candidate Connect Professional Dark Theme
   Centralized CSS variables are the source of truth for the entire UI.
   -------------------------------------------------------------------------- */
:root {
  --cc-bg-primary: #070A10;
  --cc-bg-secondary: #0B111A;
  --cc-bg-tertiary: #0E1621;
  --cc-panel: #0E1621;
  --cc-panel-alt: #131C28;
  --cc-panel-soft: #182331;
  --cc-panel-hover: #241B22;
  --cc-border: #3A2630;
  --cc-border-soft: #2A3340;
  --cc-text-primary: #F8FAFC;
  --cc-text-secondary: #CBD5E1;
  --cc-text-muted: #94A3B8;
  --cc-text-faint: #64748B;
  --cc-red: #C91F27;
  --cc-blue: #1D4ED8;
  --cc-green: #4C9A2A;
  --cc-gold: #F2B84B;
  --cc-cyan: #F2B84B;
  --cc-shadow: 0 14px 34px rgba(0, 0, 0, .35);
  --cc-shadow-soft: 0 8px 18px rgba(0, 0, 0, .22);
  --cc-radius-lg: 16px;
  --cc-radius-md: 11px;
  --cc-radius-sm: 8px;
  --cc-input-bg: #0F172A;
  --cc-input-border: #334155;
  --cc-input-focus: #F2B84B;
  --cc-button-bg: linear-gradient(180deg, #9F151C 0%, #6E0F14 100%);
  --cc-button-hover: linear-gradient(180deg, #C91F27 0%, #8F1118 100%);
  --cc-button-secondary: #1E293B;
  --cc-button-danger: linear-gradient(180deg, #B91C1C 0%, #991B1B 100%);
}

html, body, [data-testid="stAppViewContainer"], .stApp {
  background: radial-gradient(circle at top left, rgba(201, 31, 39, .16), transparent 32%),
              radial-gradient(circle at top right, rgba(242, 184, 75, .07), transparent 26%),
              linear-gradient(180deg, var(--cc-bg-secondary) 0%, var(--cc-bg-primary) 100%) !important;
  color: var(--cc-text-primary) !important;
}

.block-container {padding-top: 1.15rem; padding-bottom: 1rem; max-width: 1660px;}
[data-testid="stHeader"] {background: rgba(7,10,16,.82) !important; backdrop-filter: blur(10px);}
[data-testid="stToolbar"] {right: 1rem;}

/* Sidebar */
[data-testid="stSidebar"] {background: linear-gradient(180deg, #0A0D12 0%, #07101A 100%) !important; border-right: 1px solid rgba(201,31,39,.28);}
[data-testid="stSidebar"] * {color: var(--cc-text-secondary) !important;}
[data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3,
[data-testid="stSidebar"] .stMarkdown strong {color: var(--cc-text-primary) !important;}
.sidebar-note {font-size:10.5px; color:var(--cc-text-muted) !important; margin-top:-.25rem; margin-bottom:.55rem; line-height:1.35;}

/* Main cards/shells */
.top-shell, .section-card, .chart-card, .table-card, .metric-card, .empty-shell, .lookup-result-card {
  border: 1px solid var(--cc-border-soft);
  border-radius: var(--cc-radius-lg);
  background: linear-gradient(180deg, rgba(18,27,38,.97) 0%, rgba(10,16,24,.97) 100%);
  box-shadow: var(--cc-shadow-soft);
  color: var(--cc-text-primary);
}
.top-shell {padding: 1.05rem 1.25rem 1.05rem 1.25rem; margin-top: .35rem; margin-bottom: .95rem; overflow: visible; border-color: rgba(201,31,39,.82); background: radial-gradient(circle at 78% 20%, rgba(201,31,39,.20), transparent 32%), radial-gradient(circle at 92% 60%, rgba(242,184,75,.09), transparent 22%), linear-gradient(90deg, rgba(12,18,26,.98) 0%, rgba(9,12,18,.98) 48%, rgba(19,8,10,.98) 100%); box-shadow: 0 0 0 1px rgba(242,184,75,.10), 0 18px 40px rgba(0,0,0,.42);}
.section-card, .chart-card, .table-card {padding: .9rem .95rem; margin-bottom: .85rem;}
.metric-card {padding: .72rem .82rem; min-height: 94px; display:flex; flex-direction:column; justify-content:center; border-left: 3px solid rgba(201,31,39,.78);}
.metric-label {font-size: 11px; color: var(--cc-text-muted); margin-bottom: .12rem; letter-spacing:.02em; text-transform: uppercase;}
.metric-value {font-size: 1.58rem; font-weight: 800; color: var(--cc-text-primary); line-height: 1.1;}
.small-header {font-size: 16px; font-weight: 900; color: var(--cc-text-primary); margin-bottom: .48rem; letter-spacing:.01em;}
.tiny-muted, .stCaption, div[data-testid="stCaptionContainer"] {font-size: 10.5px; color: var(--cc-text-muted) !important;}
.empty-shell {padding: 1.2rem 1rem; text-align:center; color:var(--cc-text-secondary);}

/* Brand header */
.brand-grid {display:grid; grid-template-columns: 220px 1fr 250px; gap:22px; align-items:center;}
.brand-left {display:flex; align-items:center; justify-content:flex-start; min-height:86px; padding-right:18px; border-right:1px solid rgba(242,184,75,.28);}
.brand-center {display:flex; flex-direction:column; justify-content:center;}
.brand-right {display:flex; flex-direction:column; align-items:center; justify-content:center; min-height:86px; padding:.55rem .75rem; border-radius:14px; background:rgba(0,0,0,.24); box-shadow: inset 0 0 0 1px rgba(242,184,75,.10);}
.brand-title {font-size: 25px; font-weight: 900; color: var(--cc-text-primary); line-height:1.05; margin-bottom:.12rem; letter-spacing:.01em;}
.brand-sub {font-size: 11px; color: var(--cc-red); font-weight:900; letter-spacing:.04em; text-transform: uppercase;}
.brand-status {font-size: 11px; color: var(--cc-text-muted); margin-top:.30rem; font-weight:650;}
.powered-by {font-size:10px; color:#FFFFFF; margin-bottom:.18rem; text-align:center; font-weight:900; letter-spacing:.02em;}
.logo-cc {max-width:205px; height:auto; display:block; filter: brightness(1.18) saturate(1.22) contrast(1.10) drop-shadow(0 10px 18px rgba(0,0,0,.48));}
.logo-tss {max-width:205px; height:auto; display:block; margin:0 auto; opacity:1; filter: brightness(1.22) saturate(1.35) contrast(1.16) drop-shadow(0 10px 18px rgba(0,0,0,.48));}
.section-divider {height:1px; background:linear-gradient(to right, rgba(96,165,250,0), rgba(201,31,39,.55) 12%, rgba(242,184,75,.58) 50%, rgba(201,31,39,.55) 88%, rgba(96,165,250,0)); margin:.55rem 0 .9rem 0;}

/* Text and links */
p, li, label, span, div {color: inherit;}
a {color:#F2B84B !important;}
hr {border-color: var(--cc-border-soft) !important;}

/* Buttons */
.stButton > button, div[data-testid="stDownloadButton"] > button, button[kind="primary"] {
  width:100%; border-radius: var(--cc-radius-md) !important; min-height: 2.22rem; font-weight: 800 !important;
  color: var(--cc-text-primary) !important; background: var(--cc-button-bg) !important;
  border: 1px solid rgba(242,184,75,.38) !important; box-shadow: 0 6px 16px rgba(201,31,39,.22);
  transition: all .12s ease-in-out;
}
.stButton > button:hover, div[data-testid="stDownloadButton"] > button:hover, button[kind="primary"]:hover {
  background: var(--cc-button-hover) !important; border-color: #F2B84B !important; transform: translateY(-1px);
}
.stButton > button:active, div[data-testid="stDownloadButton"] > button:active {transform: translateY(0); box-shadow:none;}
button[kind="secondary"] {background: var(--cc-button-secondary) !important; color: var(--cc-text-secondary) !important; border:1px solid var(--cc-border) !important;}

/* Inputs, filters, multiselects, sliders */
[data-baseweb="select"] > div, [data-baseweb="input"] > div, textarea, input,
.stTextInput input, .stNumberInput input, .stDateInput input {
  background: var(--cc-input-bg) !important; color: var(--cc-text-primary) !important; border-color: var(--cc-input-border) !important; border-radius: var(--cc-radius-sm) !important;
}
[data-baseweb="select"] span, [data-baseweb="select"] div {color: var(--cc-text-primary) !important;}
[data-baseweb="tag"] {background: rgba(201,31,39,.28) !important; color: var(--cc-text-primary) !important; border:1px solid rgba(242,184,75,.28);}
[data-baseweb="popover"], [data-baseweb="menu"] {background: var(--cc-panel-soft) !important; color: var(--cc-text-primary) !important; border:1px solid var(--cc-border) !important;}
[role="option"] {background: var(--cc-panel-soft) !important; color: var(--cc-text-primary) !important;}
[role="option"]:hover {background: var(--cc-panel-hover) !important;}
.stSlider [data-baseweb="slider"] div {color: var(--cc-text-secondary) !important;}
.stSlider [data-baseweb="slider"] div[role="slider"] {background: #F2B84B !important; box-shadow: 0 0 0 4px rgba(242,184,75,.18) !important;}
.stCheckbox label, .stRadio label {color: var(--cc-text-secondary) !important;}

/* Tabs and expanders */
.stTabs [data-baseweb="tab-list"] {gap: .35rem; border-bottom:1px solid var(--cc-border-soft);}
.stTabs [data-baseweb="tab"] {background: #101B2E; border:1px solid var(--cc-border-soft); border-bottom:none; border-radius: 10px 10px 0 0; padding: .45rem .8rem; color:var(--cc-text-muted) !important;}
.stTabs [aria-selected="true"] {background: var(--cc-panel-soft) !important; color: var(--cc-text-primary) !important; border-color: rgba(242,184,75,.42) !important;}
.streamlit-expanderHeader {background: var(--cc-panel-alt) !important; color: var(--cc-text-primary) !important; border-radius: var(--cc-radius-md) !important; border:1px solid var(--cc-border-soft) !important;}
.streamlit-expanderContent {background: rgba(15,23,42,.65) !important; border:1px solid var(--cc-border-soft) !important; border-top:0 !important; border-radius: 0 0 var(--cc-radius-md) var(--cc-radius-md) !important;}

/* Tables / dataframes */
[data-testid="stDataFrame"], [data-testid="stTable"], .stDataFrame {border:1px solid var(--cc-border-soft) !important; border-radius: var(--cc-radius-md) !important; overflow:hidden; background: var(--cc-panel) !important;}
[data-testid="stDataFrame"] * {color: var(--cc-text-primary) !important;}
table {color: var(--cc-text-primary);}
thead tr th {background: #0F1A2C !important; color: var(--cc-text-primary) !important; border-bottom:1px solid var(--cc-border) !important; font-weight:800 !important;}
tbody tr:nth-child(odd) {background: rgba(15,23,42,.96) !important;}
tbody tr:nth-child(even) {background: rgba(30,41,59,.72) !important;}
tbody tr:hover {background: rgba(201,31,39,.18) !important;}
td, th {border-color: var(--cc-border-soft) !important;}

/* Mini summary tables */
.cc-mini-table {width:100%; border-collapse:collapse; font-size:11px; margin-top:.38rem; color:var(--cc-text-secondary);}
.cc-mini-table th {text-align:center; padding:5px 6px; color:var(--cc-text-primary); font-weight:900; border-bottom:1px solid var(--cc-border); background:#0F1A2C;}
.cc-mini-table td {padding:5px 6px; border-bottom:1px solid var(--cc-border-soft); color:var(--cc-text-secondary);}
.cc-mini-table td.label-cell {text-align:left;}
.cc-mini-table td.num-cell {text-align:center; font-variant-numeric: tabular-nums;}
.cc-mini-table tr.total-row td {font-weight:800; color:var(--cc-text-primary); border-top:1px solid var(--cc-border); background:rgba(201,31,39,.10);}
.cc-swatch {display:inline-block; width:9px; height:9px; border-radius:2px; vertical-align:middle; margin-right:8px; position:relative; top:-1px; border:1px solid rgba(255,255,255,.22);}

/* Alerts */
.stAlert {background: rgba(30,41,59,.96) !important; color: var(--cc-text-primary) !important; border:1px solid var(--cc-border-soft) !important; border-radius: var(--cc-radius-md) !important;}

/* Voter lookup */
.lookup-result-card {padding:.85rem .95rem; margin:.25rem 0 .4rem 0;}
.lookup-result-card.selected {border:2px solid #F2B84B; background:linear-gradient(180deg, rgba(201,31,39,.18), rgba(18,27,38,.96));}
.lookup-result-line0 {font-size:15px; font-weight:900; color:var(--cc-text-primary); margin-bottom:.22rem;}
.lookup-result-line1, .lookup-result-line2, .lookup-result-line3 {font-size:13px; color:var(--cc-text-secondary); line-height:1.35;}
.lookup-vh-wrap {margin:.35rem 0 .85rem 0;}
.lookup-vh-title {font-size:16px; font-weight:900; color:var(--cc-text-primary); margin:.15rem 0 .35rem 0;}
.lookup-vh-table {width:100%; border-collapse:collapse; font-size:12px;}
.lookup-vh-table th, .lookup-vh-table td {border:1px solid var(--cc-border-soft); padding:7px 6px; text-align:center;}
.lookup-vh-table th {background:#0F1A2C; font-weight:900; color:var(--cc-text-primary);}
.lookup-vh-rowhead {background:#101B2E; font-weight:800; text-align:left !important; min-width:76px; color:var(--cc-text-secondary);}
.lookup-vh-cell {background:#111827; font-weight:800; min-width:48px; color:var(--cc-text-primary);}
.lookup-vh-dnv {background:#1E293B; color:var(--cc-text-faint);}
.lookup-legend {display:flex; flex-wrap:wrap; gap:16px; font-size:12px; color:var(--cc-text-secondary); margin-top:.2rem; padding:.55rem .7rem; border:1px solid var(--cc-border-soft); border-radius:10px; background:#111827;}
.lookup-legend-icon {display:inline-block; min-width:18px; text-align:center; margin-right:4px;}
.lookup-legend-swatch {display:inline-block; width:14px; height:14px; vertical-align:middle; margin-right:6px; background:#1E293B; border:1px solid var(--cc-border); border-radius:3px;}

.tss-fallback {font-size:23px; line-height:1.02; font-weight:900; color:#fff; text-align:left;}
.tss-fallback span {color: var(--cc-blue) !important;}



/* Opening dashboard preview shown before a universe is applied. */
.cc-open-dashboard {border:1px solid rgba(201,31,39,.72); border-radius:18px; padding:18px; background:radial-gradient(circle at 80% 0%, rgba(201,31,39,.18), transparent 32%), linear-gradient(180deg, rgba(10,17,27,.96), rgba(5,10,16,.98)); box-shadow:0 22px 60px rgba(0,0,0,.36);}
.cc-open-title {font-size:28px; font-weight:950; color:#fff; margin-bottom:4px;}
.cc-open-subtitle {font-size:13px; color:#CBD5E1; margin-bottom:16px;}
.cc-open-filters {display:grid; grid-template-columns:repeat(5, 1fr); gap:12px; margin-bottom:16px;}
.cc-filter-preview {border:1px solid rgba(148,163,184,.24); border-radius:12px; background:#07101A; padding:10px 12px;}
.cc-filter-preview b {display:block; color:#EF4444; font-size:11px; text-transform:uppercase; margin-bottom:8px;}
.cc-filter-preview span {color:#F8FAFC; font-weight:800;}
.cc-open-metrics {display:grid; grid-template-columns:repeat(5, 1fr); gap:12px; margin-bottom:16px;}
.cc-open-metric {border:1px solid rgba(148,163,184,.24); border-radius:14px; background:linear-gradient(180deg,#0D1724,#08111B); padding:16px; min-height:100px;}
.cc-open-metric .label {font-size:12px; font-weight:900; letter-spacing:.04em; text-transform:uppercase;}
.cc-open-metric .value {font-size:27px; font-weight:950; color:#fff; margin-top:8px;}
.cc-open-metric.red {border-left:4px solid var(--cc-red);} .cc-open-metric.red .label{color:var(--cc-red);}
.cc-open-metric.blue {border-left:4px solid var(--cc-blue);} .cc-open-metric.blue .label{color:#4AA3FF;}
.cc-open-metric.green {border-left:4px solid var(--cc-green);} .cc-open-metric.green .label{color:#7ED957;}
.cc-open-metric.gold {border-left:4px solid var(--cc-gold);} .cc-open-metric.gold .label{color:var(--cc-gold);}
.cc-open-grid {display:grid; grid-template-columns:1.05fr 1.55fr; gap:16px;}
.cc-open-card {border:1px solid rgba(148,163,184,.24); border-radius:16px; background:radial-gradient(circle at 50% 0%, rgba(29,78,216,.08), transparent 35%), #07101A; padding:18px;}
.cc-open-card h3 {margin:0 0 14px 0; color:#F8FAFC; font-size:18px;}
.cc-donut {width:245px; height:245px; border-radius:50%; margin:0 auto 12px auto; background:conic-gradient(var(--cc-red) 0 48%, var(--cc-blue) 48% 89%, var(--cc-green) 89% 99%, var(--cc-gold) 99% 100%); position:relative; box-shadow:inset 0 0 0 1px rgba(255,255,255,.14), 0 14px 28px rgba(0,0,0,.35);}
.cc-donut:after {content:none !important; display:none !important;}
.cc-legend-row {display:flex; justify-content:space-between; gap:8px; border-bottom:1px solid rgba(148,163,184,.14); padding:7px 0; color:#CBD5E1;}
.cc-legend-row b {color:#fff;}
.cc-dot {display:inline-block; width:11px; height:11px; border-radius:50%; margin-right:8px;}
.cc-bars {display:grid; grid-template-columns:repeat(8,1fr); gap:12px; height:230px; align-items:end; border-bottom:1px solid rgba(148,163,184,.35); padding:0 8px;}
.cc-bar {border-radius:7px 7px 0 0; background:linear-gradient(180deg,#EF4444,#8F1118); box-shadow:0 0 18px rgba(201,31,39,.18); min-height:30px;}
.cc-bar-labels {display:grid; grid-template-columns:repeat(8,1fr); gap:12px; padding:8px 8px 0; font-size:12px; color:#CBD5E1; text-align:center;}
.cc-open-note {margin-top:16px; color:#94A3B8; font-size:13px;}
@media (max-width: 1200px) { .cc-open-filters, .cc-open-metrics {grid-template-columns:1fr 1fr;} .cc-open-grid {grid-template-columns:1fr;} }

/* Party accent helpers */
.cc-party-r, .party-r {color: var(--cc-red) !important;}
.cc-party-d, .party-d {color: var(--cc-blue) !important;}
.cc-party-o, .party-o {color: var(--cc-green) !important;}

@media (max-width: 1100px) {
  .brand-grid {grid-template-columns: 1fr; gap:10px;}
  .brand-left, .brand-right {justify-content:center;}
  .brand-center {text-align:center;}
}


/* Final red/gold executive dashboard polish: stronger logo treatment and dark chart surfaces. */
[data-testid="stVegaLiteChart"] {
  background: transparent !important;
  border-radius: 12px !important;
}
[data-testid="stVegaLiteChart"] > div {
  background: transparent !important;
}
.brand-grid {grid-template-columns: 260px 1fr 310px;}
.brand-right {
  align-items:center !important;
  justify-content:center !important;
  padding:.32rem .55rem !important;
  background:linear-gradient(135deg, rgba(0,0,0,.44), rgba(31,12,16,.32)) !important;
  border:1px solid rgba(242,184,75,.18) !important;
}
.logo-tss {max-width:285px !important; max-height:92px !important; width:100% !important; object-fit:contain; filter: brightness(1.16) saturate(1.18) contrast(1.08) drop-shadow(0 10px 18px rgba(0,0,0,.58)) !important;}
.logo-cc {filter: brightness(1.35) saturate(1.45) contrast(1.18) drop-shadow(0 10px 18px rgba(0,0,0,.58)) !important;}
.top-shell {border-color:rgba(201,31,39,.95) !important;}
.chart-card, .table-card, .section-card {background: radial-gradient(circle at 50% 0%, rgba(29,78,216,.06), transparent 30%), linear-gradient(180deg, rgba(13,22,32,.98), rgba(6,11,17,.98)) !important;}
.small-header {color:#F8FAFC !important;}
.stDataFrame, [data-testid="stDataFrame"] {background:transparent !important;}
[data-testid="stDataFrame"] div, [data-testid="stDataFrame"] span {color:#E5E7EB !important;}
.cc-mini-table {background:#0B111A !important; border:1px solid rgba(148,163,184,.22); border-radius:10px; overflow:hidden;}
.cc-mini-table th {background:#101827 !important; color:#F2B84B !important;}
.cc-mini-table td {background:#0D1521 !important; color:#E5E7EB !important;}
.cc-mini-table tr:nth-child(even) td {background:#111B2A !important;}
.cc-mini-table tr.total-row td {background:rgba(201,31,39,.18) !important; color:#F8FAFC !important;}


/* Opening preview mockup layout refinements */
.cc-opening-dashboard{display:flex;flex-direction:column;gap:16px;margin-top:4px;}
.cc-open-filters{display:grid;grid-template-columns:1.15fr 1.1fr 1fr 1fr .9fr 1fr;gap:18px;align-items:end;border:1px solid rgba(148,163,184,.24);border-radius:14px;background:rgba(7,16,26,.92);padding:18px 20px;box-shadow:0 16px 34px rgba(0,0,0,.28);}
.cc-open-filters label{display:block;color:#EF4444;font-size:12px;font-weight:950;margin-bottom:8px;letter-spacing:.04em;}
.cc-open-select{height:42px;border:1px solid #334155;border-radius:8px;background:#060C12;color:#fff;display:flex;align-items:center;padding:0 14px;font-weight:800;}
.cc-open-apply{height:44px;border:1px solid #D71920;border-radius:8px;background:rgba(79,9,13,.72);color:#fff;font-weight:950;display:flex;align-items:center;justify-content:center;box-shadow:inset 0 0 18px rgba(215,25,32,.14);}
.cc-open-metrics{display:grid;grid-template-columns:repeat(5,1fr);gap:14px;}
.cc-open-metric{min-height:102px;border:1px solid rgba(148,163,184,.24);border-radius:14px;background:linear-gradient(180deg,rgba(16,27,39,.98),rgba(7,13,20,.98));display:flex;align-items:center;gap:16px;padding:16px 18px;box-shadow:0 12px 26px rgba(0,0,0,.25);}
.cc-open-metric .icon{width:54px;height:54px;border-radius:50%;display:flex;align-items:center;justify-content:center;background:linear-gradient(180deg,#C91F27,#6E0F14);border:1px solid rgba(255,255,255,.18);font-size:24px;color:#fff;font-weight:950;box-shadow:0 0 22px rgba(201,31,39,.22);}
.cc-open-metric.blue .icon{background:linear-gradient(180deg,#2563EB,#0B3A94);}.cc-open-metric.green .icon{background:linear-gradient(180deg,#4C9A2A,#245817);}.cc-open-metric.gold .icon{background:linear-gradient(180deg,#B87333,#6B3A10);}
.cc-open-metric .label{font-size:12px;font-weight:950;color:#EF4444;letter-spacing:.04em;}.cc-open-metric.blue .label{color:#2F8CFF}.cc-open-metric.green .label{color:#7ED957}.cc-open-metric.gold .label{color:#F2B84B}
.cc-open-metric .value{font-size:28px;font-weight:950;color:#fff;line-height:1.1;}.cc-open-metric .sub{font-size:12px;color:#CBD5E1;margin-top:8px;}
.cc-open-main-grid{display:grid;grid-template-columns:1.05fr 1.55fr;gap:16px;}.cc-open-card{border:1px solid rgba(255,255,255,.20);border-radius:16px;background:radial-gradient(circle at 50% 0%,rgba(29,78,216,.08),transparent 35%),linear-gradient(180deg,rgba(9,18,28,.98),rgba(5,10,16,.98));padding:20px 22px;box-shadow:0 16px 34px rgba(0,0,0,.30);}
.cc-open-card h3{display:flex;justify-content:space-between;align-items:center;margin:0 0 18px 0;color:#F8FAFC;font-size:18px;font-weight:950;letter-spacing:.01em;}.cc-open-card h3 span{color:#fff;letter-spacing:3px}.cc-open-card p{margin:14px 0 0;color:#94A3B8;font-size:13px;}.cc-open-split{display:grid;grid-template-columns:1fr .95fr;gap:22px;align-items:center;}
.cc-donut{width:250px;height:250px;border-radius:50%;margin:0 auto;position:relative;box-shadow:inset 0 0 0 1px rgba(255,255,255,.20),0 14px 28px rgba(0,0,0,.35);}.cc-donut-center{position:absolute;inset:76px;border-radius:50%;background:#07101A;display:flex;flex-direction:column;align-items:center;justify-content:center;color:#fff;box-shadow:inset 0 0 0 1px rgba(255,255,255,.10);}.cc-donut-center b{font-size:22px;color:#fff}.cc-donut-center span{font-size:14px;color:#CBD5E1}.cc-legend-row{display:flex;justify-content:space-between;gap:12px;border-bottom:1px solid rgba(148,163,184,.14);padding:9px 0;color:#CBD5E1;font-size:14px}.cc-legend-row b{color:#fff;font-weight:800}.cc-dot{display:inline-block;width:14px;height:14px;border-radius:50%;margin-right:9px;vertical-align:-2px;box-shadow:0 0 10px rgba(255,255,255,.12)}
.cc-bars{display:grid;grid-template-columns:repeat(8,1fr);gap:16px;height:240px;align-items:end;border-left:1px solid rgba(148,163,184,.20);border-bottom:1px solid rgba(148,163,184,.35);padding:0 14px;}.cc-open-bar-wrap{display:flex;flex-direction:column;align-items:center;justify-content:flex-end;height:100%;}.cc-open-bar-value{font-size:12px;color:#fff;font-weight:800;margin-bottom:7px}.cc-bar{width:100%;border-radius:7px 7px 0 0;background:linear-gradient(180deg,#EF4444,#8F1118);box-shadow:0 0 18px rgba(201,31,39,.22);border:1px solid rgba(239,68,68,.55);}.cc-bar-labels{display:grid;grid-template-columns:repeat(8,1fr);gap:16px;padding:9px 14px 0 15px;font-size:13px;color:#CBD5E1;text-align:center;}
.gender-card{min-height:315px}.geo-card{min-height:315px}.cc-open-table{width:100%;border-collapse:separate;border-spacing:0;border:1px solid rgba(148,163,184,.22);border-radius:8px;overflow:hidden;font-size:14px}.cc-open-table th{padding:12px 10px;background:#08111B;color:#F2B84B;text-align:left;font-weight:950}.cc-open-table td{padding:12px 10px;border-top:1px solid rgba(148,163,184,.18);color:#E5E7EB}.cc-open-table tr:nth-child(even) td{background:rgba(15,23,42,.52)}.cc-open-table .red{color:#EF4444}.cc-open-table .blue{color:#2F8CFF}.cc-open-table .green{color:#7ED957}.cc-open-table .gold{color:#F2B84B}.cc-open-note{color:#94A3B8;font-size:13px;margin:2px 0 12px 0;text-align:center;}
@media(max-width:1300px){.cc-open-filters{grid-template-columns:1fr 1fr 1fr}.cc-open-metrics{grid-template-columns:1fr 1fr}.cc-open-main-grid{grid-template-columns:1fr}.cc-open-split{grid-template-columns:1fr}.cc-donut{width:220px;height:220px}.cc-donut-center{inset:66px}}



/* v3 mockup opening corrections: no fake filter bar, no cropped TSS logo, sharper KPI icons. */
.top-shell{overflow:visible !important; padding:1.05rem 1.05rem 1.05rem 1.05rem !important;}
.brand-grid{grid-template-columns:260px 1fr 350px !important; min-height:108px !important;}
.brand-right{min-height:104px !important; max-height:none !important; overflow:visible !important; padding:.35rem .7rem !important;}
.logo-tss{width:330px !important; max-width:330px !important; max-height:104px !important; object-fit:contain !important; display:block !important;}
.logo-cc{max-width:245px !important; max-height:96px !important; object-fit:contain !important;}
.cc-opening-dashboard{margin-top:14px !important;}
.cc-open-filters{display:none !important;}
.cc-open-metric{display:flex !important; align-items:center !important; gap:18px !important; min-height:118px !important; padding:18px 18px !important;}
.cc-open-metric .icon{width:58px;height:58px;border-radius:50%;display:flex;align-items:center;justify-content:center;flex:0 0 58px;font-size:25px;font-weight:950;color:#fff;box-shadow:inset 0 0 0 1px rgba(255,255,255,.25),0 8px 18px rgba(0,0,0,.38)}
.cc-open-metric.red .icon{background:radial-gradient(circle at 35% 30%,#EF4444,#8F1118);border:1px solid #EF4444;}
.cc-open-metric.blue .icon{background:radial-gradient(circle at 35% 30%,#2F8CFF,#0B3B96);border:1px solid #2F8CFF;}
.cc-open-metric.green .icon{background:radial-gradient(circle at 35% 30%,#66B845,#215A22);border:1px solid #66B845;}
.cc-open-metric.gold .icon{background:radial-gradient(circle at 35% 30%,#F2B84B,#8B5A12);border:1px solid #F2B84B;}
.cc-open-metric .icon svg{width:34px;height:34px;display:block;fill:currentColor;}
.cc-open-metric .icon .dletter{font-size:27px;font-weight:950;line-height:1;}
.cc-donut-center b{font-size:19px !important; line-height:1.05; text-align:center;}
.cc-donut-center span{font-size:13px !important;}
.cc-open-card{border-color:rgba(255,255,255,.24) !important;}
.cc-open-card h3{letter-spacing:.08em !important;}

</style>
""", unsafe_allow_html=True)


def get_reporting_tss_logo_path() -> Path:
    """Use print/reporting-safe TSS logo for PDFs, fallback to dashboard logo."""
    return TSS_REPORTING_LOGO if "TSS_REPORTING_LOGO" in globals() and TSS_REPORTING_LOGO.exists() else TSS_LOGO


def img_to_data_uri(path: Path) -> str:
    if not path.exists():
        return ""
    encoded = base64.b64encode(path.read_bytes()).decode("utf-8")
    return f"data:image/png;base64,{encoded}"


def svg_to_data_uri(svg: str) -> str:
    """Inline SVG asset helper for dark-theme-safe brand marks."""
    encoded = base64.b64encode(svg.encode("utf-8")).decode("utf-8")
    return f"data:image/svg+xml;base64,{encoded}"


def tss_dark_logo_uri() -> str:
    """Dark-theme-safe Political Technology Company header mark."""
    svg = r"""<svg xmlns="http://www.w3.org/2000/svg" width="520" height="150" viewBox="0 0 520 150">
  <defs>
    <filter id="glow" x="-30%" y="-30%" width="160%" height="160%"><feGaussianBlur stdDeviation="2.2" result="blur"/><feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter>
    <linearGradient id="blue" x1="0" x2="1" y1="0" y2="1"><stop offset="0" stop-color="#5DB2FF"/><stop offset="1" stop-color="#0B4BC8"/></linearGradient>
    <linearGradient id="red" x1="0" x2="1"><stop offset="0" stop-color="#FF4B55"/><stop offset="1" stop-color="#C91F27"/></linearGradient>
  </defs>
  <rect x="4" y="4" width="512" height="142" rx="20" fill="rgba(22,9,11,.72)" stroke="#78421B" stroke-width="2"/>
  <g transform="translate(24,25)" filter="url(#glow)">
    <path d="M62 0 L69 22 L92 22 L73 35 L80 58 L62 44 L43 58 L50 35 L31 22 L55 22 Z" fill="#F2B84B"/>
    <rect x="15" y="82" width="116" height="12" rx="4" fill="url(#red)"/>
    <rect x="26" y="69" width="94" height="12" rx="4" fill="url(#blue)"/>
    <path d="M33 65 Q73 25 113 65 Z" fill="url(#blue)"/>
    <rect x="41" y="48" width="10" height="21" rx="2" fill="#EAF4FF"/>
    <rect x="61" y="41" width="10" height="28" rx="2" fill="#EAF4FF"/>
    <rect x="81" y="48" width="10" height="21" rx="2" fill="#EAF4FF"/>
    <rect x="101" y="52" width="10" height="17" rx="2" fill="#EAF4FF"/>
    <path d="M68 23 L78 41 L58 41 Z" fill="#F2B84B"/>
    <circle cx="27" cy="106" r="3.5" fill="#F2B84B"/><circle cx="49" cy="110" r="3.5" fill="#EF4444"/><circle cx="72" cy="112" r="3.5" fill="#F2B84B"/><circle cx="95" cy="110" r="3.5" fill="#EF4444"/><circle cx="118" cy="106" r="3.5" fill="#F2B84B"/>
  </g>
  <g transform="translate(175,28)">
    <text x="0" y="14" fill="#FFFFFF" font-size="14" font-weight="800" font-family="Arial, Helvetica, sans-serif">Powered By</text>
    <text x="0" y="48" fill="#FFFFFF" font-size="27" font-weight="900" font-family="Arial, Helvetica, sans-serif">The Political</text>
    <text x="0" y="80" fill="#2F8CFF" font-size="27" font-weight="900" font-family="Arial, Helvetica, sans-serif">Technology</text>
    <text x="0" y="112" fill="#EF4444" font-size="27" font-weight="900" font-family="Arial, Helvetica, sans-serif">Company</text>
  </g>
</svg>"""
    return svg_to_data_uri(svg)

def file_modified_text(path: Path) -> str:
    if not path.exists():
        return "R2 public source"
    try:
        ts = pd.Timestamp(path.stat().st_mtime, unit="s")
        return ts.strftime("%m/%d/%Y %I:%M %p")
    except Exception:
        return "R2 public source"

def divider():
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)



# Strict PA school district whitelist. If it is not in this list, it is not shown
# as a School District filter option. This prevents school regions from leaking
# into the School District dropdown.
VALID_PA_SCHOOL_DISTRICTS = ['Abington Heights SD', 'Abington SD', 'Albert Gallatin Area SD', 'Aliquippa SD', 'Allegheny Valley SD', 'Allegheny-Clarion Valley SD', 'Allentown City SD', 'Altoona Area SD', 'Ambridge Area SD', 'Annville-Cleona SD', 'Antietam SD', 'Apollo Ridge SD', 'Armstrong SD', 'Athens Area SD', 'Austin Area SD', 'Avella Area SD', 'Avon Grove SD', 'Avonworth SD', 'Bald Eagle Area SD', 'Baldwin-Whitehall SD', 'Bangor Area SD', 'Beaver Area SD', 'Bedford Area SD', 'Belle Vernon Area SD', 'Bellefonte Area SD', 'Bellwood-Antis SD', 'Bensalem Township SD', 'Benton Area SD', 'Bentworth SD', 'Berlin Brothersvalley SD', 'Bermudian Springs SD', 'Berwick Area SD', 'Bethel Park SD', 'Bethlehem Area SD', 'Bethlehem-Center SD', 'Big Beaver Falls Area SD', 'Big Spring SD', 'Blackhawk SD', 'Blacklick Valley SD', 'Blairsville-Saltsburg SD', 'Bloomsburg Area SD', 'Blue Mountain SD', 'Blue Ridge SD', 'Boyertown Area SD', 'Bradford Area SD', 'Brandywine Heights Area SD', 'Brentwood Borough SD', 'Bristol Borough SD', 'Bristol Township SD', 'Brockway Area SD', 'Brookville Area SD', 'Brownsville Area SD', 'Bryn Athyn SD', 'Burgettstown Area SD', 'Burrell SD', 'Butler Area SD', 'California Area SD', 'Cambria Heights SD', 'Cameron County SD', 'Camp Hill SD', 'Canon-McMillan SD', 'Canton Area SD', 'Carbondale Area SD', 'Carlisle Area SD', 'Carlynton SD', 'Carmichaels Area SD', 'Catasauqua Area SD', 'Centennial SD', 'Center Valley SD', 'Central Bucks SD', 'Central Cambria SD', 'Central Columbia SD', 'Central Dauphin SD', 'Central Fulton SD', 'Central Greene SD', 'Central York SD', 'Chambersburg Area SD', 'Charleroi SD', 'Chartiers Valley SD', 'Chartiers-Houston SD', 'Cheltenham Township SD', 'Chester-Upland SD', 'Chestnut Ridge SD', 'Chichester SD', 'Clairton City SD', 'Clarion Area SD', 'Clarion-Limestone Area SD', 'Claysburg-Kimmel SD', 'Clearfield Area SD', 'Coatesville Area SD', 'Cocalico SD', 'Colonial SD', 'Columbia Borough SD', 'Commodore Perry SD', 'Conemaugh Township Area SD', 'Conemaugh Valley SD', 'Conestoga Valley SD', 'Conewago Valley SD', 'Conneaut SD', 'Connellsville Area SD', 'Conrad Weiser Area SD', 'Cornell SD', 'Cornwall-Lebanon SD', 'Corry Area SD', 'Coudersport Area SD', 'Council Rock SD', 'Cranberry Area SD', 'Crawford Central SD', 'Crestwood SD', 'Cumberland Valley SD', 'Curwensville Area SD', 'Dallas SD', 'Dallastown Area SD', 'Daniel Boone Area SD', 'Danville Area SD', 'Deer Lakes SD', 'Delaware Valley SD', 'Derry Area SD', 'Derry Township SD', 'Donegal SD', 'Dover Area SD', 'Downingtown Area SD', 'DuBois Area SD', 'Dunmore SD', 'Duquesne City SD', 'East Allegheny SD', 'East Lycoming SD', 'East Penn SD', 'East Pennsboro SD', 'East Stroudsburg Area SD', 'Eastern Lancaster Co SD', 'Eastern Lebanon County SD', 'Eastern York SD', 'Easton Area SD', 'Elizabeth Forward SD', 'Elizabethtown Area SD', 'Elk Lake SD', 'Ellwood City Area SD', 'Ephrata Area SD', 'Erie City SD', 'Everett Area SD', 'Exeter Township SD', 'Fairfield Area SD', 'Fairview SD', 'Fannett-Metal SD', 'Farrell Area SD', 'Ferndale Area SD', 'Fleetwood Area SD', 'Forbes Road SD', 'Forest Area SD', 'Forest City Regional SD', 'Forest Hills SD', 'Fort Cherry SD', 'Fort LeBoeuf SD', 'Fox Chapel Area SD', 'Franklin Area SD', 'Franklin Regional SD', 'Frazier SD', 'Freedom Area SD', 'Freeport Area SD', 'Galeton Area SD', 'Garnet Valley SD', 'Gateway SD', 'General McLane SD', 'Gettysburg Area SD', 'Girard SD', 'Glendale SD', 'Governor Mifflin SD', 'Great Valley SD', 'Greater Johnstown SD', 'Greater Latrobe SD', 'Greater Nanticoke Area SD', 'Greencastle-Antrim SD', 'Greensburg Salem SD', 'Greenville Area SD', 'Greenwood SD', 'Grove City Area SD', 'Halifax Area SD', 'Hamburg Area SD', 'Hampton Township SD', 'Hanover Area SD', 'Hanover Public SD', 'Harbor Creek SD', 'Harmony Area SD', 'Harrisburg City SD', 'Hatboro-Horsham SD', 'Haverford Township SD', 'Hazleton Area SD', 'Hempfield Area SD', 'Hempfield SD', 'Hermitage SD', 'Highlands SD', 'Hollidaysburg SD', 'Homer-Center SD', 'Hopewell Area SD', 'Huntingdon Area SD', 'Indiana Area SD', 'Interboro SD', 'Iroquois SD', 'Jamestown Area SD', 'Jeanette City SD', 'Jefferson-Morgan SD', 'Jenkintown SD', 'Jersey Shore Area SD', 'Jim Thorpe Area SD', 'Johnsonburg Area SD', 'Juniata County SD', 'Juniata Valley SD', 'Kane Area SD', 'Karns City Area SD', 'Kennett Consolidated SD', 'Keystone Central SD', 'Keystone Oaks SD', 'Keystone SD', 'Kiski Area SD', 'Kutztown Area SD', 'Lackawanna Trail SD', 'Lake Lehman SD', 'Lakeland SD', 'Lakeview SD', 'Lampeter-Strasburg SD', 'Lancaster SD', 'Laurel Area SD', 'Laurel Highlands SD', 'Lebanon SD', 'Leechburg Area SD', 'Lehighton Area SD', 'Lewisburg Area SD', 'Ligonier Valley SD', 'Line Mountain SD', 'Littlestown Area SD', 'Lower Dauphin SD', 'Lower Merion SD', 'Lower Moreland Twp SD', 'Loyalsock Township SD', 'Mahanoy Area SD', 'Manheim Central SD', 'Manheim Township SD', 'Marion Center Area SD', 'Marple Newtown SD', 'Mars Area SD', 'McGuffey SD', 'McKeesport Area SD', 'Mechanicsburg Area SD', 'Mercer Area SD', 'Methacton SD', 'Meyers Dale Area SD', 'Mid Valley SD', 'Midd-West SD', 'Middletown Area SD', 'Midland Borough SD', 'Mifflin County SD', 'Mifflinburg Area SD', 'Millcreek Township SD', 'Millersburg Area SD', 'Millville Area SD', 'Milton Area SD', 'Minersville Area SD', 'Mohawk Area SD', 'Monessen City SD', 'Moniteau SD', 'Montgomery Area SD', 'Montour SD', 'Montoursville Area SD', 'Montrose Area SD', 'Moon Area SD', 'Morrisville Borough SD', 'Moshannon Valley SD', 'Mount Carmel Area SD', 'Mount Pleasant Area SD', 'Mount Union Area SD', 'Mountain View SD', 'Mt Lebanon SD', 'Muhlenberg SD', 'Muncy SD', 'Nazareth Area SD', 'Neshaminy SD', 'Neshannock Township SD', 'New Brighton Area SD', 'New Castle Area SD', 'New Hope-Solebury SD', 'New Kensington-Arnold SD', 'Newport SD', 'Norristown Area SD', 'North Allegheny SD', 'North Clarion County SD', 'North East SD', 'North Hills SD', 'North Penn SD', 'North Pocono SD', 'North Schuylkill SD', 'North Star SD', 'Northampton Area SD', 'Northeast Bradford SD', 'Northeastern York SD', 'Northern Bedford County SD', 'Northern Cambria SD', 'Northern Lebanon SD', 'Northern Lehigh SD', 'Northern Potter SD', 'Northern Tioga SD', 'Northern York County SD', 'Northgate SD', 'Northwest Area SD', 'Northwestern Lehigh SD', 'Northwestern SD', 'Norwin SD', 'Octorara Area SD', 'Oil City Area SD', 'Old Forge SD', 'Oley Valley SD', 'Oswayo Valley SD', 'Otto-Eldred SD', 'Owen J Roberts SD', 'Oxford Area SD', 'Palisades SD', 'Palmerton Area SD', 'Palmyra Area SD', 'Panther Valley SD', 'Parkland SD', 'Pen Argyl Area SD', 'Penn Cambria SD', 'Penn Hills SD', 'Penn Manor SD', 'Penn-Delco SD', 'Penn-Trafford SD', 'Penncrest SD', 'Pennridge SD', 'Penns Manor Area SD', 'Penns Valley Area SD', 'Pennsbury SD', 'Pequea Valley SD', 'Perkiomen Valley SD', 'Peters Township SD', 'Philadelphia City SD', 'Philipsburg-Osceola Area SD', 'Phoenixville Area SD', 'Pine Grove SD', 'Pine-Richland SD', 'Pittsburgh SD', 'Pittston Area SD', 'Pleasant Valley SD', 'Plum Borough SD', 'Pocono Mountain SD', 'Port Allegany SD', 'Portage Area SD', 'Pottsgrove SD', 'Pottstown SD', 'Pottsville Area SD', 'Punxsutawney Area SD', 'Purchase Line SD', 'Quaker Valley SD', 'Quakertown Community SD', 'Radnor Township SD', 'Reading SD', 'Red Lion Area SD', 'Redbank Valley SD', 'Reynolds SD', 'Richland SD', 'Ridgway Area SD', 'Ridley SD', 'Ringgold SD', 'Riverside Beaver County SD', 'Riverside SD', 'Riverview SD', 'Rochester Area SD', 'Rockwood Area SD', 'Rose Tree Media SD', 'Saint Clair Area SD', 'Saint Marys Area SD', 'Salisbury Township SD', 'Salisbury-Elk Lick SD', 'Saucon Valley SD', 'Sayre Area SD', 'Schuylkill Haven Area SD', 'Schuylkill Valley SD', 'Scranton City SD', 'Selinsgrove Area SD', 'Seneca Valley SD', 'Shade-Central City SD', 'Shaler Area SD', 'Shamokin Area SD', 'Shanksville-Stony Creek SD', 'Sharon City SD', 'Sharpsville Area SD', 'Shenandoah Valley SD', 'Shenango Area SD', 'Shikellamy SD', 'Shippensburg Area SD', 'Slippery Rock Area SD', 'Smethport Area SD', 'Solanco SD', 'Somerset Area SD', 'Souderton Area SD', 'South Allegheny SD', 'South Butler County SD', 'South Eastern SD', 'South Fayette Twp SD', 'South Middleton SD', 'South Park SD', 'South Side Area SD', 'South Western SD', 'South Williamsport SD', 'Southeast Delco SD', 'Southeastern Greene SD', 'Southern Columbia Area SD', 'Southern Fulton SD', 'Southern Huntingdon Co SD', 'Southern Lehigh SD', 'Southern Tioga SD', 'Southern York County SD', 'Southmoreland SD', 'Spring Cove SD', 'Spring Grove Area SD', 'Spring-Ford Area SD', 'Springfield SD', 'Springfield Township SD', 'State College Area SD', 'Steel Valley SD', 'Steelton-Highspire SD', 'Sto-Rox SD', 'Stroudsburg Area SD', 'Sullivan County SD', 'Susquehanna Community SD', 'Susquehanna Township SD', 'Susquenita SD', 'Tamaqua Area SD', 'Titusville Area SD', 'Towanda Area SD', 'Tredyffrin-Easttown SD', 'Tri-Valley SD', 'Trinity Area SD', 'Troy Area SD', 'Tulpehocken Area SD', 'Tunkhannock Area SD', 'Turkeyfoot Valley Area SD', 'Tuscarora SD', 'Tussey Mountain SD', 'Twin Valley SD', 'Tyrone Area SD', 'Union Area SD', 'Union City Area SD', 'Union SD', 'Uniontown Area SD', 'Unionville Chadds-Ford SD', 'United SD', 'Upper Adams SD', 'Upper Darby SD', 'Upper Dauphin SD', 'Upper Dublin SD', 'Upper Merion Area SD', 'Upper Moreland Twp SD', 'Upper Perkiomen SD', 'Upper Saint Clair SD', 'Valley Grove SD', 'Valley View SD', 'Wallenpaupack Area SD', 'Wallingford-Swarthmore SD', 'Warren County SD', 'Warrior Run SD', 'Warwick SD', 'Washington SD', 'Wattsburg Area SD', 'Wayne Highlands SD', 'Waynesboro Area SD', 'Weatherly Area SD', 'Wellsboro Area SD', 'West Allegheny SD', 'West Branch Area SD', 'West Chester Area SD', 'West Greene SD', 'West Jefferson Hills SD', 'West Middlesex Area SD', 'West Mifflin Area SD', 'West Perry SD', 'West Shore SD', 'West York Area SD', 'Western Beaver County SD', 'Western Wayne SD', 'Westmont Hilltop SD', 'Whitehall-Coplay SD', 'Wilkes-Barre Area SD', 'Wilkinsburg Borough SD', 'William Penn SD', 'Williams Valley SD', 'Williamsburg Community SD', 'Williamsport Area SD', 'Wilmington Area SD', 'Wilson Area SD', 'Wilson SD', 'Windber Area SD', 'Wissahickon SD', 'Woodland Hills SD', 'Wyalusing Area SD', 'Wyoming Area SD', 'Wyoming Valley West SD', 'Wyomissing Area SD', 'York City SD', 'York Suburban SD', 'Yough SD']

def _norm_school_key(value: str) -> str:
    s = str(value or '').upper().strip()
    if not s or s in {'NAN','NONE','NULL','(BLANK)'}:
        return ''
    s = re.sub(r'\([^)]*\)', ' ', s)
    s = s.replace('&', ' AND ')
    s = re.sub(r'\bA\s*/\s*C\b', 'ALLEGHENY CLARION', s)
    s = re.sub(r'\bALLEGHENY\s+CLARION\b', 'ALLEGHENY CLARION', s)
    s = re.sub(r'\bSCHOOL\s+DISTRICT\b', 'SD', s)
    s = re.sub(r'\bSCHOOL\s+DIST\b', 'SD', s)
    s = re.sub(r'\bSCH\s+DIST\b', 'SD', s)
    s = re.sub(r'\bDISTRICT\b', 'SD', s)
    s = re.sub(r'\bS D\b', 'SD', s)
    s = re.sub(r'[^A-Z0-9]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

VALID_PA_SCHOOL_DISTRICT_BY_KEY = {_norm_school_key(v): v for v in VALID_PA_SCHOOL_DISTRICTS}

def normalize_school_district_option(value: str) -> str:
    key = _norm_school_key(value)
    if not key:
        return ''
    if key in VALID_PA_SCHOOL_DISTRICT_BY_KEY:
        return VALID_PA_SCHOOL_DISTRICT_BY_KEY[key]
    # Some SURE values omit SD after parenthetical cleanup; try adding it.
    if not key.endswith(' SD') and (key + ' SD') in VALID_PA_SCHOOL_DISTRICT_BY_KEY:
        return VALID_PA_SCHOOL_DISTRICT_BY_KEY[key + ' SD']
    return ''

def geo_display_case_py(value) -> str:
    """Option B display: Title Case precinct/geo names while preserving useful codes."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = str(value).strip()
    if s.lower() in {"", "nan", "none", "null", "(blank)"}:
        return ""
    s = s.replace("_", " ")
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r"\s+", " ", s).strip()
    replacements = {
        r"\bTWP\b": "Township",
        r"\bTWP\.\b": "Township",
        r"\bBORO\b": "Borough",
        r"\bBORO\.\b": "Borough",
        r"\bPREC\b": "Precinct",
        r"\bPCT\b": "Precinct",
        r"\bDIST\b": "District",
    }
    for pat, repl in replacements.items():
        s = re.sub(pat, repl, s, flags=re.IGNORECASE)

    keep_upper = {"SD", "IU", "CTC", "AVTS", "MDJ", "PA", "US", "USA", "VTD", "WD", "DIV"}

    def cap_piece(piece: str) -> str:
        if not piece:
            return piece
        up = piece.upper()
        if up in keep_upper:
            return up
        if re.fullmatch(r"\d+[A-Z]?", up):
            return up
        if re.fullmatch(r"\d+(ST|ND|RD|TH)", up):
            return up
        if re.fullmatch(r"[IVXLCM]+", up):
            return up
        return piece[:1].upper() + piece[1:].lower()

    words = []
    for word in s.split(" "):
        # preserve punctuation around the core word
        prefix = re.match(r"^[^A-Za-z0-9]*", word).group(0)
        suffix = re.search(r"[^A-Za-z0-9]*$", word).group(0)
        core = word[len(prefix): len(word)-len(suffix) if suffix else len(word)]
        if not core:
            words.append(word)
            continue
        slash_parts = []
        for slash in core.split("/"):
            hy_parts = []
            for hy in slash.split("-"):
                apos_parts = [cap_piece(part) for part in hy.split("'")]
                hy_parts.append("'".join(apos_parts))
            slash_parts.append("-".join(hy_parts))
        words.append(prefix + "/".join(slash_parts) + suffix)
    return " ".join(words).strip()

def quote_ident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'

def sql_string_literal(value: str) -> str:
    return "'" + str(value).replace("'", "''") + "'"

@st.cache_resource(show_spinner=False)
def get_conn():
    swap_dir = Path("/tmp/candidate_connect_duckdb_swap")
    swap_dir.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect(database=":memory:")
    con.execute("PRAGMA threads=2")
    con.execute("PRAGMA preserve_insertion_order=false")
    con.execute("PRAGMA temp_directory='/tmp/candidate_connect_duckdb_swap'")
    try:
        con.execute("PRAGMA memory_limit='768MB'")
    except Exception:
        pass
    try:
        con.create_function("geo_display_case", geo_display_case_py, [str], str)
    except Exception:
        pass
    return con

def first_existing(columns, candidates):
    lower_map = {str(c).strip().lower(): c for c in columns}
    for col in candidates:
        if col in columns:
            return col
        hit = lower_map.get(str(col).strip().lower())
        if hit is not None:
            return hit
    return None


def _norm_col_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def first_existing_fuzzy(columns, candidates):
    """Find a column using exact, normalized, and cautious contains matching."""
    hit = first_existing(columns, candidates)
    if hit is not None:
        return hit

    norm_map = {_norm_col_name(c): c for c in columns}
    for cand in candidates:
        n = _norm_col_name(cand)
        if not n:
            continue
        if n in norm_map:
            return norm_map[n]

    # Cautious fuzzy pass for long Athenix-style names only. Do NOT use this for
    # canonical geo fields because short aliases like Muni/City/SD can grab wrong columns.
    for cand in candidates:
        n = _norm_col_name(cand)
        if len(n) < 6:
            continue
        for col in columns:
            cn = _norm_col_name(col)
            if n == cn or n in cn or cn in n:
                return col
    return None


def first_existing_precise(columns, candidates):
    """Exact + normalized match only. Used for canonical app fields to avoid bad fuzzy aliases."""
    hit = first_existing(columns, candidates)
    if hit is not None:
        return hit
    norm_map = {_norm_col_name(c): c for c in columns}
    for cand in candidates:
        n = _norm_col_name(cand)
        if n and n in norm_map:
            return norm_map[n]
    return None

def ensure_parent(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)

def get_secret_value(*keys, default=None):
    try:
        for key in keys:
            if key in st.secrets:
                return st.secrets[key]
    except Exception:
        pass
    for key in keys:
        val = os.environ.get(key)
        if val not in (None, ""):
            return val
    return default


def get_saved_universe_store_info() -> dict:
    account_id = get_secret_value("R2_ACCOUNT_ID", "CLOUDFLARE_ACCOUNT_ID")
    access_key = get_secret_value("R2_ACCESS_KEY_ID", "AWS_ACCESS_KEY_ID")
    secret_key = get_secret_value("R2_SECRET_ACCESS_KEY", "AWS_SECRET_ACCESS_KEY")
    bucket = get_secret_value("R2_BUCKET", "SAVED_UNIVERSES_BUCKET", default=R2_BUCKET)
    endpoint_url = get_secret_value("R2_ENDPOINT_URL", "AWS_ENDPOINT_URL_S3")
    region = get_secret_value("AWS_DEFAULT_REGION", default="auto")

    if not endpoint_url and account_id:
        endpoint_url = f"https://{account_id}.r2.cloudflarestorage.com"

    ready = all([endpoint_url, access_key, secret_key, bucket])
    return {
        "ready": bool(ready),
        "endpoint_url": endpoint_url,
        "access_key": access_key,
        "secret_key": secret_key,
        "bucket": bucket,
        "region": region,
    }


def get_saved_universe_store_label() -> str:
    info = get_saved_universe_store_info()
    return "Cloudflare R2" if info.get("ready") else "Local fallback"


def get_saved_universes_r2_client():
    info = get_saved_universe_store_info()
    if not info.get("ready"):
        return None, info
    client = boto3.client(
        "s3",
        endpoint_url=info["endpoint_url"],
        aws_access_key_id=info["access_key"],
        aws_secret_access_key=info["secret_key"],
        region_name=info["region"],
    )
    return client, info


def _load_saved_universes_local() -> dict:
    if not SAVED_UNIVERSES_PATH.exists():
        return {}
    try:
        data = json.loads(SAVED_UNIVERSES_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def load_saved_universes() -> dict:
    client, info = get_saved_universes_r2_client()
    if client is None:
        return _load_saved_universes_local()
    try:
        obj = client.get_object(Bucket=info["bucket"], Key=SAVED_UNIVERSES_R2_KEY)
        data = json.loads(obj["Body"].read().decode("utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_saved_universes(data: dict):
    payload = json.dumps(data, indent=2).encode("utf-8")
    client, info = get_saved_universes_r2_client()
    if client is None:
        SAVED_UNIVERSES_PATH.write_bytes(payload)
        return
    client.put_object(
        Bucket=info["bucket"],
        Key=SAVED_UNIVERSES_R2_KEY,
        Body=payload,
        ContentType="application/json",
        CacheControl="no-store",
    )


def r2_public_url(key: str) -> str:
    return f"{R2_BASE}/{key}"

def download_public_object(key: str, local_path: Path):
    if local_path.exists():
        return
    ensure_parent(local_path)
    url = r2_public_url(key)
    with requests.get(url, stream=True, timeout=120) as resp:
        resp.raise_for_status()
        with open(local_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    f.write(chunk)

@st.cache_data(show_spinner=True)
def load_manifest():
    LOCAL_ROOT.mkdir(parents=True, exist_ok=True)
    download_public_object("dataset_manifest.json", LOCAL_MANIFEST)
    return json.loads(LOCAL_MANIFEST.read_text(encoding="utf-8"))

@st.cache_data(show_spinner=True)
def ensure_index_shards():
    local_paths, source_label = find_local_dataset_paths("index")
    if local_paths:
        return local_paths, {"source": source_label, "count": len(local_paths)}

    manifest = load_manifest()
    local_paths = []
    for shard in manifest["index"]["shards"]:
        key = shard["key"]
        local_path = LOCAL_ROOT / key
        download_public_object(key, local_path)
        local_paths.append(str(local_path))
    return local_paths, manifest

@st.cache_data(show_spinner=False)
def get_schema(local_paths):
    con = get_conn()
    df = con.execute(f"DESCRIBE SELECT * FROM {dataset_scan_sql(local_paths)}").df()
    return df["column_name"].tolist()

def build_view_sql(columns, local_paths):
    q = quote_ident
    status_col = first_existing(columns, ["VoterStatus", "voterstatus"])
    gender_col = first_existing(columns, ["Gender", "Sex"])
    age_range_col = first_existing(columns, ["Age_Range", "Age Range", "AGERANGE"])
    reg_col = first_existing(columns, ["RegistrationDate", "registrationdate"])
    party_col = first_existing(columns, ["Party"])
    hh_col = first_existing(columns, ["HH_ID"])
    email_col = first_existing(columns, ["Email"])
    landline_col = first_existing(columns, ["Landline"])
    mobile_col = first_existing(columns, ["Mobile"])
    vote_hist_col = first_existing(columns, ["V4A"])
    mib_applied_col = first_existing(columns, ["MIB_Applied"])
    mib_ballot_col = first_existing(columns, ["MIB_BALLOT"])
    mb_score_col = first_existing(columns, ["MB_Prob_Score", "mb_prob_score", "MBScore", "MB_AProp_Score", "MMB_AProp_Score"])
    mb_perm_col = first_existing(columns, ["MB_PERM", "MB_Perm", "MB_Pern"])
    source_file_col = first_existing(columns, ["Source_File", "Source File", "source_file"])
    mb_new_reg_col = first_existing(columns, ["MailBallotNewRegistrant", "Mail Ballot New Registrant", "mail_ballot_new_registrant"])
    applicant_phone_col = first_existing(columns, ["Current_ApplicantPhone", "ApplicantPhone", "Applicant Phone"])
    applicant_phone_type_col = first_existing(columns, ["ApplicantPhone_Type", "Applicant Phone Type"])
    applicant_phone_compliance_col = first_existing(columns, ["ApplicantPhone_Compliance", "Applicant Phone Compliance"])
    current_app_return_col = first_existing(columns, ["Current_App_Return_Date", "AppReturnDate", "App Return Date"])
    current_ballot_sent_col = first_existing(columns, ["Current_Ballot_Sent_Date", "BallotSentDate", "Ballot Sent Date"])
    current_ballot_returned_col = first_existing(columns, ["Current_Ballot_Returned_Date", "BallotReturnedDate", "Ballot Returned Date"])
    age_col = first_existing(columns, ["Age"])
    house_col = first_existing(columns, ["House Number"])
    street_col = first_existing(columns, ["Street Name"])
    apt_col = first_existing(columns, ["Apartment Number"])

    # --- Athenix/local CSV compatibility aliases ---
    # Rebuild canonical filter columns from all likely source columns.
    # This fixes cases where the CSV has a blank/old raw column plus the real
    # Athenix value under a different name.
    CANONICAL_ALIASES = {
        # Stability-first rule: for geography, trust the clean Step 8 canonical fields only.
        # Do NOT fuzzy-pull raw SURE zone fields here; that caused municipalities, districts,
        # precinct codes, and school regions to leak into the wrong filters.
        "County": ["County"],
        "Municipality": ["Municipality"],
        "Precinct": ["Precinct"],
        "USC": ["USC"],
        "STS": ["STS"],
        "STH": ["STH"],
        "School District": ["School District"],
        "School Region": ["School Region"],
        "Party": ["Party"],
        "CalculatedParty": ["CalculatedParty", "Calculated Party", "calculated_party", "Calculated_Party", "ModeledParty", "Modeled Party", "PartisanScore", "Partisan Score"],
        "HH-Party": ["HH-Party", "HH Party", "Household Party", "Household_Party", "HouseholdParty", "household_party"],
        "Tags": ["Tags", "tags", "TAGS"],
        "MB_Target": ["MB_Target", "MB Target", "MBTarget", "Mail Ballot Target", "Mail_Ballot_Target"],
        "MailBallotNewRegistrant": ["MailBallotNewRegistrant", "Mail Ballot New Registrant", "mail_ballot_new_registrant"],
        "Current_App_Return_Date": ["Current_App_Return_Date", "AppReturnDate", "App Return Date"],
        "Current_Ballot_Sent_Date": ["Current_Ballot_Sent_Date", "BallotSentDate", "Ballot Sent Date"],
        "Current_Ballot_Returned_Date": ["Current_Ballot_Returned_Date", "BallotReturnedDate", "Ballot Returned Date"],
        "Current_ApplicantPhone": ["Current_ApplicantPhone", "ApplicantPhone", "Applicant Phone"],
    }

    # Exclude every raw source column used by a canonical alias, not just columns that
    # already match the canonical title. This prevents raw lowercase fields like
    # precinct/school_region from confusing the UI while preserving one clean title-case field.
    canonical_raw_present = []
    for _canonical_name, _candidates in CANONICAL_ALIASES.items():
        for _cand in _candidates:
            _src = first_existing_precise(columns, [_cand])
            if _src and _src not in canonical_raw_present:
                canonical_raw_present.append(_src)

    if canonical_raw_present:
        exprs = ["* EXCLUDE (" + ", ".join(q(c) for c in canonical_raw_present) + ")"]
    else:
        exprs = ["*"]

    def canonical_expr(canonical_name, candidates):
        sources = []
        seen_src = set()
        for cand in candidates:
            src = first_existing_precise(columns, [cand])
            if src and src not in seen_src:
                seen_src.add(src)
                raw = f"nullif(trim(coalesce(cast({q(src)} as varchar), '')), '')"
                if canonical_name in {"USC", "STS", "STH"}:
                    # Numeric district filters only. This removes duplicate options like
                    # '3RD CONGRESSIONAL DISTRICT' while preserving the district number.
                    sources.append(f"nullif(regexp_extract({raw}, '([0-9]{{1,3}})', 1), '')")
                elif canonical_name in {"County", "Municipality", "Precinct", "School District", "School Region"}:
                    sources.append(f"nullif(geo_display_case({raw}), '')")
                else:
                    sources.append(raw)
        if not sources:
            return None
        return "coalesce(" + ", ".join(sources) + ", '') as " + q(canonical_name)

    for _canonical_name, _candidates in CANONICAL_ALIASES.items():
        _expr = canonical_expr(_canonical_name, _candidates)
        if _expr:
            exprs.append(_expr)

    if status_col:
        exprs.append(f"upper(trim(coalesce(cast({q(status_col)} as varchar), ''))) as _Status")
    else:
        exprs.append("'A' as _Status")

    if party_col:
        exprs.append(
            f"""case
                when upper(trim(coalesce(cast({q(party_col)} as varchar), ''))) = 'D' then 'D'
                when upper(trim(coalesce(cast({q(party_col)} as varchar), ''))) = 'R' then 'R'
                else 'O'
            end as _PartyNorm"""
        )
    else:
        exprs.append("'O' as _PartyNorm")

    if gender_col:
        exprs.append(
            f"""case
                when upper(trim(coalesce(cast({q(gender_col)} as varchar), ''))) in ('', 'NONE', 'NAN') then 'U'
                else upper(trim(cast({q(gender_col)} as varchar)))
            end as _Gender"""
        )
    else:
        exprs.append("'U' as _Gender")

    if age_col:
        exprs.append(f"try_cast({q(age_col)} as double) as _AgeNum")
    else:
        exprs.append("NULL::DOUBLE as _AgeNum")

    if age_range_col:
        exprs.append(f"nullif(trim(coalesce(cast({q(age_range_col)} as varchar), '')), '') as _AgeRange")
    else:
        exprs.append("NULL::VARCHAR as _AgeRange")

    if reg_col:
        exprs.append(
            f"""coalesce(
                try_strptime(cast({q(reg_col)} as varchar), '%m/%d/%Y'),
                try_strptime(cast({q(reg_col)} as varchar), '%m/%d/%y'),
                try_cast({q(reg_col)} as timestamp)
            ) as _RegistrationDate"""
        )
    else:
        exprs.append("NULL::TIMESTAMP as _RegistrationDate")

    for alias, src in [("_HasEmail", email_col), ("_HasLandline", landline_col), ("_HasMobile", mobile_col)]:
        if src:
            exprs.append(
                f"""case
                    when trim(coalesce(cast({q(src)} as varchar), '')) in ('', 'None', 'NONE', 'nan', 'NAN') then false
                    else true
                end as {alias}"""
            )
        else:
            exprs.append(f"false as {alias}")

    if vote_hist_col:
        exprs.append(f"upper(trim(coalesce(cast({q(vote_hist_col)} as varchar), ''))) as _VoteHistory")
    else:
        exprs.append("'' as _VoteHistory")

    if mib_applied_col:
        exprs.append(f"case when upper(trim(coalesce(cast({q(mib_applied_col)} as varchar), ''))) = '' then 'DNA' else upper(trim(coalesce(cast({q(mib_applied_col)} as varchar), ''))) end as _MIBApplied")
    else:
        exprs.append("'DNA' as _MIBApplied")

    if mib_ballot_col:
        exprs.append(f"upper(trim(coalesce(cast({q(mib_ballot_col)} as varchar), ''))) as _MIBBallot")
    else:
        exprs.append("'' as _MIBBallot")

    if mb_score_col:
        exprs.append(f"try_cast(regexp_replace(cast({q(mb_score_col)} as varchar), '[^0-9\\.-]', '', 'g') as double) as _MBScore")
    else:
        exprs.append("NULL::DOUBLE as _MBScore")

    if mb_perm_col:
        exprs.append(f"""case
            when upper(trim(coalesce(cast({q(mb_perm_col)} as varchar), ''))) in ('TRUE', 'T', 'YES', 'Y', '1') then 'Y'
            when upper(trim(coalesce(cast({q(mb_perm_col)} as varchar), ''))) in ('FALSE', 'F', 'NO', 'N', '0') then 'N'
            else upper(trim(coalesce(cast({q(mb_perm_col)} as varchar), '')))
        end as _MBPerm""")
    else:
        exprs.append("'' as _MBPerm")

    if source_file_col:
        exprs.append(f"upper(trim(coalesce(cast({q(source_file_col)} as varchar), ''))) as _SourceFile")
    else:
        exprs.append("'' as _SourceFile")

    if mb_new_reg_col:
        exprs.append(f"""case
            when upper(trim(coalesce(cast({q(mb_new_reg_col)} as varchar), ''))) in ('Y', 'YES', 'TRUE', '1') then 'Y'
            else ''
        end as _MailBallotNewRegistrant""")
    else:
        exprs.append("'' as _MailBallotNewRegistrant")

    if applicant_phone_col:
        exprs.append(f"""case
            when trim(coalesce(cast({q(applicant_phone_col)} as varchar), '')) in ('', 'None', 'NONE', 'nan', 'NAN') then false
            else true
        end as _HasApplicantPhone""")
    else:
        exprs.append("false as _HasApplicantPhone")

    if applicant_phone_type_col:
        exprs.append(f"upper(trim(coalesce(cast({q(applicant_phone_type_col)} as varchar), ''))) as _ApplicantPhoneType")
    else:
        exprs.append("'' as _ApplicantPhoneType")

    if applicant_phone_compliance_col:
        exprs.append(f"upper(trim(coalesce(cast({q(applicant_phone_compliance_col)} as varchar), ''))) as _ApplicantPhoneCompliance")
    else:
        exprs.append("'' as _ApplicantPhoneCompliance")

    if current_app_return_col:
        exprs.append(f"""coalesce(
            try_strptime(cast({q(current_app_return_col)} as varchar), '%Y-%m-%d'),
            try_strptime(cast({q(current_app_return_col)} as varchar), '%m/%d/%Y'),
            try_strptime(cast({q(current_app_return_col)} as varchar), '%m/%d/%y'),
            try_cast({q(current_app_return_col)} as timestamp)
        ) as _CurrentAppReturnDate""")
    else:
        exprs.append("NULL::TIMESTAMP as _CurrentAppReturnDate")

    if current_ballot_sent_col:
        exprs.append(f"""coalesce(
            try_strptime(cast({q(current_ballot_sent_col)} as varchar), '%Y-%m-%d'),
            try_strptime(cast({q(current_ballot_sent_col)} as varchar), '%m/%d/%Y'),
            try_strptime(cast({q(current_ballot_sent_col)} as varchar), '%m/%d/%y'),
            try_cast({q(current_ballot_sent_col)} as timestamp)
        ) as _CurrentBallotSentDate""")
    else:
        exprs.append("NULL::TIMESTAMP as _CurrentBallotSentDate")

    if current_ballot_returned_col:
        exprs.append(f"""coalesce(
            try_strptime(cast({q(current_ballot_returned_col)} as varchar), '%Y-%m-%d'),
            try_strptime(cast({q(current_ballot_returned_col)} as varchar), '%m/%d/%Y'),
            try_strptime(cast({q(current_ballot_returned_col)} as varchar), '%m/%d/%y'),
            try_cast({q(current_ballot_returned_col)} as timestamp)
        ) as _CurrentBallotReturnedDate""")
    else:
        exprs.append("NULL::TIMESTAMP as _CurrentBallotReturnedDate")

    if hh_col:
        exprs.append(f"nullif(trim(coalesce(cast({q(hh_col)} as varchar), '')), '') as _HouseholdKey")
    else:
        parts = []
        if house_col:
            parts.append(f"coalesce(cast({q(house_col)} as varchar), '')")
        if street_col:
            parts.append(f"coalesce(cast({q(street_col)} as varchar), '')")
        if apt_col:
            parts.append(f"coalesce(cast({q(apt_col)} as varchar), '')")
        if parts:
            exprs.append("concat_ws('|', " + ", ".join(parts) + ") as _HouseholdKey")
        else:
            exprs.append("NULL::VARCHAR as _HouseholdKey")

    return "CREATE OR REPLACE VIEW voters AS SELECT\n    " + ",\n    ".join(exprs) + f"\nFROM {dataset_scan_sql(local_paths)}"

def prepare_db(local_paths):
    con = get_conn()
    raw_cols = get_schema(local_paths)
    con.execute(build_view_sql(raw_cols, local_paths))
    view_cols = con.execute("DESCRIBE SELECT * FROM voters").df()["column_name"].tolist()
    return view_cols

def sql_literal_list(values):
    return ", ".join(["?"] * len(values))

def _paths_sql(paths):
    return "[" + ", ".join(sql_string_literal(p) for p in paths) + "]"

def dataset_scan_sql(paths):
    """Return a DuckDB scan expression for local CSV/parquet or R2-downloaded parquet files."""
    if not paths:
        raise ValueError("No voter data files were found.")
    suffixes = {Path(str(p)).suffix.lower() for p in paths}
    paths_sql = _paths_sql(paths)
    if suffixes <= {".csv", ".txt"}:
        return f"read_csv_auto({paths_sql}, union_by_name=True, all_varchar=True, ignore_errors=True)"
    return f"read_parquet({paths_sql}, union_by_name=True)"

def _sorted_file_paths(folder: Path, patterns):
    if not folder.exists():
        return []
    found = []
    for pattern in patterns:
        found.extend(folder.glob(pattern))
    return [str(p) for p in sorted(set(found)) if p.is_file() and not p.name.startswith("._")]

def find_local_dataset_paths(kind: str):
    """Find local DEV data. Prefer optimized index/detail shards, then legacy shards, then CSV/R2."""
    if not USE_LOCAL_DATA:
        return [], "R2"

    # New optimized Step 8 layout:
    #   data/shards/index  = fast filter/count/chart fields
    #   data/shards/detail = full voter detail for lookup/exports
    if kind == "index":
        index_paths = _sorted_file_paths(LOCAL_DEV_INDEX_SHARDS_DIR, ["*.parquet"])
        if index_paths:
            return index_paths, "LOCAL INDEX SHARDS"

    if kind == "detail":
        detail_paths = _sorted_file_paths(LOCAL_DEV_DETAIL_SHARDS_DIR, ["*.parquet"])
        if detail_paths:
            return detail_paths, "LOCAL DETAIL SHARDS"

    # Legacy fallback: old single-folder shard layout.
    shard_paths = _sorted_file_paths(LOCAL_DEV_SHARDS_DIR, ["*.parquet"])
    if shard_paths:
        return shard_paths, "LOCAL LEGACY SHARDS"

    # Slow fallback only.
    if LOCAL_DEV_VOTERS_CSV.exists():
        return [str(LOCAL_DEV_VOTERS_CSV)], "LOCAL DEV CSV"

    if LOCAL_CANDIDATE_CONNECT_CSV.exists():
        return [str(LOCAL_CANDIDATE_CONNECT_CSV)], "LOCAL CSV"

    if LOCAL_ATHENIX_FEATURED_CSV.exists():
        return [str(LOCAL_ATHENIX_FEATURED_CSV)], "LOCAL ATHENIX FEATURED CSV"

    return [], "R2"

def local_speed_path(filename: str) -> Path:
    """Local cache path for speed tables.

    In local DEV, prefer the app's data/shards/speed folder.
    In deployed Streamlit/R2 mode, use /tmp cache under LOCAL_ROOT/speed.
    """
    local_candidate = LOCAL_DEV_SPEED_DIR / filename
    if USE_LOCAL_DATA and local_candidate.exists():
        return local_candidate
    return LOCAL_ROOT / "speed" / filename


@st.cache_data(show_spinner=False)
def ensure_speed_tables():
    """Download small R2 speed tables and manifest so startup does not scan index shards."""
    LOCAL_ROOT.mkdir(parents=True, exist_ok=True)
    manifest = load_manifest()

    # Local speed tables already exist.
    if (LOCAL_DEV_SPEED_DIR / "filter_options.parquet").exists() and (LOCAL_DEV_SPEED_DIR / "count_cube.parquet").exists():
        return manifest

    speed_tables = (manifest.get("speed") or {}).get("tables") or {}
    if not speed_tables:
        speed_tables = {
            "filter_options": "filter_options.parquet",
            "filter_ranges": "filter_ranges.json",
            "geo_hierarchy": "geo_hierarchy.parquet",
            "count_cube": "count_cube.parquet",
            "mail_ballot_counts": "mail_ballot_counts.parquet",
            "speed_manifest": "speed_manifest.json",
        }

    for _, rel in speed_tables.items():
        if not rel:
            continue
        key = str(rel)
        if not key.startswith("speed/"):
            key = "speed/" + key
        local_path = LOCAL_ROOT / key
        if not local_path.exists():
            try:
                download_public_object(key, local_path)
            except Exception:
                # Keep startup resilient if a non-critical speed table is missing.
                pass
    return manifest


def speed_tables_available() -> bool:
    try:
        ensure_speed_tables()
    except Exception:
        pass
    return local_speed_path("filter_options.parquet").exists() and local_speed_path("count_cube.parquet").exists()


@st.cache_data(show_spinner=False)
def load_speed_filter_options() -> pd.DataFrame:
    path = local_speed_path("filter_options.parquet")
    if not path.exists():
        return pd.DataFrame(columns=["field", "value", "sort_order"])
    try:
        return pd.read_parquet(path)
    except Exception:
        return pd.DataFrame(columns=["field", "value", "sort_order"])


@st.cache_data(show_spinner=False)
def load_speed_count_cube() -> pd.DataFrame:
    path = local_speed_path("count_cube.parquet")
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_parquet(path)
    except Exception:
        return pd.DataFrame()


@st.cache_data(show_spinner=False)
def load_speed_ranges() -> dict:
    path = local_speed_path("filter_ranges.json")
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def speed_option_values(field: str) -> list[str]:
    opts = load_speed_filter_options()
    if opts.empty or "field" not in opts.columns or "value" not in opts.columns:
        return []
    sub = opts[opts["field"].astype(str).eq(field)].copy()
    if sub.empty:
        return []
    if "sort_order" in sub.columns:
        sub = sub.sort_values("sort_order")
    return [str(v) for v in sub["value"].tolist() if str(v).strip() and str(v).strip() != "(Blank)"]


def _speed_active_has_unsupported_filters(active: dict) -> bool:
    unsupported_keys = [
        "hh_party_pick", "calc_party_pick", "tag_pick",
        "age_slider", "new_reg_months",
        "election_years_pick", "election_types_pick", "vote_methods_pick",
    ]
    return any(bool(active.get(k)) for k in unsupported_keys)


def _speed_filter_cube(active: dict) -> pd.DataFrame | None:
    if not speed_tables_available() or _speed_active_has_unsupported_filters(active):
        return None
    cube = load_speed_count_cube()
    if cube.empty:
        return None
    df = cube

    def apply_in(field, values):
        nonlocal df
        if values and field in df.columns:
            vals = [str(v) for v in values]
            df = df[df[field].astype(str).isin(vals)]

    for col in ["County", "Municipality", "Precinct", "USC", "STS", "STH", "School District", "School Region"]:
        apply_in(col, active.get(col, []))
    apply_in("Party", active.get("party_pick", []))
    apply_in("Gender", active.get("gender_pick", []))
    apply_in("Age_Range", active.get("age_range_pick", []))
    apply_in("MIB_Applied", active.get("mib_applied_pick", []))
    apply_in("MIB_BALLOT", active.get("mib_ballot_pick", []))
    apply_in("MB_PERM", active.get("mb_perm_pick", []))

    vh_range = active.get("vote_history_range")
    if vh_range is not None:
        vh_type = active.get("vote_history_type", "All")
        vh_field = "V4G" if vh_type == "General" else "V4P" if vh_type == "Primary" else "V4A"
        if vh_field in df.columns:
            vals = pd.to_numeric(df[vh_field].replace("(Blank)", "0"), errors="coerce").fillna(0)
            df = df[(vals >= int(vh_range[0])) & (vals <= int(vh_range[1]))]
        else:
            return None

    if active.get("current_ballot_sent_status") in {"Sent", "Not Sent/Unknown"}:
        apply_in("BallotSentStatus", [active.get("current_ballot_sent_status")])
    if active.get("current_ballot_returned_status") in {"Returned", "Not Returned/Unknown"}:
        apply_in("BallotReturnedStatus", [active.get("current_ballot_returned_status")])

    if active.get("has_email") in {"Has Email", "No Email"}:
        apply_in("HasEmail", [active.get("has_email")])
    if active.get("has_landline") in {"Has Landline", "No Landline"}:
        apply_in("HasLandline", [active.get("has_landline")])
    if active.get("has_mobile") in {"Has Mobile", "No Mobile"}:
        apply_in("HasMobile", [active.get("has_mobile")])
    if active.get("has_applicant_phone") in {"Has Applicant Phone", "No Applicant Phone"}:
        apply_in("HasApplicantPhone", [active.get("has_applicant_phone")])

    mb_score = active.get("mb_score_slider")
    if mb_score is not None and "MB_Prob_Score" in df.columns:
        vals = pd.to_numeric(df["MB_Prob_Score"].replace("(Blank)", "0"), errors="coerce").fillna(0)
        df = df[(vals >= float(mb_score[0])) & (vals <= float(mb_score[1]))]

    return df


def _speed_group_field_from_chart_expr(group_expr: str) -> str | None:
    expr = str(group_expr).strip()
    mapping = {
        "_PartyNorm": "Party",
        "_Gender": "Gender",
        "_AgeRange": "Age_Range",
        '"County"': "County",
        '"Municipality"': "Municipality",
        '"Precinct"': "Precinct",
        '"USC"': "USC",
        '"STS"': "STS",
        '"STH"': "STH",
        '"School District"': "School District",
        '"School Region"': "School Region",
    }
    if expr in mapping:
        return mapping[expr]
    unquoted = expr.strip('"')
    return unquoted if unquoted in {"County", "Municipality", "Precinct", "USC", "STS", "STH", "School District", "School Region"} else None

def clean_district_display_value(value) -> str:
    """Display USC/STS/STH without trailing .0 while preserving real text values."""
    raw = normalize_export_text(value) if "normalize_export_text" in globals() else str(value).strip()
    if raw.lower() in {"", "nan", "none", "null"}:
        return ""
    try:
        f = float(raw)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return re.sub(r"\\.0+$", "", raw)


def district_sort_key(value):
    s = clean_district_display_value(value)
    try:
        return (0, int(float(s)))
    except Exception:
        return (1, s)




def detect_election_vote_method_columns(columns):
    """Detect election-history vote-method columns from many naming styles.

    Supported examples:
    - G20, P20, G2020, P2020
    - G20_11, P22_05, G24_11
    - General_2020, Primary_2022
    - 2020_General, 2022_Primary
    - GEN20, PRI22
    - VoteMethod_G_2020, VoteMethod_Primary_2022
    """
    found = []
    seen = set()

    for col in columns:
        s = str(col).strip()
        u = re.sub(r"[^A-Z0-9]+", "_", s.upper()).strip("_")

        year = None
        etype = None

        patterns = [
            r"^([GP])_?((?:20)?\d{2})(?:_|$)",                 # G20, G20_11, P2020
            r"^(GENERAL|PRIMARY|GEN|PRI|PRIM)_?((?:20)?\d{2})(?:_|$)",
            r"^((?:20)?\d{2})_?(GENERAL|PRIMARY|GEN|PRI|PRIM)(?:_|$)",
            r"(?:^|_)([GP])_?((?:20)?\d{2})(?:_|$)",
            r"(?:^|_)(GENERAL|PRIMARY|GEN|PRI|PRIM)_?((?:20)?\d{2})(?:_|$)",
            r"(?:^|_)((?:20)?\d{2})_?(GENERAL|PRIMARY|GEN|PRI|PRIM)(?:_|$)",
        ]

        for pat in patterns:
            m = re.search(pat, u)
            if not m:
                continue

            a, b = m.group(1), m.group(2)
            if a in {"G", "P", "GENERAL", "PRIMARY", "GEN", "PRI", "PRIM"}:
                type_token = a
                yy = b
            else:
                yy = a
                type_token = b

            if type_token in {"G", "GENERAL", "GEN"}:
                etype = "General"
            elif type_token in {"P", "PRIMARY", "PRI", "PRIM"}:
                etype = "Primary"
            else:
                etype = None

            year = int(yy) if len(str(yy)) == 4 else 2000 + int(yy)
            break

        if year is None or etype is None:
            continue

        # Keep reasonable PA voter-file years.
        if year < 2000 or year > 2030:
            continue

        key = (s, year, etype)
        if key not in seen:
            seen.add(key)
            found.append({"column": s, "year": year, "type": etype})

    return sorted(found, key=lambda x: (x["year"], x["type"], x["column"]))


def election_filter_options(columns):
    elections = detect_election_vote_method_columns(columns)
    years = sorted({e["year"] for e in elections})
    types = sorted({e["type"] for e in elections}, key=lambda x: 0 if x == "General" else 1)

    # UI fallback: show standard PA cycle years even if raw method fields are not in the current index.
    # Filtering still only applies when matching columns are present.
    if not years:
        years = list(range(2000, 2027))
    if not types:
        types = ["General", "Primary"]

    return elections, years, types


def build_election_vote_method_filter_sql(active, columns):
    years = active.get("election_years_pick", []) or []
    types = active.get("election_types_pick", []) or []
    methods = active.get("vote_methods_pick", []) or []

    years = [int(y) for y in years if str(y).strip().isdigit()]
    types = [t for t in types if t and t != "All"]
    methods = [m for m in methods if m and m != "All"]

    if not years and not types and not methods:
        return "", []

    elections = detect_election_vote_method_columns(columns)
    if years:
        elections = [e for e in elections if e["year"] in years]
    if types:
        elections = [e for e in elections if e["type"] in types]

    # If the current loaded voter index does not contain election method columns yet,
    # do not crash or block the rest of the filters.
    if not elections:
        return "", []

    method_values = methods or ["AP", "MB", "P", "DNV"]
    clauses = []
    params = []

    for e in elections:
        raw = f"upper(regexp_replace(regexp_replace(trim(coalesce(cast({quote_ident(e['column'])} as varchar), '')), '[ _-]', '', 'g'), '[.]0$', ''))"
        norm = f"""case
            when {raw} in ('', 'NAN', 'NONE', 'NULL', '0', 'NO', 'N', 'DIDNOTVOTE', 'DNV') then 'DNV'
            when {raw} in ('AP', 'POLL', 'ATPOLL', 'ATPOLLS', 'INPERSON', 'ELECTIONDAY', 'ED') then 'AP'
            when {raw} in ('MB', 'MAIL', 'MAILBALLOT', 'MAILIN', 'MIB', 'ABS', 'ABSENTEE') then 'MB'
            when {raw} in ('P', 'PROV', 'PROVISIONAL') then 'P'
            when {raw} like '%MAIL%' or {raw} like '%ABSENTEE%' then 'MB'
            when {raw} like '%PROV%' then 'P'
            when {raw} like '%POLL%' or {raw} like '%PERSON%' then 'AP'
            when {raw} like '%DIDNOT%' or {raw} like '%NOTVOTE%' then 'DNV'
            else {raw}
        end"""
        clauses.append(f"{norm} IN ({sql_literal_list(method_values)})")
        params.extend(method_values)

    return " AND (" + " OR ".join(clauses) + ")", params



def current_filter_clause(active, columns):
    where = ["_Status = 'A'"]
    params = []
    geo_cols = [c for c in ["County", "Municipality", "Precinct", "USC", "STS", "STH", "School District", "School Region"] if c in columns]
    for col in geo_cols:
        picked = active.get(col, [])
        if picked:
            if col in ["USC", "STS", "STH"]:
                where.append(f"regexp_replace(trim(cast({quote_ident(col)} as varchar)), '\\.0+$', '') IN ({sql_literal_list(picked)})")
            else:
                where.append(f"{quote_ident(col)} IN ({sql_literal_list(picked)})")
            params.extend(picked)
    if active.get("party_pick"):
        picked = active["party_pick"]
        where.append(f"_PartyNorm IN ({sql_literal_list(picked)})")
        params.extend(picked)
    if active.get("hh_party_pick") and "HH-Party" in columns:
        picked = active["hh_party_pick"]
        where.append(f'{quote_ident("HH-Party")} IN ({sql_literal_list(picked)})')
        params.extend(picked)
    if active.get("calc_party_pick") and "CalculatedParty" in columns:
        picked = active["calc_party_pick"]
        where.append(f'{quote_ident("CalculatedParty")} IN ({sql_literal_list(picked)})')
        params.extend(picked)
    if active.get("tag_pick") and "Tags" in columns:
        tag_conditions = []
        for tag in active["tag_pick"]:
            tag_conditions.append("regexp_matches(coalesce(cast(" + quote_ident("Tags") + " as varchar), ''), ?)")
            params.append("(^|[,;|]\\s*)" + re.escape(str(tag).strip()) + "(\\s*[,;|]|$)")
        if tag_conditions:
            where.append("(" + " OR ".join(tag_conditions) + ")")
    if active.get("gender_pick"):
        picked = active["gender_pick"]
        where.append(f"_Gender IN ({sql_literal_list(picked)})")
        params.extend(picked)
    if active.get("age_range_pick"):
        picked = active["age_range_pick"]
        where.append(f"_AgeRange IN ({sql_literal_list(picked)})")
        params.extend(picked)
    if active.get("age_slider") is not None:
        where.append("_AgeNum >= ? AND _AgeNum <= ?")
        params.extend([active["age_slider"][0], active["age_slider"][1]])
    vote_history_type = active.get("vote_history_type", "All")
    vote_history_range = active.get("vote_history_range")

    if vote_history_range is not None:
        low, high = vote_history_range

        if vote_history_type == "General" and "V4G" in columns:
            vh_col = '"V4G"'
        elif vote_history_type == "Primary" and "V4P" in columns:
            vh_col = '"V4P"'
        elif "V4A" in columns:
            vh_col = '"V4A"'
        else:
            vh_col = None

        if vh_col:
            where.append(
                f"""
                coalesce(
                    try_cast(nullif(trim(cast({vh_col} as varchar)), '') as integer),
                    0
                ) >= ?
                AND
                coalesce(
                    try_cast(nullif(trim(cast({vh_col} as varchar)), '') as integer),
                    0
                ) <= ?
                """
            )
            params.extend([int(low), int(high)])
    if active.get("mib_applied_pick"):
        picked = active["mib_applied_pick"]
        where.append(f"_MIBApplied IN ({sql_literal_list(picked)})")
        params.extend(picked)
    if active.get("mib_ballot_pick"):
        picked = active["mib_ballot_pick"]
        where.append(f"_MIBBallot IN ({sql_literal_list(picked)})")
        params.extend(picked)
    if active.get("mb_perm_pick"):
        picked = active["mb_perm_pick"]
        where.append(f"_MBPerm IN ({sql_literal_list(picked)})")
        params.extend(picked)
    # MailBallotNewRegistrant is intentionally not applied in the main Universe builder.
    # It belongs in the dedicated Mail Ballot Center once the mail ballot is verified.
    # Applicant phone type/compliance are stored as audit fields, not user-facing filters yet.
    if active.get("has_applicant_phone") == "Has Applicant Phone":
        where.append("_HasApplicantPhone = true")
    elif active.get("has_applicant_phone") == "No Applicant Phone":
        where.append("_HasApplicantPhone = false")
    if active.get("current_ballot_sent_status") == "Sent":
        where.append("_CurrentBallotSentDate IS NOT NULL")
    elif active.get("current_ballot_sent_status") == "Not Sent/Unknown":
        where.append("_CurrentBallotSentDate IS NULL")
    if active.get("current_ballot_returned_status") == "Returned":
        where.append("_CurrentBallotReturnedDate IS NOT NULL")
    elif active.get("current_ballot_returned_status") == "Not Returned/Unknown":
        where.append("_CurrentBallotReturnedDate IS NULL")
    if active.get("mb_score_slider") is not None:
        where.append("_MBScore >= ? AND _MBScore <= ?")
        params.extend([active["mb_score_slider"][0], active["mb_score_slider"][1]])
    if active.get("new_reg_months", 0) and active.get("new_reg_months", 0) > 0:
        if "_RegistrationDate" in columns:
            where.append("_RegistrationDate >= (CURRENT_DATE - (? * INTERVAL '1 month'))")
            params.append(int(active["new_reg_months"]))
        elif "RegistrationDate" in columns:
            where.append("""coalesce(
                try_strptime(cast("RegistrationDate" as varchar), '%m/%d/%Y'),
                try_strptime(cast("RegistrationDate" as varchar), '%m/%d/%y'),
                try_cast("RegistrationDate" as timestamp)
            ) >= (CURRENT_DATE - (? * INTERVAL '1 month'))""")
            params.append(int(active["new_reg_months"]))
    if active.get("has_email") == "Has Email":
        where.append("_HasEmail = true")
    elif active.get("has_email") == "No Email":
        where.append("_HasEmail = false")
    if active.get("has_landline") == "Has Landline":
        where.append("_HasLandline = true")
    elif active.get("has_landline") == "No Landline":
        where.append("_HasLandline = false")
    if active.get("has_mobile") == "Has Mobile":
        where.append("_HasMobile = true")
    elif active.get("has_mobile") == "No Mobile":
        where.append("_HasMobile = false")

    election_sql, election_params = build_election_vote_method_filter_sql(active, columns)
    if election_sql:
        where.append(election_sql.replace(" AND ", "", 1))
        params.extend(election_params)

    return " WHERE " + " AND ".join(where), params

def get_distinct_options(column: str, label_expr: str | None = None):
    con = get_conn()
    expr = label_expr or quote_ident(column)
    df = con.execute(
        f"""
        SELECT {expr} AS value
        FROM voters
        WHERE _Status = 'A' AND nullif(trim(cast({quote_ident(column)} as varchar)), '') IS NOT NULL
        GROUP BY 1
        ORDER BY 1
        """
    ).df()
    return [str(v) for v in df["value"].tolist() if str(v).strip() != ""]


def get_dependent_geo_options(target_col: str, selected_geo: dict, columns):
    """Return narrowed geography options from already-selected geography values.

    This helper is safe: it only narrows option lists and does not mutate
    Streamlit widget/session state. Apply Filters and Clear Filters remain
    controlled by the original form submit logic.
    """
    con = get_conn()
    where = ["_Status = 'A'"]
    params = []

    for col, picked in selected_geo.items():
        if col == target_col or not picked or col not in columns:
            continue
        if col in ["USC", "STS", "STH"]:
            where.append(f"regexp_replace(trim(cast({quote_ident(col)} as varchar)), '\\.0+$', '') IN ({sql_literal_list(picked)})")
        else:
            where.append(f"{quote_ident(col)} IN ({sql_literal_list(picked)})")
        params.extend(picked)

    if target_col in ["USC", "STS", "STH"]:
        expr = f"regexp_replace(trim(cast({quote_ident(target_col)} as varchar)), '\\.0+$', '')"
    else:
        expr = quote_ident(target_col)

    try:
        df = con.execute(
            f"""
            SELECT {expr} AS value
            FROM voters
            WHERE {" AND ".join(where)}
              AND nullif(trim(cast({quote_ident(target_col)} as varchar)), '') IS NOT NULL
            GROUP BY 1
            ORDER BY 1
            """,
            params,
        ).df()
        vals = [str(v).strip() for v in df["value"].tolist() if str(v).strip()]
        if target_col in ["USC", "STS", "STH"]:
            vals = sorted(set(clean_district_display_value(v) for v in vals if clean_district_display_value(v)), key=district_sort_key)
        else:
            vals = sorted(set(vals))
        return vals
    except Exception:
        return []


def clean_geo_option_list(col: str, values):
    cleaned = []
    for v in values:
        s = normalize_export_text(v).strip() if "normalize_export_text" in globals() else str(v).strip()
        if not s or s.lower() in {"nan", "none", "null", "blank", "(blank)"}:
            continue

        # Keep school district filter locked to the official PA lookup list only.
        if col == "School District":
            sd = normalize_school_district_option(s)
            if sd:
                cleaned.append(sd)
            continue

        # Do not let region-like labels appear as School Region unless they are real text.
        if col == "School Region":
            if re.fullmatch(r"\d+(?:\.0+)?", s):
                continue
            s = geo_display_case_py(s)
            if s:
                cleaned.append(s)
            continue

        if col == "Municipality" and re.fullmatch(r"\d+(?:\.0+)?", s):
            continue

        # Candidate Connect UI should show readable precinct names only.
        # Hide raw precinct codes like 001, 001A, 001B, etc.
        if col == "Precinct" and re.fullmatch(r"[0-9A-Z]{1,6}", s.upper()):
            continue

        if col in {"USC", "STS", "STH"}:
            s = clean_district_display_value(s)
            if not re.fullmatch(r"\d{1,3}", s):
                continue

        if col in {"County", "Municipality", "Precinct"}:
            s = geo_display_case_py(s)

        if s:
            cleaned.append(s)

    if col in {"USC", "STS", "STH"}:
        return sorted(set(cleaned), key=district_sort_key)
    return sorted(set(cleaned), key=lambda x: str(x).lower())



def _mb_clean_options(values, field=None):
    """Clean Mail Ballot Center dropdown values so raw/current-file junk does not leak into UI."""
    field_key = str(field or "").strip().upper()
    bad_common = {"", "(BLANK)", "NAN", "NONE", "NULL", "<NA>"}
    bad_ballot_detail = {"1", "0", "TRUE", "FALSE", "VOTE RECORDED", "VOTED", "V", "PENDING", "CANCELLED", "CANCELED"}
    bad_application = {"1", "0", "TRUE", "FALSE", "VOTE RECORDED"}
    cleaned = []
    seen = set()
    for v in values or []:
        s = str(v).strip()
        u = s.upper().strip()
        if u in bad_common:
            continue
        if field_key in {"MIB_BALLOT", "BALLOT"} and u in bad_ballot_detail:
            continue
        if field_key in {"MIB_APPLIED", "APPLICATION"} and u in bad_application:
            continue
        if field_key in {"MB_PERM", "PERMANENT", "PERMANENT_MB"}:
            if u in {"TRUE", "T", "YES", "Y", "1"}:
                s, u = "Y", "Y"
            elif u in {"FALSE", "F", "NO", "N", "0"}:
                s, u = "N", "N"
            else:
                # Guardrail: the raw/speed option table can sometimes carry unrelated
                # geo labels here. Permanent MB must only ever be Y/N.
                continue
        if u not in seen:
            cleaned.append(s)
            seen.add(u)
    return cleaned


MB_CONTROLLED_FILTER_KEYS = {
    "mib_applied_pick", "mib_ballot_pick", "mb_perm_pick",
    "has_applicant_phone", "current_ballot_sent_status", "current_ballot_returned_status",
    "mb_score_slider", "has_email", "has_mobile", "has_landline",
}

def _mb_strip_filters(filters):
    """Remove only Mail Ballot Center filters, leaving geography/party/voter universe filters intact."""
    cleaned = dict(filters or {})
    for key in MB_CONTROLLED_FILTER_KEYS:
        cleaned.pop(key, None)
    return cleaned

def _mb_clear_center_state(clear_main=False):
    """Clear MB Center controls and optional MB filters previously pushed into main Universe."""
    st.session_state.mail_ballot_center_filters = {}
    st.session_state.mail_ballot_center_use_main_universe = True
    for export_key in [
        "mb_filtered_csv_df", "mb_filtered_excel_bytes", "mb_texting_csv_df",
        "mb_mail_csv_df", "mb_labels_pdf_bytes",
    ]:
        st.session_state.pop(export_key, None)
    if clear_main:
        st.session_state.active_filters = _mb_strip_filters(st.session_state.get("active_filters", {}))
        st.session_state.filters_applied = bool(st.session_state.get("active_filters"))


def _mb_clean_application_options(values):
    opts = _mb_clean_options(values, field="MIB_Applied")
    for required in ["APP", "DNA"]:
        if required not in {str(x).upper() for x in opts}:
            opts.append(required)
    return sorted(opts, key=lambda x: (0 if str(x).upper() in {"APP", "DNA"} else 1, str(x).upper()))


def _mb_clean_ballot_detail_options(values):
    return _mb_clean_options(values, field="MIB_BALLOT")

def get_basic_options(columns):
    options = {}
    geo_cols = [c for c in ["County", "Municipality", "Precinct", "USC", "STS", "STH", "School District", "School Region"] if c in columns]

    if speed_tables_available():
        for col in geo_cols:
            options[col] = clean_geo_option_list(col, speed_option_values(col))
        options["party_vals"] = speed_option_values("Party") or ["D", "R", "O"]
        options["gender_vals"] = speed_option_values("Gender")
        options["age_range_vals"] = speed_option_values("Age_Range")
        options["hh_party_vals"] = speed_option_values("HH-Party") if "HH-Party" in columns else []
        options["calc_party_vals"] = speed_option_values("CalculatedParty") if "CalculatedParty" in columns else []
        options["tag_vals"] = ordered_tag_values(speed_option_values("Tags")) if "Tags" in columns else []
        options["vote_history_vals"] = ordered_vote_history_values(speed_option_values("V4A"))
        options["mib_applied_vals"] = _mb_clean_application_options(speed_option_values("MIB_Applied") or [])
        options["mib_ballot_vals"] = _mb_clean_ballot_detail_options(speed_option_values("MIB_BALLOT"))
        options["mb_perm_vals"] = _mb_clean_options(speed_option_values("MB_PERM"), field="MB_PERM")
        options["mb_new_reg_vals"] = []

        ranges = load_speed_ranges()
        age_range = ranges.get("Age") or ranges.get("Age_Calc") or {}
        score_range = ranges.get("MB_Prob_Score") or {}
        options["age_min"] = int(age_range.get("min")) if age_range.get("min") is not None else None
        options["age_max"] = int(age_range.get("max")) if age_range.get("max") is not None else None
        options["mb_score_min"] = float(score_range.get("min")) if score_range.get("min") is not None else None
        options["mb_score_max"] = float(score_range.get("max")) if score_range.get("max") is not None else None
        return options

    for col in geo_cols:
        if col in ["USC", "STS", "STH"]:
            vals = get_distinct_options(col, f"regexp_replace(trim(cast({quote_ident(col)} as varchar)), '\\.0+$', '')")
        else:
            vals = get_distinct_options(col)
        options[col] = clean_geo_option_list(col, vals)
    # Party filter should always use the normalized D/R/O field.
    # _PartyNorm exists even when the raw Party column is missing or aliased.
    options["party_vals"] = get_distinct_options("_PartyNorm", "_PartyNorm") if "_PartyNorm" in columns else []
    options["gender_vals"] = get_distinct_options("_Gender", "_Gender")
    options["age_range_vals"] = get_distinct_options("_AgeRange", "_AgeRange")
    options["hh_party_vals"] = get_distinct_options("HH-Party") if "HH-Party" in columns else []
    options["calc_party_vals"] = get_distinct_options("CalculatedParty") if "CalculatedParty" in columns else []
    options["tag_vals"] = ordered_tag_values(get_distinct_options("Tags")) if "Tags" in columns else []
    options["vote_history_vals"] = ordered_vote_history_values(get_distinct_options("_VoteHistory", "_VoteHistory")) if "V4A" in columns else []
    options["mib_applied_vals"] = _mb_clean_application_options(get_distinct_options("_MIBApplied", "_MIBApplied") or [])
    options["mib_ballot_vals"] = _mb_clean_ballot_detail_options(get_distinct_options("_MIBBallot", "_MIBBallot"))
    options["mb_perm_vals"] = _mb_clean_options(get_distinct_options("_MBPerm", "_MBPerm"), field="MB_PERM")
    options["mb_new_reg_vals"] = get_distinct_options("_MailBallotNewRegistrant", "_MailBallotNewRegistrant") if "_MailBallotNewRegistrant" in columns else []

    # If speed tables are not available and the DuckDB voters table has not been prepared,
    # do not crash or try to scan remote shards during startup.
    try:
        con = get_conn()
        con.execute("SELECT 1 FROM voters LIMIT 1").fetchone()
    except Exception:
        options.setdefault("age_min", None)
        options.setdefault("age_max", None)
        options.setdefault("mb_score_min", None)
        options.setdefault("mb_score_max", None)
        return options

    age_min, age_max = con.execute(
        "SELECT min(_AgeNum), max(_AgeNum) FROM voters WHERE _Status = 'A' AND _AgeNum IS NOT NULL"
    ).fetchone()
    score_min, score_max = con.execute(
        "SELECT min(_MBScore), max(_MBScore) FROM voters WHERE _Status = 'A' AND _MBScore IS NOT NULL"
    ).fetchone()
    options["age_min"] = int(age_min) if age_min is not None else None
    options["age_max"] = int(age_max) if age_max is not None else None
    options["mb_score_min"] = float(score_min) if score_min is not None else None
    options["mb_score_max"] = float(score_max) if score_max is not None else None
    return options

def query_metrics(active, columns):
    if has_global_followup_filters(active):
        return _query_metrics_from_detail(active, columns)

    try:
        speed_df = _speed_filter_cube(active)
        if speed_df is not None and "Voters" in speed_df.columns:
            voters = safe_int(speed_df["Voters"].sum())
            emails = safe_int(speed_df["Emails"].sum()) if "Emails" in speed_df.columns else 0
            landlines = safe_int(speed_df["Landlines"].sum()) if "Landlines" in speed_df.columns else 0
            mobiles = safe_int(speed_df["Mobiles"].sum()) if "Mobiles" in speed_df.columns else 0
            if "County" in speed_df.columns:
                county_vals = speed_df.loc[speed_df["Voters"].fillna(0).astype(float) > 0, "County"].replace("(Blank)", "").astype(str).str.strip()
                unique_counties = int(county_vals[county_vals.ne("")].nunique())
            else:
                unique_counties = 0
            if "Precinct" in speed_df.columns:
                precinct_vals = speed_df.loc[speed_df["Voters"].fillna(0).astype(float) > 0, "Precinct"].replace("(Blank)", "").astype(str).str.strip()
                unique_precincts = int(precinct_vals[precinct_vals.ne("")].nunique())
            else:
                unique_precincts = 0
            return {
                "voters": voters,
                "households": None,
                "emails": emails,
                "landlines": landlines,
                "mobiles": mobiles,
                "unique_counties": unique_counties,
                "unique_precincts": unique_precincts,
                "speed_mode": True,
            }
    except Exception:
        pass

    con = get_conn()
    where_sql, params = current_filter_clause(active, columns)
    return con.execute(
        f"""
        SELECT
            count(*) AS voters,
            (
                count(DISTINCT _HouseholdKey) FILTER (WHERE _HouseholdKey IS NOT NULL AND _HouseholdKey <> '')
                + count(*) FILTER (WHERE _HouseholdKey IS NULL OR _HouseholdKey = '')
            ) AS households,
            sum(CASE WHEN _HasEmail THEN 1 ELSE 0 END) AS emails,
            sum(CASE WHEN _HasLandline THEN 1 ELSE 0 END) AS landlines,
            sum(CASE WHEN _HasMobile THEN 1 ELSE 0 END) AS mobiles,
            count(DISTINCT {quote_ident("County")}) FILTER (WHERE {quote_ident("County")} IS NOT NULL) AS unique_counties,
            count(DISTINCT {quote_ident("Precinct")}) FILTER (WHERE {quote_ident("Precinct")} IS NOT NULL) AS unique_precincts
        FROM voters
        {where_sql}
        """,
        params,
    ).df().iloc[0].to_dict()

def query_chart(active, columns, group_expr, label, not_blank=True):
    if has_global_followup_filters(active):
        return _query_chart_from_detail(active, group_expr, label, not_blank=not_blank)

    try:
        group_field = _speed_group_field_from_chart_expr(group_expr)
        speed_df = _speed_filter_cube(active)
        if speed_df is not None and group_field and group_field in speed_df.columns and "Voters" in speed_df.columns:
            temp = speed_df.copy()
            if not_blank:
                temp = temp[temp[group_field].fillna("").astype(str).str.strip().ne("") & temp[group_field].astype(str).ne("(Blank)")]
            out = temp.groupby(group_field, dropna=False, as_index=False)["Voters"].sum()
            out = out.rename(columns={group_field: label, "Voters": "Count"})
            return out.sort_values(["Count", label], ascending=[False, True]).reset_index(drop=True)
    except Exception:
        pass

    con = get_conn()
    where_sql, params = current_filter_clause(active, columns)
    extra = f" AND {group_expr} IS NOT NULL AND cast({group_expr} as varchar) <> ''" if not_blank else ""
    return con.execute(
        f"""
        SELECT {group_expr} AS "{label}", count(*) AS "Count"
        FROM voters
        {where_sql}
        {extra}
        GROUP BY 1
        ORDER BY 2 DESC, 1
        """,
        params,
    ).df()

def query_area_summary(active, columns, area_col):
    if has_global_followup_filters(active):
        return _query_area_summary_from_detail(active, area_col)

    try:
        speed_df = _speed_filter_cube(active)
        if speed_df is not None and area_col in speed_df.columns and "Voters" in speed_df.columns:
            temp = speed_df.copy()
            temp[area_col] = temp[area_col].fillna("(Blank)").astype(str).replace("", "(Blank)")
            out = temp.groupby(area_col, dropna=False, as_index=False)["Voters"].sum()
            out = out.rename(columns={area_col: area_col, "Voters": "Individuals"})
            out["Households"] = "—"
            return out.sort_values(["Individuals", area_col], ascending=[False, True]).reset_index(drop=True)
    except Exception:
        pass

    con = get_conn()
    where_sql, params = current_filter_clause(active, columns)
    return con.execute(
        f"""
        SELECT
            coalesce(cast({quote_ident(area_col)} as varchar), '(Blank)') AS "{area_col}",
            count(*) AS Individuals,
            (
                count(DISTINCT _HouseholdKey) FILTER (WHERE _HouseholdKey IS NOT NULL AND _HouseholdKey <> '')
                + count(*) FILTER (WHERE _HouseholdKey IS NULL OR _HouseholdKey = '')
            ) AS Households
        FROM voters
        {where_sql}
        GROUP BY 1
        ORDER BY Individuals DESC, 1
        """,
        params,
    ).df()



def build_statewide_summary_report_bytes(active_filters, columns):
    con = get_conn()
    where_sql, params = current_filter_clause(active_filters, columns)

    def grouped_summary(group_col: str, label: str):
        if group_col not in columns:
            return pd.DataFrame(columns=[label, "Voters", "Households", "Democrats", "Republicans", "Others", "Male", "Female", "Unknown Gender", "MIB Applied", "MIB Declined", "Did Not Apply", "Not Sent", "Not Voted", "Voted", "Permanent Mail", "Emails", "Mobiles"])
        qcol = quote_ident(group_col)
        return con.execute(
            f"""
            SELECT
                coalesce(cast({qcol} as varchar), '(Blank)') AS "{label}",
                count(*) AS "Voters",
                (
                    count(DISTINCT _HouseholdKey) FILTER (WHERE _HouseholdKey IS NOT NULL AND _HouseholdKey <> '')
                    + count(*) FILTER (WHERE _HouseholdKey IS NULL OR _HouseholdKey = '')
                ) AS "Households",
                sum(CASE WHEN _PartyNorm = 'D' THEN 1 ELSE 0 END) AS "Democrats",
                sum(CASE WHEN _PartyNorm = 'R' THEN 1 ELSE 0 END) AS "Republicans",
                sum(CASE WHEN _PartyNorm NOT IN ('D','R') THEN 1 ELSE 0 END) AS "Others",
                sum(CASE WHEN _Gender = 'M' THEN 1 ELSE 0 END) AS "Male",
                sum(CASE WHEN _Gender = 'F' THEN 1 ELSE 0 END) AS "Female",
                sum(CASE WHEN _Gender NOT IN ('M','F') THEN 1 ELSE 0 END) AS "Unknown Gender",
                sum(CASE WHEN _MIBApplied = 'APP' THEN 1 ELSE 0 END) AS "MIB Applied",
                sum(CASE WHEN _MIBApplied = 'DEC' THEN 1 ELSE 0 END) AS "MIB Declined",
                sum(CASE WHEN _MIBApplied = 'DNA' THEN 1 ELSE 0 END) AS "Did Not Apply",
                sum(CASE WHEN _MIBBallot = 'NS' THEN 1 ELSE 0 END) AS "Not Sent",
                sum(CASE WHEN _MIBBallot = 'NV' THEN 1 ELSE 0 END) AS "Not Voted",
                sum(CASE WHEN _MIBBallot = 'V' THEN 1 ELSE 0 END) AS "Voted",
                sum(CASE WHEN _MBPerm = 'Y' THEN 1 ELSE 0 END) AS "Permanent Mail",
                sum(CASE WHEN _HasEmail THEN 1 ELSE 0 END) AS "Emails",
                sum(CASE WHEN _HasMobile THEN 1 ELSE 0 END) AS "Mobiles"
            FROM voters
            {where_sql}
            GROUP BY 1
            ORDER BY "Voters" DESC, 1
            """,
            params,
        ).df()

    overview = con.execute(
        f"""
        SELECT
            count(*) AS "Voters",
            (
                count(DISTINCT _HouseholdKey) FILTER (WHERE _HouseholdKey IS NOT NULL AND _HouseholdKey <> '')
                + count(*) FILTER (WHERE _HouseholdKey IS NULL OR _HouseholdKey = '')
            ) AS "Households",
            sum(CASE WHEN _PartyNorm = 'D' THEN 1 ELSE 0 END) AS "Democrats",
            sum(CASE WHEN _PartyNorm = 'R' THEN 1 ELSE 0 END) AS "Republicans",
            sum(CASE WHEN _PartyNorm NOT IN ('D','R') THEN 1 ELSE 0 END) AS "Others",
            sum(CASE WHEN _Gender = 'M' THEN 1 ELSE 0 END) AS "Male",
            sum(CASE WHEN _Gender = 'F' THEN 1 ELSE 0 END) AS "Female",
            sum(CASE WHEN _Gender NOT IN ('M','F') THEN 1 ELSE 0 END) AS "Unknown Gender",
            sum(CASE WHEN _MIBApplied = 'APP' THEN 1 ELSE 0 END) AS "MIB Applied",
            sum(CASE WHEN _MIBApplied = 'DEC' THEN 1 ELSE 0 END) AS "MIB Declined",
            sum(CASE WHEN _MIBApplied = 'DNA' THEN 1 ELSE 0 END) AS "Did Not Apply",
            sum(CASE WHEN _MIBBallot = 'NS' THEN 1 ELSE 0 END) AS "Not Sent",
            sum(CASE WHEN _MIBBallot = 'NV' THEN 1 ELSE 0 END) AS "Not Voted",
            sum(CASE WHEN _MIBBallot = 'V' THEN 1 ELSE 0 END) AS "Voted",
            sum(CASE WHEN _MBPerm = 'Y' THEN 1 ELSE 0 END) AS "Permanent Mail",
            sum(CASE WHEN _HasEmail THEN 1 ELSE 0 END) AS "Emails",
            sum(CASE WHEN _HasMobile THEN 1 ELSE 0 END) AS "Mobiles",
            count(DISTINCT "County") FILTER (WHERE "County" IS NOT NULL) AS "Unique Counties",
            count(DISTINCT "Precinct") FILTER (WHERE "Precinct" IS NOT NULL) AS "Unique Precincts"
        FROM voters
        {where_sql}
        """,
        params,
    ).df()

    filter_df = pd.DataFrame({"Applied Universe Filters": build_filter_summary_lines(active_filters)})

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Overview", index=False)
        filter_df.to_excel(writer, sheet_name="Filters", index=False)
        for group_col, label in [("County", "County"), ("USC", "Congressional"), ("STS", "State Senate"), ("STH", "State House")]:
            grouped_summary(group_col, label).to_excel(writer, sheet_name=label[:31], index=False)
    return output.getvalue()

def fmt_pct(v: float) -> str:
    rounded = round(v, 1)
    return f"{int(rounded)}%" if float(rounded).is_integer() else f"{rounded:.1f}%"

def make_summary_table(df_chart: pd.DataFrame, label_col: str, value_col: str, colors):
    total = pd.to_numeric(df_chart[value_col], errors="coerce").fillna(0).sum()
    headers = "<tr><th></th><th>{}</th><th>{}</th><th>%</th></tr>".format(label_col, value_col)
    rows = []
    for i, (_, row) in enumerate(df_chart.iterrows()):
        val = float(pd.to_numeric(row[value_col], errors="coerce"))
        pct = 0 if total == 0 else (val / total) * 100
        color = colors[i] if i < len(colors) else CC_THEME["text_muted"]
        rows.append(
            f"<tr><td class='num-cell'><span class='cc-swatch' style='background:{color};'></span></td>"
            f"<td class='label-cell'>{row[label_col]}</td><td class='num-cell'>{val:,.0f}</td><td class='num-cell'>{fmt_pct(pct)}</td></tr>"
        )
    rows.append(f"<tr class='total-row'><td></td><td class='label-cell'>Total</td><td class='num-cell'>{total:,.0f}</td><td class='num-cell'>100%</td></tr>")
    return f"<table class='cc-mini-table'><thead>{headers}</thead><tbody>{''.join(rows)}</tbody></table>"

def pie_chart_with_table(df_chart: pd.DataFrame, label_col: str, value_col: str, title: str, color_mode: str):
    st.markdown(f'<div class="small-header">{title}</div>', unsafe_allow_html=True)
    if df_chart.empty:
        st.caption("No data")
        return
    chart_df = df_chart.copy()
    chart_df[value_col] = pd.to_numeric(chart_df[value_col], errors="coerce").fillna(0)
    chart_df = chart_df.sort_values(value_col, ascending=False).reset_index(drop=True)
    total = chart_df[value_col].sum()
    chart_df["Percent"] = 0 if total == 0 else (chart_df[value_col] / total) * 100
    domain = chart_df[label_col].astype(str).tolist()
    if color_mode == "party":
        colors = [PARTY_COLOR_MAP.get(v, CC_THEME["text_muted"]) for v in domain]
    elif color_mode == "age":
        colors = AGE_COLOR_RANGE[:len(domain)]
    else:
        colors = GENDER_COLOR_RANGE[:len(domain)]
    chart = alt.Chart(chart_df).mark_arc(innerRadius=34, outerRadius=92).encode(
        theta=alt.Theta(field=value_col, type="quantitative"),
        color=alt.Color(field=label_col, type="nominal", scale=alt.Scale(domain=domain, range=colors), legend=None),
        tooltip=[alt.Tooltip(f"{label_col}:N"), alt.Tooltip(f"{value_col}:Q", format=","), alt.Tooltip("Percent:Q", format=".1f")]
    ).properties(height=280)
    st.altair_chart(chart, width="stretch")
    st.markdown(make_summary_table(chart_df, label_col, value_col, colors), unsafe_allow_html=True)


def normalize_export_text(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def normalize_numeric_string(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.lower() in {"nan", "none", ""}:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s


def safe_int(val) -> int:
    try:
        if val is None:
            return 0
        if isinstance(val, str) and val.strip().lower() in {"", "nan", "none"}:
            return 0
        if pd.isna(val):
            return 0
        return int(float(val))
    except Exception:
        return 0

def get_filtered_voter_count_fast(active_filters, columns) -> int:
    try:
        speed_df = _speed_filter_cube(active_filters)
        if speed_df is not None and "Voters" in speed_df.columns:
            return safe_int(speed_df["Voters"].sum())
    except Exception:
        pass

    con = get_conn()
    where_sql, params = current_filter_clause(active_filters, columns)
    try:
        row = con.execute(f"SELECT count(*) AS n FROM voters {where_sql}", params).fetchone()
        return safe_int(row[0] if row else 0)
    except Exception:
        return 0

# Large-universe guard
# The old threshold was 50,000, which was too strict now that the app is using
# speed tables + local/R2 shards. Full State Senate / State House / Congressional
# districts should be usable for charts and exports. Keep the guard only for
# truly massive universes, such as statewide or near-statewide selections.
LARGE_UNIVERSE_ROW_LIMIT = 1_000_000

def use_large_filter_mode(active_filters, columns) -> bool:
    try:
        return get_filtered_voter_count_fast(active_filters, columns) >= LARGE_UNIVERSE_ROW_LIMIT
    except Exception:
        return False



def clean_zip_value(val):
    s = normalize_numeric_string(val)
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if len(digits) == 9:
        return f"{digits[:5]}-{digits[5:]}"
    if len(digits) >= 5:
        return digits[:5]
    return digits

def clean_phone_value(val):
    s = normalize_numeric_string(val)
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits


USPS_SUFFIX_MAP = {
    "STREET": "ST", "ST": "ST",
    "ROAD": "RD", "RD": "RD",
    "AVENUE": "AVE", "AVE": "AVE",
    "DRIVE": "DR", "DR": "DR",
    "LANE": "LN", "LN": "LN",
    "COURT": "CT", "CT": "CT",
    "CIRCLE": "CIR", "CIR": "CIR",
    "BOULEVARD": "BLVD", "BLVD": "BLVD",
    "PLACE": "PL", "PL": "PL",
    "TERRACE": "TER", "TER": "TER",
    "PARKWAY": "PKWY", "PKWY": "PKWY",
    "HIGHWAY": "HWY", "HWY": "HWY",
    "MOUNT": "MT", "MT": "MT",
}
STATE_ABBR = {
    "ALABAMA":"AL","ALASKA":"AK","ARIZONA":"AZ","ARKANSAS":"AR","CALIFORNIA":"CA","COLORADO":"CO",
    "CONNECTICUT":"CT","DELAWARE":"DE","FLORIDA":"FL","GEORGIA":"GA","HAWAII":"HI","IDAHO":"ID",
    "ILLINOIS":"IL","INDIANA":"IN","IOWA":"IA","KANSAS":"KS","KENTUCKY":"KY","LOUISIANA":"LA",
    "MAINE":"ME","MARYLAND":"MD","MASSACHUSETTS":"MA","MICHIGAN":"MI","MINNESOTA":"MN","MISSISSIPPI":"MS",
    "MISSOURI":"MO","MONTANA":"MT","NEBRASKA":"NE","NEVADA":"NV","NEW HAMPSHIRE":"NH","NEW JERSEY":"NJ",
    "NEW MEXICO":"NM","NEW YORK":"NY","NORTH CAROLINA":"NC","NORTH DAKOTA":"ND","OHIO":"OH","OKLAHOMA":"OK",
    "OREGON":"OR","PENNSYLVANIA":"PA","RHODE ISLAND":"RI","SOUTH CAROLINA":"SC","SOUTH DAKOTA":"SD",
    "TENNESSEE":"TN","TEXAS":"TX","UTAH":"UT","VERMONT":"VT","VIRGINIA":"VA","WASHINGTON":"WA",
    "WEST VIRGINIA":"WV","WISCONSIN":"WI","WYOMING":"WY","DISTRICT OF COLUMBIA":"DC"
}
NAME_SUFFIXES = {"JR","SR","II","III","IV","V"}

def collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", normalize_export_text(value)).strip()

def proper_case_word(word: str) -> str:
    if not word:
        return ""
    up = word.upper()
    if up in NAME_SUFFIXES:
        return up
    if re.fullmatch(r"[A-Z]\.", up):
        return up
    if "'" in word:
        return "'".join(part.capitalize() if part else "" for part in word.lower().split("'"))
    if "-" in word:
        return "-".join(part.capitalize() if part else "" for part in word.lower().split("-"))
    return word.lower().capitalize()

def normalize_name_value(value: str) -> str:
    s = collapse_spaces(value)
    if not s:
        return ""
    return " ".join(proper_case_word(part) for part in s.split(" "))

def normalize_city_value(value: str) -> str:
    s = collapse_spaces(value)
    if not s:
        return ""
    return " ".join(proper_case_word(part) for part in s.split(" "))

def normalize_state_value(value: str) -> str:
    s = collapse_spaces(value).upper()
    if not s:
        return ""
    if len(s) == 2 and s.isalpha():
        return s
    return STATE_ABBR.get(s, s[:2] if len(s) >= 2 else s)

def normalize_address_value(value: str) -> str:
    s = collapse_spaces(value)
    if not s:
        return ""

    s = re.sub(r"\bApartment\b", "Apt", s, flags=re.IGNORECASE)
    s = re.sub(r"\bSuite\b", "Ste", s, flags=re.IGNORECASE)
    s = re.sub(r"\bUnit\b", "Unit", s, flags=re.IGNORECASE)

    words = s.split(" ")
    words = [proper_case_word(w) for w in words]

    if words:
        last = re.sub(r"[^A-Za-z]", "", words[-1]).upper()
        if last in USPS_SUFFIX_MAP:
            words[-1] = USPS_SUFFIX_MAP[last].title()

    return " ".join(words)


def title_case_report_value(value: str) -> str:
    s = collapse_spaces(value)
    if not s:
        return ""
    return " ".join(proper_case_word(part) for part in s.split(" "))

def normalize_mail_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "Name" in out.columns:
        out["Name"] = out["Name"].apply(normalize_name_value)
    if "Address1" in out.columns:
        out["Address1"] = out["Address1"].apply(normalize_address_value)
    if "City" in out.columns:
        out["City"] = out["City"].apply(normalize_city_value)
    if "State" in out.columns:
        out["State"] = out["State"].apply(normalize_state_value)
    if "Zip" in out.columns:
        out["Zip"] = out["Zip"].apply(clean_zip_value)
    return out

def normalize_filtered_export_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in ["FirstName", "MiddleName", "LastName", "FullName", "Name", "NameSuffix"]:
        if col in out.columns:
            out[col] = out[col].apply(normalize_name_value)
    for col in ["Street Name", "Address", "Address1", "Mailing Address", "MailAddress"]:
        if col in out.columns:
            out[col] = out[col].apply(normalize_address_value)
    for col in ["City", "MailingCity", "Mailing City", "MailCity"]:
        if col in out.columns:
            out[col] = out[col].apply(normalize_city_value)
    for col in ["State", "MailingState", "Mailing State", "MailState"]:
        if col in out.columns:
            out[col] = out[col].apply(normalize_state_value)
    for col in ["Zip", "ZIP", "ZipCode", "ZIPCODE", "MailingZip", "Mailing Zip", "MailZip"]:
        if col in out.columns:
            out[col] = out[col].apply(clean_zip_value)
    for col in ["PrimaryPhone", "Mobile", "Landline"]:
        if col in out.columns:
            out[col] = out[col].apply(clean_phone_value)
    return out

def safe_group_series(group: pd.DataFrame, column_name: str) -> pd.Series:
    if column_name not in group.columns:
        return pd.Series([""] * len(group), index=group.index, dtype="object")
    data = group[column_name]
    if isinstance(data, pd.DataFrame):
        data = data.iloc[:, 0]
    return data.fillna("").astype(str).str.strip()

def vote_history_sort_key(value: str):
    s = normalize_export_text(value).upper()
    digits = re.findall(r"\d+", s)
    if digits:
        return (0, int(digits[0]), s)
    return (1, 9999, s)

def ordered_vote_history_values(values):
    cleaned = [normalize_export_text(v) for v in values if normalize_export_text(v) != ""]
    return sorted(cleaned, key=vote_history_sort_key)

def split_tag_values(value):
    s = normalize_export_text(value)
    if not s:
        return []
    return [part.strip() for part in re.split(r"[,;|]", s) if part.strip()]

def ordered_tag_values(values):
    seen = {}
    for raw in values:
        for tag in split_tag_values(raw):
            key = tag.lower()
            if key not in seen:
                seen[key] = tag
    return sorted(seen.values(), key=lambda x: x.lower())


def build_household_mail_name(group: pd.DataFrame) -> str:
    """Create campaign-mail household names.

    Rules:
    - same last name, multiple people: "LastName Household"
    - exactly two different last names: "Full Name1 & Full Name2"
    - more than two people with different last names: "Full Name1 & Family"
    """
    names = safe_group_series(group, "Name")
    names = [normalize_name_value(x) for x in names.tolist() if normalize_name_value(x)]
    if len(names) == 0:
        return "Current Resident"
    if len(names) == 1:
        return names[0]

    last_names = safe_group_series(group, "LastName")
    unique_last = sorted({normalize_name_value(x) for x in last_names.tolist() if normalize_name_value(x)}, key=lambda x: x.lower())
    if len(unique_last) == 1:
        return f"{unique_last[0]} Household"

    full_names = []
    seen = set()
    for name in names:
        key = name.lower()
        if key not in seen:
            full_names.append(name)
            seen.add(key)

    if len(full_names) == 2:
        return f"{full_names[0]} & {full_names[1]}"
    if len(full_names) > 2:
        return f"{full_names[0]} & Family"

    return "Current Resident"

def full_name_from_row(row):
    """Build voter name for street lists, walk lists, and PDF reports using robust fallbacks."""
    def pick(candidates):
        for col in candidates:
            try:
                val = normalize_export_text(row.get(col, ""))
            except Exception:
                val = ""
            if val:
                return val
        return ""

    # Prefer complete name fields if present.
    full = pick([
        "FullName", "Full Name", "VoterName", "Voter Name", "Name",
        "DisplayName", "Display Name", "_LookupName",
        "CURRENT_NAME", "Current_Name", "Current Name",
        "voter_name", "full_name", "name"
    ])
    if full:
        return normalize_name_value(full) if "normalize_name_value" in globals() else full

    first = pick([
        "FirstName", "First Name", "FIRSTNAME", "First", "FName", "FNAME",
        "GivenName", "Given Name", "Current_FirstName", "Current First Name",
        "first_name", "firstname"
    ])
    middle = pick([
        "MiddleName", "Middle Name", "MiddleInitial", "Middle Initial",
        "MIDDLENAME", "Middle", "MName", "MNAME",
        "Current_MiddleName", "Current Middle Name",
        "middle_name", "middlename"
    ])
    last = pick([
        "LastName", "Last Name", "LASTNAME", "Last", "LName", "LNAME",
        "Surname", "Current_LastName", "Current Last Name",
        "last_name", "lastname"
    ])
    suffix = pick([
        "NameSuffix", "Name Suffix", "Suffix", "SUFFIX",
        "Current_NameSuffix", "Current Name Suffix",
        "name_suffix"
    ])

    name = " ".join([p for p in [first, middle, last, suffix] if p]).strip()
    return normalize_name_value(name) if "normalize_name_value" in globals() else name

def build_address_line1_row(row):
    parts = [
        normalize_export_text(row.get("House Number", "")),
        normalize_export_text(row.get("Street Name", "")),
    ]
    line1 = " ".join([p for p in parts if p]).strip()
    apt = normalize_export_text(row.get("Apartment Number", ""))
    if apt:
        line1 = f"{line1} Apt {apt}".strip()
    return line1

def first_existing_detail(columns, candidates):
    """Find a detail column safely using exact, lowercase, and normalized names.

    Detail shards can contain columns from different pipeline eras, for example
    `House Number`, `House_Number`, `ResidentialHouseNumber`, or `res_house_number`.
    Vendor exports need those address fields even when spelling/spacing changes.
    """
    lower_map = {str(c).strip().lower(): c for c in columns}
    for col in candidates:
        if col in columns:
            return col
        hit = lower_map.get(str(col).strip().lower())
        if hit is not None:
            return hit

    norm_map = {_norm_col_name(c): c for c in columns}
    for cand in candidates:
        n = _norm_col_name(cand)
        if n and n in norm_map:
            return norm_map[n]

    return None


# -----------------------------
# Local Voter Record Corrections
# -----------------------------
def load_voter_corrections() -> dict:
    """Persistent correction overlay keyed by PA/voter ID.

    Stored under the Athenix pipeline output folder so Step 7.9 can re-apply
    corrections after future SURE/CURRENT/contact refreshes. A local fallback is
    read for older DEV installs and migrated on next save.
    """
    path = VOTER_CORRECTIONS_PATH
    if not path.exists() and 'VOTER_CORRECTIONS_LOCAL_FALLBACK_PATH' in globals() and VOTER_CORRECTIONS_LOCAL_FALLBACK_PATH.exists():
        path = VOTER_CORRECTIONS_LOCAL_FALLBACK_PATH
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_voter_corrections(data: dict):
    VOTER_CORRECTIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    VOTER_CORRECTIONS_PATH.write_text(json.dumps(data or {}, indent=2), encoding="utf-8")


def _correction_voter_id_from_rowlike(rowlike) -> str:
    candidates = [
        "PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID",
        "VoterID", "Voter ID", "voter_id", "IDNumber", "ID Number", "_LookupPAID"
    ]
    try:
        for col in candidates:
            if col in rowlike:
                val = normalize_numeric_string(rowlike.get(col, "")) if "normalize_numeric_string" in globals() else str(rowlike.get(col, "")).strip()
                if val:
                    return val
    except Exception:
        pass
    return ""


CORRECTION_FIELD_ALIASES = {
    "FirstName": ["FirstName", "First Name", "first_name"],
    "MiddleName": ["MiddleName", "Middle Name", "middle_name"],
    "LastName": ["LastName", "Last Name", "last_name"],
    "NameSuffix": ["NameSuffix", "Name Suffix", "Suffix", "suffix"],
    "Gender": ["Gender", "Sex", "gender"],
    "Party": ["Party", "party", "Party Code", "party_raw"],
    "DOB": ["DOB", "Date of Birth", "DateOfBirth", "dob"],
    "RegistrationDate": ["RegistrationDate", "Registration Date", "registration_date"],
    "House Number": ["House Number", "HouseNumber", "house_number"],
    "House Number Suffix": ["House Number Suffix", "HouseNumberSuffix", "house_number_suffix"],
    "Street Name": ["Street Name", "StreetName", "street_name"],
    "Apartment Number": ["Apartment Number", "ApartmentNumber", "Unit", "Apt", "apartment_number"],
    "Address Line 2": ["Address Line 2", "AddressLine2", "Address2", "address_line_2"],
    "City": ["City", "res_city", "MailingCity", "Mail City"],
    "State": ["State", "res_state", "MailingState", "Mail State"],
    "Zip": ["Zip", "ZIP", "res_zip", "ZipCode", "MailingZip", "Mail Zip"],
    "County": ["County", "county"],
    "Municipality": ["Municipality", "municipality"],
    "Precinct": ["Precinct", "precinct"],
    "School District": ["School District", "school_district"],
    "School Region": ["School Region", "school_region"],
    "Mobile": ["Mobile", "Cell", "CellPhone", "Cell Phone"],
    "Landline": ["Landline", "Phone", "HomePhone", "Home Phone"],
    "Email": ["Email", "EmailAddress", "Email Address"],
    "Tags": ["Tags", "tags", "TAGS"],
    "MIB_Applied": ["MIB_Applied", "Mail Ballot Applied", "Ballot Application Status"],
    "MIB_BALLOT": ["MIB_BALLOT", "Mail Ballot Status", "Ballot Status"],
    "MB_PERM": ["MB_PERM", "MB_Perm", "Permanent Mail", "Permanent Mail Ballot"],
    "Correction Notes": ["Correction Notes"],
}


def _first_matching_col_from_aliases(columns, aliases):
    norm = {_norm_col_name(c): c for c in columns}
    for alias in aliases:
        hit = norm.get(_norm_col_name(alias))
        if hit:
            return hit
    return aliases[0] if aliases else None


def apply_voter_corrections_to_df(df: pd.DataFrame) -> pd.DataFrame:
    corrections = load_voter_corrections()
    if df is None or df.empty or not corrections:
        return df
    out = df.copy()
    id_cols = [c for c in ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID", "VoterID", "Voter ID", "voter_id", "IDNumber", "ID Number", "_LookupPAID"] if c in out.columns]
    if not id_cols:
        return out
    ids = out[id_cols[0]].astype(str).map(lambda v: normalize_numeric_string(v) if "normalize_numeric_string" in globals() else str(v).strip())
    for voter_id, corr in corrections.items():
        if not isinstance(corr, dict):
            continue
        mask = ids.eq(str(voter_id))
        if not bool(mask.any()):
            continue
        for canonical, value in corr.get("fields", {}).items():
            if canonical == "Correction Notes":
                continue
            aliases = CORRECTION_FIELD_ALIASES.get(canonical, [canonical])
            target = _first_matching_col_from_aliases(out.columns.tolist(), aliases)
            if target and target not in out.columns:
                out[target] = ""
            if target:
                out.loc[mask, target] = value
        # Rebuild common display columns after name/address edits.
        first_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["FirstName"])
        middle_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["MiddleName"])
        last_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["LastName"])
        suffix_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["NameSuffix"])
        if first_col and last_col:
            name_series = (out.loc[mask, first_col].fillna('').astype(str) + ' ' +
                           (out.loc[mask, middle_col].fillna('').astype(str) if middle_col in out.columns else '') + ' ' +
                           out.loc[mask, last_col].fillna('').astype(str) + ' ' +
                           (out.loc[mask, suffix_col].fillna('').astype(str) if suffix_col in out.columns else '')).map(lambda x: re.sub(r'\s+', ' ', x).strip())
            for full_col in ["Name", "FullName", "Full Name", "_LookupName"]:
                if full_col in out.columns:
                    out.loc[mask, full_col] = name_series.values

        house_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["House Number"])
        street_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["Street Name"])
        apt_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["Apartment Number"])
        if "_LookupAddress" in out.columns and house_col and street_col:
            apt_part = out.loc[mask, apt_col].fillna('').astype(str).map(lambda x: ("Apt " + x.strip()) if x.strip() else "") if apt_col in out.columns else ""
            addr_series = (out.loc[mask, house_col].fillna('').astype(str) + ' ' + out.loc[mask, street_col].fillna('').astype(str) + ' ' + apt_part).map(lambda x: re.sub(r'\s+', ' ', x).strip())
            out.loc[mask, "_LookupAddress"] = addr_series.values

        city_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["City"])
        state_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["State"])
        zip_col = _first_matching_col_from_aliases(out.columns.tolist(), CORRECTION_FIELD_ALIASES["Zip"])
        if "_LookupCityStateZip" in out.columns and city_col:
            city = out.loc[mask, city_col].fillna('').astype(str)
            state_zip = ((out.loc[mask, state_col].fillna('').astype(str) if state_col in out.columns else '') + ' ' + (out.loc[mask, zip_col].fillna('').astype(str) if zip_col in out.columns else '')).map(lambda x: re.sub(r'\s+', ' ', x).strip())
            csz = (city + ', ' + state_zip).map(lambda x: re.sub(r',\s*$', '', re.sub(r'\s+', ' ', x)).strip())
            out.loc[mask, "_LookupCityStateZip"] = csz.values
    return out


def apply_voter_corrections_to_row(row):
    if row is None:
        return row
    df = pd.DataFrame([row.to_dict() if hasattr(row, 'to_dict') else dict(row)])
    fixed = apply_voter_corrections_to_df(df)
    return fixed.iloc[0] if not fixed.empty else row


def render_voter_correction_editor(selected_row):
    voter_id = _correction_voter_id_from_rowlike(selected_row)
    if not voter_id:
        st.warning("This voter record has no PA/voter ID available, so corrections cannot be saved safely.")
        return
    corrections = load_voter_corrections()
    saved = corrections.get(voter_id, {}).get("fields", {}) if isinstance(corrections.get(voter_id, {}), dict) else {}
    st.markdown("#### Edit Voter Record")
    st.caption("Saved corrections are stored as a persistent pipeline overlay and can be re-applied after future SURE/CURRENT/contact refreshes.")
    if saved:
        st.info("This voter currently has a saved local correction overlay.")

    field_groups = [
        ("Name", [("FirstName", "First Name"), ("MiddleName", "Middle Name"), ("LastName", "Last Name"), ("NameSuffix", "Name Suffix")]),
        ("Voter Details", [("Gender", "Gender"), ("Party", "Party"), ("DOB", "DOB"), ("RegistrationDate", "Registration Date")]),
        ("Address", [("House Number", "House Number"), ("House Number Suffix", "House Number Suffix"), ("Street Name", "Street Name"), ("Apartment Number", "Apartment Number"), ("Address Line 2", "Address Line 2"), ("City", "City"), ("State", "State"), ("Zip", "Zip")]),
        ("Geography", [("County", "County"), ("Municipality", "Municipality"), ("Precinct", "Precinct"), ("School District", "School District"), ("School Region", "School Region")]),
        ("Contact + Mail Ballot", [("Mobile", "Mobile"), ("Landline", "Landline"), ("Email", "Email"), ("MIB_Applied", "MB Application"), ("MIB_BALLOT", "MB Ballot"), ("MB_PERM", "Permanent MB"), ("Tags", "Tags")]),
    ]
    with st.form(f"voter_correction_form_{voter_id}", clear_on_submit=False):
        values = {}
        for group_title, edit_fields in field_groups:
            st.markdown(f"**{group_title}**")
            rows = [edit_fields[i:i+4] for i in range(0, len(edit_fields), 4)]
            for row_fields in rows:
                cols = st.columns(len(row_fields), gap="small")
                for col_obj, (canonical, label) in zip(cols, row_fields):
                    aliases = CORRECTION_FIELD_ALIASES.get(canonical, [canonical])
                    current_val = saved.get(canonical, get_lookup_value(selected_row, aliases) if 'get_lookup_value' in globals() else '')
                    with col_obj:
                        values[canonical] = st.text_input(label, value=str(current_val or ""), key=f"corr_{voter_id}_{canonical}")
        notes_val = saved.get("Correction Notes", "")
        values["Correction Notes"] = st.text_area("Correction Notes", value=str(notes_val or ""), key=f"corr_{voter_id}_notes", height=90)
        c1, c2 = st.columns(2, gap="small")
        save_clicked = c1.form_submit_button("Save Voter Correction", width="stretch", type="primary")
        clear_clicked = c2.form_submit_button("Remove Saved Correction", width="stretch")
    if save_clicked:
        corrections[voter_id] = {"updated_at": datetime.now().isoformat(timespec="seconds"), "fields": values}
        save_voter_corrections(corrections)
        st.success("Voter correction saved.")
        st.rerun()
    if clear_clicked:
        if voter_id in corrections:
            corrections.pop(voter_id, None)
            save_voter_corrections(corrections)
        st.success("Saved correction removed.")
        st.rerun()

@st.cache_data(show_spinner=True)
def ensure_detail_shards():
    local_paths, source_label = find_local_dataset_paths("detail")
    if local_paths:
        return local_paths, {"source": source_label, "count": len(local_paths)}

    manifest = load_manifest()
    local_paths = []
    for shard in manifest["detail"]["shards"]:
        key = shard["key"]
        local_path = LOCAL_ROOT / key
        download_public_object(key, local_path)
        local_paths.append(str(local_path))
    return local_paths, manifest

def build_detail_export_sql(detail_paths, active_filters, include_all=True, selected_raw_cols=None):
    columns = get_conn().execute(f"DESCRIBE SELECT * FROM {dataset_scan_sql(detail_paths)}").df()["column_name"].tolist()

    q = quote_ident
    status_col = first_existing_detail(columns, ["VoterStatus", "voterstatus"])
    party_col = first_existing_detail(columns, ["Party", "party", "PARTY", "PoliticalParty", "Political Party"])
    gender_col = first_existing_detail(columns, ["Gender", "Sex"])
    age_col = first_existing_detail(columns, ["Age"])
    hh_col = first_existing_detail(columns, ["HH_ID"])
    email_col = first_existing_detail(columns, ["Email"])
    landline_col = first_existing_detail(columns, ["Landline"])
    mobile_col = first_existing_detail(columns, ["Mobile"])
    vote_hist_col = first_existing_detail(columns, ["V4A"])
    mib_applied_col = first_existing_detail(columns, ["MIB_Applied"])
    mib_ballot_col = first_existing_detail(columns, ["MIB_BALLOT"])
    mb_score_col = first_existing_detail(columns, ["MB_AProp_Score", "MMB_AProp_Score", "MB_Prob_Score"])
    mb_perm_col = first_existing_detail(columns, ["MB_PERM", "MB_Perm", "MB_Pern"])
    source_file_col = first_existing_detail(columns, ["Source_File", "Source File", "source_file"])
    mb_new_reg_col = first_existing_detail(columns, ["MailBallotNewRegistrant", "Mail Ballot New Registrant", "mail_ballot_new_registrant"])
    applicant_phone_col = first_existing_detail(columns, ["Current_ApplicantPhone", "ApplicantPhone", "Applicant Phone"])
    applicant_phone_type_col = first_existing_detail(columns, ["ApplicantPhone_Type", "Applicant Phone Type"])
    applicant_phone_compliance_col = first_existing_detail(columns, ["ApplicantPhone_Compliance", "Applicant Phone Compliance"])
    current_app_return_col = first_existing_detail(columns, ["Current_App_Return_Date", "AppReturnDate", "App Return Date"])
    current_ballot_sent_col = first_existing_detail(columns, ["Current_Ballot_Sent_Date", "BallotSentDate", "Ballot Sent Date"])
    current_ballot_returned_col = first_existing_detail(columns, ["Current_Ballot_Returned_Date", "BallotReturnedDate", "Ballot Returned Date"])

    # For heavy outputs we keep the old full-detail behavior.
    # For vendor/mail exports, include_all=False selects only needed raw columns
    # plus the computed helper fields used by filters. This avoids pulling every
    # detail column and keeps Prepare Mail CSV from tying up DuckDB/Streamlit.
    if include_all:
        exprs = ["*"]
    else:
        selected_raw_cols = selected_raw_cols or []
        raw_keep = []
        for col in selected_raw_cols:
            if col and col in columns and col not in raw_keep:
                raw_keep.append(col)
        exprs = [f"{q(col)} as {q(col)}" for col in raw_keep]

    if status_col:
        exprs.append(f"upper(trim(coalesce(cast({q(status_col)} as varchar), ''))) as _Status")
    else:
        exprs.append("'A' as _Status")

    if party_col:
        exprs.append(
            f"""case
                when upper(trim(coalesce(cast({q(party_col)} as varchar), ''))) = 'D' then 'D'
                when upper(trim(coalesce(cast({q(party_col)} as varchar), ''))) = 'R' then 'R'
                else 'O'
            end as _PartyNorm"""
        )
    else:
        exprs.append("'O' as _PartyNorm")

    if gender_col:
        exprs.append(
            f"""case
                when upper(trim(coalesce(cast({q(gender_col)} as varchar), ''))) in ('', 'NONE', 'NAN') then 'U'
                else upper(trim(cast({q(gender_col)} as varchar)))
            end as _Gender"""
        )
    else:
        exprs.append("'U' as _Gender")

    if age_col:
        exprs.append(f"try_cast({q(age_col)} as double) as _AgeNum")
    else:
        exprs.append("NULL::DOUBLE as _AgeNum")

    for alias, src in [("_HasEmail", email_col), ("_HasLandline", landline_col), ("_HasMobile", mobile_col)]:
        if src:
            exprs.append(
                f"""case
                    when trim(coalesce(cast({q(src)} as varchar), '')) in ('', 'None', 'NONE', 'nan', 'NAN') then false
                    else true
                end as {alias}"""
            )
        else:
            exprs.append(f"false as {alias}")

    if vote_hist_col:
        exprs.append(f"upper(trim(coalesce(cast({q(vote_hist_col)} as varchar), ''))) as _VoteHistory")
    else:
        exprs.append("'' as _VoteHistory")

    if mib_applied_col:
        exprs.append(f"case when upper(trim(coalesce(cast({q(mib_applied_col)} as varchar), ''))) = '' then 'DNA' else upper(trim(coalesce(cast({q(mib_applied_col)} as varchar), ''))) end as _MIBApplied")
    else:
        exprs.append("'DNA' as _MIBApplied")

    if mib_ballot_col:
        exprs.append(f"upper(trim(coalesce(cast({q(mib_ballot_col)} as varchar), ''))) as _MIBBallot")
    else:
        exprs.append("'' as _MIBBallot")

    if mb_score_col:
        exprs.append(f"try_cast(regexp_replace(cast({q(mb_score_col)} as varchar), '[^0-9\\.-]', '', 'g') as double) as _MBScore")
    else:
        exprs.append("NULL::DOUBLE as _MBScore")

    if mb_perm_col:
        exprs.append(f"""case
            when upper(trim(coalesce(cast({q(mb_perm_col)} as varchar), ''))) in ('TRUE', 'T', 'YES', 'Y', '1') then 'Y'
            when upper(trim(coalesce(cast({q(mb_perm_col)} as varchar), ''))) in ('FALSE', 'F', 'NO', 'N', '0') then 'N'
            else upper(trim(coalesce(cast({q(mb_perm_col)} as varchar), '')))
        end as _MBPerm""")
    else:
        exprs.append("'' as _MBPerm")

    if source_file_col:
        exprs.append(f"upper(trim(coalesce(cast({q(source_file_col)} as varchar), ''))) as _SourceFile")
    else:
        exprs.append("'' as _SourceFile")

    if mb_new_reg_col:
        exprs.append(f"""case
            when upper(trim(coalesce(cast({q(mb_new_reg_col)} as varchar), ''))) in ('Y', 'YES', 'TRUE', '1') then 'Y'
            else ''
        end as _MailBallotNewRegistrant""")
    else:
        exprs.append("'' as _MailBallotNewRegistrant")

    if applicant_phone_col:
        exprs.append(f"""case
            when trim(coalesce(cast({q(applicant_phone_col)} as varchar), '')) in ('', 'None', 'NONE', 'nan', 'NAN') then false
            else true
        end as _HasApplicantPhone""")
    else:
        exprs.append("false as _HasApplicantPhone")

    if applicant_phone_type_col:
        exprs.append(f"upper(trim(coalesce(cast({q(applicant_phone_type_col)} as varchar), ''))) as _ApplicantPhoneType")
    else:
        exprs.append("'' as _ApplicantPhoneType")

    if applicant_phone_compliance_col:
        exprs.append(f"upper(trim(coalesce(cast({q(applicant_phone_compliance_col)} as varchar), ''))) as _ApplicantPhoneCompliance")
    else:
        exprs.append("'' as _ApplicantPhoneCompliance")

    if current_app_return_col:
        exprs.append(f"""coalesce(
            try_strptime(cast({q(current_app_return_col)} as varchar), '%Y-%m-%d'),
            try_strptime(cast({q(current_app_return_col)} as varchar), '%m/%d/%Y'),
            try_strptime(cast({q(current_app_return_col)} as varchar), '%m/%d/%y'),
            try_cast({q(current_app_return_col)} as timestamp)
        ) as _CurrentAppReturnDate""")
    else:
        exprs.append("NULL::TIMESTAMP as _CurrentAppReturnDate")

    if current_ballot_sent_col:
        exprs.append(f"""coalesce(
            try_strptime(cast({q(current_ballot_sent_col)} as varchar), '%Y-%m-%d'),
            try_strptime(cast({q(current_ballot_sent_col)} as varchar), '%m/%d/%Y'),
            try_strptime(cast({q(current_ballot_sent_col)} as varchar), '%m/%d/%y'),
            try_cast({q(current_ballot_sent_col)} as timestamp)
        ) as _CurrentBallotSentDate""")
    else:
        exprs.append("NULL::TIMESTAMP as _CurrentBallotSentDate")

    if current_ballot_returned_col:
        exprs.append(f"""coalesce(
            try_strptime(cast({q(current_ballot_returned_col)} as varchar), '%Y-%m-%d'),
            try_strptime(cast({q(current_ballot_returned_col)} as varchar), '%m/%d/%Y'),
            try_strptime(cast({q(current_ballot_returned_col)} as varchar), '%m/%d/%y'),
            try_cast({q(current_ballot_returned_col)} as timestamp)
        ) as _CurrentBallotReturnedDate""")
    else:
        exprs.append("NULL::TIMESTAMP as _CurrentBallotReturnedDate")

    if hh_col:
        exprs.append(f"nullif(trim(coalesce(cast({q(hh_col)} as varchar), '')), '') as _HouseholdKey")
    else:
        exprs.append("NULL::VARCHAR as _HouseholdKey")

    where_sql, params = current_filter_clause(active_filters, columns)
    sql = "SELECT\n    " + ",\n    ".join(exprs) + f"\nFROM {dataset_scan_sql(detail_paths)}\n{where_sql}"
    return sql, params

def fetch_filtered_detail(active_filters):
    """Fetch detail rows for the APPLIED universe.

    The safe path applies filters to the canonical index view first, then
    semi-joins detail shards by voter id. This prevents raw detail geography
    fields from accidentally returning statewide rows for Filtered/Texting CSVs.
    """
    detail_paths, _ = ensure_detail_shards()
    con = get_conn()
    detail_columns = con.execute(f"DESCRIBE SELECT * FROM {dataset_scan_sql(detail_paths)}").df()["column_name"].tolist()
    index_columns = st.session_state.get("columns", []) or []

    detail_id_col = first_existing_detail(detail_columns, ["voter_id", "VoterID", "Voter_ID", "PAID", "PA ID Number", "PA_ID"])
    index_id_col = first_existing_precise(index_columns, ["voter_id", "VoterID", "Voter_ID", "PAID", "PA ID Number", "PA_ID"])

    if detail_id_col and index_id_col:
        q = quote_ident
        where_sql, params = current_filter_clause(active_filters, index_columns)
        sql = f"""
        WITH filtered_ids AS (
            SELECT DISTINCT {q(index_id_col)} AS _cc_join_id
            FROM voters
            {where_sql}
        ),
        detail_src AS (
            SELECT * FROM {dataset_scan_sql(detail_paths)}
        )
        SELECT d.*
        FROM detail_src d
        SEMI JOIN filtered_ids f
          ON cast(d.{q(detail_id_col)} as varchar) = cast(f._cc_join_id as varchar)
        """
        df = con.execute(sql, params).df()
    else:
        sql, params = build_detail_export_sql(detail_paths, active_filters)
        df = con.execute(sql, params).df()

    df = apply_voter_corrections_to_df(df)
    if has_global_followup_filters(active_filters):
        df = apply_global_followup_filters_df(df, active_filters)
    return df

def build_filtered_csv_export(active_filters):
    df = fetch_filtered_detail(active_filters).copy()
    return normalize_filtered_export_dataframe(df)

def _name_series_for_export(df: pd.DataFrame) -> pd.Series:
    if df.empty:
        return pd.Series([], dtype="object")
    cols = df.columns.tolist()
    first = _first_output_series(df, ["FirstName", "First Name", "first_name", "FNAME", "Current_FirstName"])
    middle = _first_output_series(df, ["MiddleName", "Middle Name", "MiddleInitial", "Middle Initial", "middle_name", "MNAME", "Current_MiddleName"])
    last = _first_output_series(df, ["LastName", "Last Name", "last_name", "LNAME", "Current_LastName"])
    suffix = _first_output_series(df, ["NameSuffix", "Name Suffix", "Suffix", "name_suffix", "Current_NameSuffix"])
    built = (first + " " + middle + " " + last + " " + suffix).fillna("").astype(str).map(lambda x: normalize_name_value(re.sub(r"\s+", " ", x).strip()))
    full_col = first_existing_detail(cols, ["Name", "FullName", "Full Name", "VoterName", "Voter Name"])
    if full_col:
        fallback = df[full_col].fillna("").astype(str).map(normalize_name_value)
        built = built.where(built.astype(str).str.strip().ne(""), fallback)
    return built

def build_texting_export(active_filters):
    df = fetch_filtered_detail(active_filters).copy()
    empty_cols = ["Name", "PA ID Number", "Mobile", "Party", "Age", "County", "Precinct"]
    if df.empty:
        return pd.DataFrame(columns=empty_cols)

    df["Name"] = _name_series_for_export(df)

    mobile_col = first_existing_detail(df.columns.tolist(), ["Mobile", "Cell", "CellPhone", "Cell Phone", "Phone", "PhoneNumber"])
    if mobile_col is None:
        df["MobileClean"] = ""
    else:
        df["MobileClean"] = df[mobile_col].apply(clean_phone_value)

    pa_id_col = first_existing_detail(
        df.columns.tolist(),
        ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID", "Voter ID", "VoterID", "voter_id"]
    )
    if pa_id_col is not None:
        df["PA ID Number"] = df[pa_id_col].apply(normalize_numeric_string)
    else:
        df["PA ID Number"] = ""

    party_col = first_existing_detail(df.columns.tolist(), ["Party", "party", "PARTY"])
    age_col = first_existing_detail(df.columns.tolist(), ["Age", "Age_Calc"])
    county_col = first_existing_detail(df.columns.tolist(), ["County"])
    precinct_col = first_existing_detail(df.columns.tolist(), ["Precinct"])

    out = pd.DataFrame({
        "Name": df["Name"],
        "PA ID Number": df["PA ID Number"],
        "Mobile": df["MobileClean"],
        "Party": df[party_col] if party_col else "",
        "Age": df[age_col] if age_col else "",
        "County": df[county_col] if county_col else "",
        "Precinct": df[precinct_col] if precinct_col else "",
    })
    out = out[out["Mobile"].astype(str).str.strip() != ""]
    return out.reset_index(drop=True)


def _first_output_series(df: pd.DataFrame, candidates, default="") -> pd.Series:
    """Return first matching column as a cleaned text series, or a default blank series."""
    col = first_existing_detail(df.columns.tolist(), candidates)
    if col is None:
        return pd.Series([default] * len(df), index=df.index, dtype="object")
    data = df[col]
    if isinstance(data, pd.DataFrame):
        data = data.iloc[:, 0]
    return data.fillna("").astype(str).map(normalize_export_text)



USPS_POST_DIRS = {"N", "S", "E", "W", "NE", "NW", "SE", "SW"}
USPS_PRE_DIR_WORD_MAP = {"NORTH": "N", "SOUTH": "S", "EAST": "E", "WEST": "W"}
USPS_UNIT_LABELS = {"APT", "APARTMENT", "UNIT", "STE", "SUITE", "#", "LOT", "FL", "FLOOR", "BLDG", "BUILDING"}

def normalize_street_type_token(value: str) -> str:
    s = normalize_export_text(value)
    if not s:
        return ""
    key = re.sub(r"[^A-Za-z]", "", s).upper()
    return USPS_SUFFIX_MAP.get(key, s).title()

def normalize_direction_token(value: str) -> str:
    s = normalize_export_text(value).upper().replace(".", "")
    if not s:
        return ""
    return USPS_PRE_DIR_WORD_MAP.get(s, s if s in USPS_POST_DIRS else s)

def normalize_unit_value(value: str) -> str:
    s = normalize_export_text(value)
    if not s:
        return ""
    s = re.sub(r"\b(APARTMENT)\b", "APT", s, flags=re.IGNORECASE)
    s = re.sub(r"\b(SUITE)\b", "STE", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    # If the value is only a unit number/letter, make it explicit for mail vendors.
    if re.fullmatch(r"[A-Za-z0-9\-]+", s) and not re.fullmatch(r"\d{5}(?:-\d{4})?", s):
        return f"Apt {s.upper()}"
    parts = s.split(" ", 1)
    if parts:
        label = parts[0].upper().replace("APARTMENT", "APT").replace("SUITE", "STE")
        if label in USPS_UNIT_LABELS:
            rest = parts[1].upper() if len(parts) > 1 else ""
            return ("# " + rest).strip() if label == "#" else (label.title() + (" " + rest if rest else ""))
    return normalize_address_value(s)

def _split_street_name_type_postdir(street_series: pd.Series):
    """Return street name WITH street type kept in Street Name, and postdirection in Street Suffix."""
    names = []
    post_dirs = []
    for raw in street_series.fillna("").astype(str):
        s = normalize_export_text(raw)
        if not s:
            names.append(""); post_dirs.append(""); continue
        parts = s.split()
        post = ""
        if parts:
            last = re.sub(r"[^A-Za-z]", "", parts[-1]).upper()
            if last in USPS_POST_DIRS:
                post = last
                parts = parts[:-1]
        # Standardize the street type token but KEEP it in Street Name.
        if parts:
            last_type_key = re.sub(r"[^A-Za-z]", "", parts[-1]).upper()
            if last_type_key in USPS_SUFFIX_MAP:
                parts[-1] = USPS_SUFFIX_MAP[last_type_key].title()
        names.append(" ".join(parts).strip())
        post_dirs.append(post)
    return pd.Series(names, index=street_series.index), pd.Series(post_dirs, index=street_series.index)

def _extract_unit_from_street_name(street_series: pd.Series):
    """Move trailing apartment/unit numbers accidentally embedded in Street Name to Address 2."""
    clean_names = []
    units = []
    for raw in street_series.fillna("").astype(str):
        s = normalize_export_text(raw)
        if not s:
            clean_names.append(""); units.append(""); continue
        parts = s.split()
        unit = ""
        # Explicit Apt/Unit/Suite inside the field.
        m = re.search(r"\b(APT|APARTMENT|UNIT|STE|SUITE|#)\s*([A-Z0-9\-]+)\b", s, flags=re.IGNORECASE)
        if m:
            unit = normalize_unit_value(m.group(0))
            s = (s[:m.start()] + " " + s[m.end():]).strip()
            parts = s.split()
        # Common pipeline issue: StreetName has trailing unit like "Freedom Way 108".
        if not unit and len(parts) >= 3 and re.fullmatch(r"\d+[A-Z]?", parts[-1]):
            unit_num = parts[-1]
            parts = parts[:-1]
            # If duplicated, e.g. "Freedom Way 203 203", remove both but keep one unit.
            if parts and parts[-1] == unit_num:
                parts = parts[:-1]
            unit = normalize_unit_value(unit_num)
        clean_names.append(" ".join(parts).strip())
        units.append(unit)
    return pd.Series(clean_names, index=street_series.index), pd.Series(units, index=street_series.index)

def _split_full_address_components(full_address_series: pd.Series):
    """Best-effort USPS-style parser: house, predir, street name+type, postdir, address2."""
    houses, pre_dirs, names, post_dirs, suites = [], [], [], [], []
    for raw in full_address_series.fillna("").astype(str):
        s = normalize_export_text(raw)
        if not s:
            houses.append(""); pre_dirs.append(""); names.append(""); post_dirs.append(""); suites.append("")
            continue
        suite = ""
        m_suite = re.search(r"\b(APT|APARTMENT|UNIT|STE|SUITE|#|LOT|FL|FLOOR|BLDG|BUILDING)\s*([A-Z0-9\-]+)\b", s, flags=re.IGNORECASE)
        if m_suite:
            suite = normalize_unit_value(m_suite.group(0))
            s = (s[:m_suite.start()] + " " + s[m_suite.end():]).strip()
        parts = s.split()
        house = ""
        if parts and re.match(r"^\d+[A-Za-z]?(?:-\d+)?$", parts[0]):
            house = parts.pop(0)
        pre_dir = ""
        if parts:
            d = re.sub(r"[^A-Za-z]", "", parts[0]).upper()
            if d in USPS_POST_DIRS or d in USPS_PRE_DIR_WORD_MAP:
                pre_dir = normalize_direction_token(d)
                parts.pop(0)
        post_dir = ""
        if parts:
            d = re.sub(r"[^A-Za-z]", "", parts[-1]).upper()
            if d in USPS_POST_DIRS:
                post_dir = d
                parts = parts[:-1]
        if parts:
            last_type_key = re.sub(r"[^A-Za-z]", "", parts[-1]).upper()
            if last_type_key in USPS_SUFFIX_MAP:
                parts[-1] = USPS_SUFFIX_MAP[last_type_key].title()
        name = " ".join(parts).strip()
        name_series, embedded_unit = _extract_unit_from_street_name(pd.Series([name]))
        name = name_series.iloc[0]
        if not suite:
            suite = embedded_unit.iloc[0]
        houses.append(house)
        pre_dirs.append(pre_dir)
        names.append(name)
        post_dirs.append(post_dir)
        suites.append(suite)
    return (
        pd.Series(houses, index=full_address_series.index),
        pd.Series(pre_dirs, index=full_address_series.index),
        pd.Series(names, index=full_address_series.index),
        pd.Series(post_dirs, index=full_address_series.index),
        pd.Series(suites, index=full_address_series.index),
    )


def _build_vendor_mail_export_rows(df: pd.DataFrame, name_override: pd.Series | None = None) -> pd.DataFrame:
    """Vendor-friendly USPS-style mail file.

    Column rule:
    - Street Direction = pre-direction (N/S/E/W/etc.)
    - Street Name = street name WITH street type, e.g. "David Dr"
    - Street Suffix = post-direction only, e.g. "NW"
    - Suite/Apt = Address 2 / unit / apartment / PO box when available
    """
    house = _first_output_series(df, [
        "House Number", "HouseNumber", "House_Number", "ResHouseNumber", "ResidenceHouseNumber",
        "ResidentialHouseNumber", "MailingHouseNumber", "Mail House Number"
    ]).map(normalize_numeric_string)

    street_dir = _first_output_series(df, [
        "Street Direction", "StreetDirection", "Street_Direction", "Street Dir", "StreetDir",
        "PreDirection", "PreDir", "Street Pre Direction", "MailingStreetDirection"
    ]).map(normalize_direction_token)

    street_name = _first_output_series(df, [
        "Street Name", "StreetName", "Street_Name", "ResidentialStreetName", "ResidenceStreetName",
        "MailingStreetName", "Mail Street Name"
    ])

    full_address = _first_output_series(df, [
        "res_address", "ResidenceAddress", "ResidentialAddress", "Address", "Address1",
        "Mailing Address", "MailAddress", "MailingAddress", "Residential Address"
    ])

    parsed_house, parsed_dir, parsed_street, parsed_postdir, parsed_suite = _split_full_address_components(full_address)
    house = house.where(house.astype(str).str.strip().ne(""), parsed_house.map(normalize_numeric_string))
    street_dir = street_dir.where(street_dir.astype(str).str.strip().ne(""), parsed_dir)
    street_name = street_name.where(street_name.astype(str).str.strip().ne(""), parsed_street)

    street_suffix = _first_output_series(df, [
        "Street Suffix", "StreetSuffix", "Street_Suffix", "PostDirection", "PostDir", "Street Post Direction",
        "Street_Post_Direction", "MailingStreetPostDirection", "Mail Street Post Direction"
    ]).map(normalize_direction_token)
    street_suffix = street_suffix.where(street_suffix.astype(str).str.strip().ne(""), parsed_postdir)

    # Keep the street type in Street Name; only split postdirection to Street Suffix.
    street_name, embedded_postdir = _split_street_name_type_postdir(street_name)
    street_suffix = street_suffix.where(street_suffix.astype(str).str.strip().ne(""), embedded_postdir)

    suite = _first_output_series(df, [
        "Suite/Apt", "Suite Apt", "Apartment Number", "ApartmentNumber", "Apartment_Number",
        "Apt", "Apt Number", "Unit", "Unit Number", "MailingApartmentNumber"
    ]).map(normalize_unit_or_address2_value)
    suite = suite.where(suite.astype(str).str.strip().ne(""), parsed_suite)

    # Fix bad source columns where unit numbers were glued onto Street Name.
    street_name, embedded_suite = _extract_unit_from_street_name(street_name)
    suite = suite.where(suite.astype(str).str.strip().ne(""), embedded_suite)

    # Final defensive cleanup after all suite fallbacks. Parsed/full-address fallback
    # values can reintroduce junk like "Apt STEEPLE", "Apt APTS", or "Apt Rd".
    suite = suite.map(normalize_unit_or_address2_value)

    city = _first_output_series(df, [
        "MailingCity", "Mailing City", "MailCity", "City", "res_city", "ResidenceCity",
        "ResidentialCity", "Residential City", "Current_City", "CurrentCity"
    ]).map(normalize_city_value)
    state = _first_output_series(df, [
        "MailingState", "Mailing State", "MailState", "State", "res_state", "ResidenceState",
        "ResidentialState", "Residential State", "Current_State", "CurrentState"
    ]).map(normalize_state_value)
    zip_code = _first_output_series(df, [
        "MailingZip", "Mailing Zip", "MailZip", "ZIP", "Zip", "ZipCode", "ZIPCODE",
        "res_zip", "ResidenceZip", "ResidentialZip", "Residential Zip", "Current_Zip", "CurrentZip"
    ]).map(clean_zip_value)

    first = _first_output_series(df, ["FirstName", "First Name", "first_name", "FNAME"]).map(normalize_name_value)
    middle = _first_output_series(df, ["MiddleName", "Middle Name", "MiddleInitial", "Middle Initial", "middle_name", "MNAME"]).map(normalize_name_value)
    last = _first_output_series(df, ["LastName", "Last Name", "last_name", "LNAME"]).map(normalize_name_value)
    suffix = _first_output_series(df, ["NameSuffix", "Name Suffix", "Suffix", "name_suffix"]).map(normalize_name_value)

    if name_override is None:
        name = (first + " " + middle + " " + last + " " + suffix).fillna("").astype(str).map(lambda x: normalize_name_value(re.sub(r"\s+", " ", x).strip()))
        fallback = _first_output_series(df, ["Name", "FullName", "Full Name", "VoterName", "Voter Name"]).map(normalize_name_value)
        name = name.where(name.astype(str).str.strip().ne(""), fallback)
    else:
        name = name_override.fillna("").astype(str).map(normalize_name_value)

    address_line_2 = _first_output_series(df, [
        "Address Line 2", "AddressLine2", "Address_Line_2", "ResidentialAddressLine2",
        "ResidenceAddressLine2", "Mail Address 2", "MailAddress2", "MailingAddress2",
        "Current_Address2", "Current Address 2"
    ]).map(normalize_unit_or_address2_value)

    # Final defensive cleanup for Address Line 2 as well.
    address_line_2 = address_line_2.map(normalize_unit_or_address2_value)

    out = pd.DataFrame({
        "House Number": house,
        "Street Direction": street_dir,
        "Street Name": street_name.map(normalize_address_value),
        "Street Suffix": street_suffix,
        "Suite/Apt": suite,
        "Address Line 2": address_line_2,
        "City": city,
        "State": state,
        "Zip": zip_code,
        "First Name": first,
        "Middle Name": middle,
        "Last Name": last,
        "Name Suffix": suffix,
        "Name": name,
    })
    return out.reset_index(drop=True)

def fetch_mail_export_detail(active_filters):
    """Fetch only vendor mail columns for the APPLIED universe.

    Important: the universe filters are applied against the fast/canonical index
    view first, then the detail shards are semi-joined by voter id. That prevents
    mail exports from accidentally falling back to the whole statewide detail set
    when detail-shard raw geography fields do not match the cleaned UI fields.
    """
    detail_paths, _ = ensure_detail_shards()
    con = get_conn()
    detail_columns = con.execute(f"DESCRIBE SELECT * FROM {dataset_scan_sql(detail_paths)}").df()["column_name"].tolist()

    candidate_groups = [
        ["House Number", "HouseNumber", "HouseNum", "HouseNo", "house_number", "res_house_number",
         "ResidenceHouseNumber", "ResidentialHouseNumber", "Current_HouseNumber", "Current_House_Number"],
        ["Street Direction", "StreetDirection", "Street Dir", "StreetDir", "PreDir", "PreDirection",
         "StreetPreDirection", "Street Pre Direction", "res_street_direction", "ResidenceStreetDirection",
         "ResidentialStreetDirection", "Current_StreetDir", "Current_Street_Direction"],
        ["Street Name", "StreetName", "Street_Name", "street_name", "res_street_name",
         "ResidenceStreetName", "ResidentialStreetName", "Current_StreetName", "Current_Street_Name"],
        ["Street Suffix", "StreetSuffix", "Street Type", "StreetType", "PostType", "StreetPostType",
         "res_street_suffix", "ResidenceStreetSuffix", "ResidentialStreetSuffix", "Current_StreetSuffix"],
        ["Suite/Apt", "Suite Apt", "Suite", "Apt", "Apartment Number", "ApartmentNumber",
         "Apartment_Number", "Unit", "Unit Number", "res_unit", "res_apt", "ResidenceApartmentNumber",
         "ResidentialApartmentNumber", "Current_ApartmentNumber", "Current_Apt"],
        ["Address Line 2", "AddressLine2", "Address_Line_2", "ResidentialAddressLine2",
         "ResidenceAddressLine2", "Mail Address 2", "MailAddress2", "MailingAddress2",
         "Current_Address2", "Current Address 2"],
        ["City", "MailingCity", "Mailing City", "MailCity", "res_city", "ResidenceCity",
         "ResidentialCity", "Residential City", "Current_City", "CurrentCity"],
        ["State", "MailingState", "Mailing State", "MailState", "res_state", "ResidenceState",
         "ResidentialState", "Residential State", "Current_State", "CurrentState"],
        ["Zip", "ZIP", "ZipCode", "ZIPCODE", "MailingZip", "Mailing Zip", "MailZip",
         "res_zip", "ResidenceZip", "ResidentialZip", "Residential Zip", "Current_Zip", "CurrentZip"],
        ["FirstName", "First Name", "first_name", "FNAME", "Current_FirstName"],
        ["MiddleName", "Middle Name", "MiddleInitial", "Middle Initial", "middle_name", "MNAME", "Current_MiddleName"],
        ["LastName", "Last Name", "last_name", "LNAME", "Current_LastName"],
        ["Name Suffix", "NameSuffix", "Suffix", "name_suffix", "Current_NameSuffix"],
        ["Name", "FullName", "Full Name", "VoterName", "Voter Name"],
        ["res_address", "ResidenceAddress", "ResidentialAddress", "Residential Address", "Residence Address",
         "Address", "Address1", "Address Line 1", "AddressLine1", "MailingAddress", "Mailing Address",
         "Current_Address", "Current_ResidentialAddress", "Current_ResidenceAddress"],
        ["HH_ID", "Household_ID", "Household ID", "_HouseholdKey"],
    ]
    selected = []
    for group in candidate_groups:
        hit = first_existing_detail(detail_columns, group)
        if hit and hit not in selected:
            selected.append(hit)

    # Include compactly-named address columns that may not match exact lists.
    address_needles = (
        "house", "street", "address", "apartment", "apt", "unit",
        "residence", "residential", "mailing", "city", "state", "zip"
    )
    for col in detail_columns:
        n = _norm_col_name(col)
        if any(x in n for x in address_needles) and col not in selected:
            selected.append(col)

    detail_id_col = first_existing_detail(detail_columns, ["voter_id", "VoterID", "Voter_ID", "PAID", "PA ID Number", "PA_ID"])
    index_columns = st.session_state.get("columns", []) or []
    index_id_col = first_existing_precise(index_columns, ["voter_id", "VoterID", "Voter_ID", "PAID", "PA ID Number", "PA_ID"])

    q = quote_ident

    # Best path: use the already-prepared canonical index view for filters, then semi-join detail rows.
    if detail_id_col and index_id_col:
        if detail_id_col not in selected:
            selected.append(detail_id_col)
        select_expr = ",\n    ".join([f"d.{q(col)} as {q(col)}" for col in selected if col in detail_columns])
        where_sql, params = current_filter_clause(active_filters, index_columns)
        sql = f"""
        WITH filtered_ids AS (
            SELECT DISTINCT {q(index_id_col)} AS _cc_join_id
            FROM voters
            {where_sql}
        ),
        detail_src AS (
            SELECT * FROM {dataset_scan_sql(detail_paths)}
        )
        SELECT
            {select_expr}
        FROM detail_src d
        SEMI JOIN filtered_ids f
          ON cast(d.{q(detail_id_col)} as varchar) = cast(f._cc_join_id as varchar)
        """
        df = con.execute(sql, params).df()
    else:
        # Fallback if voter id is unavailable: use old direct detail filtering.
        sql, params = build_detail_export_sql(detail_paths, active_filters, include_all=False, selected_raw_cols=selected)
        df = con.execute(sql, params).df()

    df = apply_voter_corrections_to_df(df)
    if has_global_followup_filters(active_filters):
        df = apply_global_followup_filters_df(df, active_filters)
    return df


def build_mail_export(active_filters, householded=False):
    df = fetch_mail_export_detail(active_filters).copy()
    vendor_cols = [
        "House Number", "Street Direction", "Street Name", "Street Suffix", "Suite/Apt", "Address Line 2",
        "City", "State", "Zip", "First Name", "Middle Name", "Last Name", "Name Suffix", "Name"
    ]
    if df.empty:
        return pd.DataFrame(columns=vendor_cols)

    base_rows = _build_vendor_mail_export_rows(df)

    if householded:
        # Group household mail records by household key when available, otherwise by full address.
        key_name = "_HouseholdKey" if "_HouseholdKey" in df.columns else None
        address_text = (
            base_rows["House Number"].astype(str) + "|" +
            base_rows["Street Direction"].astype(str) + "|" +
            base_rows["Street Name"].astype(str) + "|" +
            base_rows["Street Suffix"].astype(str) + "|" +
            base_rows["Suite/Apt"].astype(str) + "|" +
            base_rows["Address Line 2"].astype(str) + "|" +
            base_rows["City"].astype(str) + "|" +
            base_rows["State"].astype(str) + "|" +
            base_rows["Zip"].astype(str)
        )
        if key_name and key_name in df.columns:
            base_key = safe_group_series(df, key_name)
            grp_key = base_key.where(base_key != "", address_text)
        else:
            grp_key = address_text

        temp = base_rows.copy()
        temp["_grp"] = grp_key.fillna("").astype(str).values
        temp["_BaseName"] = base_rows["Name"].fillna("").astype(str).values
        temp["_LastName"] = base_rows["Last Name"].fillna("").astype(str).values

        grouped_rows = []
        grouped = temp.sort_values(by=["_grp", "_BaseName"]).groupby("_grp", dropna=False, sort=False)
        for _, grp in grouped:
            first_row = grp.iloc[0].copy()
            household_name = build_household_mail_name(pd.DataFrame({
                "Name": grp["_BaseName"],
                "LastName": grp["_LastName"],
            }))
            row = {col: first_row.get(col, "") for col in vendor_cols}
            row["First Name"] = ""
            row["Middle Name"] = ""
            row["Last Name"] = ""
            row["Name Suffix"] = ""
            row["Name"] = normalize_name_value(household_name)
            grouped_rows.append(row)

        return pd.DataFrame(grouped_rows, columns=vendor_cols).reset_index(drop=True)

    return base_rows[vendor_cols]



def _label_line(value: str) -> str:
    """Clean a value for printable mailing labels."""
    return collapse_spaces(normalize_export_text(value))


def _mailing_label_rows(active_filters, householded=False) -> pd.DataFrame:
    """Use the same cleaned mail export data for labels that CSV/Excel uses."""
    rows = build_mail_export(active_filters, householded=householded).copy()
    if rows.empty:
        return rows
    for col in rows.columns:
        rows[col] = rows[col].fillna("").astype(str).map(_label_line)

    # Drop rows that have neither a name nor enough address to print.
    addr_key = (
        rows.get("House Number", "").astype(str) + " " +
        rows.get("Street Name", "").astype(str) + " " +
        rows.get("City", "").astype(str) + " " +
        rows.get("Zip", "").astype(str)
    ).map(_label_line)
    name_key = rows.get("Name", "").astype(str).map(_label_line)
    rows = rows[(name_key != "") & (addr_key != "")].reset_index(drop=True)
    return rows


def generate_mailing_labels_pdf_bytes(active_filters, householded=False):
    """Generate Avery 5160-style mailing labels from the applied universe.

    Layout: Letter portrait, 3 columns x 10 rows, 30 labels per page.
    Uses the already-cleaned Mail CSV/Excel export fields so labels match vendor files.
    """
    label_df = _mailing_label_rows(active_filters, householded=householded)
    if label_df.empty:
        return b""

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Avery 5160 approximate geometry, in points.
    left_margin = 13.5
    top_margin = 36.0
    label_w = 189.0
    label_h = 72.0
    col_pitch = 198.0
    row_pitch = 72.0
    inner_x = 8.0
    first_y_offset = 15.0
    line_h = 10.5

    labels_per_page = 30
    printed_date = datetime.now().strftime("%m/%d/%Y")

    for i, row in label_df.iterrows():
        page_index = i % labels_per_page
        if i > 0 and page_index == 0:
            c.showPage()

        col = page_index % 3
        r = page_index // 3
        x = left_margin + col * col_pitch
        y_top = height - top_margin - r * row_pitch

        name = _label_line(row.get("Name", ""))
        house = _label_line(row.get("House Number", ""))
        street_dir = _label_line(row.get("Street Direction", ""))
        street_name = _label_line(row.get("Street Name", ""))
        street_suffix = _label_line(row.get("Street Suffix", ""))
        suite = normalize_unit_or_address2_value(row.get("Suite/Apt", ""))
        addr2 = normalize_unit_or_address2_value(row.get("Address Line 2", ""))
        city = _label_line(row.get("City", ""))
        state = _label_line(row.get("State", "")).upper()
        zip_code = _label_line(row.get("Zip", ""))

        address1 = collapse_spaces(" ".join([p for p in [house, street_dir, street_name, street_suffix] if p]))
        address2_parts = []
        if suite:
            address2_parts.append(suite)
        if addr2 and addr2.upper() != suite.upper():
            address2_parts.append(addr2)
        address2 = collapse_spaces(" ".join(address2_parts))
        city_state_zip = collapse_spaces(f"{city}, {state} {zip_code}".replace(" ,", ","))

        text_x = x + inner_x
        text_y = y_top - first_y_offset
        max_text_w = label_w - (inner_x * 2) - 3

        def _wrap_label_line(text, font_name, font_size, max_width, max_lines=2):
            """Wrap label text without using ellipses. Keeps USPS/vendor labels readable."""
            text = collapse_spaces(str(text or ""))
            if not text:
                return []
            words = text.split(" ")
            lines = []
            current = ""
            for word in words:
                test = word if not current else current + " " + word
                if c.stringWidth(test, font_name, font_size) <= max_width:
                    current = test
                else:
                    if current:
                        lines.append(current)
                    current = word
                    if len(lines) >= max_lines:
                        break
            if current and len(lines) < max_lines:
                lines.append(current)
            return lines

        c.setFillColor(colors.black)
        name_font = "Helvetica-Bold"
        name_size = 8.7
        body_font = "Helvetica"
        body_size = 8.3
        label_line_h = 9.4

        name_lines = _wrap_label_line(name, name_font, name_size, max_text_w, max_lines=2)
        y_cursor = text_y
        c.setFont(name_font, name_size)
        for line in name_lines:
            c.drawString(text_x, y_cursor, line)
            y_cursor -= label_line_h

        c.setFont(body_font, body_size)
        for line in [address1, address2, city_state_zip]:
            if not line:
                continue
            # Address lines should not use ellipses either. Wrap long apartment/building lines if needed.
            wrapped = _wrap_label_line(line, body_font, body_size, max_text_w, max_lines=2 if line == address2 else 1)
            for wline in wrapped:
                c.drawString(text_x, y_cursor, wline)
                y_cursor -= label_line_h

    c.save()
    return buffer.getvalue()

def dataframe_to_csv_bytes(df):
    return df.to_csv(index=False).encode("utf-8")


def _first_present_column(columns, candidates):
    norm = {_norm_col_name(c): c for c in columns}
    for cand in candidates:
        hit = norm.get(_norm_col_name(cand))
        if hit:
            return hit
    return None

def build_export_area_counts_sheet(export_df: pd.DataFrame) -> pd.DataFrame:
    """Build Sheet 1 counts for CSV-style exports only after Prepare is clicked."""
    if export_df is None or export_df.empty:
        return pd.DataFrame(columns=["Area Type", "Area", "Voters", "Households"])

    df = export_df.copy()
    cols = list(df.columns)
    muni_col = _first_present_column(cols, ["Municipality", "municipality", "Muni", "City", "CITY"])
    precinct_col = _first_present_column(cols, ["Precinct", "precinct", "Precinct Name", "PrecinctName"])

    muni_values = []
    if muni_col:
        muni_values = sorted({normalize_export_text(v) for v in df[muni_col].tolist() if normalize_export_text(v)})

    if muni_col and len(muni_values) > 1:
        area_col = muni_col
        area_type = "Municipality"
    elif precinct_col:
        area_col = precinct_col
        area_type = "Precinct"
    elif muni_col:
        area_col = muni_col
        area_type = "Municipality"
    else:
        df["_cc_area"] = "TOTAL"
        area_col = "_cc_area"
        area_type = "Total"

    hh_col = _first_present_column(cols, ["HH_ID", "Household_ID", "Household ID", "HouseholdKey", "_HouseholdKey"])
    if hh_col:
        df["_cc_hh"] = df[hh_col].fillna("").astype(str).map(normalize_export_text)
    else:
        address_parts = []
        for cand in ["House Number", "Street Direction", "Street Name", "Street Suffix", "Suite/Apt", "Address Line 2", "City", "State", "Zip", "PA ID Number"]:
            col = _first_present_column(cols, [cand])
            if col:
                address_parts.append(df[col].fillna("").astype(str).map(normalize_export_text))
        if address_parts:
            key = address_parts[0]
            for part in address_parts[1:]:
                key = key + "|" + part
            df["_cc_hh"] = key
        else:
            df["_cc_hh"] = df.index.astype(str)

    df["_cc_area_label"] = df[area_col].fillna("").astype(str).map(normalize_export_text)
    df.loc[df["_cc_area_label"].eq(""), "_cc_area_label"] = "(Blank)"

    rows = []
    for area, grp in df.groupby("_cc_area_label", dropna=False, sort=True):
        hh_count = grp["_cc_hh"].replace("", pd.NA).dropna().nunique()
        rows.append({"Area Type": area_type, "Area": area, "Voters": int(len(grp)), "Households": int(hh_count)})

    total_households = df["_cc_hh"].replace("", pd.NA).dropna().nunique()
    rows.append({"Area Type": "Total", "Area": "TOTAL", "Voters": int(len(df)), "Households": int(total_households)})
    return pd.DataFrame(rows, columns=["Area Type", "Area", "Voters", "Households"])

def dataframe_to_export_excel_bytes(export_df: pd.DataFrame, export_label: str = "Export") -> bytes:
    output = BytesIO()
    data_df = export_df.copy() if export_df is not None else pd.DataFrame()
    counts_df = build_export_area_counts_sheet(data_df)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        counts_df.to_excel(writer, sheet_name="Area Counts", index=False)
        data_df.to_excel(writer, sheet_name="Data", index=False)
        wb = writer.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for col_cells in ws.columns:
                max_len = 0
                letter = col_cells[0].column_letter
                for cell in col_cells[:200]:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)
    output.seek(0)
    return output.getvalue()

def sanitize_filename_part(value: str) -> str:
    s = normalize_export_text(value)
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s).strip("_")
    return s or "blank"



def turf_packet_display_name(packet_label: str, turf_id: str) -> str:
    label = normalize_export_text(packet_label)
    turf = normalize_export_text(turf_id) or "Turf"
    return f"{label} - {turf}" if label else turf


def choose_group_value(row, preferred_columns):
    for col in preferred_columns:
        if col in row and normalize_export_text(row.get(col, "")):
            return normalize_export_text(row.get(col, ""))
    return "(Blank)"


def assign_turf_ids(df: pd.DataFrame, mode: str, target_size: int) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    out = df.copy()
    out["_HouseholdKeySafe"] = out.get("_HouseholdKey", "").fillna("").astype(str).str.strip() if "_HouseholdKey" in out.columns else ""
    if "Address1" not in out.columns:
        out["Address1"] = out.apply(build_address_line1_row, axis=1)

    if mode == "By Precinct":
        group_vals = out.apply(lambda r: choose_group_value(r, ["Precinct"]), axis=1)
        out["Turf_Group"] = group_vals
        out["Turf_ID"] = out["Turf_Group"].apply(lambda v: f"Turf_{sanitize_filename_part(v)}")
    elif mode == "By Municipality":
        group_vals = out.apply(lambda r: choose_group_value(r, ["Municipality", "County"]), axis=1)
        out["Turf_Group"] = group_vals
        out["Turf_ID"] = out["Turf_Group"].apply(lambda v: f"Turf_{sanitize_filename_part(v)}")
    else:
        out["_DoorKey"] = out["_HouseholdKeySafe"]
        blank_mask = out["_DoorKey"].eq("")
        out.loc[blank_mask, "_DoorKey"] = out.loc[blank_mask, "Address1"].fillna("").astype(str)

        work = out.copy()
        household_sizes = work.groupby("_DoorKey", dropna=False).size().reset_index(name="_VoterCount")
        household_sizes["_DoorCount"] = 1
        household_sizes["_StreetSort"] = household_sizes["_DoorKey"].astype(str)
        household_sizes = household_sizes.sort_values(["_StreetSort", "_DoorKey"], kind="stable").reset_index(drop=True)

        turf_ids = []
        turf_num = 1
        current_size = 0
        for _, hh in household_sizes.iterrows():
            increment = int(hh["_DoorCount"] if mode == "Target Doors" else hh["_VoterCount"])
            if current_size > 0 and current_size + increment > int(target_size):
                turf_num += 1
                current_size = 0
            turf_ids.append(f"Turf_{turf_num:03d}")
            current_size += increment
        household_sizes["Turf_ID"] = turf_ids
        out = out.merge(household_sizes[["_DoorKey", "Turf_ID"]], on="_DoorKey", how="left")
        out["Turf_Group"] = out["Turf_ID"]

    summary = (
        out.groupby("Turf_ID", dropna=False)
        .agg(
            Voters=("Turf_ID", "size"),
            Households=("_HouseholdKeySafe", lambda s: s.replace("", pd.NA).dropna().nunique() + (s.eq("").sum())),
            Counties=("County", lambda s: ", ".join(sorted({normalize_export_text(v) for v in s if normalize_export_text(v)})[:4])),
            Municipalities=("Municipality", lambda s: ", ".join(sorted({normalize_export_text(v) for v in s if normalize_export_text(v)})[:4])),
            Precincts=("Precinct", lambda s: ", ".join(sorted({normalize_export_text(v) for v in s if normalize_export_text(v)})[:4])),
        )
        .reset_index()
        .sort_values("Turf_ID")
        .reset_index(drop=True)
    )
    out = out.merge(summary[["Turf_ID", "Voters", "Households"]], on="Turf_ID", how="left")
    out = out.rename(columns={"Voters": "Turf_Voters", "Households": "Turf_Households"})
    return out


def build_turf_packet_zip(active_filters, mode: str, target_size: int = 50, volunteer_name: str = "", packet_label: str = "", packet_date: str = "", include_walksheets: bool = True, max_turfs: int = 0):
    df = fetch_filtered_detail(active_filters).copy()
    if df.empty:
        return b""

    volunteer_name = normalize_name_value(volunteer_name)
    packet_label = collapse_spaces(packet_label)
    packet_date = normalize_export_text(packet_date) or datetime.now().strftime("%Y-%m-%d")

    df["Name"] = df.apply(full_name_from_row, axis=1)
    df["Address1"] = df.apply(build_address_line1_row, axis=1)
    city_col = first_existing_detail(df.columns.tolist(), ["MailingCity", "Mailing City", "City", "MailCity", "res_city", "ResidenceCity", "ResidentialCity"])
    state_col = first_existing_detail(df.columns.tolist(), ["MailingState", "Mailing State", "State", "MailState", "res_state", "ResidenceState", "ResidentialState"])
    zip_col = first_existing_detail(df.columns.tolist(), ["MailingZip", "Mailing Zip", "ZIP", "Zip", "ZipCode", "ZIPCODE", "MailZip", "res_zip", "ResidenceZip", "ResidentialZip"])
    if city_col and "City" not in df.columns:
        df["City"] = df[city_col]
    if state_col and "State" not in df.columns:
        df["State"] = df[state_col]
    if zip_col and "Zip" not in df.columns:
        df["Zip"] = df[zip_col]

    pa_id_col = first_existing_detail(df.columns.tolist(), ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "Voter ID", "VoterID"])
    if pa_id_col and pa_id_col != "PA_ID_Number":
        df["PA_ID_Number"] = df[pa_id_col]
    elif "PA_ID_Number" not in df.columns:
        df["PA_ID_Number"] = ""

    df = assign_turf_ids(df, mode=mode, target_size=target_size)

    export_cols = [c for c in [
        "Turf_ID", "Name", "PA_ID_Number", "Address1", "City", "State", "Zip",
        "County", "Municipality", "Precinct", "Party", "Gender", "Age", "Mobile", "Landline"
    ] if c in df.columns]

    export_df = df[export_cols].copy()
    export_df = normalize_filtered_export_dataframe(export_df)
    if "Zip" in export_df.columns:
        export_df["Zip"] = export_df["Zip"].apply(clean_zip_value)
    if "Mobile" in export_df.columns:
        export_df["Mobile"] = export_df["Mobile"].apply(clean_phone_value)
    if "Landline" in export_df.columns:
        export_df["Landline"] = export_df["Landline"].apply(clean_phone_value)

    export_df.insert(1, "Packet_Label", packet_label)
    export_df.insert(2, "Volunteer_Name", volunteer_name)
    export_df.insert(3, "Packet_Date", packet_date)

    summary_df = (
        df.groupby("Turf_ID", dropna=False)
        .agg(
            Voters=("Turf_ID", "size"),
            Households=("_HouseholdKeySafe", lambda s: s.replace("", pd.NA).dropna().nunique() + (s.eq("").sum())),
            Counties=("County", lambda s: ", ".join(sorted({normalize_export_text(v) for v in s if normalize_export_text(v)})[:4])),
            Municipalities=("Municipality", lambda s: ", ".join(sorted({normalize_export_text(v) for v in s if normalize_export_text(v)})[:4])),
            Precincts=("Precinct", lambda s: ", ".join(sorted({normalize_export_text(v) for v in s if normalize_export_text(v)})[:4])),
        )
        .reset_index()
        .sort_values("Turf_ID")
        .reset_index(drop=True)
    )
    summary_df.insert(1, "Packet_Label", packet_label)
    summary_df.insert(2, "Volunteer_Name", volunteer_name)
    summary_df.insert(3, "Packet_Date", packet_date)

    if int(max_turfs or 0) > 0:
        selected_turfs = summary_df["Turf_ID"].head(int(max_turfs)).tolist()
        df = df[df["Turf_ID"].isin(selected_turfs)].copy()
        summary_df = summary_df[summary_df["Turf_ID"].isin(selected_turfs)].copy()

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("turf_summary.csv", summary_df.to_csv(index=False))
        readme_lines = [
            "Candidate Connect Turf Packets",
            "",
            "This zip contains turf_summary.csv and one CSV per turf.",
            "Walk sheet PDFs are included only when 'CSV + Walk Sheet PDFs' is selected.",
            f"Packet Label: {packet_label or '(none)'}",
            f"Volunteer: {volunteer_name or '(unassigned)'}",
            f"Packet Date: {packet_date}",
            f"Mode: {mode}",
            f"Walk Sheets Included: {'Yes' if include_walksheets else 'No'}",
            f"Turf Limit Applied: {int(max_turfs) if int(max_turfs or 0) > 0 else 'All'}",
        ]
        zf.writestr("README.txt", "\n".join(readme_lines) + "\n")
        for turf_id, turf_df in df.groupby("Turf_ID", sort=True, dropna=False):
            safe_id = sanitize_filename_part(str(turf_id))
            packet_base = sanitize_filename_part(turf_packet_display_name(packet_label, str(turf_id)))
            csv_df = export_df[export_df["Turf_ID"] == turf_id].drop(columns=["Turf_ID"], errors="ignore")
            zf.writestr(f"turf_packets/{packet_base}.csv", csv_df.to_csv(index=False))

            turf_street_df = build_street_list_dataframe_from_detail_df(turf_df.copy())
            summary_row = summary_df[summary_df["Turf_ID"] == turf_id]
            voters = int(summary_row["Voters"].iloc[0]) if not summary_row.empty else len(turf_df)
            households = int(summary_row["Households"].iloc[0]) if not summary_row.empty else 0
            precincts = summary_row["Precincts"].iloc[0] if not summary_row.empty else ""
            title = turf_packet_display_name(packet_label, str(turf_id))
            filter_parts = [f"{voters:,} voters", f"{households:,} households"]
            if normalize_export_text(volunteer_name):
                filter_parts.append(f"Volunteer: {volunteer_name}")
            if normalize_export_text(packet_date):
                filter_parts.append(packet_date)
            if normalize_export_text(precincts):
                filter_parts.append(precincts)
            filter_desc = " | ".join(filter_parts)
            if include_walksheets:
                pdf_bytes = generate_walk_sheet_pdf_from_street_df(turf_street_df, title, filter_desc)
                if pdf_bytes:
                    zf.writestr(f"turf_walksheets/{packet_base}_walksheet.pdf", pdf_bytes)
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def normalize_mb_perm_value(val) -> str:
    s = normalize_export_text(val).upper()
    if s in {"TRUE", "T", "YES", "Y", "1"}:
        return "Y"
    if s in {"FALSE", "F", "NO", "N", "0"}:
        return "N"
    return ""

def choose_best_phone(row) -> str:
    mobile = clean_phone_value(row.get("Mobile", ""))
    landline = clean_phone_value(row.get("Landline", ""))
    primary = clean_phone_value(row.get("PrimaryPhone", ""))
    if mobile:
        return f"({mobile[:3]}) {mobile[3:6]}-{mobile[6:]}" + " (m)" if len(mobile) == 10 else mobile + " (m)"
    if landline:
        return f"({landline[:3]}) {landline[3:6]}-{landline[6:]}" + " (l)" if len(landline) == 10 else landline + " (l)"
    if primary:
        return f"({primary[:3]}) {primary[3:6]}-{primary[6:]}" if len(primary) == 10 else primary
    return ""

def parse_house_number(value) -> int:
    s = normalize_export_text(value)
    m = re.search(r"\d+", s)
    return int(m.group()) if m else 0

def parse_apartment_sort(value) -> tuple:
    s = normalize_export_text(value)
    if not s:
        return (0, "", 0)
    m = re.match(r"([A-Za-z]*)(\d*)", s.replace(" ", ""))
    if m:
        prefix, num = m.groups()
        return (1, prefix.upper(), int(num) if num else 0)
    return (1, s.upper(), 0)



def normalize_unit_or_address2_value(value) -> str:
    """Clean apartment / address-line-2 values for USPS-style exports.

    Rules:
    - Keep real units like A, 2, 12E, C47, C-47, Apt 330, Unit B.
    - Keep PO Box values.
    - Drop street types, complex/building words, and broken values like
      APT, APTS, APT RD, APT DR, APT STEEPLE, APT WATERFORD.
    """
    s = normalize_export_text(value)
    if not s:
        return ""
    s = s.replace(".", " ").replace(",", " ")
    s = collapse_spaces(s)
    up = s.upper().strip()

    street_type_words = {
        "ST", "STREET", "RD", "ROAD", "DR", "DRIVE", "LN", "LANE", "CIR", "CIRCLE",
        "CT", "COURT", "AVE", "AVENUE", "BLVD", "BOULEVARD", "WAY", "PL", "PLACE",
        "TER", "TERRACE", "PKWY", "PARKWAY", "HWY", "HIGHWAY", "PIKE", "TRL", "TRAIL",
        "RDG", "RIDGE", "RUN", "N", "S", "E", "W", "NE", "NW", "SE", "SW"
    }
    unit_prefixes = {"APT", "APARTMENT", "APARTMENTS", "APTS", "UNIT", "STE", "SUITE", "#"}
    bad_unit_words = street_type_words | unit_prefixes | {
        "BLDG", "BUILDING", "FLOOR", "FL", "ROOM", "RM", "LOT",
        "WATERFORD", "WATERTFORD", "STEEPLE", "STEEPLECHASE", "APTS", "APT"
    }

    if up in bad_unit_words:
        return ""

    # Keep real postal secondary lines.
    if re.match(r"^(P\s*O\s*BOX|PO\s+BOX|BOX)\s+", up):
        box_num = re.sub(r"^(P\s*O\s*BOX|PO\s+BOX|BOX)\s+", "", up).strip()
        return collapse_spaces(f"PO Box {box_num}")

    cleaned = up
    for _ in range(4):
        m = re.match(r"^(APT|APARTMENT|APARTMENTS|APTS|UNIT|STE|SUITE|#)\s*(.*)$", cleaned)
        if not m:
            break
        cleaned = m.group(2).strip()

    cleaned = cleaned.strip(" -#")
    cleaned = collapse_spaces(cleaned)
    if not cleaned:
        return ""
    if cleaned in bad_unit_words:
        return ""

    tokens = [t for t in re.split(r"\s+", cleaned) if t]
    if tokens and all(t in bad_unit_words for t in tokens):
        return ""

    # Reject pure words/complex names. Good units almost always contain a digit,
    # are a single letter, or are a short letter-number code like C47 / 12E / C-47.
    compact = cleaned.replace("-", "").replace(" ", "")
    if not re.fullmatch(r"[A-Z]?\d+[A-Z]?|[A-Z]\d+[A-Z]?|\d+[A-Z]|[A-Z]", compact):
        return ""

    return "Apt " + title_case_report_value(cleaned.replace(" ", "-" if "-" in cleaned else " "))


def normalize_walk_address_label(value, address_value="", street_value="") -> str:
    """Ensure walk sheet address headers read House Number + Street + Apt.
    Also repairs older labels like 'Cape Horn Rd | 850'.
    """
    s = normalize_export_text(value)
    if "|" in s:
        parts = [collapse_spaces(p) for p in s.split("|") if collapse_spaces(p)]
        if len(parts) >= 2:
            first, second = parts[0], parts[1]
            # If the first part looks like a street and second begins with a house number, reverse them.
            if re.search(r"\b(ST|STREET|RD|ROAD|DR|DRIVE|LN|LANE|CIR|CIRCLE|CT|COURT|AVE|AVENUE|BLVD|WAY|PL|PLACE)\b", first.upper()) and re.match(r"^\d", second):
                s = collapse_spaces(second + " " + first)
            else:
                s = collapse_spaces(" ".join(parts))
    if not s:
        s = collapse_spaces(" ".join([normalize_export_text(address_value), normalize_address_value(street_value)]))
    return s


def build_walk_address_display(house_value, street_value, unit_value="") -> str:
    house = normalize_export_text(house_value)
    street = normalize_address_value(street_value)
    unit = normalize_unit_or_address2_value(unit_value)
    return collapse_spaces(" ".join([x for x in [house, street, unit] if x]))


def build_area_break_counts_from_street_df(street_df: pd.DataFrame) -> pd.DataFrame:
    """Counts for Excel exports only. This runs only after the user clicks Prepare."""
    if street_df is None or street_df.empty:
        return pd.DataFrame(columns=["Area Type", "Area", "Voters", "Households"])

    df = street_df.copy()
    for col in ["Municipality", "Precinct", "StreetGroup", "AddressLine"]:
        if col not in df.columns:
            df[col] = ""

    muni_values = sorted({normalize_export_text(v) for v in df["Municipality"].tolist() if normalize_export_text(v)})
    area_col = "Municipality" if len(muni_values) > 1 else "Precinct"
    area_type = "Municipality" if area_col == "Municipality" else "Precinct"

    df["_area"] = df[area_col].apply(normalize_export_text)
    df.loc[df["_area"].eq(""), "_area"] = "(Blank)"
    df["_hh_key"] = (
        df["StreetGroup"].apply(normalize_export_text) + "|" +
        df["AddressLine"].apply(normalize_export_text)
    )

    rows = []
    for area, grp in df.groupby("_area", dropna=False, sort=True):
        households = grp["_hh_key"].replace("", pd.NA).dropna().nunique()
        rows.append({
            "Area Type": area_type,
            "Area": area,
            "Voters": int(len(grp)),
            "Households": int(households),
        })

    total_households = df["_hh_key"].replace("", pd.NA).dropna().nunique()
    rows.append({
        "Area Type": "Total",
        "Area": "TOTAL",
        "Voters": int(len(df)),
        "Households": int(total_households),
    })
    return pd.DataFrame(rows, columns=["Area Type", "Area", "Voters", "Households"])


def clear_prepared_download_state(key: str):
    st.session_state.pop(key, None)


def expand_party_label(code: str) -> str:
    mapping = {"R": "Republicans", "D": "Democrats", "O": "Others"}
    return mapping.get(normalize_export_text(code).upper(), normalize_export_text(code))

def expand_mib_application_label(code: str) -> str:
    mapping = {"APP": "Applied", "DEC": "Declined", "DNA": "None", "": "None"}
    return mapping.get(normalize_export_text(code).upper(), normalize_export_text(code).title())

def summarize_vote_history(picks: list[str]) -> str:
    vals = [normalize_export_text(v) for v in picks if normalize_export_text(v)]
    nums = []
    for v in vals:
        m = re.search(r"(\d+)", v)
        if m:
            nums.append(int(m.group(1)))
    if not nums:
        return ", ".join(vals)
    nums = sorted(set(nums))
    if nums == [4]:
        return "All of the last 4"
    if len(nums) == 1:
        return f"{nums[0]} of the last 4"
    return f"{nums[0]}-{nums[-1]} of the last 4"

def selected_area_desc(active_filters: dict) -> str:
    counties = active_filters.get("County", []) or []
    municipalities = active_filters.get("Municipality", []) or []
    if len(counties) > 1:
        return ", ".join(counties)
    if len(counties) == 1 and municipalities:
        if len(municipalities) == 1:
            return municipalities[0]
        return ", ".join(municipalities[:4]) + (" ..." if len(municipalities) > 4 else "")
    if len(counties) == 1:
        return counties[0]
    if municipalities:
        if len(municipalities) == 1:
            return municipalities[0]
        return ", ".join(municipalities[:4]) + (" ..." if len(municipalities) > 4 else "")
    return "Selected Area"


def build_filter_summary_lines(active_filters: dict) -> list[str]:
    lines = []

    municipalities = active_filters.get("Municipality", []) or []
    if municipalities:
        if len(municipalities) == 1:
            lines.append(f"Municipality: Selected precincts in {municipalities[0].title()}")
        else:
            muni_text = ", ".join(m.title() for m in municipalities[:4])
            if len(municipalities) > 4:
                muni_text += " ..."
            lines.append(f"Municipality: Selected precincts in {muni_text}")

    parties = active_filters.get("party_pick", []) or []
    if parties:
        expanded = ", ".join(expand_party_label(p) for p in parties)
        lines.append(f"Party: {expanded}")

    vote_history_type = active_filters.get("vote_history_type", "All")
    vote_history_range = active_filters.get("vote_history_range")
    if vote_history_range is not None:
        lines.append(f"Vote History ({vote_history_type}): {int(vote_history_range[0])}-{int(vote_history_range[1])} of the last 4")

    mib_app = active_filters.get("mib_applied_pick", []) or []
    if mib_app:
        expanded = ", ".join(expand_mib_application_label(v) for v in mib_app)
        lines.append(f"Mail in Ballot Application: {expanded}")

    mib_vote = active_filters.get("mib_ballot_pick", []) or []
    if mib_vote:
        expanded = ", ".join(normalize_export_text(v).title() for v in mib_vote)
        lines.append(f"Mail Ballot Vote Status: {expanded}")

    mb_perm = active_filters.get("mb_perm_pick", []) or []
    if mb_perm:
        expanded = ", ".join("Y" if normalize_export_text(v).upper() == "Y" else normalize_export_text(v) for v in mb_perm)
        lines.append(f"Permanent Mail Ballot: {expanded}")

    source_file = active_filters.get("source_file_pick", []) or []
    if source_file:
        lines.append(f"Current Mail Ballot Voters: {', '.join(map(str, source_file))}")
    # Mail Ballot New Registrant is intentionally hidden from the main universe summary.
    sent_status = active_filters.get("current_ballot_sent_status", "All")
    if sent_status and sent_status != "All":
        lines.append(f"Current Ballot Sent: {sent_status}")
    returned_status = active_filters.get("current_ballot_returned_status", "All")
    if returned_status and returned_status != "All":
        lines.append(f"Current Ballot Returned: {returned_status}")

    for key, label in [("County","County"),("Precinct","Precinct"),("USC","USC"),("STS","STS"),("STH","STH"),("School District","School District"),
                       ("hh_party_pick","Household Party"),("calc_party_pick","Calculated Party"),("gender_pick","Gender"),
                       ("age_range_pick","Age Range")]:
        val = active_filters.get(key)
        if isinstance(val, list) and val:
            lines.append(f"{label}: {', '.join(map(str, val[:8]))}" + (" ..." if len(val) > 8 else ""))

    if active_filters.get("new_reg_months", 0):
        lines.append(f"Newly Registered: within last {active_filters['new_reg_months']} month(s)")
    for key, label in [("has_email","Email"),("has_landline","Landline"),("has_mobile","Mobile"),("has_applicant_phone","Applicant Phone")]:
        val = active_filters.get(key)
        if val and val != "All":
            lines.append(f"{label}: {val}")
    for key, label in [("applicant_phone_type_pick", "Applicant Phone Type"), ("applicant_phone_compliance_pick", "Applicant Phone Compliance")]:
        val = active_filters.get(key) or []
        if val:
            lines.append(f"{label}: {', '.join(map(str, val))}")
    return lines or ["No additional filters selected"]


def summarize_universe_filters(active_filters: dict) -> str:
    parts = build_filter_summary_lines(active_filters)
    contact_status = normalize_export_text(active_filters.get("contact_status", "All"))
    if contact_status and contact_status != "All":
        parts.append(f"Contact Status: {contact_status}")
    nh_status = normalize_export_text(active_filters.get("global_nh", "All"))
    if nh_status and nh_status != "All":
        parts.append(f"Not Home: {nh_status}")
    follow_up_status = normalize_export_text(active_filters.get("global_follow_up", "All"))
    if follow_up_status and follow_up_status != "All":
        parts.append(f"Follow-Up: {follow_up_status}")
    support_level = normalize_export_text(active_filters.get("global_support_level", "All"))
    if support_level and support_level != "All":
        parts.append(f"Support Level: {support_level}")
    return " | ".join(parts) if parts else "No filters"


def apply_followup_preset(preset_name: str):
    current = dict(st.session_state.get("active_filters", {}) or {})
    current["contact_status"] = "All"
    current["global_nh"] = "All"
    current["global_follow_up"] = "All"
    current["global_support_level"] = "All"

    if preset_name == "Re-Knock List":
        current["global_nh"] = "Yes"
    elif preset_name == "Follow-Up List":
        current["global_follow_up"] = "Yes"
    elif preset_name == "GOTV Supporters":
        current["global_support_level"] = "Strong"
    elif preset_name == "Undecided Persuasion":
        current["global_support_level"] = "Undecided"
    elif preset_name == "Yard Sign Follow-Up":
        current["contact_status"] = "Contacted"

    st.session_state.active_filters = current
    st.session_state.filters_applied = True
    st.session_state.workspace_mode = "landing"
    st.session_state.lookup_view_active = False
    st.rerun()

def get_global_support_level_options() -> list[str]:
    uploaded = st.session_state.get("walk_results_df")
    if isinstance(uploaded, pd.DataFrame) and not uploaded.empty and "Support Level" in uploaded.columns:
        vals = sorted({normalize_export_text(v) for v in uploaded["Support Level"].tolist() if normalize_export_text(v)})
        return ["All"] + vals
    return ["All", "Strong", "Lean", "Undecided", "Oppose"]

def has_global_followup_filters(active_filters: dict) -> bool:
    if not isinstance(active_filters, dict):
        return False
    return any(
        normalize_export_text(active_filters.get(key, "All")) not in {"", "All"}
        for key in ["contact_status", "global_nh", "global_follow_up", "global_support_level"]
    )

def apply_global_followup_filters_df(df: pd.DataFrame, active_filters: dict) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    out = df.copy()
    out = merge_uploaded_street_results_into_detail_df(out)
    out = merge_uploaded_walk_results_into_detail_df(out)

    for field in ["F", "A", "U", "NH", "Yard Sign", "Notes", "Contacted", "Result", "Support Level", "Follow-Up", "Walk Notes"]:
        if field not in out.columns:
            out[field] = ""

    street_contact_mask = (
        out["F"].astype(str).str.strip().ne("") |
        out["A"].astype(str).str.strip().ne("") |
        out["U"].astype(str).str.strip().ne("") |
        out["NH"].astype(str).str.strip().ne("") |
        out["Yard Sign"].astype(str).str.strip().ne("") |
        out["Notes"].astype(str).str.strip().ne("")
    )
    walk_contact_mask = (
        out["Contacted"].astype(str).str.strip().ne("") |
        out["Result"].astype(str).str.strip().ne("") |
        out["Support Level"].astype(str).str.strip().ne("") |
        out["Follow-Up"].astype(str).str.strip().ne("") |
        out["Walk Notes"].astype(str).str.strip().ne("")
    )
    contact_mask = street_contact_mask | walk_contact_mask

    contact_status = normalize_export_text(active_filters.get("contact_status", "All"))
    if contact_status == "Contacted":
        out = out[contact_mask]
    elif contact_status == "Not Contacted":
        out = out[~contact_mask]

    nh_status = normalize_export_text(active_filters.get("global_nh", "All"))
    nh_mask = (
        out["NH"].astype(str).str.strip().ne("") |
        out["Result"].astype(str).str.upper().str.replace(" ", "", regex=False).isin(["NOTHOME", "NH"])
    )
    if nh_status == "Yes":
        out = out[nh_mask]
    elif nh_status == "No":
        out = out[~nh_mask]

    follow_up_status = normalize_export_text(active_filters.get("global_follow_up", "All"))
    follow_up_mask = out["Follow-Up"].astype(str).str.strip().ne("")
    if follow_up_status == "Yes":
        out = out[follow_up_mask]
    elif follow_up_status == "No":
        out = out[~follow_up_mask]

    support_level = normalize_export_text(active_filters.get("global_support_level", "All"))
    if support_level and support_level != "All":
        out = out[
            out["Support Level"].astype(str).str.strip().str.casefold() == support_level.casefold()
        ]

    return out


def query_dashboard_followup_stats(active_filters: dict) -> dict:
    if use_large_filter_mode(active_filters, columns):
        return {
            "contacted_pct": 0,
            "nh_pct": 0,
            "followup_pct": 0,
            "strong_pct": 0,
            "undecided_pct": 0,
            "contacted_count": 0,
            "nh_count": 0,
            "followup_count": 0,
            "strong_count": 0,
            "undecided_count": 0,
            "large_mode": True,
        }

    df = fetch_filtered_detail(active_filters)
    if df is None or df.empty:
        return {
            "contacted_pct": 0,
            "nh_pct": 0,
            "followup_pct": 0,
            "strong_pct": 0,
            "undecided_pct": 0,
            "contacted_count": 0,
            "nh_count": 0,
            "followup_count": 0,
            "strong_count": 0,
            "undecided_count": 0,
            "large_mode": False,
        }

    df = merge_uploaded_street_results_into_detail_df(df)
    df = merge_uploaded_walk_results_into_detail_df(df)

    for field in ["F", "A", "U", "NH", "Yard Sign", "Notes", "Contacted", "Result", "Support Level", "Follow-Up", "Walk Notes"]:
        if field not in df.columns:
            df[field] = ""

    total = max(len(df), 1)

    street_contact_mask = (
        df["F"].astype(str).str.strip().ne("") |
        df["A"].astype(str).str.strip().ne("") |
        df["U"].astype(str).str.strip().ne("") |
        df["NH"].astype(str).str.strip().ne("") |
        df["Yard Sign"].astype(str).str.strip().ne("") |
        df["Notes"].astype(str).str.strip().ne("")
    )
    walk_contact_mask = (
        df["Contacted"].astype(str).str.strip().ne("") |
        df["Result"].astype(str).str.strip().ne("") |
        df["Support Level"].astype(str).str.strip().ne("") |
        df["Follow-Up"].astype(str).str.strip().ne("") |
        df["Walk Notes"].astype(str).str.strip().ne("")
    )
    contacted_mask = street_contact_mask | walk_contact_mask

    nh_mask = (
        df["NH"].astype(str).str.strip().ne("") |
        df["Result"].astype(str).str.upper().str.replace(" ", "", regex=False).isin(["NOTHOME", "NH"])
    )
    followup_mask = df["Follow-Up"].astype(str).str.strip().ne("")
    support_series = df["Support Level"].astype(str).str.strip().str.casefold()
    strong_mask = support_series.eq("strong")
    undecided_mask = support_series.eq("undecided")

    def pct(mask):
        return round((int(mask.sum()) / total) * 100)

    return {
        "contacted_pct": pct(contacted_mask),
        "nh_pct": pct(nh_mask),
        "followup_pct": pct(followup_mask),
        "strong_pct": pct(strong_mask),
        "undecided_pct": pct(undecided_mask),
        "contacted_count": int(contacted_mask.sum()),
        "nh_count": int(nh_mask.sum()),
        "followup_count": int(followup_mask.sum()),
        "strong_count": int(strong_mask.sum()),
        "undecided_count": int(undecided_mask.sum()),
        "large_mode": False,
    }

def _query_metrics_from_detail(active_filters, columns):
    df = fetch_filtered_detail(active_filters)
    if df is None or df.empty:
        return {
            "voters": 0,
            "households": 0,
            "emails": 0,
            "landlines": 0,
            "mobiles": 0,
            "unique_counties": 0,
            "unique_precincts": 0,
        }

    hh = df["_HouseholdKey"].fillna("").astype(str) if "_HouseholdKey" in df.columns else pd.Series([""] * len(df))
    households = int(hh.replace("", pd.NA).dropna().nunique() + hh.eq("").sum())
    return {
        "voters": int(len(df)),
        "households": households,
        "emails": int(df.get("_HasEmail", pd.Series([False] * len(df))).fillna(False).astype(bool).sum()),
        "landlines": int(df.get("_HasLandline", pd.Series([False] * len(df))).fillna(False).astype(bool).sum()),
        "mobiles": int(df.get("_HasMobile", pd.Series([False] * len(df))).fillna(False).astype(bool).sum()),
        "unique_counties": int(df["County"].fillna("").astype(str).replace("", pd.NA).dropna().nunique()) if "County" in df.columns else 0,
        "unique_precincts": int(df["Precinct"].fillna("").astype(str).replace("", pd.NA).dropna().nunique()) if "Precinct" in df.columns else 0,
    }

def _query_chart_from_detail(active_filters, group_expr, label, not_blank=True):
    df = fetch_filtered_detail(active_filters)
    if df is None or df.empty:
        return pd.DataFrame(columns=[label, "Count"])

    series_name = None
    if group_expr in df.columns:
        series_name = group_expr
    elif group_expr == "_PartyNorm" and "_PartyNorm" in df.columns:
        series_name = "_PartyNorm"
    elif group_expr == "_Gender" and "_Gender" in df.columns:
        series_name = "_Gender"
    elif group_expr == "_AgeRange" and "_AgeRange" in df.columns:
        series_name = "_AgeRange"

    if series_name is None:
        return pd.DataFrame(columns=[label, "Count"])

    ser = df[series_name]
    if not_blank:
        ser = ser[ser.fillna("").astype(str).str.strip() != ""]
    out = ser.fillna("(Blank)").astype(str).value_counts(dropna=False).reset_index()
    out.columns = [label, "Count"]
    return out

def _query_area_summary_from_detail(active_filters, area_col):
    df = fetch_filtered_detail(active_filters)
    if df is None or df.empty or area_col not in df.columns:
        return pd.DataFrame(columns=[area_col, "Individuals", "Households"])

    temp = df.copy()
    temp[area_col] = temp[area_col].fillna("(Blank)").astype(str)
    hh = temp["_HouseholdKey"].fillna("").astype(str) if "_HouseholdKey" in temp.columns else pd.Series([""] * len(temp))
    temp["_HouseholdKeySafe"] = hh
    rows = []
    for area_val, grp in temp.groupby(area_col, dropna=False):
        grp_hh = grp["_HouseholdKeySafe"]
        households = int(grp_hh.replace("", pd.NA).dropna().nunique() + grp_hh.eq("").sum())
        rows.append({
            area_col: area_val if normalize_export_text(area_val) else "(Blank)",
            "Individuals": int(len(grp)),
            "Households": households,
        })
    out = pd.DataFrame(rows).sort_values(["Individuals", area_col], ascending=[False, True]).reset_index(drop=True)
    return out



def get_street_results_template_csv_bytes():
    template_df = pd.DataFrame(columns=["PA ID Number", "F", "A", "U", "NH", "Yard Sign", "Notes"])
    return template_df.to_csv(index=False).encode("utf-8")


def get_street_results_sheet_bytes(active_filters):
    street_df = build_street_list_dataframe(active_filters)
    street_df = apply_uploaded_street_result_filters(street_df)

    export_cols = [
        "Precinct", "StreetGroup", "AddressLine", "FullName", "Phone", "Party", "Sex", "Age",
        "PA ID Number", "F", "A", "U", "NH", "Yard Sign", "Notes"
    ]
    for col in export_cols:
        if col not in street_df.columns:
            street_df[col] = ""

    export_df = street_df[export_cols].copy().rename(columns={
        "StreetGroup": "Street",
        "AddressLine": "Address",
        "FullName": "Name",
        "Sex": "Gender",
    })
    counts_df = build_area_break_counts_from_street_df(street_df)

    wb = Workbook()
    ws_counts = wb.active
    ws_counts.title = "Area Counts"
    ws = wb.create_sheet("Street List Data")

    title_fill = PatternFill("solid", fgColor="7A1523")
    header_fill = PatternFill("solid", fgColor="9F2032")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D7B7BC")

    ws_counts["A1"] = "Candidate Connect Area Break Counts"
    ws_counts["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_counts["A1"].fill = title_fill
    ws_counts["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"
    ws_counts["A2"].font = Font(italic=True, size=10)
    ws_counts.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws_counts.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    for c, header in enumerate(counts_df.columns.tolist(), start=1):
        cell = ws_counts.cell(row=4, column=c, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r, row in enumerate(counts_df.itertuples(index=False), start=5):
        for c, value in enumerate(row, start=1):
            cell = ws_counts.cell(row=r, column=c, value=value)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if c in {3, 4}:
                cell.alignment = Alignment(horizontal="center")
    for letter, width in {"A": 18, "B": 36, "C": 14, "D": 14}.items():
        ws_counts.column_dimensions[letter].width = width
    ws_counts.freeze_panes = "A5"

    ws["A1"] = "Candidate Connect Street List Tracking Sheet"
    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"
    ws["A3"] = "Enter X in F, A, U, NH, and Yard Sign. Use Notes for anything important from the candidate's conversation."
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_df.columns))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(export_df.columns))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(export_df.columns))
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, size=10)
    ws["A3"].font = Font(size=10)

    headers = export_df.columns.tolist()
    header_row = 5
    check_fill = PatternFill("solid", fgColor="F9E8EA")

    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row in enumerate(export_df.itertuples(index=False), start=header_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            header = headers[c - 1]
            if header in {"F", "A", "U", "NH", "Yard Sign"}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = check_fill
                cell.font = Font(bold=True)
            elif header == "Notes":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="center")

    widths = {
        "Precinct": 18, "Street": 24, "Address": 14, "Name": 24, "Phone": 16, "Party": 8,
        "Gender": 8, "Age": 8, "PA ID Number": 16, "F": 5, "A": 5, "U": 5, "NH": 6,
        "Yard Sign": 10, "Notes": 28
    }
    for c, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(header, 14)

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(header_row, ws.max_row)}"
    for r in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

def _normalized_col_lookup(columns):
    lookup = {}
    for col in columns:
        key = re.sub(r"[^a-z0-9]+", "", str(col).strip().lower())
        if key and key not in lookup:
            lookup[key] = col
    return lookup

def _find_uploaded_results_column(columns, candidates):
    lookup = _normalized_col_lookup(columns)
    for candidate in candidates:
        key = re.sub(r"[^a-z0-9]+", "", str(candidate).strip().lower())
        if key in lookup:
            return lookup[key]
    return None

def normalize_tracking_mark(val):
    s = normalize_export_text(val).upper()
    return "X" if s in {"X", "Y", "YES", "TRUE", "T", "1", "CHECK", "CHECKED"} else ""

def standardize_uploaded_street_results(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["PA ID Number", "F", "A", "U", "NH", "Yard Sign", "Notes"])

    pa_id_col = _find_uploaded_results_column(
        df.columns.tolist(),
        ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID", "Voter ID", "VoterID"]
    )
    if pa_id_col is None:
        return pd.DataFrame(columns=["PA ID Number", "F", "A", "U", "NH", "Yard Sign", "Notes"])

    out = pd.DataFrame()
    out["PA ID Number"] = df[pa_id_col].apply(normalize_numeric_string)
    field_map = {
        "F": ["F"],
        "A": ["A"],
        "U": ["U"],
        "NH": ["NH", "Not Home", "NotHome"],
        "Yard Sign": ["Yard Sign", "YardSign", "Sign", "Yard"],
        "Notes": ["Notes", "Note", "Comments", "Comment"],
    }
    for field, candidates in field_map.items():
        col = _find_uploaded_results_column(df.columns.tolist(), candidates)
        if col is None:
            out[field] = ""
        elif field == "Notes":
            out[field] = df[col].apply(normalize_export_text)
        else:
            out[field] = df[col].apply(normalize_tracking_mark)

    out = out[out["PA ID Number"].astype(str).str.strip() != ""].copy()
    out = out.drop_duplicates(subset=["PA ID Number"], keep="last").reset_index(drop=True)
    return out

def merge_uploaded_street_results_into_detail_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    merged = df.copy()
    uploaded = st.session_state.get("street_results_df")
    if not isinstance(uploaded, pd.DataFrame) or uploaded.empty:
        for field in ["F", "A", "U", "NH", "Yard Sign", "Notes"]:
            if field not in merged.columns:
                merged[field] = ""
        return merged

    pa_id_col = first_existing_detail(
        merged.columns.tolist(),
        ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID", "Voter ID", "VoterID"]
    )
    if pa_id_col is None:
        for field in ["F", "A", "U", "NH", "Yard Sign", "Notes"]:
            if field not in merged.columns:
                merged[field] = ""
        return merged

    merged["PA ID Number"] = merged[pa_id_col].apply(normalize_numeric_string)
    merge_cols = ["PA ID Number", "F", "A", "U", "NH", "Yard Sign", "Notes"]
    merged = merged.merge(uploaded[merge_cols], on="PA ID Number", how="left")
    for field in ["F", "A", "U", "NH", "Yard Sign", "Notes"]:
        merged[field] = merged[field].fillna("").astype(str)
    return merged

def apply_uploaded_street_result_filters(street_df: pd.DataFrame) -> pd.DataFrame:
    if street_df is None or street_df.empty:
        return street_df

    filters = st.session_state.get("street_results_filters", {}) or {}
    out = street_df.copy()
    for field in ["F", "A", "U", "NH", "Yard Sign"]:
        mode = normalize_export_text(filters.get(field, "All"))
        if mode == "Marked":
            out = out[out[field].astype(str).str.strip() != ""]
        elif mode == "Unmarked":
            out = out[out[field].astype(str).str.strip() == ""]
    return out


def get_walk_sheet_tracking_template_csv_bytes():
    template_df = pd.DataFrame(columns=["PA ID Number", "Contacted", "Result", "Support Level", "Follow-Up", "Notes"])
    return template_df.to_csv(index=False).encode("utf-8")

def normalize_walk_result_value(val):
    return normalize_export_text(val).title()

def standardize_uploaded_walk_results(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["PA ID Number", "Contacted", "Result", "Support Level", "Follow-Up", "Notes"])

    pa_id_col = _find_uploaded_results_column(
        df.columns.tolist(),
        ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID", "Voter ID", "VoterID"]
    )
    if pa_id_col is None:
        return pd.DataFrame(columns=["PA ID Number", "Contacted", "Result", "Support Level", "Follow-Up", "Notes"])

    out = pd.DataFrame()
    out["PA ID Number"] = df[pa_id_col].apply(normalize_numeric_string)

    field_map = {
        "Contacted": ["Contacted", "Contact", "C"],
        "Result": ["Result", "Outcome", "Canvass Result"],
        "Support Level": ["Support Level", "Support", "SupportLevel"],
        "Follow-Up": ["Follow-Up", "Follow Up", "Followup", "F"],
        "Notes": ["Notes", "Note", "Comments", "Comment"],
    }

    for field, candidates in field_map.items():
        col = _find_uploaded_results_column(df.columns.tolist(), candidates)
        if col is None:
            out[field] = ""
        elif field in {"Contacted", "Follow-Up"}:
            out[field] = df[col].apply(normalize_tracking_mark)
        elif field == "Notes":
            out[field] = df[col].apply(normalize_export_text)
        else:
            out[field] = df[col].apply(normalize_walk_result_value)

    out = out[out["PA ID Number"].astype(str).str.strip() != ""].copy()
    out = out.drop_duplicates(subset=["PA ID Number"], keep="last").reset_index(drop=True)
    return out

def merge_uploaded_walk_results_into_detail_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    merged = df.copy()
    uploaded = st.session_state.get("walk_results_df")
    if not isinstance(uploaded, pd.DataFrame) or uploaded.empty:
        for field in ["Contacted", "Result", "Support Level", "Follow-Up", "Walk Notes"]:
            if field not in merged.columns:
                merged[field] = ""
        return merged

    pa_id_col = first_existing_detail(
        merged.columns.tolist(),
        ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID", "Voter ID", "VoterID"]
    )
    if pa_id_col is None:
        for field in ["Contacted", "Result", "Support Level", "Follow-Up", "Walk Notes"]:
            if field not in merged.columns:
                merged[field] = ""
        return merged

    merged["PA ID Number"] = merged[pa_id_col].apply(normalize_numeric_string)
    merge_cols = ["PA ID Number", "Contacted", "Result", "Support Level", "Follow-Up", "Notes"]
    uploaded_for_merge = uploaded[merge_cols].rename(columns={"Notes": "_UploadedWalkNotes"})
    merged = merged.merge(uploaded_for_merge, on="PA ID Number", how="left")
    merged["Walk Notes"] = merged["_UploadedWalkNotes"].fillna("").astype(str) if "_UploadedWalkNotes" in merged.columns else ""
    if "_UploadedWalkNotes" in merged.columns:
        merged = merged.drop(columns=["_UploadedWalkNotes"])
    for field in ["Contacted", "Result", "Support Level", "Follow-Up", "Walk Notes"]:
        if field not in merged.columns:
            merged[field] = ""
        merged[field] = merged[field].fillna("").astype(str)
    return merged

def apply_uploaded_walk_result_filters(street_df: pd.DataFrame) -> pd.DataFrame:
    if street_df is None or street_df.empty:
        return street_df

    filters = st.session_state.get("walk_results_filters", {}) or {}
    out = street_df.copy()

    for field in ["Contacted", "Follow-Up"]:
        mode = normalize_export_text(filters.get(field, "All"))
        if mode == "Marked":
            out = out[out[field].astype(str).str.strip() != ""]
        elif mode == "Unmarked":
            out = out[out[field].astype(str).str.strip() == ""]

    not_home_mode = normalize_export_text(filters.get("Not Home", "All"))
    result_upper = out["Result"].astype(str).str.upper().str.replace(" ", "", regex=False)
    if not_home_mode == "Marked":
        out = out[result_upper.isin(["NOTHOME", "NH"])]
    elif not_home_mode == "Unmarked":
        out = out[~result_upper.isin(["NOTHOME", "NH"])]

    support_level = normalize_export_text(filters.get("Support Level", "All"))
    if support_level and support_level != "All":
        out = out[out["Support Level"].astype(str).str.strip().str.casefold() == support_level.casefold()]

    return out

def build_walk_sheet_tracking_excel_bytes(active_filters):
    street_df = build_street_list_dataframe(active_filters).copy()
    street_df = apply_uploaded_walk_result_filters(street_df)
    if street_df.empty:
        export_df = pd.DataFrame(columns=[
            "Precinct", "Street", "Address", "Name", "Phone", "Party", "Gender", "Age",
            "PA ID Number", "Contacted", "Result", "Support Level", "Follow-Up", "Notes"
        ])
    else:
        export_df = pd.DataFrame({
            "Precinct": street_df["Precinct"].apply(normalize_export_text),
            "Street": street_df["StreetGroup"].apply(normalize_export_text),
            "Address": street_df["WalkAddressDisplay"].apply(normalize_export_text) if "WalkAddressDisplay" in street_df.columns else street_df["AddressLine"].apply(normalize_export_text),
            "Name": street_df["FullName"].apply(normalize_export_text),
            "Phone": street_df["Phone"].apply(normalize_export_text),
            "Party": street_df["Party"].apply(normalize_export_text),
            "Gender": street_df["Sex"].apply(normalize_export_text),
            "Age": street_df["Age"].apply(normalize_export_text),
            "PA ID Number": street_df["PA ID Number"].apply(normalize_numeric_string),
            "Contacted": street_df["Contacted"].apply(normalize_export_text) if "Contacted" in street_df.columns else pd.Series([""] * len(street_df)),
            "Result": street_df["Result"].apply(normalize_export_text) if "Result" in street_df.columns else pd.Series([""] * len(street_df)),
            "Support Level": street_df["Support Level"].apply(normalize_export_text) if "Support Level" in street_df.columns else pd.Series([""] * len(street_df)),
            "Follow-Up": street_df["Follow-Up"].apply(normalize_export_text) if "Follow-Up" in street_df.columns else pd.Series([""] * len(street_df)),
            "Notes": street_df.get("Walk Notes", pd.Series([""] * len(street_df))).apply(normalize_export_text),
        })

    counts_df = build_area_break_counts_from_street_df(street_df)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        counts_df.to_excel(writer, sheet_name="Area Counts", index=False, startrow=3)
        export_df.to_excel(writer, sheet_name="Walk Sheet Data", index=False, startrow=4)
        workbook = writer.book
        counts_ws = writer.sheets["Area Counts"]
        worksheet = writer.sheets["Walk Sheet Data"]

        title_font = Font(bold=True, size=14, color="7A1523")
        sub_font = Font(italic=True, size=10, color="555555")
        header_fill = PatternFill(fill_type="solid", fgColor="7A1523")
        header_font = Font(bold=True, color="FFFFFF")
        thin_side = Side(style="thin", color="C9B0B4")
        box_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        box_fill = PatternFill(fill_type="solid", fgColor="F8EDED")
        note_fill = PatternFill(fill_type="solid", fgColor="FFF9F9")
        center_align = Alignment(horizontal="center", vertical="center")
        wrap_align = Alignment(vertical="top", wrap_text=True)

        counts_ws["A1"] = "Candidate Connect Area Break Counts"
        counts_ws["A1"].font = title_font
        counts_ws["A2"] = f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
        counts_ws["A2"].font = sub_font
        for cell in counts_ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        for col_letter, width in {"A": 18, "B": 36, "C": 14, "D": 14}.items():
            counts_ws.column_dimensions[col_letter].width = width
        counts_ws.freeze_panes = "A5"

        worksheet["A1"] = "Candidate Connect Walk Sheet Tracking Sheet"
        worksheet["A1"].font = title_font
        worksheet["A2"] = f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
        worksheet["A2"].font = sub_font
        worksheet["A3"] = "Enter X in Contacted or Follow-Up, type Not Home or another result in Result, and fill Support Level / Notes as needed."
        worksheet["A3"].font = sub_font

        for cell in worksheet[5]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        width_map = {
            "A": 14, "B": 22, "C": 24, "D": 28, "E": 18, "F": 8, "G": 9, "H": 8,
            "I": 15, "J": 11, "K": 16, "L": 16, "M": 11, "N": 28
        }
        for col_letter, width in width_map.items():
            worksheet.column_dimensions[col_letter].width = width

        max_row = worksheet.max_row
        for row in range(6, max_row + 1):
            for col_letter in ["J", "M"]:
                cell = worksheet[f"{col_letter}{row}"]
                cell.border = box_border
                cell.fill = box_fill
                cell.alignment = center_align
            for col_letter in ["K", "L", "N"]:
                cell = worksheet[f"{col_letter}{row}"]
                cell.border = box_border
                cell.fill = note_fill
                cell.alignment = wrap_align

        worksheet.freeze_panes = "A6"

    return output.getvalue()

def build_street_list_dataframe(active_filters):
    df = fetch_filtered_detail(active_filters).copy()
    df = merge_uploaded_street_results_into_detail_df(df)
    df = merge_uploaded_walk_results_into_detail_df(df)
    if df.empty:
        return pd.DataFrame(columns=[
            "Municipality","Precinct","StreetGroup","AddressLine","UnitDisplay","WalkAddressDisplay","FullName","Phone","Party","Sex","Age","PA ID Number",
            "F","A","U","NH","Yard Sign","Notes","Contacted","Result","Support Level","Follow-Up","Walk Notes","MB_Perm","HouseNumSort","AptSort"
        ])

    municipality_col = first_existing_detail(df.columns.tolist(), ["Municipality"])
    precinct_col = first_existing_detail(df.columns.tolist(), ["Precinct"])
    street_col = first_existing_detail(df.columns.tolist(), ["Street Name"])
    house_col = first_existing_detail(df.columns.tolist(), ["House Number"])
    apt_col = first_existing_detail(df.columns.tolist(), ["Apartment Number"])
    address2_col = first_existing_detail(df.columns.tolist(), ["Address Line 2", "AddressLine2", "Address 2"])
    sex_col = first_existing_detail(df.columns.tolist(), ["Gender", "Sex"])
    age_col = first_existing_detail(df.columns.tolist(), ["Age"])
    party_col = first_existing_detail(df.columns.tolist(), ["Party"])
    pa_id_col = first_existing_detail(df.columns.tolist(), ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID", "Voter ID", "VoterID"])
    mb_perm_col = first_existing_detail(df.columns.tolist(), ["MB_PERM", "MB_Perm", "MB_Pern"])

    out = pd.DataFrame()
    out["Municipality"] = df[municipality_col].apply(normalize_export_text) if municipality_col else ""
    out["Precinct"] = df[precinct_col].apply(normalize_export_text) if precinct_col else ""
    street_vals = df[street_col].apply(normalize_address_value) if street_col else pd.Series([""] * len(df), index=df.index)
    house_vals = df[house_col].apply(normalize_export_text) if house_col else pd.Series([""] * len(df), index=df.index)
    apt_raw_vals = df[apt_col].apply(normalize_export_text) if apt_col else pd.Series([""] * len(df), index=df.index)
    addr2_raw_vals = df[address2_col].apply(normalize_export_text) if address2_col else pd.Series([""] * len(df), index=df.index)

    apt_vals = apt_raw_vals.apply(normalize_unit_or_address2_value)
    addr2_vals = addr2_raw_vals.apply(normalize_unit_or_address2_value)
    unit_vals = apt_vals.copy()
    unit_vals.loc[unit_vals.eq("") & addr2_vals.ne("")] = addr2_vals[unit_vals.eq("") & addr2_vals.ne("")]

    out["StreetGroup"] = street_vals
    out["UnitDisplay"] = unit_vals
    out["AddressLine"] = house_vals
    out.loc[unit_vals.ne(""), "AddressLine"] = out.loc[unit_vals.ne(""), "AddressLine"] + " " + unit_vals[unit_vals.ne("")]
    out["AddressLine"] = out["AddressLine"].apply(collapse_spaces).apply(normalize_address_value)
    out["WalkAddressDisplay"] = [
        build_walk_address_display(h, s, u)
        for h, s, u in zip(house_vals.tolist(), street_vals.tolist(), unit_vals.tolist())
    ]

    out["FullName"] = df.apply(full_name_from_row, axis=1).apply(normalize_name_value)
    out["Phone"] = df.apply(choose_best_phone, axis=1)
    out["Party"] = df[party_col].apply(normalize_export_text) if party_col else ""
    out["Sex"] = df[sex_col].apply(normalize_export_text) if sex_col else ""
    out["Age"] = df[age_col].apply(lambda v: normalize_numeric_string(v)) if age_col else ""
    out["PA ID Number"] = df[pa_id_col].apply(normalize_numeric_string) if pa_id_col else ""
    out["F"] = df["F"].apply(normalize_tracking_mark) if "F" in df.columns else ""
    out["A"] = df["A"].apply(normalize_tracking_mark) if "A" in df.columns else ""
    out["U"] = df["U"].apply(normalize_tracking_mark) if "U" in df.columns else ""
    out["NH"] = df["NH"].apply(normalize_tracking_mark) if "NH" in df.columns else ""
    out["Yard Sign"] = df["Yard Sign"].apply(normalize_tracking_mark) if "Yard Sign" in df.columns else ""
    out["Notes"] = df["Notes"].apply(normalize_export_text) if "Notes" in df.columns else ""
    out["Contacted"] = df["Contacted"].apply(normalize_tracking_mark) if "Contacted" in df.columns else ""
    out["Result"] = df["Result"].apply(normalize_walk_result_value) if "Result" in df.columns else ""
    out["Support Level"] = df["Support Level"].apply(normalize_export_text) if "Support Level" in df.columns else ""
    out["Follow-Up"] = df["Follow-Up"].apply(normalize_tracking_mark) if "Follow-Up" in df.columns else ""
    out["Walk Notes"] = df["Walk Notes"].apply(normalize_export_text) if "Walk Notes" in df.columns else ""
    out["MB_Perm"] = df[mb_perm_col].apply(normalize_mb_perm_value) if mb_perm_col else ""
    out["HouseNumSort"] = house_vals.apply(parse_house_number)
    out["AptSort"] = unit_vals.apply(parse_apartment_sort)

    out = out.sort_values(by=["Precinct", "StreetGroup", "HouseNumSort", "AptSort", "FullName"], kind="stable").reset_index(drop=True)
    return out


def build_precinct_summary(street_df: pd.DataFrame) -> pd.DataFrame:
    if street_df.empty:
        return pd.DataFrame(columns=["Precinct","Individuals","Households"])
    temp = street_df.copy()
    temp["_hh"] = temp["Precinct"].astype(str) + "|" + temp["AddressLine"].astype(str)
    grp = temp.groupby("Precinct", dropna=False).agg(
        Individuals=("FullName","count"),
        Households=("_hh", lambda s: s.nunique())
    ).reset_index()
    grp = grp.sort_values("Precinct").reset_index(drop=True)
    return grp



def get_mb_perm_display(row) -> str:
    try:
        for key in ["MB_Perm", "MB_PERM", "MB_Perm_Display", "_MBPerm"]:
            if key in row:
                val = str(row.get(key, "")).strip().upper()
                if val in {"TRUE", "T", "YES", "Y", "1"}:
                    return "Y"
                if val in {"FALSE", "F", "NO", "N", "0"}:
                    return "N"
                if val in {"Y", "N"}:
                    return val
    except Exception:
        return ""
    return ""

def make_precinct_bookmark_key(precinct: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9]+", "_", str(precinct)).strip("_")
    return f"precinct_{safe}" if safe else "precinct_unknown"


REPORT_NAVY = colors.HexColor("#7A1523")
REPORT_RED = colors.HexColor("#9F2032")
REPORT_LIGHT = colors.HexColor("#F9E8EA")
REPORT_GRID = colors.HexColor("#D7B7BC")
REPORT_STREET = colors.HexColor("#F2D7DB")

def truncate_text(value, max_len):
    s = normalize_export_text(value)
    return s if len(s) <= max_len else s[:max_len - 1] + "…"

def make_precinct_bookmark_key(precinct: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9]+", "_", str(precinct)).strip("_")
    return f"precinct_{safe}" if safe else "precinct_unknown"

def draw_footer(c, page_num, total_pages, printed_date):
    width, _ = c._pagesize
    c.setStrokeColor(REPORT_GRID)
    c.line(32, 28, width - 32, 28)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(width / 2, 16, f"{page_num} of {total_pages}")
    c.drawRightString(width - 36, 16, f"Updated: {printed_date}")


def draw_brand(c, y_top):
    width, _ = c._pagesize
    try:
        if CC_LOGO.exists():
            c.drawImage(ImageReader(str(CC_LOGO)), 30, y_top - 30, width=108, height=30, preserveAspectRatio=True, mask='auto')
    except Exception:
        pass
    try:
        if get_reporting_tss_logo_path().exists():
            c.drawImage(ImageReader(str(TSS_LOGO)), width - 132, y_top - 34, width=94, height=30, preserveAspectRatio=True, mask='auto')
    except Exception:
        pass
    c.setFillColor(REPORT_NAVY)
    c.setFont("Helvetica-Bold", 5.5)
    c.drawCentredString(width - 85, y_top - 5, "Powered By")

def _street_pdf_precinct_pages(street_df: pd.DataFrame):
    body_top = 480
    body_bottom = 42
    row_h = 14
    pages = 0
    for precinct, grp in street_df.groupby("Precinct", sort=False):
        current_street = None
        y = body_top - 10
        pages += 1
        for (street, address), addr_grp in grp.groupby(["StreetGroup", "AddressLine"], sort=False, dropna=False):
            need = len(addr_grp) + 1  # address row + voter rows
            if current_street != street:
                need += 1
            if y - (need * row_h) < body_bottom:
                pages += 1
                y = body_top - 10
                current_street = None
            if current_street != street:
                y -= row_h
                current_street = street
            y -= row_h  # address
            y -= row_h * len(addr_grp)
    return pages

def estimate_street_pdf_pages(summary_df: pd.DataFrame, street_df: pd.DataFrame):
    rows_per_summary_page = 26
    summary_pages = max(1, math.ceil(len(summary_df) / rows_per_summary_page)) if len(summary_df) else 1
    return 1 + summary_pages + _street_pdf_precinct_pages(street_df)


def _draw_cover_page(c, width, height, county_desc, party_desc, printed_date, totals_ind, totals_hh, filter_lines, page_num, total_pages):
    c.setFillColor(REPORT_NAVY)
    c.roundRect(34, height - 255, width - 68, 110, 14, fill=1, stroke=0)

    try:
        if CC_LOGO.exists():
            c.drawImage(ImageReader(str(CC_LOGO)), width/2 - 150, height - 105, width=300, height=84, preserveAspectRatio=True, mask='auto')
    except Exception:
        pass

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(width / 2, height - 173, "Voter Contact List")
    c.setFont("Helvetica", 11)
    c.drawCentredString(width / 2, height - 195, printed_date)
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, height - 214, f"Individuals: {totals_ind:,}   Households: {totals_hh:,}")

    c.setFillColor(REPORT_NAVY)
    c.setFont("Helvetica-Bold", 15)
    c.drawString(52, height - 305, "Selected Voters")
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 11)
    y = height - 327
    for line in filter_lines[:14]:
        c.drawString(62, y, f"• {line}")
        y -= 17
        if y < 114:
            break

    try:
        c.setFillColor(REPORT_NAVY)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width / 2, 84, "Powered By")
        if get_reporting_tss_logo_path().exists():
            c.drawImage(ImageReader(str(TSS_LOGO)), width/2 - 48, 42, width=96, height=30, preserveAspectRatio=True, mask='auto')
    except Exception:
        pass

    draw_footer(c, page_num, total_pages, printed_date)


def _draw_summary_page(c, width, height, chunk, printed_date, page_num, total_pages):
    draw_brand(c, height - 18)
    c.setFillColor(REPORT_NAVY)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(40, height - 72, "Precinct Counts Summary")

    table_x = 40
    table_y_top = height - 96
    table_w = width - 80
    row_h = 18
    precinct_w = table_w - 180

    c.setFillColor(REPORT_NAVY)
    c.rect(table_x, table_y_top - row_h, table_w, row_h, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(table_x + 8, table_y_top - 12, "Precinct")
    c.drawRightString(table_x + precinct_w + 80, table_y_top - 12, "Individuals")
    c.drawRightString(table_x + table_w - 10, table_y_top - 12, "Households")

    y = table_y_top - row_h
    for i, (_, row) in enumerate(chunk.iterrows()):
        y -= row_h
        fill = REPORT_LIGHT if i % 2 == 0 else colors.white
        if normalize_export_text(row["Precinct"]).upper() == "TOTAL":
            fill = REPORT_STREET
        c.setFillColor(fill)
        c.rect(table_x, y, table_w, row_h, fill=1, stroke=0)
        c.setStrokeColor(REPORT_GRID)
        c.rect(table_x, y, table_w, row_h, fill=0, stroke=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold" if normalize_export_text(row["Precinct"]).upper() == "TOTAL" else "Helvetica", 9)
        c.drawString(table_x + 8, y + 5, truncate_text(row["Precinct"], 42))
        c.drawRightString(table_x + precinct_w + 80, y + 5, f"{int(row['Individuals']):,}")
        c.drawRightString(table_x + table_w - 10, y + 5, f"{int(row['Households']):,}")

    draw_footer(c, page_num, total_pages, printed_date)


def _draw_precinct_page_header(c, width, height, precinct, page_in_precinct):
    draw_brand(c, height - 18)
    title = precinct if page_in_precinct == 1 else f"{precinct} (cont)"
    c.setFillColor(REPORT_NAVY)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(40, height - 74, title)

    c.setFillColor(REPORT_NAVY)
    c.roundRect(38, height - 106, width - 76, 22, 6, fill=1, stroke=0)

    cols = {
        "Full Name": 96, "Phone": 300, "Party": 448, "Sex": 478, "Age": 505,
        "F": 536, "A": 554, "U": 572, "NH": 590, "Yard Sign": 616, "MB Perm": 686
    }
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    for label, x in cols.items():
        c.drawString(x, height - 97, label)
    return cols

def generate_street_list_pdf_bytes(active_filters):
    street_df = build_street_list_dataframe(active_filters)
    street_df = apply_uploaded_street_result_filters(street_df)
    if street_df.empty:
        return b""

    street_df = street_df.fillna("")
    summary_df = build_precinct_summary(street_df)
    county_desc = selected_area_desc(active_filters)
    parties = active_filters.get("party_pick", []) or []
    party_desc = ", ".join(expand_party_label(p) for p in parties) if parties else "Filtered Voters"
    printed_date = datetime.now().strftime("%m/%d/%Y")
    filter_lines = build_filter_summary_lines(active_filters)

    summary_total = pd.DataFrame([{"Precinct":"TOTAL","Individuals":int(summary_df["Individuals"].sum()) if len(summary_df) else 0,"Households":int(summary_df["Households"].sum()) if len(summary_df) else 0}])
    summary_df_with_total = pd.concat([summary_df, summary_total], ignore_index=True)
    total_pages = estimate_street_pdf_pages(summary_df_with_total, street_df)

    buffer = BytesIO()
    page_size = landscape(letter)
    c = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size
    page_num = 1

    totals_hh = int(summary_df["Households"].sum()) if len(summary_df) else 0
    totals_ind = int(summary_df["Individuals"].sum()) if len(summary_df) else 0
    _draw_cover_page(c, width, height, county_desc, party_desc, printed_date, totals_ind, totals_hh, filter_lines, page_num, total_pages)
    c.showPage()
    page_num += 1

    rows_per_summary_page = 26
    if len(summary_df_with_total) == 0:
        _draw_summary_page(c, width, height, summary_df_with_total, printed_date, page_num, total_pages)
        c.showPage()
        page_num += 1
    else:
        for start in range(0, len(summary_df_with_total), rows_per_summary_page):
            chunk = summary_df_with_total.iloc[start:start + rows_per_summary_page]
            _draw_summary_page(c, width, height, chunk, printed_date, page_num, total_pages)
            c.showPage()
            page_num += 1

    body_top = height - 104
    body_bottom = 40
    row_h = 14

    for precinct, grp in street_df.groupby("Precinct", sort=False):
        grp = grp.sort_values(["StreetGroup", "HouseNumSort", "AptSort", "FullName"], kind="stable")
        page_in_precinct = 1
        current_street = None
        cols = _draw_precinct_page_header(c, width, height, precinct, page_in_precinct)
        bookmark_key = make_precinct_bookmark_key(precinct)
        c.bookmarkPage(bookmark_key)
        c.addOutlineEntry(str(precinct), bookmark_key, level=0, closed=False)
        y = body_top - 10

        for (street, address), addr_grp in grp.groupby(["StreetGroup", "AddressLine"], sort=False, dropna=False):
            addr_grp = addr_grp.reset_index(drop=True)
            need = len(addr_grp) + 1
            if current_street != street:
                need += 1

            if y - (need * row_h) < body_bottom:
                draw_footer(c, page_num, total_pages, printed_date)
                c.showPage()
                page_num += 1
                page_in_precinct += 1
                cols = _draw_precinct_page_header(c, width, height, precinct, page_in_precinct)
                y = body_top - 10
                current_street = None

            if current_street != street:
                c.setFillColor(REPORT_STREET)
                c.rect(40, y - 9, width - 80, 14, fill=1, stroke=0)
                c.setFillColor(REPORT_NAVY)
                c.setFont("Helvetica-Bold", 10)
                c.drawString(48, y - 5, truncate_text(street, 80))
                y -= row_h
                current_street = street

            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(58, y - 5, truncate_text(address, 18))
            y -= row_h

            c.setFont("Helvetica", 8.5)
            for row_idx, (_, row) in enumerate(addr_grp.iterrows()):
                fill = REPORT_LIGHT if row_idx % 2 == 0 else colors.white
                c.setFillColor(fill)
                c.rect(52, y - 8, width - 104, 12, fill=1, stroke=0)

                c.setFillColor(colors.black)
                c.drawString(cols["Full Name"], y - 5, truncate_text(normalize_export_text(row.get("FullName", "")) or full_name_from_row(row), 34))
                c.drawString(cols["Phone"], y - 5, truncate_text(row["Phone"], 22))
                c.drawString(cols["Party"], y - 5, truncate_text(row["Party"], 2))
                c.drawString(cols["Sex"], y - 5, truncate_text(row["Sex"], 1))
                c.drawString(cols["Age"], y - 5, truncate_text(row["Age"], 3))

                for label in ["F", "A", "U", "NH", "Yard Sign"]:
                    c.rect(cols[label], y - 7, 8, 8, fill=0, stroke=1)
                    mark_val = normalize_export_text(row.get(label, ""))
                    if mark_val:
                        c.setFont("Helvetica-Bold", 7.5)
                        c.drawCentredString(cols[label] + 4, y - 5, "X")
                        c.setFont("Helvetica", 8.5)

                mb_val = truncate_text(get_mb_perm_display(row), 1)
                if mb_val:
                    c.drawCentredString(cols["MB Perm"] + 4, y - 5, mb_val)
                y -= row_h

        draw_footer(c, page_num, total_pages, printed_date)
        if page_num < total_pages:
            c.showPage()
            page_num += 1

    c.save()
    return buffer.getvalue()



def _make_walk_sheet_groups(active_filters):
    street_df = build_street_list_dataframe(active_filters).copy()
    street_df = apply_uploaded_walk_result_filters(street_df)
    if street_df.empty:
        return street_df, []

    groups = []
    for precinct, precinct_df in street_df.groupby("Precinct", sort=False):
        precinct_df = precinct_df.sort_values(["StreetGroup", "HouseNumSort", "AptSort", "FullName"], kind="stable")
        for (street, address), addr_grp in precinct_df.groupby(["StreetGroup", "AddressLine"], sort=False, dropna=False):
            addr_grp = addr_grp.reset_index(drop=True)
            display_address = normalize_export_text(addr_grp["WalkAddressDisplay"].iloc[0]) if "WalkAddressDisplay" in addr_grp.columns and len(addr_grp) else ""
            groups.append({
                "precinct": normalize_export_text(precinct),
                "street": normalize_export_text(street),
                "address": normalize_export_text(address),
                "display_address": normalize_walk_address_label(display_address, normalize_export_text(address), normalize_export_text(street)),
                "rows": addr_grp.to_dict("records"),
            })
    return street_df, groups


def _estimate_walk_sheet_pages(groups, page_size):
    _, height = page_size
    body_top = height - 132
    body_bottom = 44
    address_h = 20
    voter_h = 20

    pages = 1 if groups else 0
    y = body_top
    last_precinct = None

    for group in groups:
        need = address_h + (len(group["rows"]) * voter_h) + 8
        if last_precinct is not None and group["precinct"] != last_precinct:
            need += 12
        if y - need < body_bottom:
            pages += 1
            y = body_top
        y -= need
        last_precinct = group["precinct"]

    return max(pages, 1)


def _draw_walk_sheet_header(c, width, height, precinct, page_in_precinct, printed_date, filter_desc):
    draw_brand(c, height - 16)
    title = precinct if precinct else "Selected Precinct"
    if page_in_precinct > 1:
        title = f"{title} (cont. {page_in_precinct})"

    c.setFillColor(REPORT_NAVY)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(24, height - 58, f"Walk Sheet – {title}")

    c.setFillColor(colors.black)
    c.setFont("Helvetica", 9)
    subtitle = truncate_text(filter_desc, 145)
    c.drawString(24, height - 74, subtitle)

    c.setFillColor(REPORT_NAVY)
    c.setFont("Helvetica", 8)
    c.drawString(24, height - 88, "C = Contact   N = Not Home   F = Follow-up")

    c.setFont("Helvetica-Bold", 8)
    c.drawString(32, height - 102, "C")
    c.drawString(51, height - 102, "N")
    c.drawString(70, height - 102, "F")
    c.drawString(94, height - 102, "Voter")
    c.drawString(300, height - 102, "Details")
    c.drawString(510, height - 102, "Notes")
    c.setStrokeColor(REPORT_GRID)
    c.line(24, height - 108, width - 24, height - 108)


def generate_walk_sheet_pdf_bytes(active_filters):
    street_df, groups = _make_walk_sheet_groups(active_filters)
    if street_df.empty or not groups:
        return b""

    page_size = landscape(letter)
    width, height = page_size
    printed_date = datetime.now().strftime("%m/%d/%Y")
    county_desc = selected_area_desc(active_filters)
    filter_lines = build_filter_summary_lines(active_filters)
    filter_desc = county_desc
    if filter_lines:
        filter_desc += " | " + " | ".join(filter_lines[:3])

    total_pages = _estimate_walk_sheet_pages(groups, page_size)
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=page_size)

    page_num = 1
    page_in_precinct = 1
    current_precinct = groups[0]["precinct"]

    _draw_walk_sheet_header(c, width, height, current_precinct, page_in_precinct, printed_date, filter_desc)

    body_top = height - 132
    body_bottom = 44
    address_h = 20
    voter_h = 20
    y = body_top

    for idx, group in enumerate(groups):
        if group["precinct"] != current_precinct:
            current_precinct = group["precinct"]
            page_in_precinct = 1

        needed = address_h + (len(group["rows"]) * voter_h) + 8
        if y - needed < body_bottom:
            draw_footer(c, page_num, total_pages, printed_date)
            c.showPage()
            page_num += 1
            if idx > 0 and groups[idx - 1]["precinct"] == group["precinct"]:
                page_in_precinct += 1
            else:
                page_in_precinct = 1
            _draw_walk_sheet_header(c, width, height, current_precinct, page_in_precinct, printed_date, filter_desc)
            y = body_top

        c.setFillColor(REPORT_LIGHT)
        c.roundRect(24, y - 15, width - 48, 17, 6, fill=1, stroke=0)
        c.setFillColor(REPORT_NAVY)
        c.setFont("Helvetica-Bold", 10)
        address_label = normalize_walk_address_label(
            group.get("display_address", ""),
            group.get("address", ""),
            group.get("street", ""),
        )
        c.drawString(32, y - 10, truncate_text(address_label, 110))
        y -= address_h

        for row in group["rows"]:
            row_y = y
            checkbox_y = row_y - 11
            c.setStrokeColor(REPORT_GRID)
            checkbox_positions = {"C": 28, "N": 47, "F": 66}
            for x in checkbox_positions.values():
                c.rect(x, checkbox_y, 10, 10, fill=0, stroke=1)

            if normalize_export_text(row.get("Contacted", "")):
                c.setFont("Helvetica-Bold", 8)
                c.drawCentredString(checkbox_positions["C"] + 5, row_y - 3, "X")
            result_key = normalize_export_text(row.get("Result", "")).upper().replace(" ", "")
            if result_key in {"NOTHOME", "NH"}:
                c.setFont("Helvetica-Bold", 8)
                c.drawCentredString(checkbox_positions["N"] + 5, row_y - 3, "X")
            if normalize_export_text(row.get("Follow-Up", "")):
                c.setFont("Helvetica-Bold", 8)
                c.drawCentredString(checkbox_positions["F"] + 5, row_y - 3, "X")

            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 10)
            c.drawString(92, row_y - 6, truncate_text(normalize_export_text(row.get("FullName", "")) or full_name_from_row(row), 32))

            detail = " / ".join(
                part for part in [
                    truncate_text(row.get("Phone", ""), 18),
                    truncate_text(row.get("Party", ""), 2),
                    truncate_text(row.get("Sex", ""), 1),
                    truncate_text(row.get("Age", ""), 3),
                    "MB " + truncate_text(get_mb_perm_display(row), 1) if truncate_text(get_mb_perm_display(row), 1) else "",
                ]
                if part
            )
            c.setFont("Helvetica", 9)
            c.drawString(300, row_y - 6, truncate_text(detail, 40))

            notes_y = row_y - 8
            c.setStrokeColor(REPORT_GRID)
            c.line(500, notes_y, width - 28, notes_y)
            y -= voter_h

        y -= 8

    draw_footer(c, page_num, total_pages, printed_date)
    c.save()
    return buffer.getvalue()



def build_street_list_dataframe_from_detail_df(df: pd.DataFrame):
    df = merge_uploaded_street_results_into_detail_df(df)
    df = merge_uploaded_walk_results_into_detail_df(df)
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "Municipality","Precinct","StreetGroup","AddressLine","UnitDisplay","WalkAddressDisplay","FullName","Phone","Party","Sex","Age","PA ID Number",
            "F","A","U","NH","Yard Sign","Notes","Contacted","Result","Support Level","Follow-Up","Walk Notes","MB_Perm","HouseNumSort","AptSort"
        ])

    municipality_col = first_existing_detail(df.columns.tolist(), ["Municipality"])
    precinct_col = first_existing_detail(df.columns.tolist(), ["Precinct"])
    street_col = first_existing_detail(df.columns.tolist(), ["Street Name"])
    house_col = first_existing_detail(df.columns.tolist(), ["House Number"])
    apt_col = first_existing_detail(df.columns.tolist(), ["Apartment Number"])
    address2_col = first_existing_detail(df.columns.tolist(), ["Address Line 2", "AddressLine2", "Address 2"])
    sex_col = first_existing_detail(df.columns.tolist(), ["Gender", "Sex"])
    age_col = first_existing_detail(df.columns.tolist(), ["Age"])
    party_col = first_existing_detail(df.columns.tolist(), ["Party"])
    pa_id_col = first_existing_detail(df.columns.tolist(), ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID", "Voter ID", "VoterID"])
    mb_perm_col = first_existing_detail(df.columns.tolist(), ["MB_PERM", "MB_Perm", "MB_Pern"])

    out = pd.DataFrame()
    out["Municipality"] = df[municipality_col].apply(normalize_export_text) if municipality_col else ""
    out["Precinct"] = df[precinct_col].apply(normalize_export_text) if precinct_col else ""
    street_vals = df[street_col].apply(normalize_address_value) if street_col else pd.Series([""] * len(df), index=df.index)
    house_vals = df[house_col].apply(normalize_export_text) if house_col else pd.Series([""] * len(df), index=df.index)
    apt_raw_vals = df[apt_col].apply(normalize_export_text) if apt_col else pd.Series([""] * len(df), index=df.index)
    addr2_raw_vals = df[address2_col].apply(normalize_export_text) if address2_col else pd.Series([""] * len(df), index=df.index)

    apt_vals = apt_raw_vals.apply(normalize_unit_or_address2_value)
    addr2_vals = addr2_raw_vals.apply(normalize_unit_or_address2_value)
    unit_vals = apt_vals.copy()
    unit_vals.loc[unit_vals.eq("") & addr2_vals.ne("")] = addr2_vals[unit_vals.eq("") & addr2_vals.ne("")]

    out["StreetGroup"] = street_vals
    out["UnitDisplay"] = unit_vals
    out["AddressLine"] = house_vals
    out.loc[unit_vals.ne(""), "AddressLine"] = out.loc[unit_vals.ne(""), "AddressLine"] + " " + unit_vals[unit_vals.ne("")]
    out["AddressLine"] = out["AddressLine"].apply(collapse_spaces).apply(normalize_address_value)
    out["WalkAddressDisplay"] = [
        build_walk_address_display(h, s, u)
        for h, s, u in zip(house_vals.tolist(), street_vals.tolist(), unit_vals.tolist())
    ]

    out["FullName"] = df.apply(full_name_from_row, axis=1).apply(normalize_name_value)
    out["Phone"] = df.apply(choose_best_phone, axis=1)
    out["Party"] = df[party_col].apply(normalize_export_text) if party_col else ""
    out["Sex"] = df[sex_col].apply(normalize_export_text) if sex_col else ""
    out["Age"] = df[age_col].apply(lambda v: normalize_numeric_string(v)) if age_col else ""
    out["PA ID Number"] = df[pa_id_col].apply(normalize_numeric_string) if pa_id_col else ""
    out["F"] = df["F"].apply(normalize_tracking_mark) if "F" in df.columns else ""
    out["A"] = df["A"].apply(normalize_tracking_mark) if "A" in df.columns else ""
    out["U"] = df["U"].apply(normalize_tracking_mark) if "U" in df.columns else ""
    out["NH"] = df["NH"].apply(normalize_tracking_mark) if "NH" in df.columns else ""
    out["Yard Sign"] = df["Yard Sign"].apply(normalize_tracking_mark) if "Yard Sign" in df.columns else ""
    out["Notes"] = df["Notes"].apply(normalize_export_text) if "Notes" in df.columns else ""
    out["Contacted"] = df["Contacted"].apply(normalize_tracking_mark) if "Contacted" in df.columns else ""
    out["Result"] = df["Result"].apply(normalize_walk_result_value) if "Result" in df.columns else ""
    out["Support Level"] = df["Support Level"].apply(normalize_export_text) if "Support Level" in df.columns else ""
    out["Follow-Up"] = df["Follow-Up"].apply(normalize_tracking_mark) if "Follow-Up" in df.columns else ""
    out["Walk Notes"] = df["Walk Notes"].apply(normalize_export_text) if "Walk Notes" in df.columns else ""
    out["MB_Perm"] = df[mb_perm_col].apply(normalize_mb_perm_value) if mb_perm_col else ""
    out["HouseNumSort"] = house_vals.apply(parse_house_number)
    out["AptSort"] = unit_vals.apply(parse_apartment_sort)
    out = out.sort_values(by=["Precinct", "StreetGroup", "HouseNumSort", "AptSort", "FullName"], kind="stable").reset_index(drop=True)
    return out


def make_walk_sheet_groups_from_street_df(street_df: pd.DataFrame):
    if street_df is None or street_df.empty:
        return []

    groups = []
    for precinct, precinct_df in street_df.groupby("Precinct", sort=False):
        precinct_df = precinct_df.sort_values(["StreetGroup", "HouseNumSort", "AptSort", "FullName"], kind="stable")
        for (street, address), addr_grp in precinct_df.groupby(["StreetGroup", "AddressLine"], sort=False, dropna=False):
            addr_grp = addr_grp.reset_index(drop=True)
            display_address = normalize_export_text(addr_grp["WalkAddressDisplay"].iloc[0]) if "WalkAddressDisplay" in addr_grp.columns and len(addr_grp) else ""
            groups.append({
                "precinct": normalize_export_text(precinct),
                "street": normalize_export_text(street),
                "address": normalize_export_text(address),
                "display_address": normalize_walk_address_label(display_address, normalize_export_text(address), normalize_export_text(street)),
                "rows": addr_grp.to_dict("records"),
            })
    return groups

def generate_walk_sheet_pdf_from_street_df(street_df: pd.DataFrame, title: str, filter_desc: str = ""):
    if street_df is None or street_df.empty:
        return b""

    groups = make_walk_sheet_groups_from_street_df(street_df)
    if not groups:
        return b""

    page_size = landscape(letter)
    width, height = page_size
    printed_date = datetime.now().strftime("%m/%d/%Y")
    total_pages = _estimate_walk_sheet_pages(groups, page_size)
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=page_size)

    page_num = 1
    page_in_precinct = 1
    current_precinct = groups[0]["precinct"]

    header_title = title or current_precinct or "Selected Turf"
    header_desc = filter_desc or "Turf packet walk sheet"
    _draw_walk_sheet_header(c, width, height, header_title, page_in_precinct, printed_date, header_desc)

    body_top = height - 132
    body_bottom = 44
    address_h = 20
    voter_h = 20
    y = body_top

    for idx, group in enumerate(groups):
        if group["precinct"] != current_precinct:
            current_precinct = group["precinct"]
            page_in_precinct = 1

        needed = address_h + (len(group["rows"]) * voter_h) + 8
        if y - needed < body_bottom:
            draw_footer(c, page_num, total_pages, printed_date)
            c.showPage()
            page_num += 1
            if idx > 0 and groups[idx - 1]["precinct"] == group["precinct"]:
                page_in_precinct += 1
            else:
                page_in_precinct = 1
            _draw_walk_sheet_header(c, width, height, header_title, page_in_precinct, printed_date, header_desc)
            y = body_top

        c.setFillColor(REPORT_LIGHT)
        c.roundRect(24, y - 15, width - 48, 17, 6, fill=1, stroke=0)
        c.setFillColor(REPORT_NAVY)
        c.setFont("Helvetica-Bold", 10)
        address_label = normalize_walk_address_label(
            group.get("display_address", ""),
            group.get("address", ""),
            group.get("street", ""),
        )
        c.drawString(32, y - 10, truncate_text(address_label, 110))
        y -= address_h

        for row in group["rows"]:
            row_y = y
            checkbox_y = row_y - 11
            c.setStrokeColor(REPORT_GRID)
            for x in (28, 47, 66):
                c.rect(x, checkbox_y, 10, 10, fill=0, stroke=1)

            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 10)
            c.drawString(92, row_y - 6, truncate_text(normalize_export_text(row.get("FullName", "")) or full_name_from_row(row), 32))

            detail = " / ".join(
                part for part in [
                    truncate_text(row.get("Phone", ""), 18),
                    truncate_text(row.get("Party", ""), 2),
                    truncate_text(row.get("Sex", ""), 1),
                    truncate_text(row.get("Age", ""), 3),
                    "MB " + truncate_text(get_mb_perm_display(row), 1) if truncate_text(get_mb_perm_display(row), 1) else "",
                ]
                if part
            )
            c.setFont("Helvetica", 9)
            c.drawString(300, row_y - 6, truncate_text(detail, 40))

            notes_y = row_y - 8
            c.setStrokeColor(REPORT_GRID)
            c.line(500, notes_y, width - 28, notes_y)
            y -= voter_h

        y -= 8

    draw_footer(c, page_num, total_pages, printed_date)
    c.save()
    return buffer.getvalue()

def _summary_count_df(active_filters, columns, group_expr, label_alias="Label", include_blank=True):
    con = get_conn()
    where_sql, params = current_filter_clause(active_filters, columns)
    blank_filter = "" if include_blank else f" AND {group_expr} IS NOT NULL AND trim(cast({group_expr} as varchar)) <> ''"
    return con.execute(
        f"""
        SELECT
            coalesce(nullif(trim(cast({group_expr} as varchar)), ''), 'Blank/Unknown') AS {quote_ident(label_alias)},
            count(*) AS Count
        FROM voters
        {where_sql}
        {blank_filter}
        GROUP BY 1
        ORDER BY Count DESC, 1
        """,
        params,
    ).df()


def _summary_age_df(active_filters, columns):
    con = get_conn()
    where_sql, params = current_filter_clause(active_filters, columns)
    return con.execute(
        f"""
        SELECT
            case
                when _AgeNum IS NULL then 'Blank/Unknown'
                when _AgeNum < 18 then 'Under 18'
                when _AgeNum <= 24 then '18-24'
                when _AgeNum <= 34 then '25-34'
                when _AgeNum <= 44 then '35-44'
                when _AgeNum <= 54 then '45-54'
                when _AgeNum <= 64 then '55-64'
                when _AgeNum <= 74 then '65-74'
                else '75+'
            end AS AgeBucket,
            count(*) AS Count,
            case
                when _AgeNum IS NULL then 99
                when _AgeNum < 18 then 1
                when _AgeNum <= 24 then 2
                when _AgeNum <= 34 then 3
                when _AgeNum <= 44 then 4
                when _AgeNum <= 54 then 5
                when _AgeNum <= 64 then 6
                when _AgeNum <= 74 then 7
                else 8
            end AS SortKey
        FROM voters
        {where_sql}
        GROUP BY 1, 3
        ORDER BY SortKey
        """,
        params,
    ).df()[["AgeBucket", "Count"]]


def generate_summary_report_pdf_bytes(active_filters, columns):
    metrics = query_metrics(active_filters, columns)
    party_df = _summary_count_df(active_filters, columns, "_PartyNorm", "Value")
    gender_df = _summary_count_df(active_filters, columns, "_Gender", "Value")
    age_df = _summary_age_df(active_filters, columns)
    filter_lines = build_filter_summary_lines(active_filters)
    printed_dt = datetime.now().strftime("%m/%d/%Y %I:%M %p")

    buffer = BytesIO()
    page_size = landscape(letter)
    width, height = page_size
    c = canvas.Canvas(buffer, pagesize=page_size)

    def section_bar(y, title):
        c.setFillColor(REPORT_NAVY)
        c.roundRect(26, y - 14, width - 52, 18, 6, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(34, y - 9, title)

    def draw_simple_table(x, y_top, headers, rows, col_widths, row_h=16, font_size=8):
        table_w = sum(col_widths)
        c.setFillColor(REPORT_NAVY)
        c.rect(x, y_top - row_h, table_w, row_h, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", font_size)
        cursor = x
        for idx, head in enumerate(headers):
            if idx == len(headers) - 1:
                c.drawRightString(cursor + col_widths[idx] - 6, y_top - 11, str(head))
            else:
                c.drawString(cursor + 6, y_top - 11, str(head))
            cursor += col_widths[idx]

        y = y_top - row_h
        for i, row in enumerate(rows):
            y -= row_h
            fill = REPORT_LIGHT if i % 2 == 0 else colors.white
            c.setFillColor(fill)
            c.rect(x, y, table_w, row_h, fill=1, stroke=0)
            c.setStrokeColor(REPORT_GRID)
            c.rect(x, y, table_w, row_h, fill=0, stroke=1)
            c.setFillColor(colors.black)
            c.setFont("Helvetica", font_size)
            cursor = x
            for idx, cell in enumerate(row):
                cell_text = truncate_text(cell, 48)
                if idx == len(row) - 1:
                    c.drawRightString(cursor + col_widths[idx] - 6, y + 4, cell_text)
                else:
                    c.drawString(cursor + 6, y + 4, cell_text)
                cursor += col_widths[idx]
        return y

    draw_brand(c, height - 18)
    c.setFillColor(REPORT_NAVY)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(28, height - 58, "Candidate Connect Summary Report")
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)
    c.drawString(28, height - 74, f"Generated: {printed_dt}")

    section_bar(height - 96, "Overview")
    overview_rows = [
        ["Total Voters", f"{int(metrics.get('voters', 0)):,}"],
        ["Total Households", f"{int(metrics.get('households', 0)):,}"],
        ["With Email", f"{int(metrics.get('emails', 0)):,}"],
        ["With Landline", f"{int(metrics.get('landlines', 0)):,}"],
        ["With Mobile", f"{int(metrics.get('mobiles', 0)):,}"],
    ]
    draw_simple_table(28, height - 104, ["Metric", "Value"], overview_rows, [180, 90])

    section_bar(height - 228, "Selected Filters")
    if not filter_lines:
        filter_lines = ["No additional filters selected"]
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 9)
    fy = height - 250
    for line in filter_lines[:10]:
        c.drawString(34, fy, u"• " + truncate_text(line, 135))
        fy -= 14

    left_x = 28
    right_x = 405
    top_y = height - 410

    section_bar(top_y, "Party Breakdown")
    party_rows = [[str(r["Value"]), f"{int(r['Count']):,}"] for _, r in party_df.iterrows()] or [["No data", "0"]]
    y_end_left = draw_simple_table(left_x, top_y - 8, ["Value", "Count"], party_rows[:10], [180, 90])

    section_bar(top_y, "Gender Breakdown")
    gender_rows = [[str(r["Value"]), f"{int(r['Count']):,}"] for _, r in gender_df.iterrows()] or [["No data", "0"]]
    y_end_right = draw_simple_table(right_x, top_y - 8, ["Value", "Count"], gender_rows[:10], [180, 90])

    lower_top = min(y_end_left, y_end_right) - 26
    section_bar(lower_top, "Age Breakdown")
    age_rows = [[str(r["AgeBucket"]), f"{int(r['Count']):,}"] for _, r in age_df.iterrows()] or [["No data", "0"]]
    draw_simple_table(28, lower_top - 8, ["Age Range", "Count"], age_rows[:10], [180, 90])

    draw_footer(c, 1, 1, datetime.now().strftime("%m/%d/%Y"))
    c.save()
    return buffer.getvalue()


cc_logo_uri = img_to_data_uri(CC_LOGO)
# Use the real uploaded The Political Technology Company logo.
tss_logo_uri = img_to_data_uri(TSS_LOGO)

header_html = f"""
<div class="top-shell">
  <div class="brand-grid">
    <div class="brand-left">{f'<img class="logo-cc" src="{cc_logo_uri}"/>' if cc_logo_uri else ''}</div>
    <div class="brand-center">
      <div class="brand-title">Candidate Connect</div>
      <div class="brand-sub">VOTER DATA & ENGAGEMENT PLATFORM</div>
      <div class="brand-status">Campaign Intelligence Workspace</div>
    </div>
    <div class="brand-right"><img class="logo-tss" src="{tss_logo_uri}" alt="The Political Technology Company"/></div>
  </div>
</div>
"""
st.markdown(header_html, unsafe_allow_html=True)




def _cc_html_escape(value) -> str:
    return str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def _cc_open_pct(part, total) -> float:
    try:
        return 0.0 if float(total) <= 0 else (float(part) / float(total)) * 100.0
    except Exception:
        return 0.0


def _cc_count_from_chart(df: pd.DataFrame, label_col: str, label_value: str) -> int:
    try:
        if df is None or df.empty or label_col not in df.columns or 'Count' not in df.columns:
            return 0
        hit = df[df[label_col].astype(str).str.upper().eq(str(label_value).upper())]
        if hit.empty:
            return 0
        return safe_int(pd.to_numeric(hit['Count'], errors='coerce').fillna(0).sum())
    except Exception:
        return 0


def _cc_open_bar_html(age_df: pd.DataFrame) -> str:
    preferred = ['18-24', '25-34', '35-44', '45-54', '55-64', '65-74', '75-84', '85+']
    rows = []
    if age_df is not None and not age_df.empty:
        temp = age_df.copy()
        temp['Count'] = pd.to_numeric(temp['Count'], errors='coerce').fillna(0)
        for bucket in preferred:
            match = temp[temp['Age Range'].astype(str).eq(bucket)] if 'Age Range' in temp.columns else pd.DataFrame()
            count = float(match['Count'].sum()) if not match.empty else 0.0
            rows.append((bucket, count))
    if not rows:
        rows = [(b, 0.0) for b in preferred]
    max_count = max([v for _, v in rows] + [1])
    bars = []
    labels = []
    for label, count in rows:
        h = max(20, int((count / max_count) * 205)) if max_count else 20
        bars.append(f'<div class="cc-open-bar-wrap"><div class="cc-open-bar-value">{fmt_pct(_cc_open_pct(count, sum(v for _, v in rows)))}</div><div class="cc-bar" style="height:{h}px"></div></div>')
        labels.append(f'<div>{_cc_html_escape(label)}</div>')
    return '<div class="cc-bars">' + ''.join(bars) + '</div><div class="cc-bar-labels">' + ''.join(labels) + '</div>'


def _cc_open_donut_html(values, center_label='Total', center_sub='Universe') -> str:
    total = sum(max(0, float(v[1])) for v in values) or 1.0
    start = 0.0
    stops = []
    for _label, count, color in values:
        pct = max(0.0, float(count)) / total * 100.0
        end = start + pct
        stops.append(f'{color} {start:.3f}% {end:.3f}%')
        start = end
    gradient = ', '.join(stops) if stops else '#334155 0 100%'
    return f'<div class="cc-donut" style="background:conic-gradient({gradient});"><div class="cc-donut-center"><b>{_cc_html_escape(center_label)}</b><span>{_cc_html_escape(center_sub)}</span></div></div>'



def _cc_open_icon_html(kind: str) -> str:
    """KPI icons loaded from PNG files in the same folder as app.py."""
    icon_map = {
        "people": Path("icon_total_voters.png"),
        "elephant": Path("icon_republican.png"),
        "donkey": Path("icon_democrat.png"),
        "person": Path("icon_other.png"),
        "unknown": Path("icon_unknown.png"),
    }
    path = icon_map.get(kind)
    if path and path.exists():
        return f'<img src="{img_to_data_uri(path)}" alt="{kind} icon"/>'
    # Fallbacks if the PNG files are accidentally missing.
    if kind == "people":
        return '<svg viewBox="0 0 64 64"><circle cx="24" cy="23" r="9"/><circle cx="43" cy="25" r="8"/><path d="M8 54c1-13 9-21 20-21s19 8 20 21z"/><path d="M36 54c1-10 7-16 16-16 6 0 11 4 13 11v5z" opacity=".75"/></svg>'
    if kind == "elephant":
        return '<svg viewBox="0 0 64 64"><path d="M10 34c0-11 9-20 23-20h12c7 0 12 5 12 12v11c0 8-7 15-15 15H23c-7 0-13-5-13-12z"/><path d="M45 25h13v14c0 5-4 9-9 9h-4z"/><path d="M13 36H5c0-7 3-13 8-17z"/><circle cx="29" cy="28" r="3" fill="#07101A"/><rect x="18" y="49" width="7" height="10" rx="2"/><rect x="39" y="49" width="7" height="10" rx="2"/></svg>'
    if kind == "donkey":
        return '<span class="dletter">D</span>'
    if kind == "person":
        return '<svg viewBox="0 0 64 64"><circle cx="32" cy="21" r="11"/><path d="M13 56c2-14 10-22 19-22s17 8 19 22z"/></svg>'
    return '<span class="dletter">?</span>'

def render_opening_dashboard_preview():
    """Fast right-pane statewide opening dashboard shown before Apply Filters is clicked."""
    active = st.session_state.get('active_filters', {}) or {}
    columns = st.session_state.get('columns', []) or []
    preview_active = dict(active)

    try:
        metrics = query_metrics(preview_active, columns)
    except Exception:
        metrics = {'voters': 0, 'households': None, 'emails': 0, 'mobiles': 0, 'unique_precincts': 0}

    try:
        party_df = query_chart(preview_active, columns, '_PartyNorm', 'Party')
    except Exception:
        party_df = pd.DataFrame(columns=['Party', 'Count'])
    try:
        gender_df = query_chart(preview_active, columns, '_Gender', 'Gender')
    except Exception:
        gender_df = pd.DataFrame(columns=['Gender', 'Count'])
    try:
        age_df = query_chart(preview_active, columns, '_AgeRange', 'Age Range')
    except Exception:
        age_df = pd.DataFrame(columns=['Age Range', 'Count'])

    voters = safe_int(metrics.get('voters'))
    r_count = _cc_count_from_chart(party_df, 'Party', 'R')
    d_count = _cc_count_from_chart(party_df, 'Party', 'D')
    o_count = _cc_count_from_chart(party_df, 'Party', 'O')
    unknown_count = max(0, voters - r_count - d_count - o_count)

    f_count = _cc_count_from_chart(gender_df, 'Gender', 'F')
    m_count = _cc_count_from_chart(gender_df, 'Gender', 'M')
    u_count = max(0, safe_int(pd.to_numeric(gender_df.get('Count', pd.Series(dtype=float)), errors='coerce').fillna(0).sum()) - f_count - m_count)

    party_donut = _cc_open_donut_html([
        ('Republican', r_count, CC_THEME['rep_red']),
        ('Democrat', d_count, CC_THEME['dem_blue']),
        ('Other', o_count, CC_THEME['other_green']),
        ('Unknown', unknown_count, CC_THEME['brand_gold']),
    ], f'{voters:,}', 'Total')
    gender_donut = _cc_open_donut_html([
        ('Female', f_count, CC_THEME['rep_red']),
        ('Male', m_count, CC_THEME['dem_blue']),
        ('Other / Unknown', u_count, CC_THEME['other_green']),
    ], f'{(f_count + m_count + u_count):,}', 'Total')

    metric_cards = [
        ('TOTAL VOTERS', f'{voters:,}', 'red', _cc_open_icon_html('people')),
        ('REPUBLICAN', f'{r_count:,}', 'red', _cc_open_icon_html('elephant')),
        ('DEMOCRAT', f'{d_count:,}', 'blue', _cc_open_icon_html('donkey')),
        ('OTHER / UNAFFILIATED', f'{o_count:,}', 'green', _cc_open_icon_html('person')),
    ]
    metric_html = ''.join([
        f'<div class="cc-open-metric {klass}"><div class="icon">{icon}</div><div><div class="label">{label}</div><div class="value">{value}</div><div class="sub">{fmt_pct(_cc_open_pct(int(value.replace(",", "")) if value.replace(",", "").isdigit() else 0, voters))} of universe</div></div></div>'
        for label, value, klass, icon in metric_cards
    ])

    party_legend = ''.join([
        f'<div class="cc-legend-row"><span><i class="cc-dot" style="background:{color}"></i>{label}</span><b>{count:,} ({fmt_pct(_cc_open_pct(count, voters))})</b></div>'
        for label, count, color in [
            ('Republican', r_count, CC_THEME['rep_red']),
            ('Democrat', d_count, CC_THEME['dem_blue']),
            ('Other / Unaffiliated', o_count, CC_THEME['other_green']),
            ('Unknown', unknown_count, CC_THEME['brand_gold']),
        ]
    ])
    gender_legend = ''.join([
        f'<div class="cc-legend-row"><span><i class="cc-dot" style="background:{color}"></i>{label}</span><b>{count:,} ({fmt_pct(_cc_open_pct(count, max(1, f_count + m_count + u_count)))})</b></div>'
        for label, count, color in [
            ('Female', f_count, CC_THEME['rep_red']),
            ('Male', m_count, CC_THEME['dem_blue']),
            ('Other / Unknown', u_count, CC_THEME['other_green']),
        ]
    ])

    geo_rows = ''.join([
        f'<tr><td>{geo}</td><td>{voters:,}</td><td class="red">{r_count:,} ({fmt_pct(_cc_open_pct(r_count, voters))})</td><td class="blue">{d_count:,} ({fmt_pct(_cc_open_pct(d_count, voters))})</td><td class="green">{o_count:,} ({fmt_pct(_cc_open_pct(o_count, voters))})</td><td class="gold">{unknown_count:,} ({fmt_pct(_cc_open_pct(unknown_count, voters))})</td></tr>'
        for geo in ['US Congress', 'State Senate', 'State House']
    ])

    html = f'''
    <div class="cc-opening-dashboard">
      <div class="cc-open-metrics">{metric_html}</div>
      <div class="cc-open-main-grid">
        <div class="cc-open-card party-card"><h3>VOTERS BY PARTY <span>•••</span></h3><div class="cc-open-split"><div>{party_donut}</div><div>{party_legend}</div></div><p>Universe: All Voters</p></div>
        <div class="cc-open-card age-card"><h3>VOTERS BY AGE RANGE <span>•••</span></h3>{_cc_open_bar_html(age_df)}<p>Universe: All Voters</p></div>
        <div class="cc-open-card gender-card"><h3>VOTERS BY GENDER <span>•••</span></h3><div class="cc-open-split"><div>{gender_donut}</div><div>{gender_legend}</div></div><p>Universe: All Voters</p></div>
        <div class="cc-open-card geo-card"><h3>VOTERS BY GEOGRAPHY <span>•••</span></h3><table class="cc-open-table"><thead><tr><th>Geography</th><th>Total Voters</th><th>Republican</th><th>Democrat</th><th>Other / Unaffiliated</th></tr></thead><tbody>{geo_rows}</tbody></table><p>Universe: All Voters</p></div>
      </div>
      <div class="cc-open-note">Use the sidebar to build a campaign universe.</div>
    </div>
    '''
    st.markdown(html, unsafe_allow_html=True)


def format_lookup_date(value) -> str:
    if value is None:
        return ""
    try:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return normalize_export_text(value)
        return ts.strftime("%m/%d/%Y")
    except Exception:
        return normalize_export_text(value)


def format_lookup_phone(value) -> str:
    digits = clean_phone_value(value)
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return normalize_export_text(value)

def format_lookup_zip(value) -> str:
    raw = normalize_export_text(value)
    if not raw:
        return ""
    if re.fullmatch(r"\d+\.0+", raw):
        raw = raw.split(".")[0]
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 9:
        return f"{digits[:5]}-{digits[5:]}"
    if len(digits) >= 5:
        return digits[:5]
    return raw


def sanitize_multiselect_defaults(default_values, option_values):
    if default_values is None:
        return []
    if not isinstance(default_values, (list, tuple, set)):
        default_values = [default_values]
    option_text = {str(v).strip(): v for v in option_values or []}
    cleaned = []
    for value in default_values:
        key = str(value).strip()
        if key in option_text:
            cleaned.append(option_text[key])
    return cleaned

def sanitize_selectbox_value(current_value, option_values, fallback=None):
    options_list = list(option_values or [])
    if not options_list:
        return fallback
    if current_value in options_list:
        return current_value
    current_key = str(current_value).strip()
    for option in options_list:
        if str(option).strip() == current_key:
            return option
    if fallback in options_list:
        return fallback
    return options_list[0]

def sanitize_slider_range(current_value, low, high, numeric_type=int):
    """Keep saved slider defaults inside the current data range.

    Streamlit raises an exception when a saved universe has slider values
    outside the available min/max for the current dataset. This helper lets
    saved filters load safely after data changes.
    """
    try:
        if not isinstance(current_value, (list, tuple)) or len(current_value) != 2:
            return (numeric_type(low), numeric_type(high))
        a = numeric_type(current_value[0])
        b = numeric_type(current_value[1])
        lo = numeric_type(low)
        hi = numeric_type(high)
        a = max(lo, min(a, hi))
        b = max(lo, min(b, hi))
        if a > b:
            a, b = b, a
        return (a, b)
    except Exception:
        return (numeric_type(low), numeric_type(high))



def get_detail_columns(detail_paths):
    con = get_conn()
    df = con.execute(f"DESCRIBE SELECT * FROM {dataset_scan_sql(detail_paths)}").df()
    return df["column_name"].tolist()


def _detail_col_expr(columns, candidates, fallback="''"):
    col = first_existing_detail(columns, candidates)
    if col is None:
        return fallback, None
    return f"coalesce(cast(src.{quote_ident(col)} as varchar), '')", col


@st.cache_data(show_spinner=False)
def get_detail_distinct_values(detail_paths, column_name: str):
    qcol = quote_ident(column_name)
    df = get_conn().execute(
        f'''
        SELECT DISTINCT trim(cast({qcol} as varchar)) AS value
        FROM {dataset_scan_sql(detail_paths)}
        WHERE nullif(trim(cast({qcol} as varchar)), '') IS NOT NULL
        ORDER BY 1
        '''
    ).df()
    return [normalize_export_text(v) for v in df["value"].tolist() if normalize_export_text(v) != ""]


def _normalize_lookup_place(value: str) -> str:
    s = normalize_export_text(value).upper()
    s = re.sub(r"\bCOUNTY\b", "", s)
    s = re.sub(r"\bCO\b", "", s)
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_lookup_search(search_text: str, detail_paths, detail_columns):
    raw = normalize_export_text(search_text)
    parsed = {
        "raw": raw,
        "email": "",
        "phone_digits": "",
        "zip5": "",
        "pa_id_digits": "",
        "county": "",
        "remaining_tokens": [],
    }
    if not raw:
        return parsed

    email_match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", raw)
    if email_match:
        parsed["email"] = email_match.group(0).strip()

    phone_match = re.search(r"(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}", raw)
    if phone_match:
        parsed["phone_digits"] = "".join(re.findall(r"\d", phone_match.group(0)))[-10:]

    zip_match = re.search(r"\b(\d{5})(?:-\d{4})?\b", raw)
    if zip_match:
        parsed["zip5"] = zip_match.group(1)

    pa_id_match = re.search(r"\b\d{6,}(?:-\d+)?\b", raw)
    if pa_id_match and not parsed["phone_digits"]:
        parsed["pa_id_digits"] = "".join(re.findall(r"\d", pa_id_match.group(0)))

    county_map = {}
    if "County" in detail_columns:
        for county in get_detail_distinct_values(detail_paths, "County"):
            norm = _normalize_lookup_place(county)
            if norm:
                county_map[norm] = county

    normalized_query = _normalize_lookup_place(raw)
    for county_norm, county_label in county_map.items():
        if county_norm and re.search(rf"(^| )({re.escape(county_norm)})( |$)", normalized_query):
            parsed["county"] = county_label
            normalized_query = re.sub(rf"(^| )({re.escape(county_norm)})( |$)", " ", normalized_query).strip()
            break

    cleaned = raw
    if parsed["email"]:
        cleaned = cleaned.replace(parsed["email"], " ")
    if parsed["phone_digits"]:
        cleaned = re.sub(r"(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}", " ", cleaned)
    if parsed["zip5"]:
        cleaned = re.sub(rf"\b{re.escape(parsed['zip5'])}(?:-\d{{4}})?\b", " ", cleaned)
    if parsed["pa_id_digits"]:
        cleaned = re.sub(r"[0-9-]{6,}", " ", cleaned)
    if parsed["county"]:
        cleaned = re.sub(re.escape(parsed["county"]), " ", cleaned, flags=re.I)

    cleaned = re.sub(r"[^A-Za-z0-9]+", " ", cleaned)
    parsed["remaining_tokens"] = [tok.upper() for tok in cleaned.split() if len(tok.strip()) >= 2]
    return parsed


def search_voters_for_lookup(active_filters, search_text: str, limit: int = 50, use_current_filters: bool = False) -> pd.DataFrame:
    detail_paths, _ = ensure_detail_shards()
    detail_columns = get_detail_columns(detail_paths)
    lookup_filters = active_filters if use_current_filters else {}
    base_sql, base_params = build_detail_export_sql(detail_paths, lookup_filters)

    first_expr, first_col = _detail_col_expr(detail_columns, ["FirstName", "First Name"])
    middle_expr, middle_col = _detail_col_expr(detail_columns, ["MiddleName", "Middle Name"])
    last_expr, last_col = _detail_col_expr(detail_columns, ["LastName", "Last Name"])
    suffix_expr, suffix_col = _detail_col_expr(detail_columns, ["NameSuffix", "Suffix", "Name Suffix"])
    full_name_expr, full_name_col = _detail_col_expr(detail_columns, ["FullName", "Full Name", "Name"], fallback=None)
    house_expr, house_col = _detail_col_expr(detail_columns, ["House Number", "HouseNumber", "Street Number"])
    street_expr, street_col = _detail_col_expr(detail_columns, ["Street Name", "StreetName", "Street"])
    apt_expr, apt_col = _detail_col_expr(detail_columns, ["Apartment Number", "ApartmentNumber", "Unit", "Apt"])
    city_expr, city_col = _detail_col_expr(detail_columns, ["MailingCity", "Mailing City", "City", "MailCity"])
    state_expr, state_col = _detail_col_expr(detail_columns, ["MailingState", "Mailing State", "State", "MailState"])
    zip_expr, zip_col = _detail_col_expr(detail_columns, ["MailingZip", "Mailing Zip", "ZIP", "Zip", "ZipCode", "ZIPCODE", "MailZip"])
    email_expr, email_col = _detail_col_expr(detail_columns, ["Email", "EmailAddress", "Email Address"])
    mobile_expr, mobile_col = _detail_col_expr(detail_columns, ["Mobile", "Cell", "CellPhone", "Cell Phone"])
    landline_expr, landline_col = _detail_col_expr(detail_columns, ["Landline", "Phone", "HomePhone", "PrimaryPhone", "Primary Phone"])
    pa_id_expr, pa_id_col = _detail_col_expr(detail_columns, ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "State Voter ID", "Voter ID", "VoterID"])
    county_expr = "coalesce(cast(src.\"County\" as varchar), '')" if "County" in detail_columns else "''"
    muni_expr = "coalesce(cast(src.\"Municipality\" as varchar), '')" if "Municipality" in detail_columns else "''"
    precinct_expr = "coalesce(cast(src.\"Precinct\" as varchar), '')" if "Precinct" in detail_columns else "''"

    if full_name_col:
        lookup_name_expr = full_name_expr
    else:
        lookup_name_expr = f"trim(concat_ws(' ', {first_expr}, {middle_expr}, {last_expr}, {suffix_expr}))"

    lookup_address_expr = f"trim(concat_ws(' ', {house_expr}, {street_expr}, case when trim({apt_expr}) <> '' then concat('Apt ', trim({apt_expr})) else '' end))"
    lookup_city_state_zip_expr = f"trim(concat_ws(', ', nullif(trim({city_expr}), ''), trim(concat_ws(' ', nullif(trim({state_expr}), ''), nullif(trim({zip_expr}), '')))))"
    lookup_key_expr = f"trim(concat_ws('|', nullif(trim({pa_id_expr}), ''), {lookup_name_expr}, {lookup_address_expr}))"

    name_haystack_expr = f"upper(concat_ws(' ', {lookup_name_expr}, {first_expr}, {middle_expr}, {last_expr}, {suffix_expr}))"
    address_haystack_expr = f"upper(concat_ws(' ', {lookup_address_expr}, {city_expr}, {state_expr}, {zip_expr}, {county_expr}, {muni_expr}, {precinct_expr}))"
    general_haystack_expr = f"upper(concat_ws(' ', {name_haystack_expr}, {address_haystack_expr}, {pa_id_expr}, {email_expr}, {mobile_expr}, {landline_expr}))"

    parsed = parse_lookup_search(search_text, detail_paths, detail_columns)
    params = list(base_params)
    where_parts = []

    if parsed["county"] and "County" in detail_columns:
        where_parts.append(f"upper(trim({county_expr})) = ?")
        params.append(parsed["county"].upper())

    if parsed["email"]:
        where_parts.append(f"upper(trim({email_expr})) = ?")
        params.append(parsed["email"].upper())

    if parsed["phone_digits"]:
        cleaned_mobile_expr = f"regexp_replace({mobile_expr}, '[^0-9]', '', 'g')"
        cleaned_landline_expr = f"regexp_replace({landline_expr}, '[^0-9]', '', 'g')"
        where_parts.append(f"({cleaned_mobile_expr} LIKE ? OR {cleaned_landline_expr} LIKE ?)")
        params.extend([f"%{parsed['phone_digits']}%", f"%{parsed['phone_digits']}%"])

    if parsed["zip5"]:
        cleaned_zip_expr = f"regexp_replace({zip_expr}, '[^0-9]', '', 'g')"
        where_parts.append(f"{cleaned_zip_expr} LIKE ?")
        params.append(f"{parsed['zip5']}%")

    if parsed["pa_id_digits"]:
        cleaned_paid_expr = f"regexp_replace({pa_id_expr}, '[^0-9]', '', 'g')"
        where_parts.append(f"{cleaned_paid_expr} = ?")
        params.append(parsed["pa_id_digits"])

    remaining_tokens = parsed["remaining_tokens"][:6]
    for tok in remaining_tokens:
        if tok.isdigit():
            where_parts.append(f"{general_haystack_expr} LIKE ?")
            params.append(f"%{tok}%")
        else:
            where_parts.append(f"({name_haystack_expr} LIKE ? OR {address_haystack_expr} LIKE ?)")
            params.extend([f"%{tok}%", f"%{tok}%"])

    if not where_parts:
        return pd.DataFrame()

    order_sql = "_LookupName, _LookupAddress"
    if remaining_tokens:
        exact_name = " ".join(remaining_tokens).upper()
        order_sql = f"case when upper(trim({lookup_name_expr})) = ? then 0 else 1 end, _LookupName, _LookupAddress"
        params.append(exact_name)

    sql = f'''
        SELECT
            src.*,
            {lookup_name_expr} AS _LookupName,
            {lookup_address_expr} AS _LookupAddress,
            {lookup_city_state_zip_expr} AS _LookupCityStateZip,
            {pa_id_expr} AS _LookupPAID,
            {lookup_key_expr} AS _LookupRowKey
        FROM ({base_sql}) src
        WHERE 1=1
        AND {' AND '.join(where_parts)}
        ORDER BY {order_sql}
        LIMIT {int(limit)}
    '''
    df = get_conn().execute(sql, params).df()
    return apply_voter_corrections_to_df(df)

def get_lookup_selected_row(results_df: pd.DataFrame):
    if results_df is None or results_df.empty:
        return None
    selected_key = st.session_state.get("lookup_selected_key", "")
    if selected_key:
        hit = results_df[results_df["_LookupRowKey"].astype(str) == str(selected_key)]
        if not hit.empty:
            return hit.iloc[0]
    return results_df.iloc[0]


def _lookup_norm_key(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())

def get_lookup_value(row, candidates, formatter=None) -> str:
    index_map = {}
    try:
        for actual_col in row.index:
            actual_str = str(actual_col)
            index_map[actual_str] = actual_col
            index_map[actual_str.strip().lower()] = actual_col
            index_map[actual_str.replace("_", "").replace(" ", "").strip().lower()] = actual_col
            index_map[_lookup_norm_key(actual_str)] = actual_col
    except Exception:
        pass

    for col in candidates:
        possible_keys = [
            col,
            str(col).strip().lower(),
            str(col).replace("_", "").replace(" ", "").strip().lower(),
            _lookup_norm_key(col),
        ]
        actual_col = None
        for key in possible_keys:
            if key in row.index:
                actual_col = key
                break
            if key in index_map:
                actual_col = index_map[key]
                break
        if actual_col is not None:
            value = row.get(actual_col)
            if formatter is not None:
                rendered = formatter(value)
            else:
                rendered = normalize_export_text(value)
            if normalize_export_text(rendered) != "":
                return rendered
    return ""

def get_lookup_dob(row) -> str:
    direct = get_lookup_value(
        row,
        [
            "DOB", "D_O_B", "Date of Birth", "DateOfBirth", "Birth Date", "BirthDate",
            "Birth Dt", "BirthDt", "Date Birth", "DateBirth", "Dob"
        ],
        formatter=format_lookup_date,
    )
    if normalize_export_text(direct):
        return direct

    try:
        for actual_col in row.index:
            norm = _lookup_norm_key(actual_col)
            if norm in {"dob", "dateofbirth", "birthdate", "birthdt", "datebirth"} or ("birth" in norm and "date" in norm):
                value = format_lookup_date(row.get(actual_col))
                if normalize_export_text(value):
                    return value
    except Exception:
        pass
    return ""

def get_lookup_registered_party(row) -> str:
    direct = get_lookup_value(
        row,
        ["Registered Party", "RegisteredParty", "Party", "Registration Party", "Voter Party"]
    )
    if normalize_export_text(direct):
        return direct
    try:
        for actual_col in row.index:
            norm = _lookup_norm_key(actual_col)
            if norm in {"party", "registeredparty", "registrationparty", "voterparty"}:
                value = normalize_export_text(row.get(actual_col))
                if value:
                    return value
    except Exception:
        pass
    return ""


def build_lookup_full_name(row) -> str:
    full_name = get_lookup_value(row, ["FullName", "Full Name", "Name"])
    if full_name:
        return normalize_name_value(full_name)
    parts = [
        get_lookup_value(row, ["FirstName", "First Name"]),
        get_lookup_value(row, ["MiddleName", "Middle Name"]),
        get_lookup_value(row, ["LastName", "Last Name"]),
        get_lookup_value(row, ["NameSuffix", "Suffix", "Name Suffix"]),
    ]
    return normalize_name_value(" ".join([p for p in parts if p]).strip())


def build_lookup_address(row) -> str:
    line1 = normalize_address_value(" ".join([
        get_lookup_value(row, ["House Number", "HouseNumber", "Street Number"]),
        get_lookup_value(row, ["Street Name", "StreetName", "Street"]),
    ]).strip())
    apt = get_lookup_value(row, ["Apartment Number", "ApartmentNumber", "Unit", "Apt"])
    if apt:
        line1 = f"{line1} Apt {apt}".strip()
    city = normalize_city_value(get_lookup_value(row, ["MailingCity", "Mailing City", "City", "MailCity"]))
    state = normalize_state_value(get_lookup_value(row, ["MailingState", "Mailing State", "State", "MailState"]))
    zip_code = clean_zip_value(get_lookup_value(row, ["MailingZip", "Mailing Zip", "ZIP", "Zip", "ZipCode", "ZIPCODE", "MailZip"]))
    line2 = " ".join([p for p in [city + "," if city else "", state, zip_code] if p]).strip().replace(" ,", ",")
    if line1 and line2:
        return f"{line1}\n{line2}"
    return line1 or line2


def render_lookup_field_block(title: str, rows: list[tuple[str, str]]):
    clean_rows = [{"Field": label, "Value": value} for label, value in rows if normalize_export_text(value) != ""]
    st.markdown(f"#### {title}")
    if not clean_rows:
        st.caption("No data available")
    else:
        st.dataframe(pd.DataFrame(clean_rows), width="stretch", hide_index=True)


def format_vote_method_label(value: str) -> str:
    raw = normalize_export_text(value).upper()
    mapping = {"AP": "At Poll", "MB": "Mail Ballot", "PROVISIONAL": "Provisional", "PV": "Provisional", "P": "Provisional", "": "DNV"}
    return mapping.get(raw, raw or "DNV")


def vote_method_icon(value: str) -> str:
    raw = normalize_export_text(value).upper()
    if raw == "MB":
        return "✉️"
    if raw == "AP":
        return "🗳️"
    if raw in {"PROVISIONAL", "PV", "P"}:
        return "🟨"
    return ""


def vote_method_title(value: str) -> str:
    raw = normalize_export_text(value).upper().strip()
    if raw == "MB":
        return "Mail Ballot"
    if raw in {"AP", "IP", "AT POLL", "AT POLLS", "POLL", "POLLING PLACE"}:
        return "At Poll"
    if raw in {"PROVISIONAL", "PV", "PROV"}:
        return "Provisional"
    return "Did Not Vote"


def is_did_not_vote_value(value: str) -> bool:
    raw = normalize_export_text(value).upper().strip()
    return raw in {"", "0", "O", "N", "NO", "DID NOT VOTE", "DNV", "NONE", "NAN", "NULL", "-", "—"}


def _lookup_election_years(row, election_prefix: str, max_year: int = 25, lookback_years: int = 10) -> list[int]:
    """Detect available election years from the actual voter row.

    Pipeline columns are usually G24_method / G24_party and P24_method / P24_party,
    while older lookup code expected G24_VM / G24_P. This keeps lookup compatible
    with both schemas and intentionally skips 2026 until that election is loaded.
    """
    prefix = str(election_prefix).upper()
    found = set()
    try:
        cols = list(row.index)
    except Exception:
        cols = []
    pat = re.compile(rf"^{re.escape(prefix)}(\d{{2}})(?:_(?:METHOD|PARTY|VM|P))$", re.IGNORECASE)
    for col in cols:
        m = pat.match(str(col).strip())
        if not m:
            continue
        yy = int(m.group(1))
        if yy <= max_year and yy >= max_year - lookback_years + 1:
            found.add(yy)
    if found:
        return sorted(found, reverse=True)
    return list(range(max_year, max_year - lookback_years, -1))


def _lookup_election_method(row, prefix: str, yy: int) -> str:
    prefix = str(prefix).upper()
    candidates = [
        f"{prefix}{yy}_method", f"{prefix}{yy}_METHOD",
        f"{prefix}{yy}_VoteMethod", f"{prefix}{yy}_Vote_Method",
        f"{prefix}{yy}_VM", f"{prefix}{yy}_Method",
        f"{prefix}{yy} method", f"{prefix}{yy} METHOD",
        f"{prefix}{yy}",
    ]
    direct = normalize_export_text(get_lookup_value(row, candidates)).upper()
    if direct:
        return direct

    # Broader legacy fallback: find any column for this election year that looks like a method/vote-method field.
    pat = re.compile(rf"^{re.escape(prefix)}0?{yy}(?:[^A-Z0-9]*(?:METHOD|VOTEMETHOD|VOTE_METHOD|VM))$", re.IGNORECASE)
    try:
        for col in row.index:
            if pat.match(str(col).strip()):
                val = normalize_export_text(row[col]).upper()
                if val:
                    return val
    except Exception:
        pass
    return ""


def _lookup_election_party(row, prefix: str, yy: int) -> str:
    prefix = str(prefix).upper()
    candidates = [
        f"{prefix}{yy}_party", f"{prefix}{yy}_PARTY",
        f"{prefix}{yy}_Party", f"{prefix}{yy}_P",
        f"{prefix}{yy} party", f"{prefix}{yy} PARTY",
    ]
    direct = normalize_export_text(get_lookup_value(row, candidates)).upper()
    if direct:
        return direct
    pat = re.compile(rf"^{re.escape(prefix)}0?{yy}(?:[^A-Z0-9]*(?:PARTY|P))$", re.IGNORECASE)
    try:
        for col in row.index:
            if pat.match(str(col).strip()):
                val = normalize_export_text(row[col]).upper()
                if val:
                    return val
    except Exception:
        pass
    return ""


def render_lookup_vote_history_matrix(row, election_prefix: str, title: str, start_year: int = 25, end_year: int = 16):
    years = _lookup_election_years(row, election_prefix, max_year=start_year, lookback_years=(start_year - end_year + 1))
    header_cells = ''.join([f'<th>{election_prefix}{yy}</th>' for yy in years])
    party_cells = []
    method_cells = []
    for yy in years:
        vm_raw = _lookup_election_method(row, election_prefix, yy)
        not_voted = is_did_not_vote_value(vm_raw)
        party_val = "" if not_voted else (_lookup_election_party(row, election_prefix, yy) or "")
        cell_class = "lookup-vh-cell lookup-vh-dnv" if not_voted else "lookup-vh-cell"
        method_text = "" if not_voted else vote_method_icon(vm_raw)
        title_attr = vote_method_title(vm_raw)
        party_cells.append(f'<td class="{cell_class}">{party_val}</td>')
        method_cells.append(f'<td class="{cell_class}" title="{title_attr}">{method_text}</td>')

    html = f'''<div class="lookup-vh-wrap">
  <div class="lookup-vh-title">{title}</div>
  <table class="lookup-vh-table">
    <thead>
      <tr><th></th>{header_cells}</tr>
    </thead>
    <tbody>
      <tr><td class="lookup-vh-rowhead">Party</td>{''.join(party_cells)}</tr>
      <tr><td class="lookup-vh-rowhead">Method</td>{''.join(method_cells)}</tr>
    </tbody>
  </table>
</div>'''
    st.markdown(html, unsafe_allow_html=True)


def render_lookup_vote_history_tables(row):
    st.markdown("#### Election History")
    render_lookup_vote_history_matrix(row, "G", "General Elections")
    render_lookup_vote_history_matrix(row, "P", "Primary Elections")
    legend_html = '''<div class="lookup-legend">
  <span><span class="lookup-legend-icon">✉️</span> Mail Ballot</span>
  <span><span class="lookup-legend-icon">🗳️</span> At Poll</span>
  <span><span class="lookup-legend-icon">🟨</span> Provisional</span>
  <span><span class="lookup-legend-swatch"></span> Did Not Vote</span>
</div>'''
    st.markdown(legend_html, unsafe_allow_html=True)




def get_selected_lookup_row(results_df: pd.DataFrame):
    if results_df is None or results_df.empty:
        return None
    selected_key = st.session_state.get("lookup_selected_key", "")
    valid_keys = set(results_df["_LookupRowKey"].tolist()) if "_LookupRowKey" in results_df.columns else set()
    if selected_key and selected_key in valid_keys:
        return results_df.loc[results_df["_LookupRowKey"] == selected_key].iloc[0]
    first_row = results_df.iloc[0]
    st.session_state["lookup_selected_key"] = first_row.get("_LookupRowKey", "")
    return first_row


def _pdf_vote_cell_fill(vm_raw: str):
    raw = normalize_export_text(vm_raw).upper()
    if raw == "MB":
        return colors.HexColor("#E8F5E9")
    if raw == "AP":
        return colors.HexColor("#E3F2FD")
    if raw in {"PROVISIONAL", "PV", "P"}:
        return colors.HexColor("#FFF3E0")
    return colors.HexColor("#ECEFF1")


def _pdf_vote_method_code(vm_raw: str) -> str:
    raw = normalize_export_text(vm_raw).upper()
    if raw == "MB":
        return "MB"
    if raw == "AP":
        return "AP"
    if raw in {"PROVISIONAL", "PV", "P"}:
        return "P"
    return "DNV"


def _draw_pdf_vote_history_table(c, row, x, y, title, prefix, start_year=25, end_year=16):
    years = _lookup_election_years(row, prefix, max_year=start_year, lookback_years=(start_year - end_year + 1))
    cell_w = 48
    row_h = 20
    label_w = 56
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x, y, title)
    top = y - 14

    c.setFillColor(colors.HexColor("#F0F2F5"))
    c.rect(x, top-row_h, label_w, row_h, stroke=1, fill=1)
    for i, yy in enumerate(years):
        cx = x + label_w + i * cell_w
        c.setFillColor(colors.HexColor("#F0F2F5"))
        c.rect(cx, top-row_h, cell_w, row_h, stroke=1, fill=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(cx + cell_w/2, top-13, f"{prefix}{yy}")

    party_y = top - row_h
    c.setFillColor(colors.white)
    c.rect(x, party_y-row_h, label_w, row_h, stroke=1, fill=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(x+6, party_y-13, "Party")
    for i, yy in enumerate(years):
        cx = x + label_w + i * cell_w
        vm_raw = _lookup_election_method(row, prefix, yy)
        party_val = _lookup_election_party(row, prefix, yy)
        c.setFillColor(_pdf_vote_cell_fill(vm_raw))
        c.rect(cx, party_y-row_h, cell_w, row_h, stroke=1, fill=1)
        c.setFillColor(colors.HexColor("#1E3A8A") if vm_raw else colors.HexColor("#667085"))
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(cx + cell_w/2, party_y-13, "" if is_did_not_vote_value(vm_raw) else (party_val or ""))

    method_y = party_y - row_h
    c.setFillColor(colors.white)
    c.rect(x, method_y-row_h, label_w, row_h, stroke=1, fill=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(x+6, method_y-13, "Method")
    for i, yy in enumerate(years):
        cx = x + label_w + i * cell_w
        vm_raw = _lookup_election_method(row, prefix, yy)
        c.setFillColor(_pdf_vote_cell_fill(vm_raw))
        c.rect(cx, method_y-row_h, cell_w, row_h, stroke=1, fill=1)
        c.setFillColor(colors.black if vm_raw else colors.HexColor("#98A2B3"))
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(cx + cell_w/2, method_y-13, _pdf_vote_method_code(vm_raw))

    return method_y - row_h - 10


def build_voter_report_pdf_bytes(row) -> bytes:
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(letter))
    width, height = landscape(letter)
    margin_x = 24

    header_top = height - 18
    header_logo_x = margin_x
    header_logo_y = header_top - 36
    if CC_LOGO.exists():
        try:
            c.drawImage(
                ImageReader(str(CC_LOGO)),
                header_logo_x,
                header_logo_y,
                width=104,
                height=34,
                preserveAspectRatio=True,
                mask='auto',
            )
        except Exception:
            pass

    title_x = header_logo_x + 116
    c.setFillColor(colors.HexColor("#173B73"))
    c.setFont("Helvetica-Bold", 13)
    c.drawString(title_x, header_top - 6, "Candidate Connect")
    c.setFont("Helvetica", 8)
    c.setFillColor(colors.HexColor("#4B5563"))
    c.drawString(title_x, header_top - 20, "Voter Lookup Report")
    c.drawString(title_x, header_top - 31, datetime.now().strftime("Generated %m/%d/%Y %I:%M %p"))

    logo_width = 56
    logo_height = 18
    logo_x = width - margin_x - logo_width
    logo_center_x = logo_x + (logo_width / 2)

    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(colors.HexColor("#4B5563"))
    c.drawCentredString(logo_center_x, header_top - 6, "Powered By")
    if get_reporting_tss_logo_path().exists():
        try:
            c.drawImage(
                ImageReader(str(TSS_LOGO)),
                logo_x,
                header_top - 26,
                width=logo_width,
                height=logo_height,
                preserveAspectRatio=True,
                mask='auto',
            )
        except Exception:
            pass

    divider_y = header_top - 42
    c.setStrokeColor(colors.HexColor("#D7DCE3"))
    c.line(margin_x, divider_y, width - margin_x, divider_y)

    voter_name = build_lookup_full_name(row) or "Unnamed voter"
    name_y = divider_y - 18
    c.setFont("Helvetica-Bold", 20)
    c.setFillColor(colors.HexColor("#8A1C1C"))
    c.drawString(margin_x, name_y, voter_name.upper())
    c.setFillColor(colors.black)

    address_title_y = name_y - 28
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin_x, address_title_y, "Address")
    c.setFont("Helvetica", 10)
    address_line_y = address_title_y - 16
    address_lines = [ln for ln in build_lookup_address(row).split("\n") if normalize_export_text(ln)]
    for line in address_lines:
        c.drawString(margin_x, address_line_y, line)
        address_line_y -= 14

    left_x, mid_x, right_x = margin_x, 285, 545
    top_y = address_line_y - 4

    c.setFont("Helvetica-Bold", 10)
    c.drawString(left_x, top_y, "Districts + Geography")
    left_end_y = top_y - 18
    for label, value in [
        ("County", get_lookup_value(row, ["County"])),
        ("Municipality", get_lookup_value(row, ["Municipality"])),
        ("Precinct", get_lookup_value(row, ["Precinct"])),
        ("USC", get_lookup_value(row, ["USC", "Congressional"], formatter=lambda v: normalize_numeric_string(v))),
        ("STS", get_lookup_value(row, ["STS", "State Senate"], formatter=lambda v: normalize_numeric_string(v))),
        ("STH", get_lookup_value(row, ["STH", "State House"], formatter=lambda v: normalize_numeric_string(v))),
        ("School District", get_lookup_value(row, ["School District"])),
    ]:
        c.setFont("Helvetica-Bold", 9)
        c.drawString(left_x, left_end_y, f"{label}:")
        c.setFont("Helvetica", 9)
        c.drawString(left_x + 88, left_end_y, normalize_export_text(value) or "—")
        left_end_y -= 14

    c.setFont("Helvetica-Bold", 10)
    c.drawString(mid_x, top_y, "Voter Snapshot")
    mid_end_y = top_y - 18
    for label, value in [
        ("DOB", get_lookup_dob(row)),
        ("Reg Date", get_lookup_value(row, ["RegistrationDate", "Registration Date"], formatter=format_lookup_date)),
        ("Last Vote", get_lookup_value(row, ["Last Vote", "LastVote"], formatter=format_lookup_date) or get_lookup_value(row, ["Last Vote", "LastVote"])),
        ("Last Change", get_lookup_value(row, ["Last Change Date", "LastChangeDate"], formatter=format_lookup_date) or get_lookup_value(row, ["Last Change", "LastChange"])),
        ("Registered Party", get_lookup_registered_party(row)),
        ("Gender", get_lookup_value(row, ["Gender", "Sex"])),
        ("Age", get_lookup_value(row, ["Age"], formatter=lambda v: normalize_numeric_string(v))),
        ("PA ID", get_lookup_value(row, ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "VoterID"], formatter=lambda v: normalize_numeric_string(v))),
    ]:
        c.setFont("Helvetica-Bold", 9)
        c.drawString(mid_x, mid_end_y, f"{label}:")
        c.setFont("Helvetica", 9)
        c.drawString(mid_x + 88, mid_end_y, normalize_export_text(value) or "—")
        mid_end_y -= 14

    c.setFont("Helvetica-Bold", 10)
    c.drawString(right_x, top_y, "Contact + Mail Ballot")
    right_end_y = top_y - 18
    for label, value in [
        ("Mobile", format_lookup_phone(get_lookup_value(row, ["Mobile"]))),
        ("Landline", format_lookup_phone(get_lookup_value(row, ["Landline", "PrimaryPhone", "Phone"]))),
        ("Email", get_lookup_value(row, ["Email"])),
        ("Applied", get_lookup_value(row, ["MIB_Applied"])),
        ("Status", get_lookup_value(row, ["MIB_BALLOT"])),
        ("Permanent", get_lookup_value(row, ["MB_PERM", "MB_Perm", "MB_Pern"])),
        ("MB Score", get_lookup_value(row, ["MB_AProp_Score", "MMB_AProp_Score"], formatter=lambda v: normalize_numeric_string(v))),
    ]:
        c.setFont("Helvetica-Bold", 9)
        c.drawString(right_x, right_end_y, f"{label}:")
        c.setFont("Helvetica", 9)
        c.drawString(right_x + 66, right_end_y, (normalize_export_text(value) or "—")[:32])
        right_end_y -= 14

    section_bottom_y = min(left_end_y, mid_end_y, right_end_y)
    table_y = max(250, section_bottom_y - 16)
    table_y = _draw_pdf_vote_history_table(c, row, margin_x, table_y, "General Elections", "G")
    table_y = _draw_pdf_vote_history_table(c, row, margin_x, table_y - 8, "Primary Elections", "P")

    legend_y = max(40, table_y - 12)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(margin_x, legend_y, "Legend:")
    c.setFont("Helvetica", 9)
    legend_items = ["MB = Mail Ballot", "AP = At Poll", "P = Provisional", "DNV = Did Not Vote"]
    lx = margin_x + 48
    for item in legend_items:
        c.drawString(lx, legend_y, item)
        lx += 128

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer.getvalue()




def get_lookup_household_members(selected_row) -> pd.DataFrame:
    """Return household members without assuming CamelCase name/address columns.

    Some detail shards use lowercase pipeline names like last_name/first_name, while
    older lookup code ordered by "LastName"/"FirstName" directly. That caused a
    Binder Error after finding a voter. This version uses detected expressions.
    """
    detail_paths, _ = ensure_detail_shards()
    detail_columns = get_conn().execute(
        f"DESCRIBE SELECT * FROM {dataset_scan_sql(detail_paths)}"
    ).df()["column_name"].tolist()

    base_sql, base_params = build_detail_export_sql(detail_paths, {})
    house_key = normalize_export_text(selected_row.get("_HouseholdKey", ""))
    pa_id_val = normalize_numeric_string(
        get_lookup_value(
            selected_row,
            ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "VoterID", "voter_id", "ID Number", "IDNumber"]
        )
    )

    # Detected source expressions from the detail shards / base export query.
    last_expr, _ = _detail_col_expr(detail_columns, ["LastName", "Last Name", "last_name", "LNAME"])
    first_expr, _ = _detail_col_expr(detail_columns, ["FirstName", "First Name", "first_name", "FNAME"])
    age_expr, _ = _detail_col_expr(detail_columns, ["Age", "age", "Age_Calc"])
    house_expr, _ = _detail_col_expr(detail_columns, ["House Number", "HouseNumber", "house_number", "Street Number"])
    street_expr, _ = _detail_col_expr(detail_columns, ["Street Name", "StreetName", "street_name", "Street"])
    apt_expr, _ = _detail_col_expr(detail_columns, ["Apartment Number", "ApartmentNumber", "apartment_number", "Unit", "Apt"])
    city_expr, _ = _detail_col_expr(detail_columns, ["MailingCity", "Mailing City", "City", "res_city", "mail_city", "MailCity"])
    county_expr, _ = _detail_col_expr(detail_columns, ["County", "county"])

    if house_key:
        member_where = "coalesce(_HouseholdKey, '') = ?"
        member_params = [house_key]
    else:
        house_num = normalize_export_text(get_lookup_value(selected_row, ["House Number", "HouseNumber", "house_number", "Street Number"]))
        street_name = normalize_export_text(get_lookup_value(selected_row, ["Street Name", "StreetName", "street_name", "Street"]))
        apt_num = normalize_export_text(get_lookup_value(selected_row, ["Apartment Number", "ApartmentNumber", "apartment_number", "Unit", "Apt"]))
        city_val = normalize_export_text(get_lookup_value(selected_row, ["MailingCity", "Mailing City", "City", "res_city", "MailCity"]))
        county_val = normalize_export_text(get_lookup_value(selected_row, ["County", "county"]))
        member_where = f"""
            upper(trim({house_expr})) = upper(trim(?))
            AND upper(trim({street_expr})) = upper(trim(?))
            AND upper(trim({apt_expr})) = upper(trim(?))
            AND upper(trim({city_expr})) = upper(trim(?))
            AND upper(trim({county_expr})) = upper(trim(?))
        """
        member_params = [house_num, street_name, apt_num, city_val, county_val]

    members_df = get_conn().execute(
        f"""
        SELECT *
        FROM ({base_sql}) src
        WHERE {member_where}
        ORDER BY upper(trim({last_expr})),
                 upper(trim({first_expr})),
                 try_cast({age_expr} as double) DESC NULLS LAST
        """,
        base_params + member_params,
    ).df()

    if members_df.empty:
        return members_df

    members_df = apply_voter_corrections_to_df(members_df)

    if pa_id_val:
        def _same_selected(row):
            row_pa = normalize_numeric_string(
                get_lookup_value(row, ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "VoterID", "voter_id", "ID Number", "IDNumber"])
            )
            return row_pa == pa_id_val
        members_df["_IsSelectedMember"] = members_df.apply(_same_selected, axis=1)
    else:
        members_df["_IsSelectedMember"] = False

    return members_df.reset_index(drop=True)

def render_lookup_empty_workspace():
    st.markdown('<div class="section-card empty-shell"><div class="small-header">Voter Lookup</div><div class="tiny-muted">Open <strong>Voter Lookup</strong> in the left menu, enter a voter search, and click <strong>Search</strong>.</div></div>', unsafe_allow_html=True)

def render_lookup_result_card(result_row, selected: bool):
    title = normalize_name_value(normalize_export_text(result_row.get("_LookupName", ""))) or "Unnamed voter"
    party = normalize_export_text(result_row.get("Party", ""))
    age_text = normalize_numeric_string(result_row.get("Age", ""))
    title_parts = [title]
    if party:
        title_parts.append(party)
    if age_text:
        title_parts.append(age_text)
    line0 = ", ".join(title_parts)
    line1 = normalize_address_value(normalize_export_text(result_row.get("_LookupAddress", "")))
    line2 = normalize_export_text(result_row.get("_LookupCityStateZip", ""))
    county = normalize_export_text(result_row.get("County", ""))
    county = f"{county} County" if county and "county" not in county.lower() else county
    card_class = "lookup-result-card selected" if selected else "lookup-result-card"
    html = f'''<div class="{card_class}">
  <div class="lookup-result-line0">{line0}</div>
  <div class="lookup-result-line1">{line1}</div>
  <div class="lookup-result-line2">{line2}</div>
  <div class="lookup-result-line3">{county}</div>
</div>'''
    st.markdown(html, unsafe_allow_html=True)


def render_voter_lookup_results():
    results_df = pd.DataFrame(st.session_state.get("lookup_results_records", []))
    lookup_query = st.session_state.get("lookup_query", "")

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="small-header">Voter Lookup</div>', unsafe_allow_html=True)
    st.caption("Showing lookup results from the full statewide active voter file.")

    if not normalize_export_text(lookup_query):
        st.info("Enter a voter search on the left and click Search.")
        st.markdown('</div>', unsafe_allow_html=True)
        return

    if results_df.empty:
        st.warning(f'No voters matched "{lookup_query}" in the statewide active voter file.')
        st.markdown('</div>', unsafe_allow_html=True)
        return

    st.caption(f"{len(results_df):,} result(s) found for: {lookup_query}")
    left_col, right_col = st.columns([1.02, 1.98], gap="large")

    with left_col:
        st.markdown("#### Search Results")
        for _, result_row in results_df.iterrows():
            row_key = result_row.get("_LookupRowKey", "")
            is_selected = st.session_state.get("lookup_selected_key", "") == row_key
            render_lookup_result_card(result_row, is_selected)
            if st.button("Selected" if is_selected else "View Voter", key=f'lookup_pick_{row_key}', width="stretch", type="primary" if is_selected else "secondary"):
                st.session_state["lookup_selected_key"] = row_key
                st.rerun()

    selected_row = get_selected_lookup_row(results_df)
    if selected_row is None:
        st.markdown('</div>', unsafe_allow_html=True)
        return
    selected_row = apply_voter_corrections_to_row(selected_row)

    with right_col:
        voter_name = normalize_name_value(normalize_export_text(selected_row.get("_LookupName", ""))) or "Unnamed voter"
        header_cols = st.columns([0.78, 0.22])
        with header_cols[0]:
            st.markdown(f"## {voter_name}")
            address_block = build_lookup_address(selected_row)
            if address_block:
                st.markdown(address_block.replace("\n", "  \n"))
        with header_cols[1]:
            pdf_bytes = build_voter_report_pdf_bytes(selected_row)
            safe_name = sanitize_filename_part(voter_name)
            st.download_button(
                "Download PDF Report",
                data=pdf_bytes,
                file_name=f"{safe_name}_voter_report.pdf",
                mime="application/pdf",
                width="stretch",
            )

        metric_cols = st.columns(4, gap="small")
        metric_cols[0].metric("Party", get_lookup_value(selected_row, ["Party"], formatter=lambda v: normalize_export_text(v)) or "—")
        metric_cols[1].metric("Gender", get_lookup_value(selected_row, ["Gender", "Sex"], formatter=lambda v: normalize_export_text(v)) or "—")
        metric_cols[2].metric("Age", get_lookup_value(selected_row, ["Age"], formatter=lambda v: normalize_numeric_string(v)) or "—")
        metric_cols[3].metric("PA ID", get_lookup_value(selected_row, ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "VoterID"], formatter=lambda v: normalize_numeric_string(v)) or "—")

        detail_cols = st.columns(2, gap="medium")
        with detail_cols[0]:
            render_lookup_field_block("Voter Details", [
                ("Date of Birth", get_lookup_dob(selected_row)),
                ("Registration Date", get_lookup_value(selected_row, ["RegistrationDate", "Registration Date"], formatter=format_lookup_date)),
                ("Registered Party", get_lookup_registered_party(selected_row)),
                ("Last Vote", get_lookup_value(selected_row, ["Last Vote", "LastVote"], formatter=format_lookup_date) or get_lookup_value(selected_row, ["Last Vote", "LastVote"])),
                ("Last Change", get_lookup_value(selected_row, ["Last Change", "LastChange"])),
                ("Last Change Date", get_lookup_value(selected_row, ["Last Change Date", "LastChangeDate"], formatter=format_lookup_date)),
                ("County", get_lookup_value(selected_row, ["County"])),
                ("Municipality", get_lookup_value(selected_row, ["Municipality"])),
                ("Precinct", get_lookup_value(selected_row, ["Precinct"])),
                ("Congressional", get_lookup_value(selected_row, ["USC", "Congressional", "Congressional District"], formatter=lambda v: normalize_numeric_string(v))),
                ("State Senate", get_lookup_value(selected_row, ["STS", "State Senate", "Senate District"], formatter=lambda v: normalize_numeric_string(v))),
                ("State House", get_lookup_value(selected_row, ["STH", "State House", "House District"], formatter=lambda v: normalize_numeric_string(v))),
                ("School District", get_lookup_value(selected_row, ["School District"])),
            ])
        with detail_cols[1]:
            render_lookup_field_block("Contact + Mail Ballot", [
                ("Mobile", format_lookup_phone(get_lookup_value(selected_row, ["Mobile"]))),
                ("Landline", format_lookup_phone(get_lookup_value(selected_row, ["Landline", "PrimaryPhone", "Phone"]))),
                ("Email", get_lookup_value(selected_row, ["Email"])),
                ("Mail Ballot Applied", get_lookup_value(selected_row, ["MIB_Applied"])),
                ("Mail Ballot Status", get_lookup_value(selected_row, ["MIB_BALLOT"])),
                ("Permanent Mail", get_lookup_value(selected_row, ["MB_PERM", "MB_Perm", "MB_Pern"])),
                ("App Return Date", get_lookup_value(selected_row, ["Current_App_Return_Date", "AppReturnDate"], formatter=format_lookup_date)),
                ("Ballot Sent Date", get_lookup_value(selected_row, ["Current_Ballot_Sent_Date", "BallotSentDate"], formatter=format_lookup_date)),
                ("Ballot Returned Date", get_lookup_value(selected_row, ["Current_Ballot_Returned_Date", "BallotReturnedDate"], formatter=format_lookup_date)),
                ("Mail Ballot Applicant Phone", format_lookup_phone(get_lookup_value(selected_row, ["Current_ApplicantPhone", "ApplicantPhone", "Applicant Phone"]))),
                ("Mail Ballot New Registrant", get_lookup_value(selected_row, ["MailBallotNewRegistrant"])),
                ("Mail Ballot Score", get_lookup_value(selected_row, ["MB_AProp_Score", "MMB_AProp_Score"], formatter=lambda v: normalize_numeric_string(v))),
            ])

        with st.expander("Edit / Correct This Voter Record", expanded=True):
            render_voter_correction_editor(selected_row)

        household_df = apply_voter_corrections_to_df(get_lookup_household_members(selected_row))
        st.markdown("#### Household Members")
        if household_df.empty:
            st.caption("No household members found.")
        else:
            for idx, member_row in household_df.iterrows():
                member_name = build_lookup_full_name(member_row) or "Unnamed voter"
                member_party = get_lookup_value(member_row, ["Party"])
                member_age = get_lookup_value(member_row, ["Age"], formatter=lambda v: normalize_numeric_string(v))
                member_line = member_name
                meta_bits = [bit for bit in [member_party, member_age] if normalize_export_text(bit)]
                if meta_bits:
                    member_line += ", " + ", ".join(meta_bits)
                is_selected_member = bool(member_row.get("_IsSelectedMember", False))
                member_cols = st.columns([5, 1.4])
                with member_cols[0]:
                    st.markdown(f"- **{member_line}**")
                with member_cols[1]:
                    if is_selected_member:
                        st.caption("Current")
                    else:
                        member_pa_id = normalize_numeric_string(
                            get_lookup_value(
                                member_row,
                                ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "VoterID"]
                            )
                        )
                        member_button_key = member_pa_id or f"member_{idx}"
                        if st.button("Open", key=f"hh_open_{member_button_key}_{idx}", width="stretch"):
                            if member_pa_id:
                                st.session_state["lookup_household_open_pa_id"] = member_pa_id
                                st.session_state.workspace_mode = "lookup"
                                st.rerun()

        render_lookup_vote_history_tables(selected_row)

    st.markdown('</div>', unsafe_allow_html=True)



GEO_FILTER_COLUMNS = ["County", "Municipality", "Precinct", "USC", "STS", "STH", "School District", "School Region"]

def _geo_value_clean(v) -> str:
    return normalize_export_text(v) if "normalize_export_text" in globals() else str(v or "").strip()


@st.cache_data(show_spinner=False)
def load_geo_dependency_table() -> pd.DataFrame:
    """Load compact geography hierarchy for interdependent geography filters."""
    candidates = [
        Path("data/shards/speed/geo_hierarchy.parquet"),
        Path("data") / "shards" / "speed" / "geo_hierarchy.parquet",
    ]
    for path in candidates:
        try:
            if path.exists():
                df = pd.read_parquet(path)
                for col in GEO_FILTER_COLUMNS:
                    if col not in df.columns:
                        df[col] = ""
                    df[col] = df[col].astype(str).map(_geo_value_clean)
                return df[GEO_FILTER_COLUMNS].drop_duplicates().reset_index(drop=True)
        except Exception:
            pass

    rows = []
    try:
        opts = st.session_state.get("options", {}) or {}
        max_len = max([len(opts.get(c, [])) for c in GEO_FILTER_COLUMNS] + [0])
        for i in range(max_len):
            row = {}
            for col in GEO_FILTER_COLUMNS:
                vals = opts.get(col, []) or []
                row[col] = vals[i] if i < len(vals) else ""
            rows.append(row)
    except Exception:
        rows = []
    df = pd.DataFrame(rows)
    for col in GEO_FILTER_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[GEO_FILTER_COLUMNS].fillna("").astype(str)



def geo_sort_key(value):
    """Sort district-style geography values numerically when possible."""
    s = str(value or "").strip()
    m = re.search(r'(\d+)', s)
    if m:
        try:
            return (0, int(m.group(1)), s)
        except Exception:
            pass
    return (1, s.lower())


def _geo_options_for_column(geo_df: pd.DataFrame, target_col: str, current_selections: dict, fallback_options: list[str]) -> list[str]:
    if geo_df is None or geo_df.empty or target_col not in geo_df.columns:
        return fallback_options

    temp = geo_df.copy()
    for col, selected in (current_selections or {}).items():
        if col == target_col or col not in temp.columns:
            continue
        selected = [_geo_value_clean(x) for x in (selected or []) if _geo_value_clean(x)]
        if selected:
            temp = temp[temp[col].astype(str).isin(selected)]

    vals = sorted(
        {str(v).strip() for v in temp[target_col].dropna().astype(str).tolist() if str(v).strip() and str(v).strip() != "(Blank)"},
        key=geo_sort_key
    )
    return vals or fallback_options


def render_interdependent_geo_filters(cols, opts) -> dict:
    """Render geography filters with mutual narrowing."""
    geo_cols = [c for c in GEO_FILTER_COLUMNS if c in cols]
    geo_df = load_geo_dependency_table()
    selections = {}

    st.markdown("<div class='cc-active-section-title'>Geography</div>", unsafe_allow_html=True)
    st.caption("Geography options narrow together as selections are made.")

    current = {}
    for col in geo_cols:
        key = f"geo_dep_{_norm_col_name(col)}"
        existing = st.session_state.get(key, st.session_state.active_filters.get(col, []))
        current[col] = existing if isinstance(existing, list) else []

    for col in geo_cols:
        key = f"geo_dep_{_norm_col_name(col)}"
        fallback = opts.get(col, []) or []
        options = _geo_options_for_column(geo_df, col, current, fallback)

        extras = [v for v in current.get(col, []) if v and v not in options]
        full_options = options + [v for v in extras if v not in options]

        selections[col] = st.multiselect(
            geo_label(col),
            full_options,
            default=sanitize_multiselect_defaults(current.get(col, []), full_options),
            key=key,
            help="Options narrow based on your other geography selections.",
        )
        current[col] = selections[col]

    return selections


def render_lookup_sidebar(active_filters, columns):
    if st.session_state.pop("lookup_clear_requested", False):
        st.session_state["lookup_query"] = ""
        st.session_state["lookup_results_records"] = []
        st.session_state["lookup_selected_key"] = ""
        st.session_state["lookup_last_query"] = ""
        st.session_state["lookup_view_active"] = False
        st.session_state["lookup_query_input"] = ""

    pending_household_open_query = normalize_numeric_string(st.session_state.pop("lookup_household_open_pa_id", ""))
    if pending_household_open_query:
        st.session_state["lookup_query_input"] = pending_household_open_query
        st.session_state["lookup_query"] = pending_household_open_query
        st.session_state["lookup_last_query"] = pending_household_open_query
        st.session_state["lookup_selected_key"] = ""

    with st.expander("Voter Lookup", expanded=st.session_state.get("workspace_mode", "landing") == "lookup"):
        st.caption("Search the full statewide active voter file by name, county, address, PA ID, phone, or email.")
        with st.form("lookup_form", clear_on_submit=False):
            lookup_query = st.text_input(
                "Search voters",
                placeholder="Example: Jane Smith Lancaster, Jane Smith 17520, PA ID, phone, or email",
                key="lookup_query_input",
            )
            result_limit = st.selectbox("Max Results", [10, 25, 50, 100], index=1, key="lookup_result_limit")
            action_cols = st.columns(2, gap="small")
            search_clicked = action_cols[0].form_submit_button("Search", width="stretch", type="primary")
            clear_clicked = action_cols[1].form_submit_button("Clear Lookup", width="stretch")

        if clear_clicked:
            st.session_state["lookup_clear_requested"] = True
            st.session_state.workspace_mode = "lookup"
            st.rerun()

        run_lookup_search = (search_clicked or bool(pending_household_open_query)) and lookup_query.strip()

        if run_lookup_search:
            with st.spinner("Searching voter detail shards..."):
                results_df = search_voters_for_lookup(active_filters, lookup_query.strip(), limit=int(result_limit), use_current_filters=False)
            st.session_state["lookup_query"] = lookup_query.strip()
            st.session_state["lookup_last_query"] = lookup_query.strip()
            st.session_state["lookup_results_records"] = results_df.to_dict("records")
            if pending_household_open_query and not results_df.empty:
                selected_match = None
                for _, _row in results_df.iterrows():
                    row_pa_id = normalize_numeric_string(
                        get_lookup_value(
                            _row,
                            ["PA ID Number", "PA_ID_Number", "PA ID", "StateVoterID", "VoterID"]
                        )
                    )
                    if row_pa_id == pending_household_open_query:
                        selected_match = _row["_LookupRowKey"]
                        break
                st.session_state["lookup_selected_key"] = selected_match or results_df.iloc[0]["_LookupRowKey"]
            else:
                st.session_state["lookup_selected_key"] = results_df.iloc[0]["_LookupRowKey"] if not results_df.empty else ""
            st.session_state["lookup_view_active"] = True
            st.session_state.workspace_mode = "lookup"
            st.rerun()




# -----------------------------
# Area Intelligence (Phase 2)
# -----------------------------
@st.cache_data(show_spinner=False)
def load_area_precinct_summary() -> pd.DataFrame:
    """Load Area Intelligence summary from authenticated R2 first, then local fallback.

    This avoids the earlier Cloudflare public-read/403 issue because the app reads
    area_intelligence/precinct_summary.csv through the R2 S3 API using Streamlit secrets.
    """
    key = "area_intelligence/precinct_summary.csv"
    local_path = Path("area_intelligence") / "precinct_summary.csv"
    errors = []

    # 1) Preferred: authenticated R2 read from the current environment bucket.
    try:
        client, info = get_saved_universes_r2_client()
        if client is not None:
            obj = client.get_object(Bucket=info["bucket"], Key=key)
            payload = obj["Body"].read()
            return pd.read_csv(BytesIO(payload), dtype=str).fillna("")
        errors.append("Authenticated R2 not configured")
    except Exception as e:
        errors.append(f"Authenticated R2: {e}")

    # 2) Fallback: local GitHub/repo file, useful if R2 credentials are missing.
    try:
        if local_path.exists():
            return pd.read_csv(local_path, dtype=str).fillna("")
        errors.append(f"Local file missing: {local_path}")
    except Exception as e:
        errors.append(f"Local: {e}")

    # 3) Last resort: public R2 URL, if the object happens to be public.
    try:
        url = r2_public_url(key)
        return pd.read_csv(url, dtype=str).fillna("")
    except Exception as e:
        errors.append(f"Public R2: {e}")

    raise FileNotFoundError(
        "Could not load area_intelligence/precinct_summary.csv from authenticated R2, local fallback, or public R2. "
        + " | ".join(errors)
    )


def _area_num(row, col, default=0.0):
    try:
        return float(str(row.get(col, default)).replace(",", "") or default)
    except Exception:
        return float(default)


def _metric_html(label: str, value: str, note: str = "") -> str:
    note_html = f'<div class="tiny-muted">{note}</div>' if note else ""
    return f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{value}</div>{note_html}</div>'


def _aggregate_area_profile(profile_df: pd.DataFrame) -> dict:
    """Aggregate one or more precinct rows into a single area profile."""
    numeric_cols = [
        "Total_Voters", "Dem_Voters", "Rep_Voters", "Other_Voters",
        "Male_Voters", "Female_Voters", "Unknown_Gender",
        "New_Registrations", "Mail_Applications", "Mail_Applications_Total", "Mail_Applications_Approved", "Mail_Applications_Declined",
        "Mail_Ballots_Sent", "Mail_Ballots_Returned", "Mail_Ballots_Outstanding", "Mail_Voters"
    ]
    out = {}
    work = profile_df.copy()
    for col in numeric_cols + ["Avg_Age"]:
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0)
        else:
            work[col] = 0

    for col in numeric_cols:
        out[col] = float(work[col].sum())

    total = out.get("Total_Voters", 0)
    if total > 0 and "Avg_Age" in work.columns:
        out["Avg_Age"] = float((work["Avg_Age"] * work["Total_Voters"]).sum() / total)
    else:
        out["Avg_Age"] = 0.0

    out["Precinct_Count"] = int(len(work))
    return out


def _strategy_badge(text: str, tone: str = "neutral") -> str:
    colors = {
        "good": ("#e8f5e9", "#1b5e20"),
        "watch": ("#fff8e1", "#8a5a00"),
        "priority": ("#ffebee", "#b71c1c"),
        "info": ("#e3f2fd", "#0d47a1"),
        "neutral": ("#f5f5f5", "#374151"),
    }
    bg, fg = colors.get(tone, colors["neutral"])
    return (
        f'<span style="display:inline-block; padding:6px 10px; margin:3px 5px 3px 0; '
        f'border-radius:999px; background:{bg}; color:{fg}; font-size:12px; font-weight:800;">{text}</span>'
    )


def _build_strategy_summary(total, dem, rep, other, new_reg, mail_apps, mail_returned, mail_outstanding, geo_issues):
    total = float(total or 0)
    dem_pct = 0 if total <= 0 else dem / total * 100
    rep_pct = 0 if total <= 0 else rep / total * 100
    new_reg_pct = 0 if total <= 0 else new_reg / total * 100
    app_pct = 0 if total <= 0 else mail_apps / total * 100
    return_rate = 0 if mail_apps <= 0 else mail_returned / mail_apps * 100
    outstanding_rate = 0 if mail_apps <= 0 else mail_outstanding / mail_apps * 100

    badges = []
    notes = []

    if rep_pct >= 55:
        badges.append(("Republican Advantage Area", "good"))
        notes.append("GOP-friendly geography. Strong area for base turnout and mail ballot chase.")
    elif dem_pct >= 55:
        badges.append(("Democratic Advantage Area", "priority"))
        notes.append("Democratic-leaning geography. Use for opposition awareness and selective persuasion.")
    elif abs(rep_pct - dem_pct) <= 8:
        badges.append(("Persuasion Opportunity", "watch"))
        notes.append("Party balance is close enough to justify persuasion and turnout monitoring.")
    else:
        badges.append(("Mixed Performance Area", "info"))
        notes.append("Not heavily one-sided. Review party mix and turnout behavior before assigning resources.")

    if mail_apps > 0:
        if return_rate < 35:
            badges.append(("Low Mail Return - Chase Priority", "priority"))
            notes.append("Mail ballot requests exist, but return rate is low. Prioritize chase calls, texts, and door contact.")
        elif return_rate < 65:
            badges.append(("Medium Mail Return - Watch", "watch"))
            notes.append("Mail return is moving but not complete. Keep this area on the chase list.")
        else:
            badges.append(("High Mail Return", "good"))
            notes.append("Many requested ballots have already returned. Reduce chase pressure on returned voters.")
    else:
        badges.append(("Low Mail Application Universe", "info"))
        notes.append("Few or no mail applications are currently visible. Consider application-growth messaging if strategically useful.")

    if mail_outstanding > 0:
        badges.append((f"{int(mail_outstanding):,} Outstanding Ballots", "priority" if outstanding_rate >= 40 else "watch"))

    if new_reg_pct >= 2:
        badges.append(("New Registration Watch", "watch"))
        notes.append("New registrations are elevated. Check whether they need education, ID, or first-time voter messaging.")

    if geo_issues > 0:
        badges.append(("Geography Update Watch", "info"))
        notes.append("Some rows required geography repair or reflect newer election geography than the base voter file.")

    return badges, notes, return_rate, outstanding_rate, app_pct



# Area Intelligence table renderer: centered values, comma formatting, sticky headers/label columns.
def _ai_format_cell_value(value, col_name=""):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    if text == "":
        return ""
    if "%" in text or text == "—":
        return text
    try:
        cleaned = text.replace(",", "")
        num = float(cleaned)
        if col_name in {"Avg_Age", "Dem_%", "Rep_%", "Other_%", "Mail_Return_%", "Outstanding_%"}:
            return f"{num:,.1f}".rstrip("0").rstrip(".")
        if abs(num - round(num)) < 0.000001:
            return f"{int(round(num)):,}"
        return f"{num:,.1f}"
    except Exception:
        return text


def _ai_clean_display_df(df):
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    drop_cols = []
    for c in out.columns:
        name = str(c).strip()
        if name == "" or name.lower().startswith("unnamed") or name.lower() in {"index", "level_0"}:
            drop_cols.append(c)
    if drop_cols:
        out = out.drop(columns=drop_cols, errors="ignore")
    for c in out.columns:
        out[c] = out[c].map(lambda v, col=c: _ai_format_cell_value(v, str(col)))
    return out


def _ai_render_table(df, height=360, sticky_cols=None, key=""):
    display = _ai_clean_display_df(df)
    if display.empty:
        st.caption("No table data available.")
        return
    sticky_cols = sticky_cols or []
    cols = [str(c) for c in display.columns]
    sticky_set = {c for c in sticky_cols if c in cols}
    import html as _html
    def esc(x):
        return _html.escape(str(x))
    sticky_positions = {cols[i]: i * 155 for i in range(min(3, len(cols))) if cols[i] in sticky_set}
    table_id = f"ai-table-{key}" if key else "ai-table"
    css = f"""
    <style>
    .{table_id}-wrap {{ width:100%; max-height:{int(height)}px; overflow:auto; border:1px solid #e5e7eb; border-radius:12px; background:white; }}
    table.{table_id} {{ border-collapse:separate; border-spacing:0; width:max-content; min-width:100%; font-size:12px; }}
    table.{table_id} th, table.{table_id} td {{ border-right:1px solid #edf0f2; border-bottom:1px solid #edf0f2; padding:8px 10px; text-align:center !important; vertical-align:middle; white-space:nowrap; min-width:110px; }}
    table.{table_id} th {{ position:sticky; top:0; z-index:5; background:#f8fafc; color:#24303f; font-weight:800; }}
    table.{table_id} td {{ background:white; color:#24303f; }}
    table.{table_id} tr:hover td {{ background:#f7fbff; }}
    table.{table_id} .sticky-col {{ position:sticky; z-index:4; background:#ffffff; box-shadow:1px 0 0 #e5e7eb; font-weight:700; }}
    table.{table_id} th.sticky-col {{ z-index:7; background:#f8fafc; }}
    </style>
    """
    header_cells = []
    for c in cols:
        cls = "sticky-col" if c in sticky_set else ""
        style = f"left:{sticky_positions.get(c, 0)}px; min-width:155px;" if c in sticky_set else ""
        header_cells.append(f'<th class="{cls}" style="{style}">{esc(c)}</th>')
    rows_html = []
    for _, r in display.iterrows():
        tds = []
        for c in cols:
            cls = "sticky-col" if c in sticky_set else ""
            style = f"left:{sticky_positions.get(c, 0)}px; min-width:155px;" if c in sticky_set else ""
            tds.append(f'<td class="{cls}" style="{style}">{esc(r[c])}</td>')
        rows_html.append("<tr>" + "".join(tds) + "</tr>")
    html_table = css + '<div class="{}-wrap"><table class="{}"><thead><tr>{}</tr></thead><tbody>{}</tbody></table></div>'.format(table_id, table_id, "".join(header_cells), "".join(rows_html))
    st.markdown(html_table, unsafe_allow_html=True)

def render_area_intelligence_workspace():
    st.markdown('<div class="section-card"><div class="small-header">Area Intelligence</div><div class="tiny-muted">Regional intelligence and campaign strategy workspace.</div></div>', unsafe_allow_html=True)

    try:
        area_df = load_area_precinct_summary()
    except Exception as e:
        st.error("Area Intelligence file could not be loaded.")
        st.caption(str(e))
        st.info("Expected file path: area_intelligence/precinct_summary.csv")
        return

    required_cols = ["County", "Municipality", "Precinct"]
    missing = [c for c in required_cols if c not in area_df.columns]
    if missing:
        st.error("The precinct summary file is missing required columns: " + ", ".join(missing))
        _ai_render_table(pd.DataFrame({"Available Columns": list(area_df.columns)}), height=300, sticky_cols=["Available Columns"], key="missingcols")
        return

    # Normalize Area Intelligence geography fields. District columns are optional,
    # but when present they power district-level reports.
    for col in ["County", "Municipality", "Precinct", "USC", "STS", "STH", "School District", "School Region"]:
        if col not in area_df.columns:
            area_df[col] = ""
        area_df[col] = area_df[col].astype(str).fillna("").replace({"nan": "", "None": ""}).str.strip()
        if col in ["USC", "STS", "STH"]:
            area_df[col] = area_df[col].map(normalize_numeric_string)

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="small-header">Select Area</div>', unsafe_allow_html=True)

    available_levels = ["County", "Municipality", "Precinct"]
    for _lvl in ["USC", "STS", "STH", "School District"]:
        if _lvl in area_df.columns and any(str(x).strip() for x in area_df[_lvl].unique().tolist()):
            available_levels.append(_lvl)

    area_level = st.selectbox(
        "Report Level",
        available_levels,
        index=available_levels.index("Precinct") if "Precinct" in available_levels else 0,
        key="ai_area_level",
        help="Choose whether this profile should summarize a county, municipality, precinct, or district."
    )

    def _district_sort_key(v):
        text = normalize_numeric_string(v)
        try:
            return (0, int(float(text)), text)
        except Exception:
            return (1, text)

    def _clean_options(series, numeric=False):
        vals = [str(x).strip() for x in series.dropna().astype(str).tolist() if str(x).strip() and str(x).strip().lower() not in {"nan", "none"}]
        if numeric:
            vals = [normalize_numeric_string(v) for v in vals]
        vals = sorted(set(vals), key=_district_sort_key if numeric else lambda x: x)
        return vals

    c1, c2, c3 = st.columns(3)
    selected_county = ""
    selected_muni = ""
    selected_precinct = ""
    selected_district = ""
    profile_df = pd.DataFrame()
    title = ""

    if area_level in ["County", "Municipality", "Precinct"]:
        counties = _clean_options(area_df["County"])
        with c1:
            selected_county = st.selectbox("County", counties, key="ai_county") if counties else ""

        county_df = area_df[area_df["County"] == selected_county].copy() if selected_county else area_df.copy()
        municipalities = _clean_options(county_df["Municipality"])
        if area_level in ["Municipality", "Precinct"]:
            with c2:
                selected_muni = st.selectbox("Municipality", municipalities, key="ai_municipality") if municipalities else ""
        else:
            with c2:
                st.caption("Municipality not needed for county report")

        muni_df = county_df[county_df["Municipality"] == selected_muni].copy() if selected_muni else county_df.copy()
        precincts = _clean_options(muni_df["Precinct"])
        if area_level == "Precinct":
            with c3:
                selected_precinct = st.selectbox("Precinct", precincts, key="ai_precinct") if precincts else ""
        else:
            with c3:
                st.caption("Precinct not needed for this report level")

        if area_level == "County":
            profile_df = county_df.copy()
            title = f"{selected_county} County"
        elif area_level == "Municipality":
            profile_df = muni_df.copy() if selected_muni else pd.DataFrame()
            title = f"{selected_muni} • {selected_county}"
        else:
            profile_df = muni_df[muni_df["Precinct"] == selected_precinct].copy() if selected_precinct else pd.DataFrame()
            title = f"{selected_precinct} • {selected_muni} • {selected_county}"

    else:
        district_col = area_level
        numeric_district = area_level in ["USC", "STS", "STH"]
        district_options = _clean_options(area_df[district_col], numeric=numeric_district)
        with c1:
            selected_district = st.selectbox(area_level, district_options, key=f"ai_district_{area_level}") if district_options else ""

        if selected_district:
            compare_series = area_df[district_col].astype(str).map(normalize_numeric_string if numeric_district else lambda x: str(x).strip())
            district_df = area_df[compare_series == selected_district].copy()
        else:
            district_df = pd.DataFrame()

        county_options = ["All Counties"] + _clean_options(district_df["County"] if not district_df.empty else area_df["County"])
        with c2:
            selected_county_filter = st.selectbox("County Filter", county_options, key=f"ai_county_filter_{area_level}") if county_options else "All Counties"
        if selected_county_filter and selected_county_filter != "All Counties" and not district_df.empty:
            district_df = district_df[district_df["County"] == selected_county_filter].copy()
        with c3:
            st.caption("Municipality/precinct are included in the breakdown below")

        profile_df = district_df.copy()
        title = f"{area_level} {selected_district}"
        if selected_county_filter and selected_county_filter != "All Counties":
            title += f" • {selected_county_filter} County"

    st.markdown('</div>', unsafe_allow_html=True)

    if profile_df.empty:
        st.warning("No Area Intelligence data found for this selection.")
        return

    row = _aggregate_area_profile(profile_df)
    total = _area_num(row, "Total_Voters", 0)
    dem = _area_num(row, "Dem_Voters", 0)
    rep = _area_num(row, "Rep_Voters", 0)
    other = _area_num(row, "Other_Voters", 0)
    male = _area_num(row, "Male_Voters", 0)
    female = _area_num(row, "Female_Voters", 0)
    unknown_gender = _area_num(row, "Unknown_Gender", 0)
    avg_age = _area_num(row, "Avg_Age", 0)
    new_reg = _area_num(row, "New_Registrations", 0)
    mail_apps_total = _area_num(row, "Mail_Applications_Total", _area_num(row, "Mail_Applications", 0))
    mail_apps_approved = _area_num(row, "Mail_Applications_Approved", _area_num(row, "Mail_Applications", 0))
    mail_apps_declined = _area_num(row, "Mail_Applications_Declined", 0)
    mail_sent = _area_num(row, "Mail_Ballots_Sent", 0)
    mail_returned = _area_num(row, "Mail_Ballots_Returned", 0)
    if mail_returned == 0:
        mail_returned = _area_num(row, "Mail_Voters", 0)
    mail_outstanding = _area_num(row, "Mail_Ballots_Outstanding", max(mail_apps_approved - mail_returned, 0))

    # Safety repair for older precinct_summary.csv files or source rows where application status
    # is missing but sent/outstanding/returned counts prove an approved application exists.
    inferred_approved = max(mail_apps_approved, mail_outstanding + mail_returned, mail_sent, mail_returned)
    if inferred_approved > mail_apps_approved:
        mail_apps_approved = inferred_approved
    inferred_total = mail_apps_approved + mail_apps_declined
    if inferred_total > mail_apps_total:
        mail_apps_total = inferred_total
    mail_outstanding = max(mail_apps_approved - mail_returned, 0)

    # Backward-compatible name used by existing strategy logic: approved applications.
    mail_apps = mail_apps_approved
    geo_issues = _area_num(row, "Geo_Issue_Rows", 0)
    precinct_count = int(_area_num(row, "Precinct_Count", len(profile_df)))

    def pct_val(n, denom=None):
        denom = total if denom is None else denom
        return 0 if float(denom or 0) <= 0 else (float(n or 0) / float(denom)) * 100
    def pct_txt(n, denom=None):
        return fmt_pct(pct_val(n, denom))

    mail_return_rate = pct_val(mail_returned, mail_apps)
    mail_outstanding_rate = pct_val(mail_outstanding, mail_apps)
    badges, strategy_notes, _, _, _ = _build_strategy_summary(
        total, dem, rep, other, new_reg, mail_apps, mail_returned, mail_outstanding, geo_issues
    )

    st.markdown(f'<div class="section-card"><div class="small-header">{area_level} Profile</div><div class="tiny-muted">{title} &nbsp;•&nbsp; {precinct_count:,} precinct row(s) included</div></div>', unsafe_allow_html=True)

    # Cleaner top snapshot: one compact row for voter universe and party split.
    top_cols = st.columns(5, gap="small")
    top_cards = [
        ("Total Voters", f"{int(total):,}", "profile universe"),
        ("Democratic", f"{int(dem):,}", pct_txt(dem)),
        ("Republican", f"{int(rep):,}", pct_txt(rep)),
        ("Other / Unaffiliated", f"{int(other):,}", pct_txt(other)),
        ("Average Age", f"{avg_age:.1f}" if avg_age else "—", "weighted" if area_level != "Precinct" else ""),
    ]
    for col, (label, value, note) in zip(top_cols, top_cards):
        with col:
            st.markdown(_metric_html(label, value, note), unsafe_allow_html=True)

    with st.expander("More profile details", expanded=False):
        detail_cols = st.columns(4, gap="small")
        more_cards = [
            ("Male", f"{int(male):,}", pct_txt(male)),
            ("Female", f"{int(female):,}", pct_txt(female)),
            ("Unknown Gender", f"{int(unknown_gender):,}", pct_txt(unknown_gender)),
            ("New Registrations", f"{int(new_reg):,}", pct_txt(new_reg)),
        ]
        for col, (label, value, note) in zip(detail_cols, more_cards):
            with col:
                st.markdown(_metric_html(label, value, note), unsafe_allow_html=True)

    # Mail Program: compact table plus two decision cards.
    st.markdown('<div class="section-card"><div class="small-header">Mail Program</div><div class="tiny-muted">Approved/declined applications, sent ballots, returned ballots, and chase universe.</div></div>', unsafe_allow_html=True)
    mail_left, mail_right = st.columns([2, 1], gap="medium")
    with mail_left:
        mail_df = pd.DataFrame({
            "Stage": ["Applications Total", "Applications Approved", "Applications Declined", "Ballots Sent", "Ballots Returned", "Outstanding Ballots"],
            "Voters": [int(mail_apps_total), int(mail_apps_approved), int(mail_apps_declined), int(mail_sent), int(mail_returned), int(mail_outstanding)],
            "% of Voters": [pct_txt(mail_apps_total), pct_txt(mail_apps_approved), pct_txt(mail_apps_declined), pct_txt(mail_sent), pct_txt(mail_returned), pct_txt(mail_outstanding)],
            "% of Approved": ["—", "100%" if mail_apps_approved else "—", "—", pct_txt(mail_sent, mail_apps_approved) if mail_apps_approved else "—", pct_txt(mail_returned, mail_apps_approved) if mail_apps_approved else "—", pct_txt(mail_outstanding, mail_apps_approved) if mail_apps_approved else "—"],
        })
        _ai_render_table(mail_df, height=240, sticky_cols=["Stage"], key="mail")
    with mail_right:
        st.markdown(_metric_html("Outstanding Ballots", f"{int(mail_outstanding):,}", f"{mail_outstanding_rate:.1f}% of approved applications" if mail_apps else "No chase universe visible"), unsafe_allow_html=True)
        st.markdown(_metric_html("Return Rate", f"{mail_return_rate:.1f}%" if mail_apps else "—", "Returned / Approved"), unsafe_allow_html=True)

    # Strategy Summary gets a visual block and stays above charts.
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="small-header">Strategy Summary</div>', unsafe_allow_html=True)
    st.markdown("".join(_strategy_badge(text, tone) for text, tone in badges), unsafe_allow_html=True)
    if strategy_notes:
        st.markdown("<ul>" + "".join(f"<li>{note}</li>" for note in strategy_notes[:4]) + "</ul>", unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    chart_tab, breakdown_tab, debug_tab = st.tabs(["Charts", "Area Breakdown", "Debug"])

    with chart_tab:
        chart_col1, chart_col2 = st.columns(2, gap="medium")
        with chart_col1:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.markdown('<div class="small-header">Party Breakdown</div>', unsafe_allow_html=True)
            party_chart = pd.DataFrame({"Party": ["Democratic", "Republican", "Other"], "Voters": [dem, rep, other]})
            party_chart["Voters"] = pd.to_numeric(party_chart["Voters"], errors="coerce").fillna(0)
            party_chart["Percent"] = party_chart["Voters"].apply(lambda x: pct_val(x))
            party_colors = [PARTY_NAME_COLOR_MAP["Democratic"], PARTY_NAME_COLOR_MAP["Republican"], PARTY_NAME_COLOR_MAP["Other"]]
            if party_chart["Voters"].sum() > 0:
                chart = alt.Chart(party_chart).mark_arc(innerRadius=62, outerRadius=98).encode(
                    theta=alt.Theta(field="Voters", type="quantitative"),
                    color=alt.Color(field="Party", type="nominal", scale=alt.Scale(domain=party_chart["Party"].tolist(), range=party_colors), legend=alt.Legend(title="Party")),
                    tooltip=[alt.Tooltip("Party:N"), alt.Tooltip("Voters:Q", format=","), alt.Tooltip("Percent:Q", format=".1f", title="Percent")],
                ).properties(height=265)
                st.altair_chart(chart, width="stretch")
                st.markdown(make_summary_table(party_chart, "Party", "Voters", party_colors), unsafe_allow_html=True)
            else:
                st.caption("No party data available.")
            st.markdown('</div>', unsafe_allow_html=True)
        with chart_col2:
            st.markdown('<div class="chart-card">', unsafe_allow_html=True)
            st.markdown('<div class="small-header">Gender Breakdown</div>', unsafe_allow_html=True)
            gender_chart = pd.DataFrame({"Gender": ["Male", "Female", "Unknown"], "Voters": [male, female, unknown_gender]})
            gender_chart["Voters"] = pd.to_numeric(gender_chart["Voters"], errors="coerce").fillna(0)
            gender_chart["Percent"] = gender_chart["Voters"].apply(lambda x: pct_val(x))
            gender_colors = ["#4b4f54", "#b98088", "#9b9da1"]
            if gender_chart["Voters"].sum() > 0:
                chart = alt.Chart(gender_chart).mark_arc(innerRadius=62, outerRadius=98).encode(
                    theta=alt.Theta(field="Voters", type="quantitative"),
                    color=alt.Color(field="Gender", type="nominal", scale=alt.Scale(domain=gender_chart["Gender"].tolist(), range=gender_colors), legend=alt.Legend(title="Gender")),
                    tooltip=[alt.Tooltip("Gender:N"), alt.Tooltip("Voters:Q", format=","), alt.Tooltip("Percent:Q", format=".1f", title="Percent")],
                ).properties(height=265)
                st.altair_chart(chart, width="stretch")
                st.markdown(make_summary_table(gender_chart, "Gender", "Voters", gender_colors), unsafe_allow_html=True)
            else:
                st.caption("No gender data available.")
            st.markdown('</div>', unsafe_allow_html=True)

    # Prepare breakdown once and display in breakdown tab.
    breakdown_df = profile_df.copy()
    for col in ["Total_Voters", "Dem_Voters", "Rep_Voters", "Other_Voters", "Male_Voters", "Female_Voters", "Unknown_Gender", "New_Registrations", "Mail_Applications", "Mail_Applications_Total", "Mail_Applications_Approved", "Mail_Applications_Declined", "Mail_Ballots_Sent", "Mail_Ballots_Returned", "Mail_Ballots_Outstanding", "Mail_Voters", "Geo_Issue_Rows", "Avg_Age"]:
        if col in breakdown_df.columns:
            breakdown_df[col] = pd.to_numeric(breakdown_df[col], errors="coerce").fillna(0)
        else:
            breakdown_df[col] = 0
    breakdown_df["Mail_Ballots_Returned"] = breakdown_df["Mail_Ballots_Returned"].where(breakdown_df["Mail_Ballots_Returned"] > 0, breakdown_df["Mail_Voters"])

    if area_level == "County":
        with breakdown_tab:
            breakdown_mode = st.radio("Breakdown View", ["By Municipality", "By Precinct"], horizontal=True, key="ai_county_breakdown_mode")
        group_cols = ["County", "Municipality"] if breakdown_mode == "By Municipality" else ["County", "Municipality", "Precinct"]
    elif area_level == "Municipality":
        group_cols = ["County", "Municipality", "Precinct"]
    elif area_level == "Precinct":
        group_cols = ["County", "Municipality", "Precinct"]
    else:
        with breakdown_tab:
            breakdown_mode = st.radio("Breakdown View", ["By County", "By Municipality", "By Precinct"], horizontal=True, key=f"ai_district_breakdown_mode_{area_level}")
        if breakdown_mode == "By County":
            group_cols = [area_level, "County"]
        elif breakdown_mode == "By Municipality":
            group_cols = [area_level, "County", "Municipality"]
        else:
            group_cols = [area_level, "County", "Municipality", "Precinct"]

    display_df = (
        breakdown_df.groupby(group_cols, dropna=False)
        .agg(
            Total_Voters=("Total_Voters", "sum"),
            Dem_Voters=("Dem_Voters", "sum"),
            Rep_Voters=("Rep_Voters", "sum"),
            Other_Voters=("Other_Voters", "sum"),
            New_Registrations=("New_Registrations", "sum"),
            Mail_Applications=("Mail_Applications", "sum"),
            Mail_Applications_Total=("Mail_Applications_Total", "sum"),
            Mail_Applications_Approved=("Mail_Applications_Approved", "sum"),
            Mail_Applications_Declined=("Mail_Applications_Declined", "sum"),
            Mail_Ballots_Sent=("Mail_Ballots_Sent", "sum"),
            Mail_Ballots_Returned=("Mail_Ballots_Returned", "sum"),
            Mail_Ballots_Outstanding=("Mail_Ballots_Outstanding", "sum"),
            Geo_Issue_Rows=("Geo_Issue_Rows", "sum"),
        )
        .reset_index()
    )
    weighted_age = (
        breakdown_df.assign(_AgeWeight=breakdown_df["Avg_Age"] * breakdown_df["Total_Voters"])
        .groupby(group_cols, dropna=False)
        .agg(_AgeWeight=("_AgeWeight", "sum"), _AgeTotal=("Total_Voters", "sum"))
        .reset_index()
    )
    weighted_age["Avg_Age"] = weighted_age.apply(lambda r: 0 if r["_AgeTotal"] <= 0 else round(float(r["_AgeWeight"] / r["_AgeTotal"]), 1), axis=1)
    display_df = display_df.merge(weighted_age[group_cols + ["Avg_Age"]], on=group_cols, how="left")
    for col in ["Total_Voters", "Dem_Voters", "Rep_Voters", "Other_Voters", "New_Registrations", "Mail_Applications", "Mail_Applications_Total", "Mail_Applications_Approved", "Mail_Applications_Declined", "Mail_Ballots_Sent", "Mail_Ballots_Returned", "Mail_Ballots_Outstanding", "Geo_Issue_Rows"]:
        display_df[col] = pd.to_numeric(display_df[col], errors="coerce").fillna(0).astype(int)
    display_df["Dem_%"] = display_df.apply(lambda r: 0 if r["Total_Voters"] <= 0 else round((r["Dem_Voters"] / r["Total_Voters"]) * 100, 1), axis=1)
    display_df["Rep_%"] = display_df.apply(lambda r: 0 if r["Total_Voters"] <= 0 else round((r["Rep_Voters"] / r["Total_Voters"]) * 100, 1), axis=1)
    display_df["Other_%"] = display_df.apply(lambda r: 0 if r["Total_Voters"] <= 0 else round((r["Other_Voters"] / r["Total_Voters"]) * 100, 1), axis=1)
    display_df["Mail_Return_%"] = display_df.apply(lambda r: 0 if r["Mail_Applications_Approved"] <= 0 else round((r["Mail_Ballots_Returned"] / r["Mail_Applications_Approved"]) * 100, 1), axis=1)
    display_df["Outstanding_%"] = display_df.apply(lambda r: 0 if r["Mail_Applications_Approved"] <= 0 else round((r["Mail_Ballots_Outstanding"] / r["Mail_Applications_Approved"]) * 100, 1), axis=1)
    display_df = display_df.sort_values("Total_Voters", ascending=False).reset_index(drop=True)

    with breakdown_tab:
        st.markdown('<div class="section-card"><div class="small-header">Area Breakdown</div><div class="tiny-muted">Summarized areas included in this profile.</div></div>', unsafe_allow_html=True)
        _ai_render_table(display_df, height=420, sticky_cols=["USC", "STS", "STH", "School District", "County", "Municipality", "Precinct"], key="breakdown")

    with debug_tab:
        st.caption("Raw precinct_summary.csv source rows for troubleshooting.")
        _ai_render_table(profile_df, height=420, sticky_cols=["USC", "STS", "STH", "School District", "County", "Municipality", "Precinct"], key="debug")


if "data_loaded" not in st.session_state:
    st.session_state.data_loaded = False
if "filters_applied" not in st.session_state:
    st.session_state.filters_applied = False

# Locked startup: load only tiny R2 speed metadata on app startup.
# Never call ensure_index_shards() or prepare_db() here on Streamlit Cloud.
# Full index/detail shards are loaded later only when lookup/export workflows need them.
if not st.session_state.data_loaded:
    with st.spinner("Opening Candidate Connect data..."):
        _manifest = ensure_speed_tables()
        try:
            st.session_state.data_source_label = _manifest.get("source", "R2 speed tables") if isinstance(_manifest, dict) else "R2 speed tables"
        except Exception:
            st.session_state.data_source_label = "R2 speed tables"
        st.session_state.columns = (
            (_manifest.get("index") or {}).get("columns")
            or (_manifest.get("schema") or {}).get("index_columns")
            or [
                "County", "Municipality", "Precinct", "USC", "STS", "STH",
                "School District", "School Region", "Party", "CalculatedParty", "HH-Party", "Gender",
                "Age", "Age_Calc", "Age_Range", "V4A", "V4G", "V4P",
                "MIB_Applied", "MIB_BALLOT", "MB_PERM", "MB_Prob_Score", "Tags",
                "Email", "Landline", "Mobile", "MailBallotNewRegistrant",
            ]
            if isinstance(_manifest, dict) else [
                "County", "Municipality", "Precinct", "USC", "STS", "STH",
                "School District", "School Region", "Party", "Gender", "Age_Range",
                "V4A", "V4G", "V4P", "MIB_Applied", "MIB_BALLOT", "MB_PERM",
                "MB_Prob_Score", "Tags", "Email", "Landline", "Mobile", "MailBallotNewRegistrant",
            ]
        )
        st.session_state.options = get_basic_options(st.session_state.columns)
        st.session_state.data_loaded = True
        st.session_state.filters_applied = False
if "active_filters" not in st.session_state:
    st.session_state.active_filters = {}
if "columns" not in st.session_state:
    st.session_state.columns = []
if "options" not in st.session_state:
    st.session_state.options = {}


def _mb_clean_options(values, field=None):
    """Clean Mail Ballot Center dropdown values so raw/current-file junk does not leak into UI."""
    field_key = str(field or "").strip().upper()
    bad_common = {"", "(BLANK)", "NAN", "NONE", "NULL", "<NA>"}
    # These are source artifacts or redundant/raw statuses that confused the MB Center dropdowns.
    bad_ballot_detail = {"1", "0", "TRUE", "FALSE", "VOTE RECORDED", "VOTED", "V", "PENDING", "CANCELLED", "CANCELED"}
    bad_application = {"1", "0", "TRUE", "FALSE", "VOTE RECORDED"}
    cleaned = []
    seen = set()
    for v in values or []:
        s = str(v).strip()
        u = s.upper().strip()
        if u in bad_common:
            continue
        if field_key in {"MIB_BALLOT", "BALLOT"} and u in bad_ballot_detail:
            continue
        if field_key in {"MIB_APPLIED", "APPLICATION"} and u in bad_application:
            continue
        if field_key in {"MB_PERM", "PERMANENT", "PERMANENT_MB"}:
            if u in {"TRUE", "T", "YES", "Y", "1"}:
                s, u = "Y", "Y"
            elif u in {"FALSE", "F", "NO", "N", "0"}:
                s, u = "N", "N"
            else:
                # Guardrail: the raw/speed option table can sometimes carry unrelated
                # geo labels here. Permanent MB must only ever be Y/N.
                continue
        if u not in seen:
            cleaned.append(s)
            seen.add(u)
    return cleaned


def _mb_clean_application_options(values):
    opts = _mb_clean_options(values, field="MIB_Applied")
    for required in ["APP", "DNA"]:
        if required not in {str(x).upper() for x in opts}:
            opts.append(required)
    return sorted(opts, key=lambda x: (0 if str(x).upper() in {"APP", "DNA"} else 1, str(x).upper()))


def _mb_clean_ballot_detail_options(values):
    # Ballot Sent/Returned are already handled by dedicated controls. This optional
    # detail dropdown should only show genuinely useful county/cure/problem values.
    return _mb_clean_options(values, field="MIB_BALLOT")


def _mb_first_present(options, preferred):
    options_set = {str(x): str(x) for x in options or []}
    for p in preferred:
        if p in options_set:
            return p
    return "All"


def _mb_apply_preset_to_filters(preset, filters):
    """Apply operational mail-ballot presets without touching geography/base universe filters."""
    out = dict(filters or {})

    # Clear only Mail Ballot Center-controlled fields.
    out = _mb_strip_filters(out)

    if preset == "Chase outstanding ballots":
        out["mib_applied_pick"] = ["APP"]
        out["current_ballot_sent_status"] = "Sent"
        out["current_ballot_returned_status"] = "Not Returned/Unknown"
    elif preset == "Cure / problem ballots":
        out["mib_applied_pick"] = ["APP"]
        out["current_ballot_returned_status"] = "Not Returned/Unknown"
    elif preset == "Non-applicant targeting":
        out["mib_applied_pick"] = ["DNA"]
        out["mb_score_slider"] = (2, 4)
    elif preset == "mail ballot targeting":
        out["mib_applied_pick"] = ["DNA"]
        out["mb_score_slider"] = (2, 4)
    elif preset == "Permanent mail voters":
        out["mb_perm_pick"] = ["Y"]
    return out


def _mb_metric_count(base_filters, columns, **overrides):
    f = dict(base_filters or {})
    for k, v in overrides.items():
        if v in (None, "All", [], ""):
            f.pop(k, None)
        else:
            f[k] = v
    try:
        return int(query_metrics(f, columns).get("voters", 0) or 0)
    except Exception:
        return 0



@st.cache_data(show_spinner=False, ttl=30, max_entries=64)
def _mb_summary_metrics_cached(filter_json: str, columns_tuple: tuple):
    active = json.loads(filter_json or "{}")
    columns = list(columns_tuple or [])
    try:
        con = get_conn()
        where_sql, params = current_filter_clause(active, columns)
        row = con.execute(
            f"""
            SELECT
                count(*) AS total,
                sum(CASE WHEN _MIBApplied = 'APP' THEN 1 ELSE 0 END) AS apps,
                sum(CASE WHEN _MIBApplied = 'DNA' THEN 1 ELSE 0 END) AS non_applicants,
                sum(CASE WHEN _CurrentBallotSentDate IS NOT NULL THEN 1 ELSE 0 END) AS sent,
                sum(CASE WHEN _CurrentBallotReturnedDate IS NOT NULL THEN 1 ELSE 0 END) AS returned,
                sum(CASE WHEN _CurrentBallotSentDate IS NOT NULL AND _CurrentBallotReturnedDate IS NULL THEN 1 ELSE 0 END) AS outstanding,
                sum(CASE WHEN _HasEmail THEN 1 ELSE 0 END) AS emails,
                sum(CASE WHEN _HasMobile THEN 1 ELSE 0 END) AS mobiles,
                sum(CASE WHEN _HasLandline THEN 1 ELSE 0 END) AS landlines,
                sum(CASE WHEN _HasApplicantPhone THEN 1 ELSE 0 END) AS applicant_phones
            FROM voters
            {where_sql}
            """,
            params,
        ).df().iloc[0].to_dict()
        return {k: int(v or 0) for k, v in row.items()}
    except Exception:
        return {
            "total": 0, "apps": 0, "non_applicants": 0, "sent": 0, "returned": 0,
            "outstanding": 0, "emails": 0, "mobiles": 0, "landlines": 0, "applicant_phones": 0,
        }


def _mb_summary_metrics(active_filters, columns):
    try:
        filter_json = json.dumps(active_filters or {}, sort_keys=True, default=str)
    except Exception:
        filter_json = "{}"
    return _mb_summary_metrics_cached(filter_json, tuple(columns or []))


def _mb_party_chart(chart_df: pd.DataFrame):
    if chart_df is None or chart_df.empty:
        return pd.DataFrame(columns=["Party", "Count"])
    out = chart_df.copy()
    out["Party"] = out["Party"].astype(str).str.upper().map({"R": "Republican", "D": "Democrat", "O": "Other"}).fillna("Other")
    out["Count"] = pd.to_numeric(out["Count"], errors="coerce").fillna(0)
    out = out.groupby("Party", as_index=False)["Count"].sum()
    order = {"Republican": 0, "Democrat": 1, "Other": 2}
    out["_order"] = out["Party"].map(order).fillna(9)
    return out.sort_values("_order").drop(columns=["_order"])


def _mb_format_metric(value):
    try:
        return f"{int(value):,}"
    except Exception:
        return "0"


def _mb_status_note(active_filters):
    parts = []
    if active_filters.get("mib_applied_pick"):
        parts.append("Application: " + ", ".join(active_filters.get("mib_applied_pick", [])))
    if active_filters.get("current_ballot_sent_status"):
        parts.append("Sent: " + active_filters.get("current_ballot_sent_status"))
    if active_filters.get("current_ballot_returned_status"):
        parts.append("Returned: " + active_filters.get("current_ballot_returned_status"))
    if active_filters.get("mb_score_slider") is not None:
        lo, hi = active_filters.get("mb_score_slider")
        parts.append(f"MB Score: {lo}-{hi}")
    if active_filters.get("has_applicant_phone"):
        parts.append("Applicant Phone: " + active_filters.get("has_applicant_phone"))
    return " | ".join(parts) if parts else "No Mail Ballot Center filters applied yet."


def _mb_download_bytes_csv(df: pd.DataFrame) -> bytes:
    if df is None:
        df = pd.DataFrame()
    return df.to_csv(index=False).encode("utf-8")


def render_mail_ballot_center_workspace():
    columns = st.session_state.get("columns", []) or []
    opts = st.session_state.get("options", {}) or {}
    main_active = st.session_state.get("active_filters", {}) or {}

    if "mail_ballot_center_filters" not in st.session_state:
        st.session_state.mail_ballot_center_filters = {}
    if "mail_ballot_center_use_main_universe" not in st.session_state:
        st.session_state.mail_ballot_center_use_main_universe = True

    use_main_universe = bool(st.session_state.get("mail_ballot_center_use_main_universe", True))
    mb_only_filters = st.session_state.get("mail_ballot_center_filters", {}) or {}
    base_filters = dict(main_active) if use_main_universe else {}
    active_mb_filters = dict(base_filters)
    active_mb_filters.update(mb_only_filters)

    st.markdown(
        '<div class="section-card"><div class="small-header">Mail Ballot Center</div>'
        '<div class="tiny-muted">Strategic mail ballot operations, targeting, and follow-up workspace.</div></div>',
        unsafe_allow_html=True,
    )

    mb_summary = _mb_summary_metrics(active_mb_filters, columns)
    total = mb_summary.get("total", 0)
    apps = mb_summary.get("apps", 0)
    sent = mb_summary.get("sent", 0)
    returned = mb_summary.get("returned", 0)
    outstanding = mb_summary.get("outstanding", 0)
    non_applicants = mb_summary.get("non_applicants", 0)

    metric_cols = st.columns(6)
    metric_data = [
        ("Current MB Universe", total),
        ("Applications", apps),
        ("Ballots Sent", sent),
        ("Ballots Returned", returned),
        ("Outstanding", outstanding),
        ("Did Not Apply", non_applicants),
    ]
    for col, (label, value) in zip(metric_cols, metric_data):
        with col:
            st.markdown(
                f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{_mb_format_metric(value)}</div></div>',
                unsafe_allow_html=True,
            )

    contact_metrics = mb_summary
    contact_cols = st.columns(3)
    for col, label, key in [
        (contact_cols[0], "With Email", "emails"),
        (contact_cols[1], "With Mobile", "mobiles"),
        (contact_cols[2], "With Landline", "landlines"),
    ]:
        with col:
            st.markdown(
                f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{_mb_format_metric(contact_metrics.get(key, 0))}</div></div>',
                unsafe_allow_html=True,
            )

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="small-header">Build Mail Ballot Universe</div>', unsafe_allow_html=True)
    st.caption("Use this to create chase, cure, non-applicant, and MB-probability universes without changing the main dashboard until you choose to export or save.")

    app_options = _mb_clean_application_options(opts.get("mib_applied_vals") or opts.get("MIB_Applied") or (speed_option_values("MIB_Applied") if "speed_option_values" in globals() else []))
    ballot_options = _mb_clean_ballot_detail_options(opts.get("mib_ballot_vals") or opts.get("MIB_BALLOT") or (speed_option_values("MIB_BALLOT") if "speed_option_values" in globals() else []))
    perm_options = _mb_clean_options(opts.get("mb_perm_vals") or opts.get("MB_PERM") or (speed_option_values("MB_PERM") if "speed_option_values" in globals() else []), field="MB_PERM")

    with st.form("mail_ballot_center_form", clear_on_submit=False):
        top_cols = st.columns([1.2, 1, 1], gap="medium")
        with top_cols[0]:
            preset = st.selectbox(
                "Operational Preset",
                ["Custom", "Chase outstanding ballots", "Cure / problem ballots", "Non-applicant targeting", "mail ballot targeting", "Permanent mail voters"],
                index=0,
                help="Preset fills the core mail-ballot filters. You can still adjust the controls below before applying.",
            )
        with top_cols[1]:
            use_main = st.checkbox(
                "Start from current main universe",
                value=use_main_universe,
                help="When checked, the Mail Ballot Center respects the geography/party/voter filters already applied in Create Universe.",
            )
        with top_cols[2]:
            st.markdown("<br>", unsafe_allow_html=True)
            clear_mb = st.checkbox("Clear MB filters", value=False)

        working_defaults = _mb_apply_preset_to_filters(preset, mb_only_filters) if preset != "Custom" else dict(mb_only_filters)

        row1 = st.columns(4, gap="small")
        with row1[0]:
            app_pick = st.multiselect(
                "Application Status",
                app_options,
                default=sanitize_multiselect_defaults(working_defaults.get("mib_applied_pick", []), app_options),
                help="APP = applied/approved. DNA = did not apply.",
            )
        with row1[1]:
            sent_status = st.selectbox(
                "Ballot Sent",
                ["All", "Sent", "Not Sent/Unknown"],
                index=["All", "Sent", "Not Sent/Unknown"].index(working_defaults.get("current_ballot_sent_status", "All") if working_defaults.get("current_ballot_sent_status", "All") in ["All", "Sent", "Not Sent/Unknown"] else "All"),
            )
        with row1[2]:
            returned_status = st.selectbox(
                "Ballot Returned",
                ["All", "Returned", "Not Returned/Unknown"],
                index=["All", "Returned", "Not Returned/Unknown"].index(working_defaults.get("current_ballot_returned_status", "All") if working_defaults.get("current_ballot_returned_status", "All") in ["All", "Returned", "Not Returned/Unknown"] else "All"),
            )
        with row1[3]:
            perm_pick = st.multiselect(
                "Permanent MB",
                perm_options,
                default=sanitize_multiselect_defaults(working_defaults.get("mb_perm_pick", []), perm_options),
            )

        row2 = st.columns(5, gap="small")
        with row2[0]:
            mb_score_default = working_defaults.get("mb_score_slider", (0, 4))
            try:
                mb_score_default = (int(mb_score_default[0]), int(mb_score_default[1]))
            except Exception:
                mb_score_default = (0, 4)
            mb_score = st.slider("MB Probability Score", 0, 4, mb_score_default, 1)
        with row2[1]:
            applicant_phone = st.selectbox(
                "Applicant Phone",
                ["All", "Has Applicant Phone", "No Applicant Phone"],
                index=["All", "Has Applicant Phone", "No Applicant Phone"].index(working_defaults.get("has_applicant_phone", "All") if working_defaults.get("has_applicant_phone", "All") in ["All", "Has Applicant Phone", "No Applicant Phone"] else "All"),
            )
        with row2[2]:
            mobile_status = st.selectbox(
                "Mobile",
                ["All", "Has Mobile", "No Mobile"],
                index=["All", "Has Mobile", "No Mobile"].index(working_defaults.get("has_mobile", "All") if working_defaults.get("has_mobile", "All") in ["All", "Has Mobile", "No Mobile"] else "All"),
            )
        with row2[3]:
            landline_status = st.selectbox(
                "Landline",
                ["All", "Has Landline", "No Landline"],
                index=["All", "Has Landline", "No Landline"].index(working_defaults.get("has_landline", "All") if working_defaults.get("has_landline", "All") in ["All", "Has Landline", "No Landline"] else "All"),
            )
        with row2[4]:
            email_status = st.selectbox(
                "Email",
                ["All", "Has Email", "No Email"],
                index=["All", "Has Email", "No Email"].index(working_defaults.get("has_email", "All") if working_defaults.get("has_email", "All") in ["All", "Has Email", "No Email"] else "All"),
            )

        if ballot_options:
            ballot_pick = st.multiselect(
                "Ballot Status / Cure Status Detail",
                ballot_options,
                default=sanitize_multiselect_defaults(working_defaults.get("mib_ballot_pick", []), ballot_options),
                help="Use this for county-specific ballot status or cure/problem status values coming from CURRENT.",
            )
        else:
            ballot_pick = []

        submit_cols = st.columns(3, gap="small")
        with submit_cols[0]:
            apply_mb = st.form_submit_button("Apply Mail Ballot Filters", width="stretch")
        with submit_cols[1]:
            save_to_main = st.form_submit_button("Send to Main Universe", width="stretch")
        with submit_cols[2]:
            reset_mb = st.form_submit_button("Reset Mail Ballot Center", width="stretch")

    if apply_mb or save_to_main:
        if clear_mb:
            # Clear the center and any MB filters previously pushed into Create Universe.
            _mb_clear_center_state(clear_main=True)
            st.session_state["mb_run_analysis"] = False
            st.success("Mail Ballot filters were cleared.")
            st.stop()

        new_mb_filters = {}
        if app_pick:
            new_mb_filters["mib_applied_pick"] = app_pick
        if ballot_pick:
            new_mb_filters["mib_ballot_pick"] = ballot_pick
        if perm_pick:
            new_mb_filters["mb_perm_pick"] = perm_pick
        if sent_status != "All":
            new_mb_filters["current_ballot_sent_status"] = sent_status
        if returned_status != "All":
            new_mb_filters["current_ballot_returned_status"] = returned_status
        if applicant_phone != "All":
            new_mb_filters["has_applicant_phone"] = applicant_phone
        if mobile_status != "All":
            new_mb_filters["has_mobile"] = mobile_status
        if landline_status != "All":
            new_mb_filters["has_landline"] = landline_status
        if email_status != "All":
            new_mb_filters["has_email"] = email_status
        if tuple(mb_score) != (0, 4):
            new_mb_filters["mb_score_slider"] = tuple(mb_score)

        st.session_state.mail_ballot_center_use_main_universe = bool(use_main)
        st.session_state.mail_ballot_center_filters = new_mb_filters
        st.session_state["mb_run_analysis"] = False

        if save_to_main:
            # Important performance fix: do not st.rerun() after this submit.
            # Streamlit already reruns once for the form submit; forcing another rerun
            # could trap the app in a long refresh loop on large MB universes.
            merged = dict(main_active if use_main else {})
            merged.update(new_mb_filters)
            st.session_state.active_filters = merged
            st.session_state.filters_applied = True
            st.session_state.workspace_mode = "universe"
            st.success("Mail Ballot Center filters were sent to the main Universe. The next dashboard view will use the updated universe.")
            st.stop()

        st.success("Mail Ballot Center filters applied.")
        st.stop()

    if reset_mb:
        # Reset returns the center to the full selected universe and removes any
        # MB filters that were previously pushed into Create Universe.
        _mb_clear_center_state(clear_main=True)
        st.session_state["mb_run_analysis"] = False
        st.success("Mail Ballot Center was reset.")
        st.stop()

    st.caption("Active Mail Ballot Center filters: " + _mb_status_note(active_mb_filters))
    st.markdown('</div>', unsafe_allow_html=True)

    analysis_tabs = st.tabs(["Operations", "Exports", "Analysis", "Notes"])

    with analysis_tabs[0]:
        op_cols = st.columns(4, gap="medium")
        chase_count = _mb_metric_count(active_mb_filters, columns, mib_applied_pick=["APP"], current_ballot_sent_status="Sent", current_ballot_returned_status="Not Returned/Unknown")
        nonapp_score_count = _mb_metric_count(active_mb_filters, columns, mib_applied_pick=["DNA"], mb_score_slider=(2, 4))
        phone_chase_count = _mb_metric_count(active_mb_filters, columns, mib_applied_pick=["APP"], current_ballot_sent_status="Sent", current_ballot_returned_status="Not Returned/Unknown", has_applicant_phone="Has Applicant Phone")
        no_contact_count = _mb_metric_count(active_mb_filters, columns, mib_applied_pick=["APP"], current_ballot_returned_status="Not Returned/Unknown", has_mobile="No Mobile", has_email="No Email")
        for col, title, value, note in [
            (op_cols[0], "Chase Universe", chase_count, "Sent but not returned"),
            (op_cols[1], "Growth Universe", nonapp_score_count, "DNA with MB score 2-4"),
            (op_cols[2], "Phone Chase", phone_chase_count, "Outstanding with applicant phone"),
            (op_cols[3], "Needs Contact Append", no_contact_count, "Outstanding with no mobile/email"),
        ]:
            with col:
                st.markdown(f'<div class="metric-card"><div class="metric-label">{title}</div><div class="metric-value">{_mb_format_metric(value)}</div><div class="tiny-muted">{note}</div></div>', unsafe_allow_html=True)

        st.info("Daily mail ballot operations and follow-up workspace.")

    with analysis_tabs[1]:
        st.caption("Exports use the same stabilized detail-shard engine as the main Output Center, so they remain universe-safe.")
        export_cols = st.columns(3, gap="medium")
        with export_cols[0]:
            if st.button("Prepare MB Filtered CSV", width="stretch"):
                with st.spinner("Building filtered CSV from detail shards..."):
                    st.session_state["mb_filtered_csv_df"] = build_filtered_csv_export(active_mb_filters)
            if isinstance(st.session_state.get("mb_filtered_csv_df"), pd.DataFrame):
                df = st.session_state["mb_filtered_csv_df"]
                st.download_button("Download MB Filtered CSV", data=_mb_download_bytes_csv(df), file_name="mail_ballot_center_filtered.csv", mime="text/csv", width="stretch")
        with export_cols[1]:
            if st.button("Prepare MB Excel Workbook", width="stretch"):
                with st.spinner("Building Excel workbook from detail shards..."):
                    df = build_filtered_csv_export(active_mb_filters)
                    st.session_state["mb_filtered_excel_bytes"] = dataframe_to_export_excel_bytes(df, "Mail Ballot Center")
            if st.session_state.get("mb_filtered_excel_bytes"):
                st.download_button("Download MB Excel Workbook", data=st.session_state["mb_filtered_excel_bytes"], file_name="mail_ballot_center_filtered.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch")
        with export_cols[2]:
            if st.button("Prepare MB Texting CSV", width="stretch"):
                with st.spinner("Building texting CSV from detail shards..."):
                    st.session_state["mb_texting_csv_df"] = build_texting_export(active_mb_filters)
            if isinstance(st.session_state.get("mb_texting_csv_df"), pd.DataFrame):
                df = st.session_state["mb_texting_csv_df"]
                st.download_button("Download MB Texting CSV", data=_mb_download_bytes_csv(df), file_name="mail_ballot_center_texting.csv", mime="text/csv", width="stretch")

        mail_cols = st.columns(3, gap="medium")
        with mail_cols[0]:
            household_mail = st.checkbox("Household mail export", value=True, key="mb_household_mail_export")
        with mail_cols[1]:
            if st.button("Prepare USPS Mail CSV", width="stretch"):
                with st.spinner("Building USPS mail export..."):
                    st.session_state["mb_mail_csv_df"] = build_mail_export(active_mb_filters, householded=household_mail)
            if isinstance(st.session_state.get("mb_mail_csv_df"), pd.DataFrame):
                df = st.session_state["mb_mail_csv_df"]
                st.download_button("Download USPS Mail CSV", data=_mb_download_bytes_csv(df), file_name="mail_ballot_center_usps_mail.csv", mime="text/csv", width="stretch")
        with mail_cols[2]:
            if st.button("Prepare Mailing Labels PDF", width="stretch"):
                with st.spinner("Building mailing labels PDF..."):
                    st.session_state["mb_labels_pdf_bytes"] = generate_mailing_labels_pdf_bytes(active_mb_filters, householded=household_mail)
            if st.session_state.get("mb_labels_pdf_bytes"):
                st.download_button("Download Mailing Labels PDF", data=st.session_state["mb_labels_pdf_bytes"], file_name="mail_ballot_center_labels.pdf", mime="application/pdf", width="stretch")

    with analysis_tabs[2]:
        st.markdown('<div class="section-card"><div class="small-header">Mail Ballot Analysis</div><div class="tiny-muted">Detailed analysis is intentionally run on demand so normal filter changes stay fast.</div></div>', unsafe_allow_html=True)
        if st.button("Run / Refresh Mail Ballot Analysis", width="stretch", key="run_mb_analysis_now"):
            st.session_state["mb_run_analysis"] = True

        if not st.session_state.get("mb_run_analysis", False):
            st.info("Click Run / Refresh Mail Ballot Analysis when you want the deeper charts and county table. This keeps day-to-day filter changes from re-running heavier analysis queries every time.")
        else:
            try:
                contactable = int(mb_summary.get("emails", 0) or 0)
                applicant_phone_count = int(mb_summary.get("applicant_phones", 0) or 0)
                high_prob_nonapp = _mb_metric_count(active_mb_filters, columns, mib_applied_pick=["DNA"], mb_score_slider=(3, 4))
                return_rate = 0 if apps == 0 else round((returned / apps) * 100, 1)
                outstanding_rate = 0 if sent == 0 else round((outstanding / sent) * 100, 1)

                insight_cols = st.columns(5, gap="small")
                for col, title, value, note in [
                    (insight_cols[0], "Return Rate", f"{return_rate}%", "Returned / applications"),
                    (insight_cols[1], "Outstanding Rate", f"{outstanding_rate}%", "Outstanding / sent"),
                    (insight_cols[2], "With Email", _mb_format_metric(contactable), "Email contact available"),
                    (insight_cols[3], "Applicant Phone", _mb_format_metric(applicant_phone_count), "Phone from CURRENT"),
                    (insight_cols[4], "High-Prob DNA", _mb_format_metric(high_prob_nonapp), "DNA with score 3-4"),
                ]:
                    with col:
                        st.markdown(f'<div class="metric-card"><div class="metric-label">{title}</div><div class="metric-value">{value}</div><div class="tiny-muted">{note}</div></div>', unsafe_allow_html=True)

                return_chart = pd.DataFrame([
                    {"Status": "Returned", "Voters": returned},
                    {"Status": "Outstanding", "Voters": outstanding},
                    {"Status": "Not Sent / Unknown", "Voters": max(apps - sent, 0)},
                ])
                contact_chart = pd.DataFrame([
                    {"Contact Type": "Email", "Voters": int(mb_summary.get("emails", 0) or 0)},
                    {"Contact Type": "Mobile", "Voters": int(mb_summary.get("mobiles", 0) or 0)},
                    {"Contact Type": "Landline", "Voters": int(mb_summary.get("landlines", 0) or 0)},
                    {"Contact Type": "Applicant Phone", "Voters": applicant_phone_count},
                ])
                party_chart = _mb_party_chart(query_chart(active_mb_filters, columns, "_PartyNorm", "Party"))
                score_chart = query_chart(active_mb_filters, columns, "_MBScore", "MB Score", not_blank=False)

                chart_cols = st.columns(2, gap="medium")
                with chart_cols[0]:
                    st.markdown('<div class="chart-card"><div class="small-header">Mail Ballot Return Status</div>', unsafe_allow_html=True)
                    if not return_chart.empty:
                        st.altair_chart(alt.Chart(return_chart).mark_arc(innerRadius=45).encode(theta="Voters:Q", color="Status:N", tooltip=["Status:N", "Voters:Q"]), width="stretch")
                    st.markdown('</div>', unsafe_allow_html=True)
                with chart_cols[1]:
                    st.markdown('<div class="chart-card"><div class="small-header">Contact Coverage</div>', unsafe_allow_html=True)
                    st.altair_chart(alt.Chart(contact_chart).mark_bar().encode(x="Contact Type:N", y="Voters:Q", tooltip=["Contact Type:N", "Voters:Q"]), width="stretch")
                    st.markdown('</div>', unsafe_allow_html=True)

                chart_cols2 = st.columns(2, gap="medium")
                with chart_cols2[0]:
                    st.markdown('<div class="chart-card"><div class="small-header">Party Mix</div>', unsafe_allow_html=True)
                    if party_chart is not None and not party_chart.empty:
                        st.altair_chart(
                            alt.Chart(party_chart).mark_bar().encode(
                                x=alt.X("Party:N", sort=["Republican", "Democrat", "Other"]),
                                y="Count:Q",
                                color=alt.Color(
                                    "Party:N",
                                    scale=alt.Scale(
                                        domain=["Republican", "Democrat", "Other"],
                                        range=[PARTY_NAME_COLOR_MAP["Republican"], PARTY_NAME_COLOR_MAP["Democrat"], PARTY_NAME_COLOR_MAP["Other"]],
                                    ),
                                    legend=None,
                                ),
                                tooltip=["Party:N", "Count:Q"],
                            ),
                            width="stretch",
                        )
                    st.markdown('</div>', unsafe_allow_html=True)
                with chart_cols2[1]:
                    st.markdown('<div class="chart-card"><div class="small-header">MB Probability Score</div>', unsafe_allow_html=True)
                    if score_chart is not None and not score_chart.empty:
                        score_chart["MB Score"] = score_chart["MB Score"].astype(str)
                        st.altair_chart(alt.Chart(score_chart).mark_bar().encode(x="MB Score:N", y="Count:Q", tooltip=["MB Score:N", "Count:Q"]), width="stretch")
                    st.markdown('</div>', unsafe_allow_html=True)

                st.markdown('<div class="table-card"><div class="small-header">Top Counties in Current MB Universe</div>', unsafe_allow_html=True)
                county_df = query_area_summary(active_mb_filters, columns, "County") if "County" in columns else pd.DataFrame()
                if county_df is not None and not county_df.empty:
                    st.dataframe(county_df.head(25), width="stretch", hide_index=True, height=360)
                else:
                    st.caption("County summary is not available for this selection.")
                st.markdown('</div>', unsafe_allow_html=True)
            except Exception as exc:
                st.warning("Mail Ballot Center analysis could not load, but filters and exports are still available.")
                st.caption(str(exc))

    with analysis_tabs[3]:
        st.markdown("""
        **What is included in this first operational patch**

        - Dedicated Mail Ballot Center workspace
        - Chase, cure, non-applicant, MB probability, and permanent MB presets
        - mail ballot fields: application status, ballot sent, ballot returned, applicant phone
        - MB probability score slider
        - Shared CSV, Excel, texting, USPS mail, and label export engine
        - Safe option to send Mail Ballot Center filters back into the main Universe

        **Still queued next**: county mailing-start tracking, daily mail ballot deltas, deeper cure classification, and saved MB report packs.
        """)


if "saved_universes" not in st.session_state:
    st.session_state.saved_universes = load_saved_universes()
if "street_results_df" not in st.session_state:
    st.session_state.street_results_df = pd.DataFrame(columns=["PA ID Number", "F", "A", "U", "NH", "Yard Sign", "Notes"])
if "street_results_filters" not in st.session_state:
    st.session_state.street_results_filters = {}
if "walk_results_df" not in st.session_state:
    st.session_state.walk_results_df = pd.DataFrame(columns=["PA ID Number", "Contacted", "Result", "Support Level", "Follow-Up", "Notes"])
if "walk_results_filters" not in st.session_state:
    st.session_state.walk_results_filters = {}
if "lookup_view_active" not in st.session_state:
    st.session_state.lookup_view_active = False
if "workspace_mode" not in st.session_state:
    st.session_state.workspace_mode = "landing"
if "lookup_query_input" not in st.session_state:
    st.session_state.lookup_query_input = st.session_state.get("lookup_query", "")


# Clean first-load sidebar state: show navigation only until user selects a workspace.
if "cc_initial_workspace_set" not in st.session_state:
    st.session_state.cc_initial_workspace_set = True
    st.session_state.workspace_mode = "landing"
    st.session_state.lookup_view_active = False

with st.sidebar:
    if APP_ENV == "DEV":
        st.header("Candidate Connect")
        st.markdown(
            "<div style='background:#374151;color:#f9fafb;border:1px solid #4b5563;border-radius:10px;padding:10px 12px;font-weight:800;margin:8px 0 12px 0;'>Internal DEV Workspace</div>",
            unsafe_allow_html=True
        )
    else:
        st.header("Candidate Connect")

    if not st.session_state.data_loaded:
        st.info("Opening Candidate Connect...")
        st.stop()
    else:
        cols = st.session_state.columns
        opts = st.session_state.options

        st.markdown("<div class='cc-nav-menu-title'>Navigation</div>", unsafe_allow_html=True)
        if st.button("▦  Create Universe", key="nav_create_universe_single", width="stretch"):
            st.session_state.workspace_mode = "universe"
            st.session_state.lookup_view_active = False
            st.rerun()
        if st.button("⌕  Voter Lookup", key="nav_voter_lookup_single", width="stretch"):
            st.session_state.workspace_mode = "lookup"
            st.session_state.lookup_view_active = False
            st.rerun()
        if st.button("✉  Mail Ballot Center", key="nav_mail_ballot_single", width="stretch"):
            st.session_state.workspace_mode = "mail_ballot_center"
            st.session_state.lookup_view_active = False
            st.rerun()
        if st.button("⌂  Area Intelligence", key="nav_area_intel_single", width="stretch"):
            st.session_state.workspace_mode = "area_intelligence"
            st.session_state.lookup_view_active = False
            st.rerun()

        if st.session_state.get("workspace_mode", "landing") == "landing":
            st.caption("Select a workspace above to begin.")

        if st.session_state.get("workspace_mode", "landing") == "universe":
            st.markdown("<div class='cc-active-section-title'>Create Universe</div>", unsafe_allow_html=True)
            geo_selections = render_interdependent_geo_filters(cols, opts)

            with st.form("filter_form", clear_on_submit=False):
                with st.expander("Voter Details", expanded=False):
                    party_options = [v for v in ["D", "R", "O"] if v in (opts.get("party_vals", []) or [])]
                    for v in opts.get("party_vals", []) or []:
                        if v not in party_options:
                            party_options.append(v)
                    party_pick = st.multiselect(
                        "Party",
                        party_options,
                        default=sanitize_multiselect_defaults(st.session_state.active_filters.get("party_pick", []), party_options),
                        help="D = Democratic, R = Republican, O = every other party/blank."
                    )

                    hh_party_options = opts.get("hh_party_vals", []) or []
                    if hh_party_options:
                        hh_party_pick = st.multiselect(
                            "Household Party",
                            hh_party_options,
                            default=sanitize_multiselect_defaults(st.session_state.active_filters.get("hh_party_pick", []), hh_party_options),
                        )
                    else:
                        hh_party_pick = []

                    calc_party_options = opts.get("calc_party_vals", []) or []
                    if "CalculatedParty" in cols:
                        calc_party_pick = st.multiselect(
                            "Calculated Party",
                            calc_party_options,
                            default=sanitize_multiselect_defaults(st.session_state.active_filters.get("calc_party_pick", []), calc_party_options),
                            help="Modeled/calculated party category from the pipeline.",
                        )
                    else:
                        calc_party_pick = []

                    gender_options = opts.get("gender_vals", []) or []
                    gender_pick = st.multiselect(
                        "Gender",
                        gender_options,
                        default=sanitize_multiselect_defaults(st.session_state.active_filters.get("gender_pick", []), gender_options),
                    )

                    age_range_options = opts.get("age_range_vals", []) or []
                    age_range_pick = st.multiselect(
                        "Age Range",
                        age_range_options,
                        default=sanitize_multiselect_defaults(st.session_state.active_filters.get("age_range_pick", []), age_range_options),
                    )

                    age_slider = None
                    if opts.get("age_min") is not None and opts.get("age_max") is not None:
                        age_default = sanitize_slider_range(
                            st.session_state.active_filters.get("age_slider", (opts["age_min"], opts["age_max"])),
                            opts["age_min"],
                            opts["age_max"],
                            int,
                        )
                        age_slider = st.slider("Age", opts["age_min"], opts["age_max"], age_default)

                with st.expander("Vote History", expanded=False):
                    vh_type_options = ["All", "General", "Primary"]
                    current_vh_type = st.session_state.active_filters.get("vote_history_type", "All")
                    if current_vh_type not in vh_type_options:
                        current_vh_type = "All"
                    vote_history_type = st.selectbox(
                        "Vote History Type",
                        vh_type_options,
                        index=vh_type_options.index(current_vh_type),
                        help="All uses V4A, General uses V4G, and Primary uses V4P.",
                    )
                    current_range = st.session_state.active_filters.get("vote_history_range", (0, 4))
                    if not isinstance(current_range, (list, tuple)) or len(current_range) != 2:
                        current_range = (0, 4)
                    vote_history_range = st.slider(
                        "Vote History Range",
                        min_value=0,
                        max_value=4,
                        value=(int(current_range[0]), int(current_range[1])),
                        help="0-4 elections in the selected vote history field.",
                    )

                    new_reg_months = st.slider(
                        "Newly Registered (within last N months; 0 = all)",
                        min_value=0,
                        max_value=24,
                        value=max(0, min(24, int(st.session_state.active_filters.get("new_reg_months", 0) or 0))),
                        step=1,
                    )


                    st.markdown("**Election & Vote Method**")
                    election_columns_found, election_year_options, election_type_options = election_filter_options(cols)

                    election_years_pick = st.multiselect(
                        "Election Years",
                        election_year_options,
                        default=sanitize_multiselect_defaults(st.session_state.active_filters.get("election_years_pick", []), election_year_options),
                        help="Select election years available in the voter history data.",
                    )

                    election_types_available = [t for t in ["General", "Primary"] if t in election_type_options]
                    election_types_pick = st.multiselect(
                        "Election Type",
                        election_types_available,
                        default=sanitize_multiselect_defaults(st.session_state.active_filters.get("election_types_pick", []), election_types_available),
                        help="Choose General, Primary, or leave blank for all available election types.",
                    )

                    vote_method_options = ["AP", "MB", "P", "DNV"]
                    vote_methods_pick = st.multiselect(
                        "Vote Method",
                        vote_method_options,
                        default=sanitize_multiselect_defaults(st.session_state.active_filters.get("vote_methods_pick", []), vote_method_options),
                        help="AP = at poll, MB = mail ballot, P = provisional, DNV = did not vote.",
                    )

                    if not election_columns_found:
                        st.caption("Use this to match voters by election year, election type, and how they voted.")

                    # Mail ballot filters live in their own section below.
                    if "election_years_pick" not in locals():
                        election_years_pick = []
                    if "election_types_pick" not in locals():
                        election_types_pick = []
                    if "vote_methods_pick" not in locals():
                        vote_methods_pick = []
                    mib_applied_pick = []
                    mib_ballot_pick = []
                    mb_perm_pick = []
                    source_file_pick = []
                    mb_new_reg_pick = []
                    current_ballot_sent_status = "All"
                    current_ballot_returned_status = "All"
                    mb_score_slider = None

                with st.expander("Mail Ballots", expanded=False):
                    mib_applied_pick = st.multiselect(
                        "Mail Ballot Application Status",
                        opts.get("mib_applied_vals", []),
                        default=sanitize_multiselect_defaults(st.session_state.active_filters.get("mib_applied_pick", []), opts.get("mib_applied_vals", [])),
                        help="APP = applied/approved, DEC = declined/rejected, DNA = did not apply when present in the source data.",
                    )
                    # Mail Ballot Vote Status removed; Ballot Returned covers this use case.
                    mib_ballot_pick = []
                    mb_perm_pick = st.multiselect(
                        "Permanent Mail Ballot",
                        opts.get("mb_perm_vals", []),
                        default=sanitize_multiselect_defaults(st.session_state.active_filters.get("mb_perm_pick", []), opts.get("mb_perm_vals", [])),
                    )

                    source_file_pick = []
                    # Mail Ballot New Registrant removed from the main Universe builder.
                    # It will return inside the dedicated Mail Ballot Center after the
                    # CURRENT.txt overlay is isolated and verified at the pipeline level.
                    mb_new_reg_pick = []
                    current_ballot_sent_status = st.selectbox(
                        "Ballot Sent",
                        ["All", "Sent", "Not Sent/Unknown"],
                        index=["All", "Sent", "Not Sent/Unknown"].index(sanitize_selectbox_value(st.session_state.active_filters.get("current_ballot_sent_status", "All"), ["All", "Sent", "Not Sent/Unknown"], "All")),
                    )
                    current_ballot_returned_status = st.selectbox(
                        "Ballot Returned",
                        ["All", "Returned", "Not Returned/Unknown"],
                        index=["All", "Returned", "Not Returned/Unknown"].index(sanitize_selectbox_value(st.session_state.active_filters.get("current_ballot_returned_status", "All"), ["All", "Returned", "Not Returned/Unknown"], "All")),
                    )

                    if opts.get("mb_score_min") is not None and opts.get("mb_score_max") is not None:
                        lo = float(opts["mb_score_min"])
                        hi = float(opts["mb_score_max"])
                        default_score = sanitize_slider_range(
                            st.session_state.active_filters.get("mb_score_slider", (lo, hi)),
                            lo,
                            hi,
                            float,
                        )
                        mb_score_slider = st.slider("MB Probability Score", min_value=lo, max_value=hi, value=default_score)

                with st.expander("Contact Filters", expanded=False):
                    tag_pick = st.multiselect(
                        "Tags",
                        opts.get("tag_vals", []),
                        default=sanitize_multiselect_defaults(st.session_state.active_filters.get("tag_pick", []), opts.get("tag_vals", [])),
                        help="Matches comma-separated values in the Tags column."
                    ) if "Tags" in cols else []
                    email_opts = ["All", "Has Email", "No Email"]
                    landline_opts = ["All", "Has Landline", "No Landline"]
                    mobile_opts = ["All", "Has Mobile", "No Mobile"]
                    has_email = st.selectbox("Email", email_opts, index=email_opts.index(sanitize_selectbox_value(st.session_state.active_filters.get("has_email", "All"), email_opts, "All")))
                    has_landline = st.selectbox("Landline", landline_opts, index=landline_opts.index(sanitize_selectbox_value(st.session_state.active_filters.get("has_landline", "All"), landline_opts, "All")))
                    has_mobile = st.selectbox("Mobile", mobile_opts, index=mobile_opts.index(sanitize_selectbox_value(st.session_state.active_filters.get("has_mobile", "All"), mobile_opts, "All")))
                    applicant_phone_opts = ["All", "Has Applicant Phone", "No Applicant Phone"]
                    has_applicant_phone = st.selectbox("Mail Ballot Applicant Phone", applicant_phone_opts, index=applicant_phone_opts.index(sanitize_selectbox_value(st.session_state.active_filters.get("has_applicant_phone", "All"), applicant_phone_opts, "All")))
                    applicant_phone_type_pick = []
                    applicant_phone_compliance_pick = []

                with st.expander("Smart Follow-Up", expanded=False):
                    contact_status_opts = ["All", "Not Contacted", "Contacted"]
                    global_yes_no_opts = ["All", "Yes", "No"]
                    support_level_opts = get_global_support_level_options()

                    contact_status = st.selectbox(
                        "Contact Status",
                        contact_status_opts,
                        index=contact_status_opts.index(sanitize_selectbox_value(st.session_state.active_filters.get("contact_status", "All"), contact_status_opts, "All")),
                        help="Uses uploaded candidate Street List and Walk Sheet results.",
                    )
                    global_nh = st.selectbox(
                        "Not Home",
                        global_yes_no_opts,
                        index=global_yes_no_opts.index(sanitize_selectbox_value(st.session_state.active_filters.get("global_nh", "All"), global_yes_no_opts, "All")),
                    )
                    global_follow_up = st.selectbox(
                        "Follow-Up",
                        global_yes_no_opts,
                        index=global_yes_no_opts.index(sanitize_selectbox_value(st.session_state.active_filters.get("global_follow_up", "All"), global_yes_no_opts, "All")),
                    )
                    current_support = st.session_state.active_filters.get("global_support_level", "All")
                    if current_support not in support_level_opts:
                        current_support = "All"
                    global_support_level = st.selectbox(
                        "Support Level",
                        support_level_opts,
                        index=support_level_opts.index(current_support),
                    )
                    st.caption("Use these filters to organize follow-up and field activity.")

                st.caption("")
                cols2 = st.columns(2)
                apply_filters = cols2[0].form_submit_button("Apply Filters", width="stretch", type="primary")
                clear_filters = cols2[1].form_submit_button("Clear Filters", width="stretch")

            if clear_filters:
                st.session_state.active_filters = {}
                for _geo_col in GEO_FILTER_COLUMNS:
                    _geo_key = f"geo_dep_{_norm_col_name(_geo_col)}"
                    if _geo_key in st.session_state:
                        st.session_state[_geo_key] = []
                st.session_state.filters_applied = False
                st.session_state.workspace_mode = "landing"
                st.session_state.lookup_view_active = False
                st.rerun()

            if apply_filters:
                st.session_state.workspace_mode = "universe"
                st.session_state.lookup_view_active = False
                st.session_state.active_filters = {
                    **geo_selections,
                    "party_pick": party_pick,
                    "hh_party_pick": hh_party_pick,
                    "calc_party_pick": calc_party_pick,
                    "tag_pick": tag_pick,
                    "gender_pick": gender_pick,
                    "age_range_pick": age_range_pick,
                    "age_slider": age_slider,
                    "vote_history_type": vote_history_type,
                    "vote_history_range": vote_history_range,
                    "mib_applied_pick": mib_applied_pick,
                    "mib_ballot_pick": mib_ballot_pick,
                    "mb_perm_pick": mb_perm_pick,
                    "source_file_pick": [],
                    "mb_new_reg_pick": mb_new_reg_pick,
                    "current_ballot_sent_status": current_ballot_sent_status,
                    "current_ballot_returned_status": current_ballot_returned_status,
                    "mb_score_slider": mb_score_slider,
                    "new_reg_months": new_reg_months,
                    "election_years_pick": election_years_pick,
                    "election_types_pick": election_types_pick,
                    "vote_methods_pick": vote_methods_pick,
                    "has_email": has_email,
                    "has_landline": has_landline,
                    "has_mobile": has_mobile,
                    "has_applicant_phone": has_applicant_phone,
                    "applicant_phone_type_pick": [],
                    "applicant_phone_compliance_pick": [],
                    "contact_status": contact_status,
                    "global_nh": global_nh,
                    "global_follow_up": global_follow_up,
                    "global_support_level": global_support_level,
                }
                st.session_state.filters_applied = True
                st.rerun()
            divider()
            with st.expander("⚡ Quick Select Campaign Lists", expanded=False):
                st.caption("These buttons keep your existing geography and voter filters, but quickly set the Smart Follow-Up filters.")
                qs_row1 = st.columns(2, gap="small")
                with qs_row1[0]:
                    if st.button("Re-Knock List", width="stretch", key="qs_reknock"):
                        apply_followup_preset("Re-Knock List")
                with qs_row1[1]:
                    if st.button("Follow-Up List", width="stretch", key="qs_followup"):
                        apply_followup_preset("Follow-Up List")

                qs_row2 = st.columns(2, gap="small")
                with qs_row2[0]:
                    if st.button("GOTV Supporters", width="stretch", key="qs_gotv"):
                        apply_followup_preset("GOTV Supporters")
                with qs_row2[1]:
                    if st.button("Undecided Persuasion", width="stretch", key="qs_undecided"):
                        apply_followup_preset("Undecided Persuasion")

                qs_row3 = st.columns(2, gap="small")
                with qs_row3[0]:
                    if st.button("Yard Sign Follow-Up", width="stretch", key="qs_yardsign"):
                        apply_followup_preset("Yard Sign Follow-Up")
                with qs_row3[1]:
                    if st.button("Clear Quick Select", width="stretch", key="qs_clear"):
                        apply_followup_preset("Clear")

            with st.expander("💾 Saved Universes", expanded=False):
                store_label = get_saved_universe_store_label()
                st.caption("Saved universes are available for quick reuse.")

                saved_universes = load_saved_universes()
                st.session_state["saved_universes"] = saved_universes
                universe_names = list(saved_universes.keys())

                if universe_names:
                    selected_sidebar_universe = st.selectbox(
                        "Saved Universes",
                        universe_names,
                        key="sidebar_saved_universe_name",
                    )
                    universe_info = saved_universes[selected_sidebar_universe]
                    st.caption(
                        f"Saved: {universe_info.get('saved_at', '')} | Count: {int(universe_info.get('count', 0)):,}"
                    )
                    st.caption(universe_info.get("summary", "No filters"))
                    load_col, delete_col = st.columns(2, gap="small")
                    with load_col:
                        if st.button("Load Universe", width="stretch", key="load_sidebar_universe"):
                            loaded_filters = universe_info.get("filters", {}) or {}
                            st.session_state.active_filters = loaded_filters
                            for _geo_col in GEO_FILTER_COLUMNS:
                                _geo_key = f"geo_dep_{_norm_col_name(_geo_col)}"
                                st.session_state[_geo_key] = loaded_filters.get(_geo_col, []) or []
                            st.session_state.filters_applied = False
                            st.session_state.workspace_mode = "universe"
                            st.session_state.lookup_view_active = False
                            st.success(f"Loaded universe: {selected_sidebar_universe}")
                            st.rerun()
                    with delete_col:
                        if st.button("Delete Universe", width="stretch", key="delete_sidebar_universe"):
                            saved_universes.pop(selected_sidebar_universe, None)
                            save_saved_universes(saved_universes)
                            st.session_state["saved_universes"] = saved_universes
                            st.success(f"Deleted universe: {selected_sidebar_universe}")
                            st.rerun()
                else:
                    st.caption("No saved universes yet.")

                save_name = st.text_input(
                    "Save current filters as",
                    key="save_universe_name_sidebar",
                    placeholder="Example: GOTV Democrats Week 1",
                )
                if st.button("Save Current Universe", width="stretch", key="save_sidebar_universe"):
                    universe_name = save_name.strip()
                    if universe_name:
                        current_filters = st.session_state.get("active_filters", {})
                        saved_universes = load_saved_universes()
                        saved_universes[universe_name] = {
                            "filters": current_filters,
                            "saved_at": datetime.now().strftime("%Y-%m-%d %I:%M %p"),
                            "count": int(query_metrics(current_filters, st.session_state.get("columns", [])).get("voters", 0)),
                            "summary": summarize_universe_filters(current_filters),
                        }
                        save_saved_universes(saved_universes)
                        st.session_state["saved_universes"] = saved_universes
                        st.success(f"Saved universe: {universe_name}")
                        st.rerun()
                    else:
                        st.warning("Enter a universe name first.")

        if st.session_state.get("workspace_mode") == "lookup":
            render_lookup_sidebar(st.session_state.active_filters, cols)

        if st.session_state.get("workspace_mode") == "mail_ballot_center":
            st.markdown("<div class='cc-active-section-title'>Mail Ballot Center</div>", unsafe_allow_html=True)
            st.caption("Operations, analysis, follow-up, and reporting for the daily CURRENT.txt mail ballot overlay.")

        if st.session_state.get("workspace_mode") == "area_intelligence":
            st.markdown("<div class='cc-active-section-title'>Area Intelligence</div>", unsafe_allow_html=True)
            st.caption("Precinct profile, summary metrics, and strategy foundation.")

if not st.session_state.data_loaded:
    st.markdown('<div class="section-card empty-shell"><div class="small-header">Ready to load</div><div class="tiny-muted">Click <strong>Load Voter Data</strong> in the sidebar to open the R2 index shards with DuckDB.</div></div>', unsafe_allow_html=True)
    st.stop()

active = st.session_state.active_filters
columns = st.session_state.columns
workspace_mode = st.session_state.get("workspace_mode", "landing")

if workspace_mode == "lookup":
    if st.session_state.get("lookup_view_active", False):
        try:
            render_voter_lookup_results()
        except Exception as e:
            st.error("Voter Lookup hit a data-field error instead of crashing the dashboard.")
            st.info("The rest of Candidate Connect is still available.")
    else:
        render_lookup_empty_workspace()
elif workspace_mode == "mail_ballot_center":
    render_mail_ballot_center_workspace()
elif workspace_mode == "area_intelligence":
    render_area_intelligence_workspace()
else:
    if not st.session_state.filters_applied:
        st.markdown("<div style='font-size:28px;font-weight:900;letter-spacing:.08em;color:#ffffff;margin:0 0 18px 6px;text-transform:uppercase;'>Voters Statewide</div>", unsafe_allow_html=True)
        render_opening_dashboard_preview()
        st.stop()

    with st.spinner("Loading dashboard from speed tables..."):
        metrics = query_metrics(active, columns)
        large_filter_mode = use_large_filter_mode(active, columns)
        # Speed-table dashboard rule: do not scan detail shards during Apply Filters.
        # Contact-tracking aggregates will be added as their own lightweight table later.
        followup_stats = {
            "contacted_pct": 0, "nh_pct": 0, "followup_pct": 0, "undecided_pct": 0,
            "contacted_count": 0, "nh_count": 0, "followup_count": 0, "undecided_count": 0,
            "strong_pct": 0, "strong_count": 0, "large_mode": True, "speed_placeholder": True,
        }

        if large_filter_mode:
            party_df = pd.DataFrame(columns=["Party", "Count"])
            gender_df = pd.DataFrame(columns=["Gender", "Count"])
            age_df = pd.DataFrame(columns=["Age Range", "Count"])
            area_choices = []
        else:
            party_df = query_chart(active, columns, "_PartyNorm", "Party")
            gender_df = query_chart(active, columns, "_Gender", "Gender")
            age_df = query_chart(active, columns, "_AgeRange", "Age Range")
            area_choices = [c for c in ["County", "Municipality", "Precinct", "USC", "STS", "STH", "School District", "School Region"] if c in columns]

    metric_cols = st.columns(5, gap="small")
    metric_values = [
        ("Voters", f"{safe_int(metrics.get('voters')):,}"),
        ("Households", "—" if metrics.get("households") is None else f"{safe_int(metrics.get('households')):,}"),
        ("Emails", f"{safe_int(metrics.get('emails')):,}"),
        ("Mobiles", f"{safe_int(metrics.get('mobiles')):,}"),
        ("Unique Precincts", f"{safe_int(metrics.get('unique_precincts')):,}"),
    ]
    for col, (label, value) in zip(metric_cols, metric_values):
        with col:
            st.markdown(f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{value}</div></div>', unsafe_allow_html=True)
    if metrics.get("speed_mode"):
        st.caption("")
    campaign_cols = st.columns(4, gap="small")
    campaign_values = [
        ("Contacted", f"{safe_int(followup_stats.get('contacted_pct'))}%", f"{safe_int(followup_stats.get('contacted_count')):,} voters"),
        ("Not Home", f"{safe_int(followup_stats.get('nh_pct'))}%", f"{safe_int(followup_stats.get('nh_count')):,} voters"),
        ("Follow-Up", f"{safe_int(followup_stats.get('followup_pct'))}%", f"{safe_int(followup_stats.get('followup_count')):,} voters"),
        ("Undecided", f"{safe_int(followup_stats.get('undecided_pct'))}%", f"{safe_int(followup_stats.get('undecided_count')):,} voters"),
    ]
    for col, (label, value, subvalue) in zip(campaign_cols, campaign_values):
        with col:
            st.markdown(
                f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{value}</div><div class="tiny-muted">{subvalue}</div></div>',
                unsafe_allow_html=True
            )

    divider()

    if large_filter_mode:
        st.warning("Large universe detected. Summary mode is active only for very large selections to keep the app stable. Full districts under the safety limit now stay interactive.")

    dashboard_tabs = st.tabs(["Overview", "Contact Tracking", "Output Center"])

    with dashboard_tabs[0]:
        if large_filter_mode:
            st.info("Summary view is active for this very large universe. Narrow by geography or voter filters for deeper charts and grouped tables.")
            summary_only_df = pd.DataFrame([
                {"Metric": "Voters", "Value": f"{safe_int(metrics.get('voters')):,}"},
                {"Metric": "Households", "Value": "—" if metrics.get("households") is None else f"{safe_int(metrics.get('households')):,}"},
                {"Metric": "Emails", "Value": f"{safe_int(metrics.get('emails')):,}"},
                {"Metric": "Mobiles", "Value": f"{safe_int(metrics.get('mobiles')):,}"},
                {"Metric": "Unique Precincts", "Value": f"{safe_int(metrics.get('unique_precincts')):,}"},
            ])
            st.dataframe(summary_only_df, width="stretch", hide_index=True)
        else:
            chart_cols = st.columns(3, gap="medium")
            with chart_cols[0]:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                pie_chart_with_table(party_df, "Party", "Count", "Party Breakdown", "party")
                st.markdown('</div>', unsafe_allow_html=True)
            with chart_cols[1]:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                pie_chart_with_table(gender_df, "Gender", "Count", "Gender Breakdown", "gender")
                st.markdown('</div>', unsafe_allow_html=True)
            with chart_cols[2]:
                st.markdown('<div class="chart-card">', unsafe_allow_html=True)
                pie_chart_with_table(age_df, "Age Range", "Count", "Age Range Breakdown", "age")
                st.markdown('</div>', unsafe_allow_html=True)

            divider()

            st.markdown('<div class="table-card">', unsafe_allow_html=True)
            st.markdown('<div class="small-header">Counts by Area</div>', unsafe_allow_html=True)
            if area_choices:
                selected_area = st.selectbox("Area", area_choices, format_func=geo_label, label_visibility="collapsed", key="overview_area_group")
                area_df = query_area_summary(active, columns, selected_area).copy()
                area_df["Individuals"] = pd.to_numeric(area_df["Individuals"], errors="coerce").fillna(0).map(lambda x: f"{x:,.0f}")
                if "Households" in area_df.columns and not area_df["Households"].astype(str).eq("—").all():
                    area_df["Households"] = pd.to_numeric(area_df["Households"], errors="coerce").fillna(0).map(lambda x: f"{x:,.0f}")
                st.dataframe(area_df, width="stretch", hide_index=True)
            else:
                st.caption("No area fields available.")
            st.markdown('</div>', unsafe_allow_html=True)

    with dashboard_tabs[1]:
        if large_filter_mode:
            st.info("Summary-only mode is active for this very large universe. Narrow by geography or voter filters to load Contact Tracking details.")
        else:
            tracking_cols = st.columns(2, gap="medium")
            with tracking_cols[0]:
                st.markdown('<div class="table-card">', unsafe_allow_html=True)
                st.markdown('<div class="small-header">Contact Tracking</div>', unsafe_allow_html=True)
                tracking_summary_df = pd.DataFrame([
                    {"Metric": "Contacted", "Percent": f"{safe_int(followup_stats.get('contacted_pct'))}%", "Voters": f"{safe_int(followup_stats.get('contacted_count')):,}"},
                    {"Metric": "Not Home", "Percent": f"{safe_int(followup_stats.get('nh_pct'))}%", "Voters": f"{safe_int(followup_stats.get('nh_count')):,}"},
                    {"Metric": "Follow-Up", "Percent": f"{safe_int(followup_stats.get('followup_pct'))}%", "Voters": f"{safe_int(followup_stats.get('followup_count')):,}"},
                ])
                st.dataframe(tracking_summary_df, width="stretch", hide_index=True)
                st.markdown('</div>', unsafe_allow_html=True)
            with tracking_cols[1]:
                st.markdown('<div class="table-card">', unsafe_allow_html=True)
                st.markdown('<div class="small-header">Support Snapshot</div>', unsafe_allow_html=True)
                support_summary_df = pd.DataFrame([
                    {"Metric": "Strong Support", "Percent": f"{safe_int(followup_stats.get('strong_pct'))}%", "Voters": f"{safe_int(followup_stats.get('strong_count')):,}"},
                    {"Metric": "Undecided", "Percent": f"{safe_int(followup_stats.get('undecided_pct'))}%", "Voters": f"{safe_int(followup_stats.get('undecided_count')):,}"},
                ])
                st.dataframe(support_summary_df, width="stretch", hide_index=True)
                st.markdown('</div>', unsafe_allow_html=True)

    with dashboard_tabs[2]:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="small-header">Output Center</div>', unsafe_allow_html=True)

        if large_filter_mode:
            st.warning("Very large universe detected. Full detail exports are still protected here to keep the app stable. District-level universes under the safety limit now use the normal Output Center.")
            if st.button("Prepare Large Universe Summary Report", width="stretch"):
                with st.spinner("Building statewide summary report..."):
                    st.session_state["statewide_summary_report_bytes"] = build_statewide_summary_report_bytes(active, columns)
            if "statewide_summary_report_bytes" in st.session_state and st.session_state["statewide_summary_report_bytes"]:
                st.download_button(
                    "Download Large Universe Summary Report",
                    data=st.session_state["statewide_summary_report_bytes"],
                    file_name="candidate_connect_large_universe_summary_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width="stretch",
                )
            st.caption("This workbook includes Overview, Filters, County, Congressional, State Senate, and State House sheets for the current universe. Full Senate/House/Congressional district exports under 1,000,000 voters now use the normal Output Center.")
        else:
            output_tabs = st.tabs(["Exports", "Reports"])
            with output_tabs[0]:
                st.markdown('<div class="small-header">Exports</div>', unsafe_allow_html=True)
                st.caption("CSV files are only built when you click the button for that export type.")
        
                mail_mode = st.radio(
                    "Mailing Mode",
                    ["Not Householded", "Householded"],
                    horizontal=True,
                    key="mail_mode_radio",
                )
        
                exp_cols = st.columns(3, gap="medium")
        
                with exp_cols[0]:
                    if st.button("Prepare Filtered CSV", width="stretch"):
                        with st.spinner("Building filtered CSV from detail shards..."):
                            export_df = build_filtered_csv_export(active)
                            st.session_state["filtered_export_df"] = export_df
                            st.session_state.pop("filtered_export_xlsx", None)
                    if "filtered_export_df" in st.session_state:
                        st.download_button(
                            "Download Filtered CSV",
                            data=dataframe_to_csv_bytes(st.session_state["filtered_export_df"]),
                            file_name="candidate_connect_filtered.csv",
                            mime="text/csv",
                            width="stretch",
                            on_click=clear_prepared_download_state,
                            args=("filtered_export_df",),
                        )
        
                with exp_cols[1]:
                    if st.button("Prepare Texting CSV", width="stretch"):
                        with st.spinner("Building texting CSV from detail shards..."):
                            export_df = build_texting_export(active)
                            st.session_state["texting_export_df"] = export_df
                            st.session_state.pop("texting_export_xlsx", None)
                    if "texting_export_df" in st.session_state:
                        st.download_button(
                            "Download Texting CSV",
                            data=dataframe_to_csv_bytes(st.session_state["texting_export_df"]),
                            file_name="candidate_connect_texting.csv",
                            mime="text/csv",
                            width="stretch",
                            on_click=clear_prepared_download_state,
                            args=("texting_export_df",),
                        )
        
                with exp_cols[2]:
                    if st.button("Prepare Mail CSV", width="stretch"):
                        with st.spinner("Building mail CSV from selected detail fields..."):
                            export_df = build_mail_export(active, householded=(mail_mode == "Householded"))
                            st.session_state["mail_export_df"] = export_df
                            st.session_state["mail_export_mode"] = mail_mode
                            st.session_state.pop("mail_export_xlsx", None)
                    if "mail_export_df" in st.session_state:
                        suffix = "householded" if st.session_state.get("mail_export_mode") == "Householded" else "individual"
                        st.download_button(
                            "Download Mail CSV",
                            data=dataframe_to_csv_bytes(st.session_state["mail_export_df"]),
                            file_name=f"candidate_connect_mail_{suffix}.csv",
                            mime="text/csv",
                            width="stretch",
                            on_click=clear_prepared_download_state,
                            args=("mail_export_df",),
                        )

                st.divider()
                st.caption("Excel versions of the CSV exports are built only after you click Prepare, so they do not slow the dashboard or counts panel. Each workbook includes Sheet 1: Area Counts and Sheet 2: Data.")
                excel_cols = st.columns(3, gap="medium")
                with excel_cols[0]:
                    if st.button("Prepare Filtered Excel", width="stretch"):
                        with st.spinner("Building filtered Excel workbook..."):
                            export_df = st.session_state.get("filtered_export_df")
                            if export_df is None:
                                export_df = build_filtered_csv_export(active)
                                st.session_state["filtered_export_df"] = export_df
                            st.session_state["filtered_export_xlsx"] = dataframe_to_export_excel_bytes(export_df, "Filtered")
                    if st.session_state.get("filtered_export_xlsx"):
                        st.download_button(
                            "Download Filtered Excel",
                            data=st.session_state["filtered_export_xlsx"],
                            file_name="candidate_connect_filtered.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            width="stretch",
                            on_click=clear_prepared_download_state,
                            args=("filtered_export_xlsx",),
                        )
                with excel_cols[1]:
                    if st.button("Prepare Texting Excel", width="stretch"):
                        with st.spinner("Building texting Excel workbook..."):
                            export_df = st.session_state.get("texting_export_df")
                            if export_df is None:
                                export_df = build_texting_export(active)
                                st.session_state["texting_export_df"] = export_df
                            st.session_state["texting_export_xlsx"] = dataframe_to_export_excel_bytes(export_df, "Texting")
                    if st.session_state.get("texting_export_xlsx"):
                        st.download_button(
                            "Download Texting Excel",
                            data=st.session_state["texting_export_xlsx"],
                            file_name="candidate_connect_texting.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            width="stretch",
                            on_click=clear_prepared_download_state,
                            args=("texting_export_xlsx",),
                        )
                with excel_cols[2]:
                    if st.button("Prepare Mail Excel", width="stretch"):
                        with st.spinner("Building mail Excel workbook..."):
                            export_df = st.session_state.get("mail_export_df")
                            if export_df is None or st.session_state.get("mail_export_mode") != mail_mode:
                                export_df = build_mail_export(active, householded=(mail_mode == "Householded"))
                                st.session_state["mail_export_df"] = export_df
                                st.session_state["mail_export_mode"] = mail_mode
                            st.session_state["mail_export_xlsx"] = dataframe_to_export_excel_bytes(export_df, "Mail")
                            st.session_state["mail_export_xlsx_mode"] = mail_mode
                    if st.session_state.get("mail_export_xlsx"):
                        suffix = "householded" if st.session_state.get("mail_export_xlsx_mode") == "Householded" else "individual"
                        st.download_button(
                            "Download Mail Excel",
                            data=st.session_state["mail_export_xlsx"],
                            file_name=f"candidate_connect_mail_{suffix}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            width="stretch",
                            on_click=clear_prepared_download_state,
                            args=("mail_export_xlsx",),
                        )
        
            with output_tabs[1]:
                st.markdown('<div class="small-header">Reports</div>', unsafe_allow_html=True)
                st.caption("Prepare PDFs only when needed to keep the app responsive.")
        
                report_sections = st.tabs(["Summary", "Street List", "Walk Sheet", "Mailing Labels"])
        
                with report_sections[0]:
                    st.caption("Builds a clean PDF summary of the current filtered universe with overview counts, selected filters, and party/gender/age breakdowns.")
                    summary_cols = st.columns(2, gap="medium")
                    with summary_cols[0]:
                        if st.button("Prepare Summary Report PDF", width="stretch"):
                            with st.spinner("Building Summary Report PDF from current filtered universe..."):
                                pdf_bytes = generate_summary_report_pdf_bytes(active, cols)
                                st.session_state["summary_report_pdf_bytes"] = pdf_bytes
                    with summary_cols[1]:
                        if "summary_report_pdf_bytes" in st.session_state and st.session_state["summary_report_pdf_bytes"]:
                            st.download_button(
                                "Download Summary Report PDF",
                                data=st.session_state["summary_report_pdf_bytes"],
                                file_name="candidate_connect_summary_report.pdf",
                                mime="application/pdf",
                                width="stretch",
                            )
        
                with report_sections[1]:
                    st.caption("Builds a compact precinct-grouped PDF and also supports a Street List Excel tracking sheet so the same list can be used to record F, A, U, NH, and Yard Sign results.")
                    upload_cols = st.columns([1, 1.2, 1], gap="medium")
                    with upload_cols[0]:
                        st.download_button(
                            "Download Street Results CSV Template",
                            data=get_street_results_template_csv_bytes(),
                            file_name="candidate_connect_street_results_template.csv",
                            mime="text/csv",
                            width="stretch",
                            on_click="ignore",
                        )
                        if st.button("Prepare Street List Excel Tracking Sheet", width="stretch"):
                            with st.spinner("Building Street List Excel tracking sheet from filtered detail shards..."):
                                st.session_state["street_results_sheet_bytes"] = get_street_results_sheet_bytes(active)
                        if st.session_state.get("street_results_sheet_bytes"):
                            st.download_button(
                                "Download Street List Excel Tracking Sheet",
                                data=st.session_state["street_results_sheet_bytes"],
                                file_name="candidate_connect_street_list_tracking.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                width="stretch",
                                on_click=clear_prepared_download_state,
                                args=("street_results_sheet_bytes",),
                            )
                    with upload_cols[1]:
                        uploaded_results_file = st.file_uploader(
                            "Upload Street Results File",
                            type=["csv", "xlsx"],
                            key="street_results_upload",
                            help="Upload either the Street List Excel tracking sheet or a CSV using PA ID Number plus F, A, U, NH, and Yard Sign columns.",
                        )
                        if uploaded_results_file is not None:
                            upload_sig = f"{uploaded_results_file.name}:{getattr(uploaded_results_file, 'size', 0)}"
                            if st.session_state.get("street_results_upload_sig") != upload_sig:
                                try:
                                    if str(uploaded_results_file.name).lower().endswith(".xlsx"):
                                        raw_upload_df = pd.read_excel(uploaded_results_file, dtype=str).fillna("")
                                        normalized_cols = [re.sub(r"[^a-z0-9]+", "", str(c).strip().lower()) for c in raw_upload_df.columns]
                                        if "paidnumber" not in normalized_cols:
                                            try:
                                                raw_upload_df = pd.read_excel(uploaded_results_file, dtype=str, header=4).fillna("")
                                            except Exception:
                                                uploaded_results_file.seek(0)
                                                raw_upload_df = pd.read_excel(uploaded_results_file, dtype=str).fillna("")
                                        uploaded_results_file.seek(0)
                                    else:
                                        raw_upload_df = pd.read_csv(uploaded_results_file, dtype=str).fillna("")
                                    standardized_upload_df = standardize_uploaded_street_results(raw_upload_df)
                                    if standardized_upload_df.empty:
                                        st.warning("No usable PA ID Number column was found in the uploaded file.")
                                    else:
                                        st.session_state["street_results_df"] = standardized_upload_df
                                        st.session_state["street_results_upload_sig"] = upload_sig
                                        st.session_state["street_results_upload_name"] = uploaded_results_file.name
                                        st.success(f"Loaded {len(standardized_upload_df):,} street-result rows.")
                                except Exception as exc:
                                    st.error(f"Could not read the street results file: {exc}")
                    with upload_cols[2]:
                        loaded_results = st.session_state.get("street_results_df")
                        if isinstance(loaded_results, pd.DataFrame) and not loaded_results.empty:
                            st.caption(f"Loaded rows: {len(loaded_results):,}")
                            st.caption(f"Source: {st.session_state.get('street_results_upload_name', 'uploaded CSV')}")
                            if st.button("Clear Uploaded Street Results", width="stretch"):
                                st.session_state["street_results_df"] = pd.DataFrame(columns=["PA ID Number", "F", "A", "U", "NH", "Yard Sign", "Notes"])
                                st.session_state["street_results_filters"] = {}
                                st.session_state.pop("street_results_upload_sig", None)
                                st.session_state.pop("street_results_upload_name", None)
                                st.rerun()
                        else:
                            st.caption("No street results uploaded yet.")
        
                    loaded_results = st.session_state.get("street_results_df")
                    if isinstance(loaded_results, pd.DataFrame) and not loaded_results.empty:
                        st.caption("These tracking filters only affect the Street List outputs, so you can reprint or re-export candidate follow-up lists without changing the dashboard counts.")
                        filter_defaults = st.session_state.get("street_results_filters", {}) or {}
                        street_filter_cols = st.columns(5, gap="small")
                        street_results_filters = {}
                        for col, field in zip(street_filter_cols, ["F", "A", "U", "NH", "Yard Sign"]):
                            with col:
                                street_results_filters[field] = st.selectbox(
                                    field,
                                    ["All", "Marked", "Unmarked"],
                                    index=["All", "Marked", "Unmarked"].index(filter_defaults.get(field, "All")),
                                    key=f"street_results_filter_{field}",
                                )
                        st.session_state["street_results_filters"] = street_results_filters
                    else:
                        st.caption("Download the Street List Excel tracking sheet if you want a ready-to-use file with F, A, U, NH, Yard Sign, and Notes columns, then upload it back after results are entered.")
        
                    pdf_cols = st.columns(2, gap="medium")
                    with pdf_cols[0]:
                        if st.button("Prepare Street List PDF", width="stretch"):
                            with st.spinner("Building Street List PDF from filtered detail shards..."):
                                pdf_bytes = generate_street_list_pdf_bytes(active)
                                st.session_state["street_pdf_bytes"] = pdf_bytes
                    with pdf_cols[1]:
                        if "street_pdf_bytes" in st.session_state and st.session_state["street_pdf_bytes"]:
                            st.download_button(
                                "Download Street List PDF",
                                data=st.session_state["street_pdf_bytes"],
                                file_name="candidate_connect_street_list.pdf",
                                mime="application/pdf",
                                width="stretch",
                                on_click=clear_prepared_download_state,
                                args=("street_pdf_bytes",),
                            )
        
                with report_sections[2]:
                    st.caption("Builds a volunteer-friendly walk sheet and supports a tracking workbook that can be uploaded back by PA ID.")
                    upload_cols = st.columns([1, 1.15, 1], gap="medium")
                    with upload_cols[0]:
                        st.download_button(
                            "Download Walk Sheet Tracking Template",
                            data=get_walk_sheet_tracking_template_csv_bytes(),
                            file_name="candidate_connect_walk_sheet_tracking_template.csv",
                            mime="text/csv",
                            width="stretch",
                            on_click="ignore",
                        )
                        if st.button("Prepare Walk Sheet Excel Tracking Sheet", width="stretch"):
                            with st.spinner("Building Walk Sheet Excel tracking sheet from filtered detail shards..."):
                                excel_bytes = build_walk_sheet_tracking_excel_bytes(active)
                                st.session_state["walk_sheet_tracking_excel_bytes"] = excel_bytes
                        if "walk_sheet_tracking_excel_bytes" in st.session_state and st.session_state["walk_sheet_tracking_excel_bytes"]:
                            st.download_button(
                                "Download Walk Sheet Excel Tracking Sheet",
                                data=st.session_state["walk_sheet_tracking_excel_bytes"],
                                file_name="candidate_connect_walk_sheet_tracking.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                width="stretch",
                                on_click=clear_prepared_download_state,
                                args=("walk_sheet_tracking_excel_bytes",),
                            )
                    with upload_cols[1]:
                        uploaded_walk_file = st.file_uploader(
                            "Upload Walk Sheet Results",
                            type=["csv", "xlsx"],
                            key="walk_results_upload",
                            help="Upload a completed Walk Sheet tracking workbook or CSV using PA ID Number plus Contacted, Result, Support Level, Follow-Up, and Notes columns.",
                        )
                        if uploaded_walk_file is not None:
                            upload_sig = f"{uploaded_walk_file.name}:{getattr(uploaded_walk_file, 'size', 0)}"
                            if st.session_state.get("walk_results_upload_sig") != upload_sig:
                                try:
                                    if str(uploaded_walk_file.name).lower().endswith(".xlsx"):
                                        raw_upload_df = pd.read_excel(uploaded_walk_file, dtype=str).fillna("")
                                        normalized_cols = [re.sub(r"[^a-z0-9]+", "", str(c).strip().lower()) for c in raw_upload_df.columns]
                                        if "paidnumber" not in normalized_cols:
                                            try:
                                                raw_upload_df = pd.read_excel(uploaded_walk_file, dtype=str, header=4).fillna("")
                                            except Exception:
                                                uploaded_walk_file.seek(0)
                                                raw_upload_df = pd.read_excel(uploaded_walk_file, dtype=str).fillna("")
                                        uploaded_walk_file.seek(0)
                                    else:
                                        raw_upload_df = pd.read_csv(uploaded_walk_file, dtype=str).fillna("")
                                    standardized_upload_df = standardize_uploaded_walk_results(raw_upload_df)
                                    if standardized_upload_df.empty:
                                        st.warning("No usable PA ID Number column was found in the uploaded Walk Sheet file.")
                                    else:
                                        st.session_state["walk_results_df"] = standardized_upload_df
                                        st.session_state["walk_results_upload_sig"] = upload_sig
                                        st.session_state["walk_results_upload_name"] = uploaded_walk_file.name
                                        st.success(f"Loaded {len(standardized_upload_df):,} walk-result rows.")
                                except Exception as exc:
                                    st.error(f"Could not read the Walk Sheet results file: {exc}")
                    with upload_cols[2]:
                        loaded_walk_results = st.session_state.get("walk_results_df")
                        if isinstance(loaded_walk_results, pd.DataFrame) and not loaded_walk_results.empty:
                            st.caption(f"Loaded rows: {len(loaded_walk_results):,}")
                            st.caption(f"Source: {st.session_state.get('walk_results_upload_name', 'uploaded file')}")
                            if st.button("Clear Uploaded Walk Sheet Results", width="stretch"):
                                st.session_state["walk_results_df"] = pd.DataFrame(columns=["PA ID Number", "Contacted", "Result", "Support Level", "Follow-Up", "Notes"])
                                st.session_state["walk_results_filters"] = {}
                                st.session_state.pop("walk_results_upload_sig", None)
                                st.session_state.pop("walk_results_upload_name", None)
                                st.rerun()
                        else:
                            st.caption("No Walk Sheet results uploaded yet.")
        
                    loaded_walk_results = st.session_state.get("walk_results_df")
                    if isinstance(loaded_walk_results, pd.DataFrame) and not loaded_walk_results.empty:
                        st.caption("These tracking filters apply only to the Walk Sheet PDF, so you can rebuild volunteer re-knock or follow-up sheets without changing the dashboard counts.")
                        filter_defaults = st.session_state.get("walk_results_filters", {}) or {}
                        walk_filter_cols = st.columns(4, gap="small")
                        with walk_filter_cols[0]:
                            contacted_filter = st.selectbox(
                                "Contacted",
                                ["All", "Marked", "Unmarked"],
                                index=["All", "Marked", "Unmarked"].index(filter_defaults.get("Contacted", "All")),
                                key="walk_results_filter_contacted",
                            )
                        with walk_filter_cols[1]:
                            not_home_filter = st.selectbox(
                                "Not Home",
                                ["All", "Marked", "Unmarked"],
                                index=["All", "Marked", "Unmarked"].index(filter_defaults.get("Not Home", "All")),
                                key="walk_results_filter_not_home",
                            )
                        with walk_filter_cols[2]:
                            followup_filter = st.selectbox(
                                "Follow-Up",
                                ["All", "Marked", "Unmarked"],
                                index=["All", "Marked", "Unmarked"].index(filter_defaults.get("Follow-Up", "All")),
                                key="walk_results_filter_followup",
                            )
                        support_options = ["All"] + sorted(
                            {normalize_export_text(v) for v in loaded_walk_results["Support Level"].tolist() if normalize_export_text(v)}
                        )
                        default_support = filter_defaults.get("Support Level", "All")
                        if default_support not in support_options:
                            default_support = "All"
                        with walk_filter_cols[3]:
                            support_filter = st.selectbox(
                                "Support Level",
                                support_options,
                                index=support_options.index(default_support),
                                key="walk_results_filter_support",
                            )
                        st.session_state["walk_results_filters"] = {
                            "Contacted": contacted_filter,
                            "Not Home": not_home_filter,
                            "Follow-Up": followup_filter,
                            "Support Level": support_filter,
                        }
                    else:
                        st.caption("Download the Walk Sheet Excel tracking sheet if you want a ready-to-use file with Contacted, Result, Support Level, Follow-Up, and Notes columns, then upload it back after results are entered.")
        
                    walk_cols = st.columns(2, gap="medium")
                    with walk_cols[0]:
                        if st.button("Prepare Walk Sheet PDF", width="stretch"):
                            with st.spinner("Building Walk Sheet PDF from filtered detail shards..."):
                                pdf_bytes = generate_walk_sheet_pdf_bytes(active)
                                st.session_state["walk_sheet_pdf_bytes"] = pdf_bytes
                    with walk_cols[1]:
                        if "walk_sheet_pdf_bytes" in st.session_state and st.session_state["walk_sheet_pdf_bytes"]:
                            st.download_button(
                                "Download Walk Sheet PDF",
                                data=st.session_state["walk_sheet_pdf_bytes"],
                                file_name="candidate_connect_walk_sheet.pdf",
                                mime="application/pdf",
                                width="stretch",
                                on_click=clear_prepared_download_state,
                                args=("walk_sheet_pdf_bytes",),
                            )
        
                with report_sections[3]:
                    st.caption("Builds a print-ready Avery 5160-style PDF label sheet from the current mail export universe.")
                    label_mode = st.radio(
                        "Label Mode",
                        ["Householded", "Individual"],
                        horizontal=True,
                        key="mail_labels_mode",
                    )
                    label_cols = st.columns(2, gap="medium")
                    with label_cols[0]:
                        if st.button("Prepare Mailing Labels PDF", width="stretch"):
                            with st.spinner("Building mailing labels PDF from filtered detail shards..."):
                                pdf_bytes = generate_mailing_labels_pdf_bytes(active, householded=(label_mode == "Householded"))
                                st.session_state["mailing_labels_pdf_bytes"] = pdf_bytes
                                st.session_state["mailing_labels_pdf_mode"] = label_mode
                    with label_cols[1]:
                        if "mailing_labels_pdf_bytes" in st.session_state and st.session_state["mailing_labels_pdf_bytes"]:
                            suffix = "householded" if st.session_state.get("mailing_labels_pdf_mode") == "Householded" else "individual"
                            st.download_button(
                                "Download Mailing Labels PDF",
                                data=st.session_state["mailing_labels_pdf_bytes"],
                                file_name=f"candidate_connect_mailing_labels_{suffix}.pdf",
                                mime="application/pdf",
                                width="stretch",
                                on_click=clear_prepared_download_state,
                                args=("mailing_labels_pdf_bytes",),
                            )
        
        

        
            st.markdown('</div>', unsafe_allow_html=True)